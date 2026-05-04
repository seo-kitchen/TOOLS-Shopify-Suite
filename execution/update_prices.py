"""
Serax prijzen updaten vanuit nieuwe prijslijst.

Leest de prijsverhogings-Excel, matcht op SKU tegen Supabase,
past giftbox-logica toe, en genereert:
  1. Database update (prijzen in seo_products)
  2. Hextom bulk update Excel (alleen prijskolommen voor Shopify)
  3. CRX prijsverhogingsoverzicht (oud vs nieuw, met correcte giftbox-prijzen)

Gebruik:
    python execution/update_prices.py --file "Master Files/SERAX_Price Increase April 2026 (EUR1) (2).xlsx" [--dry-run]
"""

import argparse
import os
import sys
from pathlib import Path
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
OUTPUT_DIR = Path(os.getenv("SHOPIFY_OUTPUT_DIR", "./exports/"))


def get_supabase():
    from supabase import create_client
    return create_client(SUPABASE_URL, SUPABASE_KEY)


def load_price_file(file_path: str) -> pd.DataFrame:
    """Laad Serax prijsverhogings-Excel (header op rij 2, pandas header=1)."""
    df = pd.read_excel(file_path, header=1, dtype=str)
    print(f"Prijsbestand geladen: {len(df)} rijen")
    return df


def parse_price(val) -> float | None:
    if pd.isna(val) or str(val).strip() in ("", "nan", "-"):
        return None
    try:
        return float(str(val).replace(",", "."))
    except ValueError:
        return None


def parse_int(val) -> int | None:
    if pd.isna(val) or str(val).strip() in ("", "nan", "-"):
        return None
    try:
        return int(float(str(val)))
    except ValueError:
        return None


def extract_prices(row: pd.Series) -> dict:
    """Extract prijzen uit een rij met giftbox-logica.

    REGEL: als giftbox_qty > 1, gebruik ALTIJD giftbox-prijzen.
    """
    sku = str(row.get("Product code", "")).strip()
    gb_qty = parse_int(row.get("Giftbox Quantity"))

    rrp_stuk = parse_price(row.get("RRP EUR"))
    inkoop_stuk = parse_price(row.get("Purchase Price EUR"))
    rrp_gb = parse_price(row.get("RRP EUR.1"))
    inkoop_gb = parse_price(row.get("Purchase Price EUR.1"))

    is_giftbox_set = gb_qty is not None and gb_qty > 1

    if is_giftbox_set and rrp_gb is not None:
        effectieve_rrp = rrp_gb
        effectieve_inkoop = inkoop_gb
    else:
        effectieve_rrp = rrp_stuk
        effectieve_inkoop = inkoop_stuk

    return {
        "sku": sku,
        "rrp_stuk_eur_nieuw": rrp_stuk,
        "inkoopprijs_stuk_eur_nieuw": inkoop_stuk,
        "rrp_gb_eur_nieuw": rrp_gb,
        "inkoopprijs_gb_eur_nieuw": inkoop_gb,
        "giftbox_qty": gb_qty,
        "is_giftbox_set": is_giftbox_set,
        "effectieve_rrp": effectieve_rrp,
        "effectieve_inkoop": effectieve_inkoop,
        "product_name": str(row.get("Product name / piece", "")).strip(),
        "product_status": str(row.get("Product Status", "")).strip(),
        "collection": str(row.get("Collection", "")).strip(),
    }


def fetch_db_products(sb) -> dict:
    """Haal alle Serax producten op uit Supabase. Geeft {sku: product}."""
    all_products = []
    offset = 0
    batch_size = 1000

    while True:
        result = sb.table("seo_products").select(
            "id, sku, ean_shopify, product_name_raw, giftbox, giftbox_qty, "
            "rrp_stuk_eur, rrp_gb_eur, inkoopprijs_stuk_eur, inkoopprijs_gb_eur, "
            "verkoopprijs, inkoopprijs, status, fase, handle, product_title_nl"
        ).range(offset, offset + batch_size - 1).execute()

        if not result.data:
            break
        all_products.extend(result.data)
        if len(result.data) < batch_size:
            break
        offset += batch_size

    print(f"Database: {len(all_products)} producten opgehaald")
    return {p["sku"]: p for p in all_products if p.get("sku")}


def update_database(sb, updates: list[dict], dry_run: bool = False):
    """Update prijzen in Supabase. Vult rrp, inkoop, verkoopprijs, en gb-velden."""
    if dry_run:
        print(f"\n[DRY RUN] Zou {len(updates)} producten updaten in Supabase")
        vp_missing = sum(1 for u in updates if u["db"].get("verkoopprijs") is None)
        gb_missing = sum(1 for u in updates if u["db"].get("rrp_gb_eur") is None)
        print(f"  Waarvan verkoopprijs nieuw invullen: {vp_missing}")
        print(f"  Waarvan rrp_gb_eur nieuw invullen: {gb_missing}")
        return

    success = 0
    errors = []

    for i in range(0, len(updates), 50):
        batch = updates[i:i + 50]
        for upd in batch:
            try:
                update_data = {
                    "rrp_stuk_eur": upd["effectieve_rrp"],
                    "inkoopprijs_stuk_eur": upd["effectieve_inkoop"],
                    "verkoopprijs": upd["effectieve_rrp"],
                    "inkoopprijs": upd["effectieve_inkoop"],
                }
                if upd.get("rrp_gb_eur_nieuw") is not None:
                    update_data["rrp_gb_eur"] = upd["rrp_gb_eur_nieuw"]
                if upd.get("inkoopprijs_gb_eur_nieuw") is not None:
                    update_data["inkoopprijs_gb_eur"] = upd["inkoopprijs_gb_eur_nieuw"]
                if upd.get("rrp_stuk_eur_nieuw") is not None and upd.get("is_giftbox_set"):
                    update_data["rrp_gb_eur"] = upd["rrp_gb_eur_nieuw"]

                sb.table("seo_products").update(update_data).eq("id", upd["id"]).execute()
                success += 1
            except Exception as e:
                errors.append(f"SKU {upd['sku']}: {e}")

        pct = min(100, int((i + len(batch)) / len(updates) * 100))
        print(f"  Batch {i // 50 + 1}: {len(batch)} updates ({pct}%)")

    print(f"\nDatabase update: {success} succesvol, {len(errors)} fouten")
    if errors:
        for err in errors[:10]:
            print(f"  FOUT: {err}")


def write_hextom_price_update(matched: list[dict], output_path: Path):
    """Genereer Hextom bulk price update Excel.

    Alleen producten met status=ready (online op Shopify).
    Kolommen per Hextom-formaat: SKU voor matching, prijs + inkoop.
    Bij giftbox-sets: prijs = setprijs (NOOIT stuksprijs).
    """
    online = [m for m in matched if m["db"].get("status") == "ready" and m["db"].get("handle")]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hextom Price Update"

    columns = [
        "Variant SKU",
        "Product Handle",
        "Product Title",
        "Variant Barcode",
        "Variant Price",
        "Variant Cost",
        "Giftbox Set",
        "GB Qty",
    ]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    gb_fill = PatternFill("solid", fgColor="DAEEF3")

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, item in enumerate(online, start=2):
        db = item["db"]
        is_gb = item.get("is_giftbox_set", False)
        vals = [
            db.get("sku", ""),
            db.get("handle", ""),
            db.get("product_title_nl", ""),
            db.get("ean_shopify", ""),
            item["effectieve_rrp"],
            item["effectieve_inkoop"],
            "SET" if is_gb else "",
            item.get("giftbox_qty", ""),
        ]
        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val else "")
            if col_idx == 4 and val:
                cell.number_format = "@"
            if is_gb:
                cell.fill = gb_fill

    col_widths = {1: 20, 2: 45, 3: 55, 4: 18, 5: 14, 6: 14, 7: 10, 8: 8}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_path)
    gb_online = sum(1 for m in online if m.get("is_giftbox_set"))
    print(f"Hextom update: {output_path}")
    print(f"  {len(online)} producten online (waarvan {gb_online} giftbox-sets)")
    print(f"  Kolommen: Variant SKU, Handle, Title, Barcode, Price, Cost")
    print(f"  Bij sets: Price = giftbox-prijs, Cost = giftbox-inkoop")


def clean_decimal(val) -> str:
    if val is None:
        return ""
    try:
        f = float(val)
        return f"{f:.10f}".rstrip("0").rstrip(".")
    except (ValueError, TypeError):
        return str(val)


def write_crx_overview(matched: list[dict], not_in_db: list[dict], output_path: Path):
    """Genereer CRX prijsoverzicht: complete FKI-prijzen met giftbox-logica."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Alle prijzen (FKI overzicht) ──
    ws1 = wb.active
    ws1.title = "FKI Prijsoverzicht"

    columns = [
        "SKU", "Productnaam", "Collection",
        "Giftbox Set?", "GB Qty",
        "RRP Stuk EUR", "Inkoop Stuk EUR",
        "RRP Giftbox EUR", "Inkoop Giftbox EUR",
        "Prijs FKI (verkoopprijs)", "Inkoop FKI",
        "Verkoopprijs DB (oud)", "Verschil",
        "Status DB", "Fase",
    ]

    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    gb_fill = PatternFill("solid", fgColor="DAEEF3")
    missing_fill = PatternFill("solid", fgColor="FFF2CC")
    thin_border = Border(bottom=Side(style="thin", color="DDDDDD"))

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws1.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    increases = 0
    decreases = 0
    unchanged = 0
    newly_filled = 0

    for row_idx, item in enumerate(matched, start=2):
        db = item["db"]
        is_gb = item.get("is_giftbox_set", False)

        verkoopprijs_oud = db.get("verkoopprijs")
        rrp_oud_effectief = float(db.get("rrp_stuk_eur") or 0)
        rrp_nieuw = item["effectieve_rrp"] or 0

        rrp_diff = rrp_nieuw - rrp_oud_effectief

        if verkoopprijs_oud is None:
            newly_filled += 1
            verschil_label = "NIEUW"
        elif abs(rrp_diff) > 0.01 and rrp_diff > 0:
            increases += 1
            verschil_label = f"+{rrp_diff:.2f}"
        elif abs(rrp_diff) > 0.01 and rrp_diff < 0:
            decreases += 1
            verschil_label = f"{rrp_diff:.2f}"
        else:
            unchanged += 1
            verschil_label = "="

        vals = [
            item["sku"],
            item.get("product_name", ""),
            item.get("collection", ""),
            "SET" if is_gb else "",
            item.get("giftbox_qty", ""),
            item.get("rrp_stuk_eur_nieuw"),
            item.get("inkoopprijs_stuk_eur_nieuw"),
            item.get("rrp_gb_eur_nieuw"),
            item.get("inkoopprijs_gb_eur_nieuw"),
            item["effectieve_rrp"],
            item["effectieve_inkoop"],
            float(verkoopprijs_oud) if verkoopprijs_oud is not None else None,
            verschil_label,
            db.get("status", ""),
            db.get("fase", ""),
        ]

        row_fill = None
        if is_gb:
            row_fill = gb_fill
        if verkoopprijs_oud is None:
            row_fill = missing_fill

        for col_idx, val in enumerate(vals, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val if val is not None else "")
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill
            if col_idx in (6, 7, 8, 9, 10, 11, 12):
                cell.number_format = '#,##0.00'

    col_widths = {1: 16, 2: 42, 3: 30, 4: 10, 5: 8,
                  6: 14, 7: 14, 8: 14, 9: 14,
                  10: 18, 11: 14, 12: 16, 13: 12, 14: 12, 15: 8}
    for col_idx, width in col_widths.items():
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Sheet 2: Samenvatting ──
    ws2 = wb.create_sheet("Samenvatting")
    summary_data = [
        ["Serax Prijsupdate April 2026 -- Overzicht", ""],
        ["", ""],
        ["Totaal producten in prijslijst", len(matched) + len(not_in_db)],
        ["Waarvan in database (matched)", len(matched)],
        ["Waarvan NIET in database", len(not_in_db)],
        ["", ""],
        ["Verkoopprijs nieuw ingevuld (was leeg)", newly_filled],
        ["Prijsverhogingen (vs. DB)", increases],
        ["Prijsverlagingen (vs. DB)", decreases],
        ["Ongewijzigd", unchanged],
        ["", ""],
        ["Giftbox-sets in update", sum(1 for m in matched if m.get("is_giftbox_set"))],
        ["", ""],
        ["Legenda:", ""],
        ["  Blauw = giftbox-set (prijs = setprijs)", ""],
        ["  Geel = verkoopprijs was leeg, nu ingevuld", ""],
        ["  'Prijs FKI' = de prijs die op Shopify staat", ""],
        ["  Bij sets: FKI = giftbox-prijs, NOOIT stuksprijs", ""],
    ]
    title_font = Font(bold=True, size=14, color="1F4E79")
    label_font = Font(bold=True, size=11)

    for row_idx, (label, value) in enumerate(summary_data, start=1):
        cell_a = ws2.cell(row=row_idx, column=1, value=label)
        cell_b = ws2.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            cell_a.font = title_font
        elif value != "":
            cell_a.font = label_font

    ws2.column_dimensions["A"].width = 55
    ws2.column_dimensions["B"].width = 20

    # ── Sheet 3: Niet in database ──
    if not_in_db:
        ws3 = wb.create_sheet("Niet in database")
        cols3 = ["SKU", "Productnaam", "Collection", "Status Serax",
                 "RRP Stuk", "RRP Giftbox", "GB Qty", "Prijs FKI"]
        for col_idx, col_name in enumerate(cols3, start=1):
            cell = ws3.cell(row=1, column=col_idx, value=col_name)
            cell.fill = PatternFill("solid", fgColor="C0504D")
            cell.font = Font(bold=True, color="FFFFFF", size=10)

        for row_idx, item in enumerate(not_in_db, start=2):
            vals = [
                item["sku"], item.get("product_name", ""), item.get("collection", ""),
                item.get("product_status", ""),
                item.get("rrp_stuk_eur_nieuw", ""), item.get("rrp_gb_eur_nieuw", ""),
                item.get("giftbox_qty", ""),
                item.get("effectieve_rrp", ""),
            ]
            for col_idx, val in enumerate(vals, start=1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=val if val else "")
                if col_idx in (5, 6, 8):
                    cell.number_format = '#,##0.00'

        for col_idx, w in enumerate([16, 42, 30, 14, 12, 14, 8, 14], start=1):
            ws3.column_dimensions[get_column_letter(col_idx)].width = w

    wb.save(output_path)
    print(f"CRX overzicht: {output_path}")
    print(f"  Nieuw ingevuld: {newly_filled} | Verhogingen: {increases} | "
          f"Verlagingen: {decreases} | Ongewijzigd: {unchanged}")
    print(f"  Niet in database: {len(not_in_db)} producten")


def main():
    parser = argparse.ArgumentParser(description="Serax prijzen updaten")
    parser.add_argument("--file", required=True, help="Pad naar prijsverhogings-Excel")
    parser.add_argument("--dry-run", action="store_true", help="Alleen analyseren, niet updaten")
    parser.add_argument("--output", default=str(OUTPUT_DIR), help="Outputmap")
    args = parser.parse_args()

    file_path = Path(args.file)
    if not file_path.exists():
        print(f"FOUT: Bestand niet gevonden: {file_path}", file=sys.stderr)
        sys.exit(1)

    output = Path(args.output)
    output.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")

    # 1. Prijsbestand laden
    print("=" * 60)
    print("STAP 1: Prijsbestand laden")
    print("=" * 60)
    df = load_price_file(str(file_path))

    # 2. Prijzen extracten met giftbox-logica
    print("\n" + "=" * 60)
    print("STAP 2: Prijzen extracten (giftbox-logica)")
    print("=" * 60)
    price_data = {}
    gb_count = 0
    for _, row in df.iterrows():
        extracted = extract_prices(row)
        if not extracted["sku"]:
            continue
        price_data[extracted["sku"]] = extracted
        if extracted["is_giftbox_set"]:
            gb_count += 1

    print(f"  {len(price_data)} unieke SKUs in prijslijst")
    print(f"  {gb_count} giftbox-sets (qty > 1 -> giftbox-prijs gebruikt)")

    # 3. Database ophalen
    print("\n" + "=" * 60)
    print("STAP 3: Database producten ophalen")
    print("=" * 60)
    sb = get_supabase()
    db_products = fetch_db_products(sb)

    # 4. Matchen
    print("\n" + "=" * 60)
    print("STAP 4: SKU matching")
    print("=" * 60)
    matched = []
    not_in_db = []

    for sku, price_info in price_data.items():
        if sku in db_products:
            db_prod = db_products[sku]
            matched.append({
                **price_info,
                "db": db_prod,
                "id": db_prod["id"],
                "rrp_stuk_eur": price_info["effectieve_rrp"],
                "rrp_gb_eur": price_info["rrp_gb_eur_nieuw"],
                "inkoopprijs_stuk_eur": price_info["effectieve_inkoop"],
                "inkoopprijs_gb_eur": price_info["inkoopprijs_gb_eur_nieuw"],
            })
        else:
            not_in_db.append(price_info)

    print(f"  Matched: {len(matched)} producten")
    print(f"  Niet in database: {len(not_in_db)} producten")

    gb_matched = sum(1 for m in matched if m["is_giftbox_set"])
    print(f"  Waarvan giftbox-sets: {gb_matched}")

    # 5. Voorbeelden tonen
    print("\n--- Voorbeeld giftbox-sets (prijs = giftbox-prijs) ---")
    gb_examples = [m for m in matched if m["is_giftbox_set"]][:5]
    for ex in gb_examples:
        db = ex["db"]
        print(f"  {ex['sku']} | {ex['product_name'][:40]}")
        print(f"    GB qty: {ex['giftbox_qty']} | Stukprijs: €{ex['rrp_stuk_eur_nieuw']} | "
              f"GB-prijs: €{ex['rrp_gb_eur_nieuw']} | → Effectief: €{ex['effectieve_rrp']}")
        print(f"    DB oud: RRP €{db.get('rrp_stuk_eur', '?')} | Verkoop €{db.get('verkoopprijs', '?')}")

    # 6. Database updaten
    print("\n" + "=" * 60)
    print("STAP 5: Database update")
    print("=" * 60)
    update_database(sb, matched, dry_run=args.dry_run)

    # 7. Hextom export
    print("\n" + "=" * 60)
    print("STAP 6: Hextom bulk update Excel")
    print("=" * 60)
    hextom_path = output / f"Hextom_Price_Update_{timestamp}.xlsx"
    write_hextom_price_update(matched, hextom_path)

    # 8. CRX overzicht
    print("\n" + "=" * 60)
    print("STAP 7: CRX prijsverhogingsoverzicht")
    print("=" * 60)
    crx_path = output / f"CRX_Prijsverhoging_Serax_{timestamp}.xlsx"
    write_crx_overview(matched, not_in_db, crx_path)

    print("\n" + "=" * 60)
    print("KLAAR")
    print("=" * 60)
    if args.dry_run:
        print("[DRY RUN] Database NIET gewijzigd. Draai zonder --dry-run om te updaten.")
    print(f"Exports: {output.resolve()}")


if __name__ == "__main__":
    main()
