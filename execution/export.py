"""
Stap 5: Shopify importbestanden genereren voor Hextom Bulk Product Edit.
Zie directives/export.md voor volledige instructies.

Gebruik:
    python execution/export.py --fase 3 [--output ./exports/]
"""

import argparse
import os
from pathlib import Path
from dotenv import load_dotenv

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
OUTPUT_DIR   = os.getenv("SHOPIFY_OUTPUT_DIR", "./exports/")

# Exacte Hextom-kolomstructuur per SOP Stap 14
# Lege strings = gereserveerde kolommen (Hextom gebruikt positie, niet alleen naam)
HEXTOM_COLUMNS = [
    "Variant SKU",                                              # Col 1
    "",                                                          # Col 2 (gereserveerd)
    "",                                                          # Col 3 (gereserveerd)
    "Product Handle",                                            # Col 4
    "Product Title",                                             # Col 5
    "Product Vendor",                                            # Col 6
    "Product Type",                                              # Col 7
    "Variant Barcode",                                           # Col 8 — EAN als TEKST
    "Variant Price",                                             # Col 9
    "Variant Cost",                                              # Col 10
    "Product Description",                                       # Col 11
    "",                                                          # Col 12
    "",                                                          # Col 13
    "",                                                          # Col 14
    "Product Tags",                                              # Col 15
    "Variant Metafield custom.collectie",                        # Col 16
    "Product Metafield custom.designer",                         # Col 17
    "Product Metafield custom.materiaal",                        # Col 18
    "Product Metafield custom.kleur",                            # Col 19
    "Product Metafield custom.hoogte_filter",                    # Col 20
    "Product Metafield custom.lengte_filter",                    # Col 21
    "Product Metafield custom.breedte_filter",                   # Col 22
    "photo_packshot_1",                                          # Col 23
    "photo_packshot_2",                                          # Col 24
    "photo_packshot_3",                                          # Col 25
    "photo_packshot_4",                                          # Col 26
    "photo_packshot_5",                                          # Col 27
    "photo_lifestyle_1",                                         # Col 28
    "photo_lifestyle_2",                                         # Col 29
    "photo_lifestyle_3",                                         # Col 30
    "photo_lifestyle_4",                                         # Col 31
    "photo_lifestyle_5",                                         # Col 32
    # Extra metafields
    "Product Metafield custom.ean",                              # EAN als tekst
    "Product Metafield custom.artikelnummer",                    # SKU
    "Product Metafield custom.meta_description",                 # Meta description
]

# Kolommen die als tekst opgeslagen worden (geen getal-afronding)
TEXT_FORMAT_COLUMNS = {"Variant Barcode", "Product Metafield custom.ean"}

# Rijkleuren op basis van status_shopify
STATUS_FILL = {
    "actief":  PatternFill("solid", fgColor="FFCCCC"),   # licht rood
    "archief": PatternFill("solid", fgColor="FFE4B5"),   # licht oranje
    "nieuw":   None,
}


def get_supabase():
    from supabase import create_client
    return create_client(SUPABASE_URL, SUPABASE_KEY)


def clean_decimal(value) -> str:
    """22.50 -> '22.5', 4.00 -> '4'."""
    if value is None:
        return ""
    import re
    s = str(value).replace(",", ".")
    try:
        f = float(s)
        result = f"{f:.10f}".rstrip("0").rstrip(".")
        return result
    except ValueError:
        return s


def product_to_row(product: dict) -> dict:
    """Zet één product om naar één Hextom-rij (alle 10 foto's in aparte kolommen)."""
    return {
        "Variant SKU":                                  product.get("sku", ""),
        "":                                             "",
        "Product Handle":                               product.get("handle", ""),
        "Product Title":                                product.get("product_title_nl", ""),
        "Product Vendor":                               "Serax",
        "Product Type":                                 product.get("hoofdcategorie", ""),
        "Variant Barcode":                              product.get("ean_shopify", ""),
        "Variant Price":                                clean_decimal(product.get("verkoopprijs")),
        "Variant Cost":                                 clean_decimal(product.get("inkoopprijs")),
        "Product Description":                          product.get("meta_description", ""),
        "Product Tags":                                 product.get("tags", ""),
        "Variant Metafield custom.collectie":           product.get("collectie", ""),
        "Product Metafield custom.designer":            product.get("designer", ""),
        "Product Metafield custom.materiaal":           product.get("materiaal_nl", ""),
        "Product Metafield custom.kleur":               product.get("kleur_nl", ""),
        "Product Metafield custom.hoogte_filter":       clean_decimal(product.get("hoogte_cm")),
        "Product Metafield custom.lengte_filter":       clean_decimal(product.get("lengte_cm")),
        "Product Metafield custom.breedte_filter":      clean_decimal(product.get("breedte_cm")),
        "photo_packshot_1":                             product.get("photo_packshot_1", ""),
        "photo_packshot_2":                             product.get("photo_packshot_2", ""),
        "photo_packshot_3":                             product.get("photo_packshot_3", ""),
        "photo_packshot_4":                             product.get("photo_packshot_4", ""),
        "photo_packshot_5":                             product.get("photo_packshot_5", ""),
        "photo_lifestyle_1":                            product.get("photo_lifestyle_1", ""),
        "photo_lifestyle_2":                            product.get("photo_lifestyle_2", ""),
        "photo_lifestyle_3":                            product.get("photo_lifestyle_3", ""),
        "photo_lifestyle_4":                            product.get("photo_lifestyle_4", ""),
        "photo_lifestyle_5":                            product.get("photo_lifestyle_5", ""),
        "Product Metafield custom.ean":                 product.get("ean_piece", ""),
        "Product Metafield custom.artikelnummer":       product.get("sku", ""),
        "Product Metafield custom.meta_description":    product.get("meta_description", ""),
    }


def write_excel(products: list[dict], path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Header
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col_idx, col_name in enumerate(HEXTOM_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name if col_name else "")
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=False)

    # Data
    for row_idx, product in enumerate(products, start=2):
        row_data  = product_to_row(product)
        row_fill  = STATUS_FILL.get(product.get("status_shopify") or "nieuw")

        for col_idx, col_name in enumerate(HEXTOM_COLUMNS, start=1):
            value = row_data.get(col_name, "") if col_name else ""
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)

            # EAN als tekst forceren
            if col_name in TEXT_FORMAT_COLUMNS and value:
                cell.value         = str(value)
                cell.number_format = "@"

            if row_fill:
                cell.fill = row_fill

    # Kolombreedte
    col_widths = {1: 18, 4: 40, 5: 50, 8: 16, 11: 60, 15: 50}
    for col_idx in range(1, len(HEXTOM_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_idx, 20)

    wb.save(path)


def export(fase: str, output_dir: str):
    sb  = get_supabase()
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    result   = sb.table("seo_products").select("*").eq("status", "ready").eq("fase", fase).execute()
    products = result.data

    if not products:
        print(f"Geen producten met status='ready' voor fase {fase}.")
        print("Draai eerst validate.py en los de review-items op.")
        return

    print(f"Exporteren: {len(products)} producten (fase {fase})")

    # Splits op status_shopify
    nieuw   = [p for p in products if (p.get("status_shopify") or "nieuw") != "archief"]
    archief = [p for p in products if p.get("status_shopify") == "archief"]

    for label, subset, filename in [
        ("Nieuw",   nieuw,   f"Shopify_Nieuw_fase{fase}.xlsx"),
        ("Archief", archief, f"Shopify_Archief_fase{fase}.xlsx"),
    ]:
        if not subset:
            print(f"  {label}: geen producten, bestand overgeslagen")
            continue
        path = out / filename
        write_excel(subset, path)
        print(f"  {label}: {len(subset)} producten -> {path}")

    print(f"\nExport klaar. Bestanden staan in: {out.resolve()}")
    print("Controleer voor import:")
    print("  - EAN in kolom 8 als tekst (geen wetenschappelijke notatie)")
    print("  - Decimalen correct (22.5 niet 22.50)")
    print("  - Rijkleuren: actief=rood, archief=oranje, nieuw=geen kleur")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--fase",   required=True, help="Fasecode, bijv. 3")
    parser.add_argument("--output", default=OUTPUT_DIR, help="Outputmap")
    args = parser.parse_args()

    export(args.fase, args.output)
