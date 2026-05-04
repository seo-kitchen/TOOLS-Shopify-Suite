"""
Vul meta_title en meta_description in Serax_Batch Excel via de meta-bot
(execution/meta_audit_generate.py). Werkt product-voor-product, gebruikt
Claude voor description en build_title voor title.

Gebruik:
    python execution/fill_batch_meta.py --input "exports/Serax_Batch_MET_FOTOS.xlsx" \
        --output "exports/Serax_Batch_20260414_MET_META.xlsx" --limit 3
    python execution/fill_batch_meta.py --input "..." --output "..."  # alle producten
"""

import argparse
import os
import sys
import time
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

load_dotenv()

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

# Hergebruik functies uit de meta-bot
sys.path.insert(0, str(Path(__file__).parent))
from meta_audit_generate import (
    get_supabase, get_anthropic, make_title, generate_desc, fetch_context,
    TITLE_MAX, DESC_MAX, DESC_MIN,
)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--sheet", default="Shopify_Nieuw")
    ap.add_argument("--limit", type=int, help="Alleen eerste N producten")
    ap.add_argument("--skus", help="Comma-separated specifieke SKUs")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.input)
    ws = wb[args.sheet]

    # Header mapping
    headers = [c.value for c in ws[1]]
    header_idx = {h: i for i, h in enumerate(headers)}

    sku_col = header_idx.get("Variant SKU")
    pid_col = header_idx.get("Product ID")
    handle_col = header_idx.get("Product handle")
    title_col = header_idx.get("Product title")
    vendor_col = header_idx.get("Product vendor")

    # meta_title toevoegen als kolom nog niet bestaat (naast meta_description)
    if "meta_title" not in header_idx:
        new_col = ws.max_column + 1
        ws.cell(row=1, column=new_col, value="meta_title")
        header_idx["meta_title"] = new_col - 1
        print(f"Kolom 'meta_title' toegevoegd op positie {new_col}")
    meta_title_col = header_idx["meta_title"]
    meta_desc_col = header_idx.get("meta_description")
    if meta_desc_col is None:
        raise RuntimeError("Geen 'meta_description' kolom gevonden")

    # Filter rows
    filter_skus = set(s.strip() for s in args.skus.split(",")) if args.skus else None

    rows_to_process = []
    for row_idx in range(2, ws.max_row + 1):
        sku = ws.cell(row=row_idx, column=sku_col + 1).value
        if not sku:
            continue
        sku = str(sku).strip()
        if filter_skus and sku not in filter_skus:
            continue
        rows_to_process.append(row_idx)
    if args.limit:
        rows_to_process = rows_to_process[: args.limit]
    print(f"Te verwerken: {len(rows_to_process)} producten")

    sb = get_supabase()
    client = get_anthropic()

    filled_title = 0
    filled_desc = 0
    errors = 0

    for i, row_idx in enumerate(rows_to_process, 1):
        sku = str(ws.cell(row=row_idx, column=sku_col + 1).value).strip()
        pid = ws.cell(row=row_idx, column=pid_col + 1).value
        handle = ws.cell(row=row_idx, column=handle_col + 1).value
        title = ws.cell(row=row_idx, column=title_col + 1).value
        vendor = ws.cell(row=row_idx, column=vendor_col + 1).value or "Serax"

        print(f"\n[{i}/{len(rows_to_process)}] {sku}  {(title or '')[:60]}")
        if not title:
            print(f"  SKIP: geen product title")
            continue

        try:
            # Context ophalen uit seo_products via handle of shopify_product_id
            product_dict = {
                "shopify_product_id": str(pid) if pid else "",
                "handle": handle or "",
            }
            ctx = fetch_context(sb, product_dict)

            # Title
            new_title, title_reason = make_title(client, str(title), str(vendor))
            ws.cell(row=row_idx, column=meta_title_col + 1, value=new_title)
            filled_title += 1
            print(f"  title ({len(new_title)}ch, {title_reason}): {new_title}")

            # Description — gebruik SKU als pid voor USP/CTA rotatie als geen Product ID
            desc_pid = str(pid) if pid else sku
            new_desc = generate_desc(client, str(title), ctx, desc_pid)
            ws.cell(row=row_idx, column=meta_desc_col + 1, value=new_desc)
            filled_desc += 1
            print(f"  desc  ({len(new_desc)}ch): {new_desc}")

        except Exception as e:
            errors += 1
            print(f"  FOUT: {e}")
            continue

        # Rate limit safety
        if i % 10 == 0:
            time.sleep(0.5)

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"\n{'='*60}\nKLAAR")
    print(f"  Titles ingevuld: {filled_title}")
    print(f"  Descriptions ingevuld: {filled_desc}")
    print(f"  Fouten: {errors}")
    print(f"  Output: {args.output}")


if __name__ == "__main__":
    main()
