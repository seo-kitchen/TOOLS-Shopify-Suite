"""
Standaard export: één Excel-bestand met drie tabs.

  Tab 1: Shopify_Nieuw     — nieuwe producten (template-format)
  Tab 2: Shopify_Archief   — te reactiveren producten (zelfde format, incl. Product ID / Variant ID)
  Tab 3: Analyse           — overzicht van de batch (voor review)

Gebruik:
    python execution/export_standaard.py --fase 3 [--output ./exports/]
"""

import argparse
import os
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
OUTPUT_DIR   = os.getenv("SHOPIFY_OUTPUT_DIR", "./exports/")

# Exacte kolomvolgorde uit template (33 kolommen)
TEMPLATE_COLS = [
    ("Variant SKU",              "sku"),
    ("Product ID",               "shopify_product_id"),
    ("Variant ID",               "shopify_variant_id"),
    ("Product handle",           "handle"),
    ("Product title",            "product_title_nl"),
    ("Product vendor",           "_vendor"),
    ("Product type",             "hoofdcategorie"),
    ("EAN Code",                 "ean_shopify"),      # tekst-format!
    ("Verkoopprijs Shopify",     "verkoopprijs"),
    ("Inkoopprijs Shopify",      "inkoopprijs"),
    ("Product description",      "meta_description"),
    ("Nieuwe hoofdcategorie",    "hoofdcategorie"),
    ("Nieuwe subcategorie",      "subcategorie"),
    ("Nieuwe sub-subcategorie",  "sub_subcategorie"),
    ("Nieuwe tag",               "tags"),
    ("collectie",                "collectie"),
    ("designer",                 "designer"),
    ("materiaal",                "materiaal_nl"),
    ("kleur",                    "kleur_nl"),
    ("hoogte_cm",                "_hoogte"),
    ("lengte_cm",                "_lengte"),
    ("breedte_cm",               "_breedte"),
    ("meta_description",         "meta_description"),
    ("photo_packshot1",          "photo_packshot_1"),
    ("photo_packshot2",          "photo_packshot_2"),
    ("photo_packshot3",          "photo_packshot_3"),
    ("photo_packshot4",          "photo_packshot_4"),
    ("photo_packshot5",          "photo_packshot_5"),
    ("photo_lifestyle1",         "photo_lifestyle_1"),
    ("photo_lifestyle2",         "photo_lifestyle_2"),
    ("photo_lifestyle3",         "photo_lifestyle_3"),
    ("photo_lifestyle4",         "photo_lifestyle_4"),
    ("photo_lifestyle5",         "photo_lifestyle_5"),
]

HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")  # lichtblauw, identiek aan template
EAN_COL_IDX  = 8   # 1-based index van "EAN Code" kolom


def get_supabase():
    from supabase import create_client
    return create_client(SUPABASE_URL, SUPABASE_KEY)


def clean_decimal(value) -> str:
    if value is None:
        return ""
    try:
        f = float(str(value).replace(",", "."))
        return f"{f:.10f}".rstrip("0").rstrip(".")
    except ValueError:
        return str(value)


def get_veld(product: dict, veld: str | None) -> str:
    if veld is None:
        return ""
    if veld == "_vendor":
        return "Serax"
    if veld in ("_hoogte", "_lengte", "_breedte"):
        key = {"_hoogte": "hoogte_cm", "_lengte": "lengte_cm", "_breedte": "breedte_cm"}[veld]
        return clean_decimal(product.get(key))
    if veld in ("verkoopprijs", "inkoopprijs"):
        return clean_decimal(product.get(veld))
    return str(product.get(veld) or "")


def schrijf_product_tab(ws, producten: list[dict]):
    """Schrijf producten naar een worksheet in het template-format."""
    header_font = Font(bold=True, size=10)

    # Header rij
    for col_idx, (col_naam, _) in enumerate(TEMPLATE_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_naam)
        cell.fill      = HEADER_FILL
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Data rijen
    for row_idx, product in enumerate(producten, start=2):
        for col_idx, (col_naam, veld) in enumerate(TEMPLATE_COLS, start=1):
            waarde = get_veld(product, veld)
            cell   = ws.cell(row=row_idx, column=col_idx, value=waarde)
            cell.font      = Font(size=9)
            cell.alignment = Alignment(vertical="top")

            # EAN als tekst zodat Excel geen getal maakt
            if col_idx == EAN_COL_IDX and waarde:
                cell.value         = str(waarde)
                cell.number_format = "@"

    # Kolombreedte
    breedte_map = {
        1: 18,   # Variant SKU
        4: 40,   # Product handle
        5: 50,   # Product title
        8: 16,   # EAN Code
        9: 14,   # Verkoopprijs
        10: 14,  # Inkoopprijs
        11: 60,  # Product description
        15: 50,  # Nieuwe tag
    }
    for col_idx in range(1, len(TEMPLATE_COLS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = breedte_map.get(col_idx, 16)

    return len(producten)


def schrijf_analyse_tab(ws, fase: str, alle_producten: list[dict], sb):
    """Overzicht-tab met samenvatting, nieuwe filterwaarden en review-items."""

    def sectie_header(row: int, tekst: str) -> int:
        cel = ws.cell(row=row, column=1, value=tekst)
        cel.font  = Font(bold=True, size=11)
        cel.fill  = PatternFill("solid", fgColor="4472C4")
        cel.font  = Font(bold=True, color="FFFFFF", size=11)
        cel.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 22
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        return row + 1

    def kolom_header(row: int, headers: list) -> int:
        for col, h in enumerate(headers, start=1):
            cel = ws.cell(row=row, column=col, value=h)
            cel.font  = Font(bold=True, size=9)
            cel.fill  = HEADER_FILL
            cel.alignment = Alignment(horizontal="center")
        return row + 1

    rij = 1

    # Titel
    ws.cell(row=rij, column=1, value=f"Serax Product Onboarding — Analyse Rapport").font = Font(bold=True, size=14)
    rij += 1
    ws.cell(row=rij, column=1, value=f"Fase: {fase}  |  Gegenereerd: {datetime.now().strftime('%d-%m-%Y %H:%M')}").font = Font(size=9, color="666666")
    rij += 2

    # Samenvatting
    rij = sectie_header(rij, "SAMENVATTING")
    rij = kolom_header(rij, ["", "Aantal"])

    status_counts  = {}
    shopify_counts = {}
    for p in alle_producten:
        status_counts[p.get("status") or "?"]         = status_counts.get(p.get("status") or "?", 0) + 1
        shopify_counts[p.get("status_shopify") or "?"] = shopify_counts.get(p.get("status_shopify") or "?", 0) + 1

    for label, val in [
        ("Totaal producten", len(alle_producten)),
        ("Klaar voor import (ready)",            status_counts.get("ready", 0)),
        ("Handmatige controle nodig (review)",   status_counts.get("review", 0)),
        ("Nieuw aan te maken in Shopify",        shopify_counts.get("nieuw", 0)),
        ("Reactiveren uit archief",              shopify_counts.get("archief", 0)),
        ("Al actief op webshop",                 shopify_counts.get("actief", 0)),
    ]:
        ws.cell(row=rij, column=1, value=label).font = Font(size=9)
        ws.cell(row=rij, column=2, value=val).font   = Font(size=9, bold=True)
        rij += 1
    rij += 1

    # Nieuwe filterwaarden
    filter_result = sb.table("seo_filter_values").select("type,waarde").execute()
    bekende = {(r["type"], r["waarde"]) for r in filter_result.data}

    nieuwe_kleuren    = sorted({p.get("kleur_nl") or "" for p in alle_producten if p.get("kleur_nl") and ("kleur", p["kleur_nl"]) not in bekende})
    nieuwe_materialen = sorted({p.get("materiaal_nl") or "" for p in alle_producten if p.get("materiaal_nl") and ("materiaal", p["materiaal_nl"]) not in bekende})

    if nieuwe_kleuren or nieuwe_materialen:
        rij = sectie_header(rij, "NIEUWE FILTERWAARDEN — AANMAKEN IN SHOPIFY VOOR IMPORT")
        rij = kolom_header(rij, ["Type", "Waarde", "Actie"])
        for k in nieuwe_kleuren:
            ws.cell(row=rij, column=1, value="Kleur").font    = Font(size=9)
            ws.cell(row=rij, column=2, value=k).font          = Font(size=9, bold=True)
            ws.cell(row=rij, column=3, value="Aanmaken in Shopify > Metavelden > Kleur filter").font = Font(size=9)
            rij += 1
        for m in nieuwe_materialen:
            ws.cell(row=rij, column=1, value="Materiaal").font = Font(size=9)
            ws.cell(row=rij, column=2, value=m).font           = Font(size=9, bold=True)
            ws.cell(row=rij, column=3, value="Aanmaken in Shopify > Metavelden > Materiaal filter").font = Font(size=9)
            rij += 1
        rij += 1

    # Review-producten
    review_producten = [p for p in alle_producten if p.get("status") == "review"]
    if review_producten:
        rij = sectie_header(rij, f"HANDMATIGE CONTROLE VEREIST ({len(review_producten)} producten)")
        rij = kolom_header(rij, ["SKU", "Productnaam", "Status Shopify", "Reden", "Prijs", "EAN"])
        for p in review_producten:
            for col, val in enumerate([
                p.get("sku", ""),
                p.get("product_title_nl") or p.get("product_name_raw", ""),
                p.get("status_shopify", ""),
                p.get("review_reden", ""),
                clean_decimal(p.get("verkoopprijs")),
                p.get("ean_shopify", ""),
            ], start=1):
                ws.cell(row=rij, column=col, value=val).font = Font(size=9)
            rij += 1
        rij += 1

    # Alle producten
    rij = sectie_header(rij, "ALLE PRODUCTEN")
    rij = kolom_header(rij, ["SKU", "Producttitel NL", "Status Shopify", "Categorie", "Kleur", "Materiaal", "Prijs", "EAN", "Foto?"])
    for p in alle_producten:
        vals = [
            p.get("sku", ""),
            p.get("product_title_nl", ""),
            p.get("status_shopify", ""),
            f"{p.get('hoofdcategorie','')} > {p.get('sub_subcategorie','')}",
            p.get("kleur_nl", ""),
            p.get("materiaal_nl", ""),
            clean_decimal(p.get("verkoopprijs")),
            p.get("ean_shopify", ""),
            "Ja" if p.get("photo_packshot_1") else "Nee",
        ]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=rij, column=col, value=val)
            cell.font = Font(size=9)
            if col == 8:
                cell.number_format = "@"
        rij += 1

    # Kolombreedte analyse-tab
    for col, b in [(1, 18), (2, 45), (3, 14), (4, 35), (5, 16), (6, 18), (7, 12), (8, 16), (9, 8)]:
        ws.column_dimensions[get_column_letter(col)].width = b

    ws.freeze_panes = "A5"


def export_standaard(fase: str, output_dir: str) -> str | None:
    sb  = get_supabase()
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    result_ready = sb.table("seo_products").select("*").eq("status", "ready").eq("fase", fase).execute()
    producten    = result_ready.data or []

    result_all = sb.table("seo_products").select("*").eq("fase", fase).execute()
    alle       = result_all.data or []

    if not producten:
        print(f"Geen producten met status='ready' voor fase {fase}.")
        return None

    # Splits op status_shopify
    nieuw   = [p for p in producten if p.get("status_shopify") == "nieuw"]
    archief = [p for p in producten if p.get("status_shopify") == "archief"]
    # Actieve producten die worden bijgewerkt gaan ook in de archief-tab (hebben al Product ID)
    actief  = [p for p in producten if p.get("status_shopify") == "actief"]

    bestandsnaam = f"Serax_Import_Fase{fase}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    pad = out / bestandsnaam

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Tab 1: Nieuw
    ws_nieuw = wb.create_sheet("Shopify_Nieuw")
    n_nieuw  = schrijf_product_tab(ws_nieuw, nieuw)

    # Tab 2: Archief (incl. actieve producten die bijgewerkt worden)
    ws_archief = wb.create_sheet("Shopify_Archief")
    n_archief  = schrijf_product_tab(ws_archief, archief + actief)

    # Tab 3: Analyse
    ws_analyse = wb.create_sheet("Analyse")
    schrijf_analyse_tab(ws_analyse, fase, alle, sb)

    wb.save(pad)

    print(f"\nStandaard export klaar:")
    print(f"  Bestand:  {pad}")
    print(f"  Nieuw:    {n_nieuw} producten (tab Shopify_Nieuw)")
    print(f"  Archief:  {n_archief} producten (tab Shopify_Archief)")
    print(f"  Analyse:  {len(alle)} producten totaal (tab Analyse)")

    return str(pad)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--fase",   required=True, help="Fasecode, bijv. 3")
    parser.add_argument("--output", default=OUTPUT_DIR)
    args = parser.parse_args()

    export_standaard(args.fase, args.output)
