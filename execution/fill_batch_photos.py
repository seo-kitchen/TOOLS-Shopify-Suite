"""
Vul foto-kolommen (photo_packshot1..5, photo_lifestyle1..5) in een Serax_Batch-Excel
op basis van de foto-URLs die in seo_products staan (Supabase Storage).

Gebruik:
    python execution/fill_batch_photos.py \
        --input ".tmp/Serax_Batch_20260414_0139 (1).xlsx" \
        --output "exports/Serax_Batch_20260414_MET_FOTOS.xlsx"
"""

import argparse
import os
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

load_dotenv()
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY")

# Excel-kolomnamen (zonder underscore) -> DB-kolomnamen (met underscore)
COL_MAP = {
    "photo_packshot1": "photo_packshot_1",
    "photo_packshot2": "photo_packshot_2",
    "photo_packshot3": "photo_packshot_3",
    "photo_packshot4": "photo_packshot_4",
    "photo_packshot5": "photo_packshot_5",
    "photo_lifestyle1": "photo_lifestyle_1",
    "photo_lifestyle2": "photo_lifestyle_2",
    "photo_lifestyle3": "photo_lifestyle_3",
    "photo_lifestyle4": "photo_lifestyle_4",
    "photo_lifestyle5": "photo_lifestyle_5",
}


def get_supabase():
    from supabase import create_client
    return create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--sheet", default="Shopify_Nieuw")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.input)
    ws = wb[args.sheet]

    # Build header -> column index map
    headers = [c.value for c in ws[1]]
    header_idx = {h: i for i, h in enumerate(headers)}
    sku_col = header_idx.get("Variant SKU")
    if sku_col is None:
        raise RuntimeError("Geen 'Variant SKU' kolom gevonden")

    photo_cols_excel = [c for c in COL_MAP if c in header_idx]
    if not photo_cols_excel:
        raise RuntimeError("Geen photo_* kolommen gevonden in sheet")
    print(f"Foto-kolommen: {photo_cols_excel}")

    # Collect SKUs
    skus = []
    for row_idx in range(2, ws.max_row + 1):
        sku = ws.cell(row=row_idx, column=sku_col + 1).value
        if sku:
            skus.append(str(sku).strip())
    print(f"SKUs in sheet: {len(skus)}")

    # Batch fetch from Supabase
    sb = get_supabase()
    db_photos: dict[str, dict] = {}
    db_cols = list(COL_MAP.values())
    for i in range(0, len(skus), 100):
        chunk = skus[i:i+100]
        r = sb.table("seo_products").select("sku, " + ", ".join(db_cols)).in_("sku", chunk).execute()
        for row in (r.data or []):
            db_photos[row["sku"]] = row
    print(f"Gevonden in seo_products: {len(db_photos)}/{len(skus)}")

    # Fill sheet
    filled_rows = 0
    filled_cells = 0
    skus_without_photos = []
    for row_idx in range(2, ws.max_row + 1):
        sku = ws.cell(row=row_idx, column=sku_col + 1).value
        if not sku:
            continue
        sku = str(sku).strip()
        prod = db_photos.get(sku)
        if not prod:
            continue
        any_filled = False
        for excel_col, db_col in COL_MAP.items():
            if excel_col not in header_idx:
                continue
            url = prod.get(db_col)
            cell_idx = header_idx[excel_col] + 1
            # Schrijf alleen als URL bestaat; laat bestaande waarde staan als geen nieuwe URL
            if url:
                ws.cell(row=row_idx, column=cell_idx).value = url
                filled_cells += 1
                any_filled = True
            else:
                # Leeg maken als er niets staat (zorg dat we geen oude rommel laten staan)
                current = ws.cell(row=row_idx, column=cell_idx).value
                if not current:
                    ws.cell(row=row_idx, column=cell_idx).value = None
        if any_filled:
            filled_rows += 1
        else:
            skus_without_photos.append(sku)

    print(f"\nProducten met foto's ingevuld: {filled_rows}")
    print(f"Totaal foto-cellen ingevuld: {filled_cells}")
    print(f"Producten zonder foto's: {len(skus_without_photos)}")

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"\nOpgeslagen: {args.output}")


if __name__ == "__main__":
    main()
