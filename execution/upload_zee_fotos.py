"""Upload ZEE-foto's naar Supabase Storage en update de Hextom Excel.

Verwachte mapstructuur in .tmp/zee_fotos/:
  .tmp/zee_fotos/
    mirtoon10bk/
      packshot_1.jpg
      packshot_2.jpg
      lifestyle_1.jpg
      ...
    mirtoon30bk/
      ...
    tasmanbk/
      ...

Na afloop worden de URLs ingevuld in .tmp/hextom_ZEE_2026_FINAL.xlsx

Gebruik:
    python execution/upload_zee_fotos.py
    python execution/upload_zee_fotos.py --dry-run
    python execution/upload_zee_fotos.py --fotos-dir pad/naar/fotos
"""
from __future__ import annotations

import argparse
import io
import os
import sys
from pathlib import Path

from dotenv import load_dotenv
from PIL import Image

load_dotenv()

BUCKET = "product-photos"
MAX_DIM = 4999      # Shopify max 5000×5000 = 25 MP
JPEG_QUALITY = 90
FOTOS_DIR = Path(".tmp/zee_fotos")
EXCEL_IN  = Path(".tmp/hextom_ZEE_2026_FINAL.xlsx")
EXCEL_OUT = Path(".tmp/hextom_ZEE_2026_MET_FOTOS.xlsx")

# Hoe herkennen we packshots vs lifestyle?
# Alles met 'packshot', 'pack', 'white', 'studio', 'plain' → packshot
# Alles met 'lifestyle', 'mood', 'sfeer', 'context' → lifestyle
# Overige volgorde: gewoon op bestandsnaam sorteren, eerste 5 = packshot, rest = lifestyle

PACKSHOT_HINTS   = {"packshot", "pack", "studio", "white_bg", "white bg", "plain", "background"}
LIFESTYLE_HINTS  = {"lifestyle", "mood", "sfeer", "context", "ambiance", "scene"}

HEXTOM_PHOTO_COLS = [
    "photo_packshot_1","photo_packshot_2","photo_packshot_3","photo_packshot_4","photo_packshot_5",
    "photo_lifestyle_1","photo_lifestyle_2","photo_lifestyle_3","photo_lifestyle_4","photo_lifestyle_5",
]


def get_supabase():
    from supabase import create_client
    url = os.getenv("SUPABASE_NEW_URL","")
    key = os.getenv("SUPABASE_NEW_SERVICE_KEY","") or os.getenv("SUPABASE_NEW_KEY","")
    return create_client(url, key)


def resize_and_compress(img_path: Path) -> bytes:
    """Resize naar max MAX_DIM×MAX_DIM, sla op als JPEG, geef bytes terug."""
    with Image.open(img_path) as img:
        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")
        w, h = img.size
        if w > MAX_DIM or h > MAX_DIM:
            ratio = min(MAX_DIM / w, MAX_DIM / h)
            img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=JPEG_QUALITY, optimize=True)
        return buf.getvalue()


def classificeer_fotos(bestanden: list[Path]) -> tuple[list[Path], list[Path]]:
    """Splits op packshots en lifestyle op basis van bestandsnaam."""
    packshots, lifestyle = [], []
    for f in sorted(bestanden):
        naam = f.stem.lower()
        is_lifestyle = any(h in naam for h in LIFESTYLE_HINTS)
        is_packshot  = any(h in naam for h in PACKSHOT_HINTS)
        if is_lifestyle and not is_packshot:
            lifestyle.append(f)
        else:
            packshots.append(f)
    # Als geen onderscheid gemaakt: eerste 5 = packshot, rest = lifestyle
    if not lifestyle and len(packshots) > 5:
        lifestyle  = packshots[5:]
        packshots  = packshots[:5]
    return packshots[:5], lifestyle[:5]


def upload_foto(sb, sku: str, bestand: Path, slot: str, dry_run: bool) -> str | None:
    """Upload één foto, return publieke URL."""
    storage_pad = f"zee/{sku}/{slot}.jpg"
    if dry_run:
        print(f"  [dry] zou uploaden: {bestand.name} → {storage_pad}")
        return f"https://PLACEHOLDER/{storage_pad}"
    try:
        data = resize_and_compress(bestand)
        sb.storage.from_(BUCKET).upload(
            path=storage_pad,
            file=data,
            file_options={"content-type": "image/jpeg", "upsert": "true"},
        )
        supabase_url = os.getenv("SUPABASE_NEW_URL","").rstrip("/")
        return f"{supabase_url}/storage/v1/object/public/{BUCKET}/{storage_pad}"
    except Exception as e:
        print(f"  ❌ Upload fout voor {bestand.name}: {e}")
        return None


def verwerk_sku(sb, sku: str, dry_run: bool) -> dict[str, str]:
    """Verwerk alle foto's voor één SKU, return {photo_slot: url}."""
    sku_dir = FOTOS_DIR / sku
    if not sku_dir.exists():
        print(f"  ⚠️ Map niet gevonden: {sku_dir}")
        return {}

    foto_exts = {".jpg", ".jpeg", ".png", ".webp", ".tiff", ".tif"}
    bestanden = [f for f in sku_dir.iterdir() if f.suffix.lower() in foto_exts]
    if not bestanden:
        print(f"  ⚠️ Geen foto's gevonden in {sku_dir}")
        return {}

    packshots, lifestyle = classificeer_fotos(bestanden)
    print(f"  {len(packshots)} packshots, {len(lifestyle)} lifestyle")

    urls: dict[str, str] = {}
    for i, foto in enumerate(packshots, 1):
        slot = f"photo_packshot_{i}"
        url = upload_foto(sb, sku, foto, slot, dry_run)
        if url:
            urls[slot] = url
            print(f"  ✅ {slot} → {foto.name}")

    for i, foto in enumerate(lifestyle, 1):
        slot = f"photo_lifestyle_{i}"
        url = upload_foto(sb, sku, foto, slot, dry_run)
        if url:
            urls[slot] = url
            print(f"  ✅ {slot} → {foto.name}")

    return urls


def update_excel(foto_urls: dict[str, dict[str, str]], dry_run: bool) -> None:
    """Schrijf foto-URLs terug in de Hextom Excel."""
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_IN)
    ws = wb.active

    # Bouw kolom-index
    header = {ws.cell(row=1, column=ci).value: ci for ci in range(1, ws.max_column + 1)}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        sku_ci = header.get("Variant SKU")
        if not sku_ci:
            break
        sku = ws.cell(row=row[0].row, column=sku_ci).value
        if not sku or sku not in foto_urls:
            continue
        for slot, url in foto_urls[sku].items():
            ci = header.get(slot)
            if ci:
                ws.cell(row=row[0].row, column=ci).value = url

    if not dry_run:
        wb.save(EXCEL_OUT)
        print(f"\n✅ Excel opgeslagen: {EXCEL_OUT}")
    else:
        print(f"\n[dry] Excel zou opgeslagen worden als: {EXCEL_OUT}")


def main(dry_run: bool = False, fotos_dir: str | None = None) -> None:
    global FOTOS_DIR
    if fotos_dir:
        FOTOS_DIR = Path(fotos_dir)

    if not FOTOS_DIR.exists():
        print(f"❌ Foto-map niet gevonden: {FOTOS_DIR}")
        print(f"   Maak de map aan en zet per SKU een submap met foto's.")
        print(f"   Voorbeeld: .tmp/zee_fotos/mirtoon30bk/packshot_1.jpg")
        return

    if not EXCEL_IN.exists():
        print(f"❌ Excel niet gevonden: {EXCEL_IN}")
        return

    sb = get_supabase()

    # Detecteer welke SKU-mappen aanwezig zijn
    sku_mappen = [d for d in FOTOS_DIR.iterdir() if d.is_dir()]
    if not sku_mappen:
        print(f"Geen SKU-mappen gevonden in {FOTOS_DIR}")
        print("Verwacht: één map per SKU, bv. .tmp/zee_fotos/mirtoon30bk/")
        return

    print(f"{'DRY RUN — ' if dry_run else ''}Verwerken: {len(sku_mappen)} SKU-mappen\n")

    alle_urls: dict[str, dict[str, str]] = {}
    for sku_dir in sorted(sku_mappen):
        sku = sku_dir.name
        print(f"SKU: {sku}")
        urls = verwerk_sku(sb, sku, dry_run)
        if urls:
            alle_urls[sku] = urls
        print()

    if alle_urls:
        update_excel(alle_urls, dry_run)
        print(f"\nSamenvatting: {sum(len(v) for v in alle_urls.values())} foto-URLs ingevuld voor {len(alle_urls)} SKUs")
    else:
        print("Geen foto's verwerkt.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--fotos-dir", help="Pad naar foto-mappen (default: .tmp/zee_fotos)")
    args = parser.parse_args()
    main(dry_run=args.dry_run, fotos_dir=args.fotos_dir)
