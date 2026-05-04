"""
Serax 86 producten herverwerking
Verwerkt 86 Serax producten vanuit nieuwe masterdata en pricelist naar Supabase.
"""
import sys
import os

# Voeg execution map toe aan sys.path
sys.path.insert(0, os.path.join(os.path.dirname(__file__)))

from dotenv import load_dotenv
import openpyxl
import pandas as pd
from supabase import create_client
import re

# Laad .env
env_path = os.path.join(os.path.dirname(__file__), '..', '.env')
load_dotenv(env_path)

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY")

BASE_DIR = os.path.join(os.path.dirname(__file__), '..')

# ─────────────────────────────────────────────
# KLEUR VERTALING EN/NL
# ─────────────────────────────────────────────
KLEUR_MAP = {
    "white": "Wit",
    "black": "Zwart",
    "blue": "Blauw",
    "green": "Groen",
    "yellow": "Geel",
    "red": "Rood",
    "grey": "Grijs",
    "gray": "Grijs",
    "beige": "Beige",
    "transparent": "Transparant",
    "light blue": "Lichtblauw",
    "ocher": "Oker",
    "ochre": "Oker",
    "purple": "Paars",
    "orange": "Oranje",
    "pink": "Roze",
    "brown": "Bruin",
    "mix": "Mix",
    "natural": "Naturel",
    "sand": "Zand",
    "ivory": "Ivoor",
    "cream": "Crème",
    "gold": "Goud",
    "silver": "Zilver",
    "copper": "Koper",
    "dark blue": "Donkerblauw",
    "dark green": "Donkergroen",
    "light green": "Lichtgroen",
    "olive": "Olijf",
    "terracotta": "Terracotta",
    "bordeaux": "Bordeaux",
    "burgundy": "Bordeaux",
    "mustard": "Mosterd",
    "turquoise": "Turquoise",
    "multicolor": "Multicolor",
    "multi": "Multicolor",
    "off white": "Gebroken Wit",
    "off-white": "Gebroken Wit",
    "rust": "Roestbruin",
    "taupe": "Taupe",
    "charcoal": "Antraciet",
    "anthracite": "Antraciet",
    "dark grey": "Donkergrijs",
    "light grey": "Lichtgrijs",
    "navy": "Marineblauw",
    "ecru": "Ecru",
    "amber": "Amber",
}

# ─────────────────────────────────────────────
# MATERIAAL VERTALING EN/NL
# ─────────────────────────────────────────────
MATERIAAL_MAP = {
    "porcelain": "Porselein",
    "glass": "Glas",
    "borosilicate glass": "Glas",
    "stoneware": "Steengoed",
    "ceramic": "Keramiek",
    "wood": "Hout",
    "metal": "Metaal",
    "earthenware": "Aardewerk",
    "terracotta": "Aardewerk",
    "stainless steel": "RVS",
    "marble": "Marmer",
    "bamboo": "Bamboe",
    "cotton": "Katoen",
    "linen": "Linnen",
    "leather": "Leer",
    "rattan": "Riet",
    "acrylic": "Acryl",
    "silicone": "Siliconen",
    "maple": "Esdoorn",
    "oak": "Eiken",
    "walnut": "Walnoot",
    "concrete": "Beton",
    "resin": "Hars",
    "faience": "Faïence",
    "enamel": "Emaille",
    "cast iron": "Gietijzer",
    "aluminium": "Aluminium",
    "aluminum": "Aluminium",
    "brass": "Messing",
    "copper": "Koper",
}


def vertaal_kleur(kleur_en: str) -> str:
    """Vertaal Engelse kleur naar Nederlands."""
    if not kleur_en:
        return ""
    kleur_lower = kleur_en.strip().lower()
    # Exacte match
    if kleur_lower in KLEUR_MAP:
        return KLEUR_MAP[kleur_lower]
    # Gedeeltelijke match (eerste woord)
    for en, nl in KLEUR_MAP.items():
        if kleur_lower.startswith(en):
            return nl
    # Terugval: eerste letter hoofdletter
    return kleur_en.strip().title()


def vertaal_materiaal(mat_en: str) -> str:
    """Vertaal Engels materiaal naar Nederlands."""
    if not mat_en:
        return ""
    mat_lower = mat_en.strip().lower()
    if mat_lower in MATERIAAL_MAP:
        return MATERIAAL_MAP[mat_lower]
    for en, nl in MATERIAAL_MAP.items():
        if en in mat_lower:
            return nl
    return mat_en.strip().title()


def parse_dimensions(dim_str: str) -> dict:
    """
    Parse dimensies uit formaat "L x W x H" of "L 13,8 W 13,8 H 7".
    Geeft dict met lengte_cm, breedte_cm, hoogte_cm of lege dict.
    """
    if not dim_str:
        return {}

    dim_str = str(dim_str).strip()

    # Formaat: "L 14 W 2 H 0.4 CM" (pricelist formaat)
    m = re.search(r'L\s*([\d.,]+)\s*W\s*([\d.,]+)\s*H\s*([\d.,]+)', dim_str, re.IGNORECASE)
    if m:
        try:
            def parse_num(s):
                return float(s.replace(',', '.'))
            return {
                'lengte_cm': parse_num(m.group(1)),
                'breedte_cm': parse_num(m.group(2)),
                'hoogte_cm': parse_num(m.group(3)),
            }
        except (ValueError, AttributeError):
            pass

    # Formaat: "13.8 x 13.8 x 7" of "13,8 x 13,8 x 7"
    parts = re.split(r'\s*[xX]\s*', dim_str.replace(',', '.'))
    if len(parts) >= 3:
        try:
            nums = [float(p.strip().split()[0]) for p in parts[:3]]
            return {
                'lengte_cm': nums[0],
                'breedte_cm': nums[1],
                'hoogte_cm': nums[2],
            }
        except (ValueError, IndexError):
            pass

    return {}


def bouw_titel(designer: str, dutch_name: str) -> str:
    """Bouw producttitel: Serax - [Designer] - [Title Case van Dutch name]."""
    if not dutch_name:
        return ""
    title_dutch = dutch_name.strip().title()
    if designer and designer.strip():
        return f"Serax - {designer.strip()} - {title_dutch}"
    else:
        return f"Serax - {title_dutch}"


def load_categorie_mapping() -> dict:
    """
    Laad Website indeling en bouw een mapping:
    product_category (Serax) -> (hoofdcategorie, subcategorie, sub_subcategorie)

    De structuur van het bestand:
    - Rij 1: hoofdcategorie (gevuld bij eerste kolom van blok)
    - Rij 2: subcategorie per kolom
    - Rij 3+: sub_subcategorieën per kolom

    We bouwen een dict van alle sub_subcategorie-waarden naar hun categorie-pad.
    """
    path = os.path.join(BASE_DIR, 'Master Files', 'Website indeling (1).xlsx')
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Blad1']

    col_to_hoofdcat = {}
    col_to_subcat = {}
    current_hoofdcat = None

    for col in range(1, ws.max_column + 1):
        val_r1 = ws.cell(row=1, column=col).value
        val_r2 = ws.cell(row=2, column=col).value
        if val_r1:
            current_hoofdcat = val_r1
        col_to_hoofdcat[col] = current_hoofdcat
        if val_r2:
            col_to_subcat[col] = val_r2

    # sub_sub -> (hoofdcat, subcat, sub_sub)
    sub_sub_mapping = {}
    # subcat -> (hoofdcat, subcat)
    subcat_mapping = {}

    for col in range(1, ws.max_column + 1):
        hoofdcat = col_to_hoofdcat.get(col)
        subcat = col_to_subcat.get(col)
        if not (hoofdcat and subcat):
            continue

        subcat_mapping[subcat.lower()] = (hoofdcat, subcat)

        for row in range(3, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                sub_sub_mapping[val.lower()] = (hoofdcat, subcat, val)

    return sub_sub_mapping, subcat_mapping


# Serax Product Category -> website-categorie mapping (handmatig aangevuld)
SERAX_CAT_MAP = {
    # Masterdata col 18 waarden -> (hoofdcategorie, subcategorie, sub_subcategorie)
    "dinnerware": ("servies", "Serviessets", "Design serviezen"),
    "cups & mugs": ("servies", "Kommen, Mokken & Bekers", "Koffiemokken"),
    "cups and mugs": ("servies", "Kommen, Mokken & Bekers", "Koffiemokken"),
    "glasses": ("Glazen", "Water & Thee", "Drinkglazen"),
    "wine glasses": ("Glazen", "Wijn & Champagne", "Wijnglazen"),
    "vases": ("Vazen & Potten", "Vazen", "Kleine vazen"),
    "pots": ("Vazen & Potten", "Potten", "Bloempotten binnen"),
    "bowls": ("servies", "Kommen, Mokken & Bekers", "Kommen"),
    "plates": ("servies", "Borden", "Borden"),
    "cutlery": ("servies", "Bestek", "Besteksets"),
    "knives": ("servies", "Bestek", "Steakmessen"),
    "serving": ("servies", "Serveergoed", "Theepotten"),
    "candles": ("Keuken & Eetkamer", "Tafel & Sfeer", "Dinerkaarsen"),
    "trays": ("Keuken & Eetkamer", "Serveren", "Dienbladen"),
    "kitchen": ("Keuken & Eetkamer", "Keuken & Bereiding", "Ovenschalen"),
    "lighting": ("Wonen & badkamer", "Verlichting & Meubels", "Tafellampen"),
    "interior": ("Wonen & badkamer", "Interieur & Styling", "Vloerkleden"),
    "teapots": ("servies", "Serveergoed", "Theepotten"),
    "carafes": ("Glazen", "Karaffen & Flessen", "Karaffen"),
    "mugs": ("servies", "Kommen, Mokken & Bekers", "Koffiemokken"),
    "espresso cups": ("servies", "Kommen, Mokken & Bekers", "Espressokopjes"),
    "espresso": ("servies", "Kommen, Mokken & Bekers", "Espressokopjes"),
}


def get_categorie(product_category: str, sub_sub_mapping: dict, subcat_mapping: dict) -> tuple:
    """
    Haal categorie-pad op voor een Serax product category.
    Geeft (hoofdcategorie, subcategorie, sub_subcategorie).
    """
    if not product_category:
        return ("", "", "")

    cat_lower = product_category.strip().lower()

    # Directe lookup in onze mapping
    if cat_lower in SERAX_CAT_MAP:
        return SERAX_CAT_MAP[cat_lower]

    # Gedeeltelijke match
    for key, val in SERAX_CAT_MAP.items():
        if key in cat_lower or cat_lower in key:
            return val

    # Sub_sub lookup
    if cat_lower in sub_sub_mapping:
        return sub_sub_mapping[cat_lower]

    # Subcat lookup
    if cat_lower in subcat_mapping:
        h, s = subcat_mapping[cat_lower]
        return (h, s, "")

    return ("", "", "")


def safe_float(val) -> float | None:
    """Converteer naar float, None als niet mogelijk."""
    if val is None or val == "":
        return None
    try:
        return float(str(val).replace(",", "."))
    except (ValueError, TypeError):
        return None


def safe_str(val) -> str:
    """Converteer naar string, leeg als None."""
    if val is None:
        return ""
    return str(val).strip()


def load_masterdata(skus_set: set) -> dict:
    """Laad masterdata voor de 86 SKUs."""
    path = os.path.join(BASE_DIR, 'Master Files', 'Masterdata serax new items_2026_Interieur-Shop (2).xlsx')
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Collectie Serax 2026']

    data = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        sku = safe_str(row[0])
        if not sku or sku not in skus_set:
            continue

        data[sku] = {
            'sku': sku,
            'dutch_name': safe_str(row[2]),       # Col 3
            'english_name': safe_str(row[3]),      # Col 4
            'brand': safe_str(row[6]),              # Col 7
            'giftbox': safe_str(row[8]),            # Col 9
            'giftbox_qty': safe_str(row[9]),        # Col 10
            'ean_piece': safe_str(row[10]),         # Col 11
            'ean_shopify': safe_str(row[11]),       # Col 12
            'kleur_en': safe_str(row[13]),          # Col 14
            'designer': safe_str(row[14]),          # Col 15
            'product_category': safe_str(row[17]), # Col 18
            'collectie': safe_str(row[18]),         # Col 19
            'materiaal_en': safe_str(row[20]),      # Col 21
        }

    return data


def load_pricelist(skus_set: set) -> dict:
    """Laad pricelist voor de 86 SKUs."""
    path = os.path.join(BASE_DIR, 'Master Files', 'SERAX_Price Increase April 2026 (EUR1) (2).xlsx')
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Sheet1']

    data = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        sku = safe_str(row[0])
        if not sku or sku not in skus_set:
            continue

        data[sku] = {
            'sku': sku,
            'materiaal_en_pl': safe_str(row[10]),        # Col 11
            'dimensions': safe_str(row[11]),              # Col 12
            'ean_piece_pl': safe_str(row[14]),            # Col 15
            'inkoop_stuk': safe_float(row[16]),           # Col 17
            'rrp_stuk': safe_float(row[17]),              # Col 18
            'ean_giftbox_pl': safe_str(row[23]),          # Col 24
            'giftbox_qty_pl': safe_str(row[25]),          # Col 26
            'inkoop_gb': safe_float(row[26]),             # Col 27
            'rrp_gb': safe_float(row[27]),                # Col 28
        }

    return data


def bereken_prijzen(md_row: dict, pl_row: dict) -> dict:
    """
    Bereken verkoopprijs en inkoopprijs op basis van giftbox-logica.
    Regel: Als giftbox=Yes EN qty>1 => giftbox prijs, anders stuksprijs.
    """
    giftbox = md_row.get('giftbox', '').strip().lower()
    giftbox_qty_str = md_row.get('giftbox_qty', '').strip()

    try:
        giftbox_qty = int(float(giftbox_qty_str)) if giftbox_qty_str else 0
    except (ValueError, TypeError):
        giftbox_qty = 0

    is_giftbox = (giftbox == 'yes') and (giftbox_qty > 1)

    if is_giftbox:
        verkoopprijs = pl_row.get('rrp_gb')
        inkoopprijs = pl_row.get('inkoop_gb')
    else:
        verkoopprijs = pl_row.get('rrp_stuk')
        inkoopprijs = pl_row.get('inkoop_stuk')

    return {
        'verkoopprijs': verkoopprijs,
        'inkoopprijs': inkoopprijs,
        'rrp_stuk_eur': pl_row.get('rrp_stuk'),
        'rrp_gb_eur': pl_row.get('rrp_gb'),
        'inkoopprijs_stuk_eur': pl_row.get('inkoop_stuk'),
        'inkoopprijs_gb_eur': pl_row.get('inkoop_gb'),
        'is_giftbox': is_giftbox,
    }


def main():
    print("=" * 60)
    print("Serax 86 Producten Herverwerking")
    print("=" * 60)

    # Verbinding met Supabase
    print("\n[1/6] Verbinding Supabase...")
    sb = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)
    print("OK")

    # Laad 86 SKUs
    print("\n[2/6] Laden 86 SKUs met afwijkingen...")
    path_check = os.path.join(BASE_DIR, 'exports', 'Serax_MD_vs_Supabase_check.xlsx')
    wb_check = openpyxl.load_workbook(path_check, data_only=True)
    ws_check = wb_check['MD vs Supabase']

    skus_86 = []
    for row in ws_check.iter_rows(min_row=2, values_only=True):
        sku = safe_str(row[0])
        afwijkingen = row[16] if len(row) > 16 else None
        if sku and afwijkingen:
            skus_86.append(sku)

    skus_set = set(skus_86)
    print(f"Gevonden: {len(skus_86)} SKUs")

    # Laad categorie mapping
    print("\n[3/6] Laden categorie mapping...")
    sub_sub_mapping, subcat_mapping = load_categorie_mapping()
    print(f"OK ({len(sub_sub_mapping)} sub_subcategorieën, {len(subcat_mapping)} subcategorieën)")

    # Laad masterdata
    print("\n[4/6] Laden masterdata...")
    md_data = load_masterdata(skus_set)
    print(f"OK ({len(md_data)} SKUs geladen)")

    # Laad pricelist
    print("\n[5/6] Laden pricelist...")
    pl_data = load_pricelist(skus_set)
    print(f"OK ({len(pl_data)} SKUs geladen)")

    # Verwerk en update
    print("\n[6/6] Verwerken en updaten Supabase...")
    print("-" * 60)

    log_rows = []
    succes = 0
    gefaald = 0

    for sku in skus_86:
        fouten = []

        # Haal data op
        md = md_data.get(sku)
        pl = pl_data.get(sku)

        if not md:
            print(f"  [SKIP] {sku}: NIET in masterdata")
            fouten.append("Niet in masterdata")
            log_rows.append({
                'SKU': sku,
                'Actie': 'SKIP',
                'Titel_nieuw': '',
                'EAN_shopify_nieuw': '',
                'Designer_nieuw': '',
                'Verkoopprijs_nieuw': '',
                'Fouten': '; '.join(fouten),
            })
            gefaald += 1
            continue

        if not pl:
            print(f"  [SKIP] {sku}: NIET in pricelist")
            fouten.append("Niet in pricelist")
            log_rows.append({
                'SKU': sku,
                'Actie': 'SKIP',
                'Titel_nieuw': '',
                'EAN_shopify_nieuw': '',
                'Designer_nieuw': '',
                'Verkoopprijs_nieuw': '',
                'Fouten': '; '.join(fouten),
            })
            gefaald += 1
            continue

        # Bouw producttitel
        designer = md.get('designer', '')
        dutch_name = md.get('dutch_name', '')
        titel = bouw_titel(designer, dutch_name)

        # EAN velden
        ean_piece = md.get('ean_piece', '') or pl.get('ean_piece_pl', '')
        ean_shopify = md.get('ean_shopify', '') or pl.get('ean_giftbox_pl', '')

        # Als ean_shopify leeg is, gebruik dan ean_piece
        if not ean_shopify:
            ean_shopify = ean_piece

        # Kleur
        kleur_en = md.get('kleur_en', '')
        kleur_nl = vertaal_kleur(kleur_en)

        # Materiaal
        materiaal_en = md.get('materiaal_en', '') or pl.get('materiaal_en_pl', '')
        materiaal_nl = vertaal_materiaal(materiaal_en)

        # Giftbox info
        giftbox_val = md.get('giftbox', '')
        giftbox_qty_str = md.get('giftbox_qty', '')
        try:
            giftbox_qty = int(float(giftbox_qty_str)) if giftbox_qty_str else None
        except (ValueError, TypeError):
            giftbox_qty = None

        # Prijzen
        prijs = bereken_prijzen(md, pl)

        # Collectie
        collectie = md.get('collectie', '')

        # Categorie
        product_category = md.get('product_category', '')
        hoofdcat, subcat, sub_subcat = get_categorie(product_category, sub_sub_mapping, subcat_mapping)

        # Dimensies
        dims = parse_dimensions(pl.get('dimensions', ''))

        # Stel update dict op
        update = {
            'product_title_nl': titel,
            'designer': designer if designer else None,
            'ean_shopify': ean_shopify if ean_shopify else None,
            'ean_piece': ean_piece if ean_piece else None,
            'giftbox': giftbox_val if giftbox_val else None,
            'giftbox_qty': str(giftbox_qty) if giftbox_qty is not None else None,
            'kleur_nl': kleur_nl if kleur_nl else None,
            'kleur_en': kleur_en if kleur_en else None,
            'materiaal_nl': materiaal_nl if materiaal_nl else None,
            'collectie': collectie if collectie else None,
            'verkoopprijs': prijs['verkoopprijs'],
            'inkoopprijs': prijs['inkoopprijs'],
            'rrp_stuk_eur': prijs['rrp_stuk_eur'],
            'rrp_gb_eur': prijs['rrp_gb_eur'],
            'inkoopprijs_stuk_eur': prijs['inkoopprijs_stuk_eur'],
            'inkoopprijs_gb_eur': prijs['inkoopprijs_gb_eur'],
        }

        # Voeg dimensies toe als aanwezig
        if dims:
            update['hoogte_cm'] = dims.get('hoogte_cm')
            update['lengte_cm'] = dims.get('lengte_cm')
            update['breedte_cm'] = dims.get('breedte_cm')

        # Voeg categorie toe als aanwezig
        if hoofdcat:
            update['hoofdcategorie'] = hoofdcat
        if subcat:
            update['subcategorie'] = subcat
        if sub_subcat:
            update['sub_subcategorie'] = sub_subcat

        # Verwijder None-waarden om Supabase niet te overriden met NULL
        update_clean = {k: v for k, v in update.items() if v is not None}

        # Print wat er geupdate wordt
        print(f"\n  [{sku}]")
        print(f"    Titel:       {titel}")
        print(f"    EAN shopify: {ean_shopify}")
        print(f"    EAN piece:   {ean_piece}")
        print(f"    Designer:    {designer}")
        print(f"    Kleur:       {kleur_en} -> {kleur_nl}")
        print(f"    Materiaal:   {materiaal_en} -> {materiaal_nl}")
        print(f"    Giftbox:     {giftbox_val} qty={giftbox_qty} (is_gb={prijs['is_giftbox']})")
        print(f"    Prijs:       RRP={prijs['verkoopprijs']} Inkoop={prijs['inkoopprijs']}")
        print(f"    Collectie:   {collectie}")
        print(f"    Categorie:   {hoofdcat} / {subcat} / {sub_subcat}")
        if dims:
            print(f"    Dimensies:   L={dims.get('lengte_cm')} W={dims.get('breedte_cm')} H={dims.get('hoogte_cm')}")

        # Update Supabase
        try:
            result = sb.table("seo_products").update(update_clean).eq("sku", sku).execute()

            if result.data:
                print(f"    -> UPDATE OK ({len(result.data)} rij(en))")
                actie = "UPDATE_OK"
                succes += 1
            else:
                print(f"    -> WAARSCHUWING: 0 rijen geupdate (SKU niet in DB?)")
                actie = "UPDATE_0_ROWS"
                fouten.append("0 rijen geupdate - SKU mogelijk niet in DB")
                gefaald += 1
        except Exception as e:
            print(f"    -> FOUT: {e}")
            actie = "FOUT"
            fouten.append(str(e))
            gefaald += 1

        log_rows.append({
            'SKU': sku,
            'Actie': actie,
            'Titel_nieuw': titel,
            'EAN_shopify_nieuw': ean_shopify,
            'Designer_nieuw': designer,
            'Verkoopprijs_nieuw': prijs['verkoopprijs'],
            'Fouten': '; '.join(fouten) if fouten else '',
        })

    # Sla log op
    print("\n" + "=" * 60)
    print("LOG OPSLAAN...")
    log_path = os.path.join(BASE_DIR, 'exports', 'Serax_86_correctie_log.xlsx')
    df_log = pd.DataFrame(log_rows)
    df_log.to_excel(log_path, index=False)
    print(f"Log opgeslagen: {log_path}")

    # Samenvatting
    print("\n" + "=" * 60)
    print("SAMENVATTING")
    print("=" * 60)
    print(f"  Totaal verwerkt:  {len(skus_86)}")
    print(f"  Succesvol:        {succes}")
    print(f"  Gefaald/gewaarschuwd: {gefaald}")
    print("=" * 60)


if __name__ == "__main__":
    main()
