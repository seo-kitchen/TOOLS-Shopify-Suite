"""
meta_audit_export.py — Exporteer meta-audit suggesties naar Excel voor review.

Schrijft een side-by-side Excel (.tmp/meta_audit_review.xlsx) met huidige vs
voorstel title + description, lengtes, vendor en status.

Gebruik:
    python execution/meta_audit_export.py                      # alle producten met suggested_*
    python execution/meta_audit_export.py --vendor "Pottery Pots"
    python execution/meta_audit_export.py --only-suggested     # alleen rijen met voorstellen
"""

import argparse
import os
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_KEY") or os.getenv("SUPABASE_KEY")


def get_supabase():
    from supabase import create_client
    return create_client(SUPABASE_URL, SUPABASE_KEY)


def fetch_rows(sb, vendor: str | None, only_suggested: bool) -> list[dict]:
    q = sb.table("shopify_meta_audit").select("*")
    if vendor:
        q = q.eq("vendor", vendor)
    if only_suggested:
        q = q.not_.is_("suggested_meta_title", "null")
    # pagination
    rows, offset = [], 0
    page = 1000
    while True:
        batch = q.range(offset, offset + page - 1).execute().data
        if not batch:
            break
        rows.extend(batch)
        if len(batch) < page:
            break
        offset += page
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--vendor")
    ap.add_argument("--only-suggested", action="store_true", default=True)
    ap.add_argument("--out", default=".tmp/meta_audit_review.xlsx")
    args = ap.parse_args()

    import pandas as pd
    sb = get_supabase()
    rows = fetch_rows(sb, args.vendor, args.only_suggested)
    print(f"{len(rows)} rijen opgehaald")

    if not rows:
        print("Geen rijen. Heb je de generator al met --write gedraaid?")
        return

    df = pd.DataFrame(rows)
    cols = [
        "vendor", "product_title", "handle",
        "current_meta_title", "current_title_length", "title_status",
        "suggested_meta_title", "suggested_title_length",
        "current_meta_description", "current_desc_length", "desc_status",
        "suggested_meta_description", "suggested_desc_length",
        "review_status", "shopify_product_id",
    ]
    cols = [c for c in cols if c in df.columns]
    df = df[cols].sort_values(["vendor", "product_title"])

    Path(args.out).parent.mkdir(exist_ok=True)
    with pd.ExcelWriter(args.out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Review")
        ws = writer.sheets["Review"]
        # Kolombreedtes
        widths = {
            "vendor": 14, "product_title": 45, "handle": 40,
            "current_meta_title": 50, "suggested_meta_title": 55,
            "current_meta_description": 60, "suggested_meta_description": 60,
        }
        for idx, col in enumerate(df.columns, start=1):
            letter = ws.cell(row=1, column=idx).column_letter
            ws.column_dimensions[letter].width = widths.get(col, 15)
        # Wrap tekst voor lange kolommen
        from openpyxl.styles import Alignment
        wrap_cols = {"current_meta_title", "suggested_meta_title",
                     "current_meta_description", "suggested_meta_description",
                     "product_title"}
        for idx, col in enumerate(df.columns, start=1):
            if col in wrap_cols:
                letter = ws.cell(row=1, column=idx).column_letter
                for cell in ws[letter][1:]:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        # Hoogte voor data rijen
        for r in range(2, len(df) + 2):
            ws.row_dimensions[r].height = 60

    print(f"Geschreven: {args.out}")


if __name__ == "__main__":
    main()
