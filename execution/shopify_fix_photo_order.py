# -*- coding: utf-8 -*-
"""
Fix foto-volgorde in Shopify via Admin API.
Strategie: check per product of packshot_1 als EERSTE foto staat.
Als niet: verwijder alle media, voeg opnieuw toe in juiste volgorde.
Veiligheid: alleen producten aanpassen waar WIJ fotos in Supabase hebben.

Gebruik:
    python execution/shopify_fix_photo_order.py --dry-run     # alleen checken
    python execution/shopify_fix_photo_order.py --limit 3     # test 3 producten
    python execution/shopify_fix_photo_order.py               # alle producten
"""

import argparse, os, sys, time
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))
from dotenv import load_dotenv
load_dotenv(Path(__file__).parent.parent / ".env")

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

import requests
from supabase import create_client

STORE   = os.getenv("SHOPIFY_STORE")
TOKEN   = os.getenv("SHOPIFY_ACCESS_TOKEN")
API_VER = "2024-01"
BASE    = f"https://{STORE}/admin/api/{API_VER}"
HEADERS = {"X-Shopify-Access-Token": TOKEN, "Content-Type": "application/json"}

PHOTO_COLS = [
    "photo_packshot_1","photo_packshot_2","photo_packshot_3",
    "photo_packshot_4","photo_packshot_5",
    "photo_lifestyle_1","photo_lifestyle_2","photo_lifestyle_3",
    "photo_lifestyle_4","photo_lifestyle_5",
]

def get_supabase():
    return create_client(os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_SERVICE_KEY"))

def shopify_get(path):
    r = requests.get(f"{BASE}{path}", headers=HEADERS, timeout=15)
    r.raise_for_status()
    return r.json()

def shopify_delete(path):
    r = requests.delete(f"{BASE}{path}", headers=HEADERS, timeout=15)
    r.raise_for_status()

def shopify_post(path, data):
    r = requests.post(f"{BASE}{path}", headers=HEADERS, json=data, timeout=30)
    r.raise_for_status()
    return r.json()

def get_product_media(product_id):
    data = shopify_get(f"/products/{product_id}/media.json")
    return data.get("media", [])

def is_correct_order(media_list, our_packshot_url):
    """Check of eerste Shopify-foto overeenkomt met onze packshot_1."""
    if not media_list:
        return False
    first = media_list[0]
    first_src = first.get("src","") or first.get("preview_image",{}).get("src","")
    # Vergelijk bestandsnaam
    our_fname  = our_packshot_url.split("/")[-1].split("?")[0].lower()
    shop_fname = first_src.split("/")[-1].split("?")[0].lower()
    return our_fname == shop_fname or our_packshot_url in first_src

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--limit", type=int)
    ap.add_argument("--sku")
    args = ap.parse_args()

    sb = get_supabase()

    # Haal producten op met Supabase-fotos en een shopify_product_id
    # Product ID map uit fase4 lijst (Supabase heeft geen shopify_product_id)
    pid_map = {}
    fase4_path = Path(__file__).parent.parent / ".tmp" / "aa5c1f0e-5baa-41bc-938d-76c6b3bf4fdd.xlsx"
    if fase4_path.exists():
        import openpyxl
        wb = openpyxl.load_workbook(str(fase4_path))
        ws_f = wb.active
        h_f = [c.value for c in ws_f[1]]
        for row_f in range(2, ws_f.max_row+1):
            sku_v = str(ws_f.cell(row_f, h_f.index("Variant SKU")+1).value or "").strip()
            pid_v = str(ws_f.cell(row_f, h_f.index("Product ID")+1).value or "").strip()
            if sku_v and pid_v:
                pid_map[sku_v] = pid_v
        print(f"Product ID map geladen: {len(pid_map)} SKUs")

    if args.sku:
        r = sb.table("seo_products").select("sku," + ",".join(PHOTO_COLS)).eq("sku", args.sku).execute()
    else:
        rows_all = []
        for prefix in ["B","V"]:
            r = (sb.table("seo_products")
                   .select("sku," + ",".join(PHOTO_COLS))
                   .like("sku", f"{prefix}%")
                   .not_.is_("photo_packshot_1","null")
                   .limit(1000)
                   .execute())
            rows_all.extend(r.data or [])
        r = type("R",(),{"data": [row for row in rows_all
                                   if "supabase.co/storage" in (row.get("photo_packshot_1","") or "")
                                   and row["sku"] in pid_map]})()

    rows = r.data or []
    if args.limit:
        rows = rows[:args.limit]

    print(f"Te controleren: {len(rows)} producten{'  (DRY RUN)' if args.dry_run else ''}\n")

    ok = already_correct = fixed = errors = 0

    for i, row in enumerate(rows, 1):
        sku       = row["sku"]
        product_id= pid_map.get(sku, row.get("shopify_product_id",""))
        if not product_id:
            print(f"  [{i}] {sku}: geen shopify_product_id — skip")
            continue

        # Onze fotos in volgorde
        our_urls = [row[c] for c in PHOTO_COLS if row.get(c)]
        if not our_urls:
            continue

        try:
            media = get_product_media(product_id)
            time.sleep(0.3)  # rate limit

            if is_correct_order(media, our_urls[0]):
                already_correct += 1
                print(f"  [{i}/{len(rows)}] {sku}: OK ({len(media)} fotos, volgorde klopt)")
                continue

            print(f"  [{i}/{len(rows)}] {sku}: FOUT ({len(media)} fotos in Shopify, packshot staat niet eerst)")

            if args.dry_run:
                print(f"    [DRY RUN] zou {len(media)} media verwijderen en {len(our_urls)} opnieuw toevoegen")
                continue

            # Stap 1: Verwijder alle bestaande media
            for m in media:
                mid = m.get("id")
                if mid:
                    shopify_delete(f"/products/{product_id}/media/{mid}.json")
                    time.sleep(0.15)

            time.sleep(0.5)

            # Stap 2: Voeg toe in juiste volgorde
            for url in our_urls:
                shopify_post(f"/products/{product_id}/media.json", {
                    "media": {"media_content_type": "IMAGE", "src": url}
                })
                time.sleep(0.3)

            fixed += 1
            print(f"    Gereset: {len(media)} verwijderd, {len(our_urls)} toegevoegd in juiste volgorde")

        except Exception as e:
            errors += 1
            print(f"  [{i}] {sku}: FOUT — {e}")

        if i % 20 == 0:
            time.sleep(2)

    print(f"\n{'='*50}")
    print(f"Al correct:  {already_correct}")
    print(f"Gefixed:     {fixed}")
    print(f"Fouten:      {errors}")

if __name__ == "__main__":
    main()
