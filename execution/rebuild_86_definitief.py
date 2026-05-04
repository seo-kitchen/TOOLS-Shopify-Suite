# -*- coding: utf-8 -*-
import sys, os, re, time
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.path.insert(0, str(__file__).replace("\\execution\\rebuild_86_definitief.py", "\\execution"))
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))
from supabase import create_client
from meta_audit_generate import get_anthropic, make_title, generate_desc, fetch_context
import openpyxl
from pathlib import Path
from datetime import datetime

BASE = Path(__file__).parent.parent

sb = create_client(os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_SERVICE_KEY"))
client = get_anthropic()

CATEGORY_MAP = [
    (["champagneglas", "champagneglazen", "coupe"],        "Glazen", "Wijn & Champagne", "Champagneglazen"),
    (["wijnglas", "wine glass"],                            "Glazen", "Wijn & Champagne", "Wijnglazen"),
    (["karaf"],                                             "Glazen", "Karaffen & Flessen", "Karaffen"),
    (["glas"],                                              "Glazen", "Water & Thee", "Drinkglazen"),
    (["mueslikom"],                                         "Servies", "Kommen, Mokken & Bekers", "Kommen"),
    (["serveerkom", "serveerschaal"],                       "Servies", "Schalen", "Serveerschalen"),
    (["serveerbord"],                                       "Servies", "Schalen", "Serveerschalen"),
    (["gebakschaal"],                                       "Servies", "Schalen", "Gebakschalen"),
    (["saladeschaal"],                                      "Servies", "Schalen", "Saladeschalen"),
    (["schaal"],                                            "Servies", "Schalen", "Serveerschalen"),
    (["theepot"],                                           "Servies", "Serveergoed", "Theepotten"),
    (["suikerpot"],                                         "Servies", "Serveergoed", "Suikerpotten"),
    (["melkkan"],                                           "Servies", "Serveergoed", "Melkkannen"),
    (["kan "],                                              "Servies", "Serveergoed", "Melkkannen"),
    (["espressokopje", "kop + schotel", "kopje + schotel"],"Servies", "Kommen, Mokken & Bekers", "Espressokopjes"),
    (["koffiekopje", "cappuccinokopje", "theekopje"],       "Servies", "Kommen, Mokken & Bekers", "Koffiemokken"),
    (["kom"],                                               "Servies", "Kommen, Mokken & Bekers", "Kommen"),
    (["ontbijtbord"],                                       "Servies", "Borden", "Ontbijtborden"),
    (["dinerbord", "dinnerbord"],                           "Servies", "Borden", "Dinerborden"),
    (["pastabord"],                                         "Servies", "Borden", "Pastaborden"),
    (["dessertbord"],                                       "Servies", "Borden", "Dessertbordjes"),
    (["bord"],                                              "Servies", "Borden", "Dinerborden"),
    (["dinner set", "serviesset", "set 12", "set 16"],      "Servies", "Serviessets", "Dinersets"),
    (["vaas"],                                              "Vazen & Potten", "Vazen", "Design vazen"),
    (["kandelaar"],                                         "Keuken & Eetkamer", "Tafel & Sfeer", "Kandelaars"),
    (["voorraadpot"],                                       "Keuken & Eetkamer", "Keukenorganisatie", "Voorraadpotten"),
    (["tafellamp", "tafel lamp"],                           "Wonen & badkamer", "Verlichting & Meubels", "Tafellampen"),
    (["staande lamp"],                                      "Wonen & badkamer", "Verlichting & Meubels", "Vloerlampen"),
    (["wandlamp"],                                          "Wonen & badkamer", "Verlichting & Meubels", "Wandlampen"),
    (["bank "],                                             "Wonen & badkamer", "Verlichting & Meubels", "Stoelen"),
]

def get_category(title):
    t = title.lower()
    parts = t.split(" - ")
    core = parts[2] if len(parts) >= 3 else t
    for keywords, hcat, sub, subsub in CATEGORY_MAP:
        if any(k in core for k in keywords):
            return hcat, sub, subsub
    return "Servies", "Borden", "Dinerborden"

# Handle map
wb_ap = openpyxl.load_workbook(str(BASE / "Master Files" / "Alle Active Producten.xlsx"))
ws_ap = wb_ap.active
h_ap = [c.value for c in ws_ap[1]]
handle_map = {}
for r in range(2, ws_ap.max_row + 1):
    sku = str(ws_ap.cell(r, h_ap.index("Variant SKU") + 1).value or "")
    handle = ws_ap.cell(r, h_ap.index("Product handle") + 1).value
    if sku and handle:
        handle_map[sku] = handle

def make_handle(title):
    h = title.lower()
    h = re.sub(r'[^a-z0-9]+', '-', h)
    return h.strip('-')

# Load 86 SKUs
wb_check = openpyxl.load_workbook(str(BASE / "exports" / "Serax_MD_vs_Supabase_check.xlsx"))
ws_check = wb_check["MD vs Supabase"]
h_c = [c.value for c in ws_check[1]]
skus_86 = [
    str(ws_check.cell(r, h_c.index("SKU") + 1).value or "")
    for r in range(2, ws_check.max_row + 1)
    if ws_check.cell(r, h_c.index("Afwijkingen") + 1).value
]

fields = "sku,product_title_nl,ean_shopify,ean_piece,verkoopprijs,inkoopprijs,designer,kleur_nl,materiaal_nl,collectie,giftbox,giftbox_qty,hoogte_cm,lengte_cm,breedte_cm,fase,handle,shopify_product_id"
supa_rows = {}
for i in range(0, len(skus_86), 50):
    r = sb.table("seo_products").select(fields).in_("sku", skus_86[i:i+50]).execute()
    for row in (r.data or []):
        supa_rows[row["sku"]] = row

out_headers = [
    "Variant SKU", "Product ID", "Variant ID", "Product handle", "Product title",
    "Product vendor", "Product type", "EAN Code", "Verkoopprijs Shopify",
    "Inkoopprijs Shopify", "Product description", "Nieuwe hoofdcategorie",
    "Nieuwe subcategorie", "Nieuwe sub-subcategorie", "Nieuwe tag", "collectie",
    "designer", "materiaal", "kleur", "hoogte_cm", "lengte_cm", "breedte_cm",
    "meta_title", "meta_description"
]

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Shopify_Nieuw"
ws_out.append(out_headers)

total = len(skus_86)
for i, sku in enumerate(skus_86, 1):
    row = supa_rows.get(sku)
    if not row:
        print(f"  [{i}/{total}] {sku}: NIET IN SUPABASE")
        continue

    title = row.get("product_title_nl") or ""
    vendor = "Serax"
    handle = handle_map.get(sku) or make_handle(title)
    hcat, sub, subsub = get_category(title)

    cat_tag = hcat.lower().replace(" & ", "_en_").replace(" ", "_")
    sub_tag = sub.lower().replace(" & ", "_en_").replace(", ", "_").replace(" ", "_")
    subsub_tag = subsub.lower().replace(" ", "_")
    tags = f"cat_{cat_tag},cat_{sub_tag},cat_{subsub_tag},structuur_fase4"

    try:
        ctx = fetch_context(sb, {
            "shopify_product_id": str(row.get("shopify_product_id") or ""),
            "handle": handle,
        })
        desc_pid = str(row.get("shopify_product_id") or sku)
        meta_t, _ = make_title(client, title, vendor)
        meta_d = generate_desc(client, title, ctx, desc_pid)
        print(f"  [{i}/{total}] {sku} | {hcat} > {sub} > {subsub} | {len(meta_t)}ch / {len(meta_d)}ch")
    except Exception as e:
        meta_t, meta_d = "", ""
        print(f"  [{i}/{total}] {sku}: FOUT — {e}")

    ws_out.append([
        sku, row.get("shopify_product_id") or "", "", handle, title,
        vendor, hcat,
        row.get("ean_shopify") or "",
        row.get("verkoopprijs") or "",
        row.get("inkoopprijs") or "",
        "",
        hcat, sub, subsub, tags,
        row.get("collectie") or "",
        row.get("designer") or "",
        row.get("materiaal_nl") or "",
        row.get("kleur_nl") or "",
        row.get("hoogte_cm") or "",
        row.get("lengte_cm") or "",
        row.get("breedte_cm") or "",
        meta_t, meta_d,
    ])

    if i % 10 == 0:
        time.sleep(0.5)

ts = datetime.now().strftime("%Y%m%d_%H%M")
out_path = BASE / "exports" / f"Serax_86_DEFINITIEF_v2_{ts}.xlsx"
wb_out.save(str(out_path))
print(f"\nKlaar: {out_path} ({ws_out.max_row - 1} producten)")
