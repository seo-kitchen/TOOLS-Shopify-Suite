"""
Serax Masterdata vs Supabase vergelijking
Vergelijkt alle Serax new items masterdata met wat er in Supabase staat.
Output: exports/Serax_MD_vs_Supabase_check.xlsx
"""

import os
import sys
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv
from supabase import create_client

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_KEY")

MD_PATH = "Master Files/Masterdata serax new items_2026_Interieur-Shop (2).xlsx"
MD_SHEET = "Collectie Serax 2026"
OUTPUT_PATH = "exports/Serax_MD_vs_Supabase_check.xlsx"

# Kolom indices (0-based) in masterdata
COL_SKU = 0       # col 1
COL_NL_NAME = 2   # col 3
COL_EN_NAME = 3   # col 4
COL_BRAND = 6     # col 7
COL_GIFTBOX = 8   # col 9
COL_GIFTBOX_QTY = 9  # col 10
COL_EAN_PIECE = 10    # col 11
COL_EAN_SHOPIFY = 11  # col 12
COL_COLOR = 13    # col 14
COL_DESIGNER = 14 # col 15
COL_ITEM_CAT = 15 # col 16
COL_PROD_CAT = 17 # col 18 (Product Category / collectie)
COL_MATERIAL = 20 # col 21


def normalize(val):
    """Normalize waarden voor vergelijking: strip, lowercase, None → ''"""
    if val is None:
        return ""
    return str(val).strip()


def normalize_ean(val):
    """EAN normalisatie: verwijder decimalen, strip"""
    if val is None:
        return ""
    s = str(val).strip()
    # Verwijder .0 suffix als het een float was
    if s.endswith(".0"):
        s = s[:-2]
    return s


def normalize_giftbox(val):
    """Giftbox ja/nee → 'true'/'false' voor vergelijking met Supabase boolean"""
    if val is None:
        return ""
    s = str(val).strip().lower()
    if s in ("yes", "ja", "true", "1"):
        return "true"
    if s in ("no", "nee", "false", "0"):
        return "false"
    return s


def normalize_giftbox_qty(val):
    """Giftbox qty normalisatie: int vergelijking"""
    if val is None:
        return ""
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def normalize_giftbox_db(val):
    """Normaliseer giftbox waarde uit Supabase (kan string 'Yes'/'No' of bool zijn)"""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "true" if val else "false"
    s = str(val).strip().lower()
    if s in ("yes", "ja", "true", "1"):
        return "true"
    if s in ("no", "nee", "false", "0"):
        return "false"
    return s


def compare_fields(md_row, db_row):
    """
    Vergelijk velden. Geeft lijst van afwijkende veldnamen terug.
    Let op: product_title_nl en kleur worden NIET vergeleken in Afwijkingen
    omdat Supabase gecureerde NL-titels heeft vs raw MD namen (EN kleur vs NL kleur).
    Deze worden wel getoond ter informatie.
    Echte afwijkingen: ean_shopify, ean_piece, designer, giftbox, giftbox_qty.
    """
    differences = []

    # ean_shopify vs col12
    md_ean_shopify = normalize_ean(md_row["ean_shopify"])
    db_ean_shopify = normalize_ean(db_row.get("ean_shopify"))
    if md_ean_shopify != db_ean_shopify:
        differences.append("ean_shopify")

    # ean_piece vs col11
    md_ean_piece = normalize_ean(md_row["ean_piece"])
    db_ean_piece = normalize_ean(db_row.get("ean_piece"))
    if md_ean_piece != db_ean_piece:
        differences.append("ean_piece")

    # designer vs col15 (echte mismatch)
    md_designer = normalize(md_row["designer"])
    db_designer = normalize(db_row.get("designer"))
    if md_designer.lower() != db_designer.lower():
        differences.append("designer")

    # giftbox vs col9
    md_giftbox = normalize_giftbox(md_row["giftbox"])
    db_giftbox = normalize_giftbox_db(db_row.get("giftbox"))
    if md_giftbox != db_giftbox:
        differences.append("giftbox")

    # giftbox_qty vs col10
    md_qty = normalize_giftbox_qty(md_row["giftbox_qty"])
    db_qty_raw = db_row.get("giftbox_qty")
    db_qty = normalize_giftbox_qty(db_qty_raw)
    if md_qty != db_qty:
        differences.append("giftbox_qty")

    return differences


def main():
    print("=== Serax MD vs Supabase vergelijking ===\n")

    # 1. Laad masterdata
    print(f"Masterdata laden: {MD_PATH}")
    wb = openpyxl.load_workbook(MD_PATH, read_only=True, data_only=True)
    ws = wb[MD_SHEET]

    md_rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        sku = row[COL_SKU]
        if not sku:
            continue
        md_rows.append({
            "sku": str(sku).strip(),
            "naam_nl": row[COL_NL_NAME],
            "naam_en": row[COL_EN_NAME],
            "giftbox": row[COL_GIFTBOX],
            "giftbox_qty": row[COL_GIFTBOX_QTY],
            "ean_piece": row[COL_EAN_PIECE],
            "ean_shopify": row[COL_EAN_SHOPIFY],
            "kleur": row[COL_COLOR],
            "designer": row[COL_DESIGNER],
        })

    print(f"  {len(md_rows)} SKUs gevonden in masterdata\n")

    # 2. Haal alle Serax producten op uit Supabase via SKU-lijst
    print("Supabase verbinding maken...")
    client = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Verzamel alle MD SKUs
    md_skus = [r["sku"] for r in md_rows]

    # Haal in batches van 200 op via SKU IN-filter
    all_db_rows = []
    batch_size = 200
    for i in range(0, len(md_skus), batch_size):
        batch_skus = md_skus[i:i + batch_size]
        response = client.table("seo_products").select(
            "sku, product_title_nl, ean_shopify, ean_piece, kleur_nl, designer, "
            "materiaal_nl, hoofdcategorie, collectie, giftbox, giftbox_qty"
        ).in_("sku", batch_skus).execute()

        batch = response.data
        all_db_rows.extend(batch)
        print(f"  Batch {i//batch_size + 1}: {len(batch)} rijen opgehaald ({i+len(batch_skus)}/{len(md_skus)})")

    # Maak SKU lookup dict
    db_by_sku = {row["sku"]: row for row in all_db_rows}
    print(f"\n  {len(all_db_rows)} Serax producten gevonden in Supabase (van {len(md_skus)} MD SKUs)\n")

    # 3. Vergelijking uitvoeren
    print("Vergelijking uitvoeren...")
    results = []
    stats = {
        "gevonden": 0,
        "niet_gevonden": 0,
        "met_afwijkingen": 0,
        "fields": {
            "ean_shopify": 0,
            "ean_piece": 0,
            "designer": 0,
            "giftbox": 0,
            "giftbox_qty": 0,
        }
    }

    for md in md_rows:
        sku = md["sku"]
        db = db_by_sku.get(sku)

        if db is None:
            stats["niet_gevonden"] += 1
            results.append({
                "sku": sku,
                "in_supabase": "nee",
                "naam_md": normalize(md["naam_nl"]),
                "naam_db": "",
                "ean_shopify_md": normalize_ean(md["ean_shopify"]),
                "ean_shopify_db": "",
                "ean_piece_md": normalize_ean(md["ean_piece"]),
                "ean_piece_db": "",
                "designer_md": normalize(md["designer"]),
                "designer_db": "",
                "kleur_md": normalize(md["kleur"]),
                "kleur_db": "",
                "giftbox_md": normalize_giftbox(md["giftbox"]),
                "giftbox_db": "",
                "giftbox_qty_md": normalize_giftbox_qty(md["giftbox_qty"]),
                "giftbox_qty_db": "",
                "afwijkingen": "NIET GEVONDEN IN SUPABASE",
            })
        else:
            stats["gevonden"] += 1
            diffs = compare_fields(md, db)

            for field in diffs:
                if field in stats["fields"]:
                    stats["fields"][field] += 1

            if diffs:
                stats["met_afwijkingen"] += 1

            # DB giftbox formatting
            db_giftbox_raw = db.get("giftbox")
            if db_giftbox_raw is None:
                db_giftbox_display = ""
            elif isinstance(db_giftbox_raw, bool):
                db_giftbox_display = "true" if db_giftbox_raw else "false"
            else:
                db_giftbox_display = normalize(db_giftbox_raw)

            results.append({
                "sku": sku,
                "in_supabase": "ja",
                "naam_md": normalize(md["naam_nl"]),
                "naam_db": normalize(db.get("product_title_nl")),
                "ean_shopify_md": normalize_ean(md["ean_shopify"]),
                "ean_shopify_db": normalize_ean(db.get("ean_shopify")),
                "ean_piece_md": normalize_ean(md["ean_piece"]),
                "ean_piece_db": normalize_ean(db.get("ean_piece")),
                "designer_md": normalize(md["designer"]),
                "designer_db": normalize(db.get("designer")),
                "kleur_md": normalize(md["kleur"]),
                "kleur_db": normalize(db.get("kleur_nl")),
                "giftbox_md": normalize_giftbox(md["giftbox"]),
                "giftbox_db": db_giftbox_display,
                "giftbox_qty_md": normalize_giftbox_qty(md["giftbox_qty"]),
                "giftbox_qty_db": normalize_giftbox_qty(db.get("giftbox_qty")),
                "afwijkingen": ", ".join(diffs) if diffs else "",
            })

    # 4. Excel output maken
    print(f"\nExcel output schrijven: {OUTPUT_PATH}")
    os.makedirs("exports", exist_ok=True)

    out_wb = Workbook()
    ws_out = out_wb.active
    ws_out.title = "MD vs Supabase"

    # Header stijlen
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    diff_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    not_found_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ok_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    headers = [
        "SKU", "In_Supabase",
        "Naam_MD", "Naam_DB",
        "EAN_shopify_MD", "EAN_shopify_DB",
        "EAN_piece_MD", "EAN_piece_DB",
        "Designer_MD", "Designer_DB",
        "Kleur_MD", "Kleur_DB",
        "Giftbox_MD", "Giftbox_DB",
        "Giftbox_qty_MD", "Giftbox_qty_DB",
        "Afwijkingen"
    ]

    ws_out.append(headers)
    for cell in ws_out[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Kolombreedtes
    col_widths = [20, 14, 50, 50, 20, 20, 20, 20, 25, 25, 20, 20, 12, 12, 14, 14, 60]
    for i, width in enumerate(col_widths, 1):
        ws_out.column_dimensions[chr(64 + i)].width = width

    for r in results:
        row_data = [
            r["sku"], r["in_supabase"],
            r["naam_md"], r["naam_db"],
            r["ean_shopify_md"], r["ean_shopify_db"],
            r["ean_piece_md"], r["ean_piece_db"],
            r["designer_md"], r["designer_db"],
            r["kleur_md"], r["kleur_db"],
            r["giftbox_md"], r["giftbox_db"],
            r["giftbox_qty_md"], r["giftbox_qty_db"],
            r["afwijkingen"],
        ]
        ws_out.append(row_data)
        row_idx = ws_out.max_row

        if r["in_supabase"] == "nee":
            for cell in ws_out[row_idx]:
                cell.fill = not_found_fill
        elif r["afwijkingen"]:
            for cell in ws_out[row_idx]:
                cell.fill = diff_fill
        else:
            ws_out[row_idx][0].fill = ok_fill  # alleen SKU cel groen voor overzicht

    # Freeze header
    ws_out.freeze_panes = "A2"

    # Samenvatting sheet
    ws_sum = out_wb.create_sheet("Samenvatting")
    ws_sum.column_dimensions["A"].width = 40
    ws_sum.column_dimensions["B"].width = 15

    sum_header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    sum_header_font = Font(color="FFFFFF", bold=True)

    ws_sum.append(["Samenvatting Serax MD vs Supabase", ""])
    ws_sum["A1"].font = Font(bold=True, size=13)
    ws_sum.append(["", ""])
    ws_sum.append(["Metric", "Aantal"])
    for cell in ws_sum[3]:
        cell.fill = sum_header_fill
        cell.font = sum_header_font

    ws_sum.append(["Totaal SKUs in masterdata", len(md_rows)])
    ws_sum.append(["Gevonden in Supabase", stats["gevonden"]])
    ws_sum.append(["NIET gevonden in Supabase", stats["niet_gevonden"]])
    ws_sum.append(["Gevonden MET afwijkingen", stats["met_afwijkingen"]])
    ws_sum.append(["", ""])
    ws_sum.append(["Afwijkingen per veld (echte fouten)", ""])
    ws_sum["A9"].font = Font(bold=True)
    ws_sum.append(["ean_shopify", stats["fields"]["ean_shopify"]])
    ws_sum.append(["ean_piece", stats["fields"]["ean_piece"]])
    ws_sum.append(["designer", stats["fields"]["designer"]])
    ws_sum.append(["giftbox", stats["fields"]["giftbox"]])
    ws_sum.append(["giftbox_qty", stats["fields"]["giftbox_qty"]])
    ws_sum.append(["", ""])
    ws_sum.append(["Noot", ""])
    ws_sum["A15"].font = Font(bold=True)
    ws_sum.append(["Naam (product_title_nl)", "Ter info getoond. DB heeft gecureerde NL-titels, MD heeft raw namen. Geen directe fout."])
    ws_sum.append(["Kleur (kleur_nl)", "Ter info getoond. MD heeft Engelse kleuren, DB heeft Nederlandse. Geen directe fout."])

    out_wb.save(OUTPUT_PATH)

    # 5. Print samenvatting
    print("\n" + "=" * 50)
    print("SAMENVATTING")
    print("=" * 50)
    print(f"Totaal SKUs in masterdata : {len(md_rows)}")
    print(f"Gevonden in Supabase      : {stats['gevonden']}")
    print(f"NIET gevonden in Supabase : {stats['niet_gevonden']}")
    print(f"Met afwijkingen (gevonden): {stats['met_afwijkingen']}")
    print()
    print("Afwijkingen per veld (echte fouten):")
    for field, count in stats["fields"].items():
        print(f"  {field:<25}: {count}")
    print()
    print("Noot: Naam (product_title_nl) en Kleur (kleur_nl) worden ter info getoond")
    print("  maar tellen NIET als afwijking: Supabase heeft gecureerde NL-titels,")
    print("  MD heeft raw namen; MD heeft Engelse kleuren, Supabase heeft NL.")
    print()
    print(f"Output opgeslagen: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
