"""
Export een Hextom Foto Update bestand met ALLEEN nieuwe foto's (Supabase Storage URLs).

Bestaande Shopify CDN foto's worden weggelaten, zodat Hextom's 'Product Media
(add to existing) at last position' geen duplicaten aanmaakt.

Gebruik:
    python execution/export_hextom_new_photos_only.py --skus B0126104-602,B0219404,...
    python execution/export_hextom_new_photos_only.py --all-from-run 1020   # tijdstamp van vorige run
    python execution/export_hextom_new_photos_only.py --share 1             # alle share-1 SKU's
"""

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
import openpyxl

load_dotenv()

ROOT = Path(__file__).resolve().parent.parent
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY")

PHOTO_SLOTS = [
    "photo_packshot_1","photo_packshot_2","photo_packshot_3","photo_packshot_4","photo_packshot_5",
    "photo_lifestyle_1","photo_lifestyle_2","photo_lifestyle_3","photo_lifestyle_4","photo_lifestyle_5",
]


def get_supabase():
    from supabase import create_client
    return create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)


def is_supabase_url(url: str | None) -> bool:
    return bool(url) and "supabase.co/storage" in url


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--share", help="Filter op share-nummer: gebruikt .tmp/bynder_fotos/share{N}_parsed.json")
    ap.add_argument("--skus", help="Comma-separated lijst")
    ap.add_argument("--output", help="Output xlsx pad")
    args = ap.parse_args()

    # Bepaal welke SKUs
    if args.skus:
        skus = [s.strip() for s in args.skus.split(",")]
    elif args.share:
        parsed = ROOT / ".tmp" / "bynder_fotos" / f"share{args.share}_parsed.json"
        if not parsed.exists():
            print(f"FOUT: {parsed} bestaat niet"); sys.exit(1)
        with parsed.open() as f:
            data = json.load(f)
        skus = sorted(data["by_sku"].keys())
        # Voeg multi-SKU lifestyle targets + merges toe
        from process_bynder_photos import MULTI_SKU_TARGETS, SKU_MERGE
        for s in MULTI_SKU_TARGETS:
            if s not in skus:
                skus.append(s)
        skus = [SKU_MERGE.get(s, s) for s in skus]
        skus = sorted(set(skus))
    else:
        print("Kies --share of --skus"); sys.exit(1)

    sb = get_supabase()
    print(f"Ophalen {len(skus)} producten uit seo_products...")
    r = sb.table("seo_products").select(
        "sku, handle, product_title_nl, status, ean_shopify, " + ", ".join(PHOTO_SLOTS)
    ).in_("sku", skus).execute()
    rows = r.data or []
    print(f"  Gevonden: {len(rows)}")

    # Filter: alleen status=ready, alleen Supabase URLs
    export_rows = []
    skipped_raw = 0
    skipped_no_new = 0
    for row in rows:
        if row.get("status") != "ready":
            skipped_raw += 1
            continue
        new_urls = [row[s] for s in PHOTO_SLOTS if is_supabase_url(row.get(s))]
        if not new_urls:
            skipped_no_new += 1
            continue
        export_rows.append({
            "sku": row["sku"],
            "handle": row.get("handle") or "",
            "title": row.get("product_title_nl") or "",
            "ean": row.get("ean_shopify") or "",
            "urls": new_urls,
        })

    print(f"\nExport-samenvatting:")
    print(f"  {len(export_rows)} producten met NIEUWE foto's (alleen deze in export)")
    print(f"  {skipped_raw} raw producten overgeslagen")
    print(f"  {skipped_no_new} producten overgeslagen (alleen bestaande Shopify foto's, geen nieuwe)")

    # Schrijf Excel
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    output = Path(args.output) if args.output else ROOT / "exports" / f"Hextom_Foto_Update_Bynder_NIEUW_{ts}.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hextom Foto Update NIEUW"
    headers = ["Variant SKU", "Product Handle", "Product Title", "Variant Barcode"] + \
              [f"Image Src {i}" for i in range(1, 11)]
    ws.append(headers)

    for er in export_rows:
        row = [er["sku"], er["handle"], er["title"], er["ean"]]
        urls = er["urls"] + [""] * (10 - len(er["urls"]))
        row += urls[:10]
        ws.append(row)

    # Barcode als tekst
    for rownum in range(2, ws.max_row + 1):
        cell = ws.cell(row=rownum, column=4)
        cell.number_format = "@"
        if cell.value is not None:
            cell.value = str(cell.value)

    wb.save(output)
    print(f"\nHextom-NIEUW export: {output}")
    print(f"  {len(export_rows)} producten, {sum(len(e['urls']) for e in export_rows)} nieuwe foto-URLs")


if __name__ == "__main__":
    main()
