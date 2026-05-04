"""
Verwerk Bynder-foto's: resize -> upload naar Supabase Storage -> update seo_products -> Hextom export.

Regels (bevestigd door user 2026-04-20):
  - B1318105 en B1318105S worden samengevoegd op SKU B1318105
  - Multi-SKU lifestyle (B7222009 ... _N.jpg) wordt gekoppeld aan elk van de 5 bestaande SKUs;
    B7222013 bestaat niet in Supabase en wordt overgeslagen.
  - Producten met status='raw' krijgen wel foto's in DB+Storage maar niet in Hextom-export
    (raw = nog niet live op Shopify, Hextom-update zou falen).
  - Duplicaten (Windows copy-markers als ' 2', '-1') worden eruit gefilterd via MD5-hash.
  - ALLE photo-slots worden overschreven met het nieuwe Bynder-set (max 5 pack + 2 life).
    Slots waar geen nieuwe foto voor is worden gewist (voorkomt duplicate foto's op Shopify).
  - Optimalisatie: als een nieuwe Bynder-foto dezelfde foto is als een al bestaande
    (filename-normalisatie match) wordt de bestaande URL hergebruikt - geen resize, geen upload.
  - Max 5 packshots + 2 lifestyles per product (=7 slots).
  - Resize: max 4999px langste zijde, aspect ratio behouden, JPEG q90.

Gebruik:
    python execution/process_bynder_photos.py --share 1 --dry-run
    python execution/process_bynder_photos.py --share 1 --execute
    python execution/process_bynder_photos.py --share 1 --execute --limit 3
    python execution/process_bynder_photos.py --share 1 --execute --skus B0219404,B0816781
"""

import argparse
import hashlib
import json
import os
import re
import sys
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from pathlib import Path

from dotenv import load_dotenv
from PIL import Image

load_dotenv()

ROOT = Path(__file__).resolve().parent.parent
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY")

BUCKET = "product-photos"
MAX_DIMENSION = 4999
JPEG_QUALITY = 90
MAX_PACKSHOTS = 5
MAX_LIFESTYLES = 2

# Multi-SKU lifestyle: één bestand hoort bij meerdere producten
MULTI_SKU_LIFESTYLE_FILES = {
    "B7222009 B7222006 B7222007 B7222012 B7222013 B7222011_1.jpg",
    "B7222009 B7222006 B7222007 B7222012 B7222013 B7222011_2.jpg",
    "B7222009 B7222006 B7222007 B7222012 B7222013 B7222011_3.jpg",
    "B7222009 B7222006 B7222007 B7222012 B7222013 B7222011_4.jpg",
}
MULTI_SKU_TARGETS = ["B7222009", "B7222006", "B7222007", "B7222011", "B7222012"]  # B7222013 skip
# SKUs die met elkaar versmolten worden (filename-SKU -> DB-SKU)
SKU_MERGE = {"B1318105S": "B1318105"}


# ----- Filename parsing + ranking -----
SKU_RE = re.compile(r"B\d{7}(?:-\d{3})?[A-Z]?")
PREFIX_RE = re.compile(r"^(?:ECOM_|Ecom_|ECOM_ |Ecom_ |B2B_|b2b_|B2B_ |b2b_ )")
DUP_SUFFIX_RE = re.compile(r"(?:[-_ ]\(?[0-9]+\)?)$")


def classify(filename: str) -> tuple[str, int, str]:
    """
    Return (type, rank_score, tail) voor één bestand.
    type: 'packshot' of 'lifestyle'
    rank_score: lager = belangrijker (positie 1 = highest priority)
    tail: hint-string voor debug
    """
    base = re.sub(r"\.(jpg|jpeg|gif|png)$", "", filename, flags=re.I)
    # strip prefix
    no_pref = PREFIX_RE.sub("", base)
    has_ecom = bool(re.match(r"^(?:ECOM_|Ecom_|ECOM_ |Ecom_ )", base))
    has_b2b = bool(re.match(r"^(?:B2B_|b2b_|B2B_ |b2b_ )", base))

    lower = no_pref.lower()
    # type detection
    if "lifestyle" in lower:
        ftype = "lifestyle"
    else:
        ftype = "packshot"

    # numeric hint: neem ALTIJD de LAATSTE digit-groep (dat is de foto-positie, niet het SKU-nummer)
    all_nums = re.findall(r"\d+", no_pref)
    num = int(all_nums[-1]) if all_nums else 99

    # Priority ranking:
    # 1 = Ecom_BSKU_NN (ecommerce, genummerd)
    # 2 = bare BSKU.jpg (main hoofdfoto)
    # 3 = BSKUsN or BSKU_sN (styling packshot)
    # 4 = B2B_BSKU_packshotN
    # 5 = B2B_BSKU_lifestyleN
    # 6 = BSKU-b.jpg (back)
    priority = 99
    tail = no_pref
    if has_ecom:
        priority = 1
    elif "lifestyle" in lower:
        priority = 5
    elif has_b2b and "packshot" in lower:
        priority = 4
    elif re.search(r"[-_ ]?[sS]\d*$", no_pref):
        priority = 3
    elif re.search(r"-b$", lower):
        priority = 6
    elif re.fullmatch(r"B\d{7}(?:-\d{3})?[A-Z]?", no_pref):
        priority = 2
        num = 0  # bare = hoofdfoto, geen positie in filename
    else:
        priority = 4  # fallback, behandel als packshot

    # rank_score = priority * 100 + num (lager = eerder)
    score = priority * 100 + num
    return ftype, score, tail


def list_photos(share_dir: Path) -> list[Path]:
    photos = []
    for p in share_dir.rglob("*"):
        if p.is_file() and p.suffix.lower() in {".jpg", ".jpeg", ".png"}:
            # .gif bewust overgeslagen (animatie gaat verloren bij JPEG-conversie)
            photos.append(p)
    return photos


def file_hash(path: Path, length: int = 256 * 1024) -> str:
    """MD5 van eerste blok (snel, voldoende voor dedup van identieke kopieën)."""
    h = hashlib.md5()
    with path.open("rb") as f:
        h.update(f.read(length))
    return h.hexdigest()


def plan_per_sku(share_dir: Path) -> dict:
    """Bouw {sku: {'packshots': [Path..], 'lifestyles': [Path..], 'skipped': [...]}}."""
    photos = list_photos(share_dir)
    # sku -> list of (path, type, score)
    groups: dict[str, list[tuple[Path, str, int, str]]] = defaultdict(list)
    for p in photos:
        fn = p.name
        if fn in MULTI_SKU_LIFESTYLE_FILES:
            for sku in MULTI_SKU_TARGETS:
                groups[sku].append((p, "lifestyle", 500, fn))
            continue
        skus_found = SKU_RE.findall(fn)
        if not skus_found:
            continue
        sku_raw = skus_found[0]
        sku = SKU_MERGE.get(sku_raw, sku_raw)
        ftype, score, tail = classify(fn)
        groups[sku].append((p, ftype, score, tail))

    # Per sku: dedup via MD5 van eerste 256KB, rank, cap
    plan = {}
    for sku, items in groups.items():
        seen_hashes: dict[str, tuple[Path, str, int, str]] = {}
        skipped_dup = []
        for p, t, s, tail in items:
            h = file_hash(p)
            if h in seen_hashes:
                skipped_dup.append((p, seen_hashes[h][0]))
                continue
            seen_hashes[h] = (p, t, s, tail)
        items = list(seen_hashes.values())

        # Sort binnen type op score asc
        packshots = sorted([i for i in items if i[1] == "packshot"], key=lambda x: (x[2], x[0].name))
        lifestyles = sorted([i for i in items if i[1] == "lifestyle"], key=lambda x: (x[2], x[0].name))

        # Cap
        overflow = packshots[MAX_PACKSHOTS:] + lifestyles[MAX_LIFESTYLES:]
        packshots = packshots[:MAX_PACKSHOTS]
        lifestyles = lifestyles[:MAX_LIFESTYLES]

        plan[sku] = {
            "packshots": [{"path": str(p.path) if hasattr(p, 'path') else str(p[0]), "score": p[2], "tail": p[3]} for p in packshots],
            "lifestyles": [{"path": str(p[0]), "score": p[2], "tail": p[3]} for p in lifestyles],
            "skipped_cap": [{"path": str(p[0]), "type": p[1], "score": p[2]} for p in overflow],
            "skipped_dup": [{"path": str(a), "dup_of": str(b)} for a, b in skipped_dup],
        }
        # Fix packshot serialization
        plan[sku]["packshots"] = [{"path": str(p[0]), "score": p[2], "tail": p[3]} for p in packshots]
    return plan


# ----- Supabase helpers -----
def get_supabase():
    from supabase import create_client
    return create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)


def fetch_products(sb, skus: list[str]) -> dict[str, dict]:
    """Batch fetch seo_products op exacte sku."""
    out: dict[str, dict] = {}
    chunk = 100
    for i in range(0, len(skus), chunk):
        sub = skus[i:i+chunk]
        r = sb.table("seo_products").select(
            "id, sku, handle, product_title_nl, fase, status, ean_shopify, "
            "photo_packshot_1, photo_packshot_2, photo_packshot_3, photo_packshot_4, photo_packshot_5, "
            "photo_lifestyle_1, photo_lifestyle_2, photo_lifestyle_3, photo_lifestyle_4, photo_lifestyle_5"
        ).in_("sku", sub).execute()
        for row in (r.data or []):
            out[row["sku"]] = row
    return out


def storage_path(sku: str, slot: str) -> str:
    clean = sku.replace("/", "_").replace(" ", "_")
    return f"serax/{clean}_{slot}.jpg"


def public_url(path: str) -> str:
    return f"{SUPABASE_URL}/storage/v1/object/public/{BUCKET}/{path}"


def resize_to_bytes(path: Path) -> tuple[bytes, int, int, int, int]:
    """Return (jpeg_bytes, orig_w, orig_h, new_w, new_h)."""
    with Image.open(path) as img:
        orig_w, orig_h = img.size
        img = img.convert("RGB")
        img.thumbnail((MAX_DIMENSION, MAX_DIMENSION), Image.LANCZOS)
        new_w, new_h = img.size
        buf = BytesIO()
        img.save(buf, format="JPEG", quality=JPEG_QUALITY, optimize=True)
        return buf.getvalue(), orig_w, orig_h, new_w, new_h


def upload(sb, path: str, data: bytes) -> bool:
    try:
        sb.storage.from_(BUCKET).upload(
            path, data,
            {"content-type": "image/jpeg", "upsert": "true"}
        )
        return True
    except Exception as e:
        print(f"    UPLOAD FOUT: {e}")
        return False


# ----- Orchestration -----
def print_plan(plan: dict, db_products: dict, filter_skus: list[str] | None):
    lines = []
    skus_to_show = filter_skus if filter_skus else sorted(plan.keys())
    for sku in skus_to_show:
        if sku not in plan:
            continue
        p = plan[sku]
        prod = db_products.get(sku)
        head = f"[{sku}]"
        if prod:
            head += f" status={prod.get('status')} | {prod.get('product_title_nl') or '(geen titel)'}"
        else:
            head += " NIET IN SUPABASE -> OVERSLAAN"
        lines.append(head)
        lines.append(f"  packshots ({len(p['packshots'])}):")
        for i, ph in enumerate(p["packshots"], 1):
            lines.append(f"    {i}. {Path(ph['path']).name}   (score {ph['score']})")
        lines.append(f"  lifestyles ({len(p['lifestyles'])}):")
        for i, lf in enumerate(p["lifestyles"], 1):
            lines.append(f"    {i}. {Path(lf['path']).name}   (score {lf['score']})")
        if p["skipped_cap"]:
            lines.append(f"  CAP-overflow ({len(p['skipped_cap'])} te veel, niet gebruikt):")
            for sk in p["skipped_cap"]:
                lines.append(f"    - {Path(sk['path']).name}")
        if p["skipped_dup"]:
            lines.append(f"  duplicaten overgeslagen ({len(p['skipped_dup'])}):")
            for sk in p["skipped_dup"]:
                lines.append(f"    - {Path(sk['path']).name}  (identiek aan {Path(sk['dup_of']).name})")
        lines.append("")
    return "\n".join(lines)


PHOTO_SLOTS = (
    "photo_packshot_1","photo_packshot_2","photo_packshot_3","photo_packshot_4","photo_packshot_5",
    "photo_lifestyle_1","photo_lifestyle_2","photo_lifestyle_3","photo_lifestyle_4","photo_lifestyle_5",
)


def normalize_photo_key(name_or_url: str) -> str:
    """Reduceer filename/URL tot canonieke id voor dup-foto detectie.
    Matcht bv 'Ecom_B0126104-602_01-jpg.jpg' <-> 'Ecom_B0126104-602_01.jpg' -> 'b0126104-602_01'
    """
    s = name_or_url.split("/")[-1].split("?")[0].lower()
    s = re.sub(r"\.(jpg|jpeg|gif|png)$", "", s)
    s = re.sub(r"-jpg$", "", s)                    # Shopify's redundant -jpg suffix
    s = re.sub(r"^(ecom_ ?|b2b_ ?)", "", s)        # strip prefix
    s = re.sub(r"[ -]\d+$", "", s)                 # Windows copy-markers ' 2' / '-1' (NIET '_01')
    return s.strip()


def execute_plan(sb, plan: dict, db_products: dict, filter_skus: list[str] | None, dry: bool):
    """Override-all met same-photo detectie. Hergebruik bestaande URL als de foto
    dezelfde is (filename match na normalisatie); anders resize+upload naar Supabase.
    Slots waar geen nieuwe Bynder-foto voor is worden gewist."""
    report = []
    skus = filter_skus if filter_skus else sorted(plan.keys())
    for idx, sku in enumerate(skus, 1):
        if sku not in plan:
            continue
        if sku not in db_products:
            print(f"[{idx}/{len(skus)}] {sku}  -> niet in seo_products, skip")
            continue
        entry = plan[sku]
        prod = db_products[sku]

        # Registry: norm_key -> existing URL (van ALLE slots, cross-slot reuse toegestaan)
        existing_by_key: dict[str, str] = {}
        for slot in PHOTO_SLOTS:
            url = prod.get(slot)
            if url:
                existing_by_key[normalize_photo_key(url)] = url

        final_slots: dict[str, str | None] = {k: None for k in PHOTO_SLOTS}

        def process(ph_info: dict, slot_type: str, slot_idx: int) -> None:
            slot_key = f"photo_{slot_type}_{slot_idx}"
            src = Path(ph_info["path"])
            norm = normalize_photo_key(src.name)
            if norm in existing_by_key:
                url = existing_by_key[norm]
                host = "shopify" if "shopify" in url else ("supabase" if "supabase" in url else "other")
                print(f"[{idx}/{len(skus)}] {sku} {slot_type}_{slot_idx}  REUSE ({host}): {src.name}")
                final_slots[slot_key] = url
                return
            if dry:
                print(f"[{idx}/{len(skus)}] {sku} {slot_type}_{slot_idx}  [DRY] UPLOAD {src.name}")
                final_slots[slot_key] = public_url(storage_path(sku, f"{slot_type}_{slot_idx}"))
                return
            data, ow, oh, nw, nh = resize_to_bytes(src)
            target = storage_path(sku, f"{slot_type}_{slot_idx}")
            ok = upload(sb, target, data)
            note = f"{ow}x{oh}->{nw}x{nh}" if (ow != nw or oh != nh) else f"{ow}x{oh}"
            print(f"[{idx}/{len(skus)}] {sku} {slot_type}_{slot_idx}  UPLOAD {note}  {'OK' if ok else 'FAIL'}: {src.name}")
            if ok:
                final_slots[slot_key] = public_url(target)

        for i, ph in enumerate(entry["packshots"], 1):
            process(ph, "packshot", i)
        for i, ph in enumerate(entry["lifestyles"], 1):
            process(ph, "lifestyle", i)

        if not dry:
            db_updates = {k: final_slots[k] for k in PHOTO_SLOTS}  # explicit None's om te wissen
            try:
                sb.table("seo_products").update(db_updates).eq("id", prod["id"]).execute()
            except Exception as e:
                print(f"    DB UPDATE FOUT voor {sku}: {e}")
                continue

        report.append({
            "sku": sku,
            "status": prod.get("status"),
            "handle": prod.get("handle"),
            "title": prod.get("product_title_nl"),
            "ean": prod.get("ean_shopify"),
            "photos": [final_slots[k] for k in PHOTO_SLOTS if final_slots.get(k)],
        })
    return report


def write_hextom_export(report: list[dict], output: Path):
    """Hextom Foto Update format: SKU/Handle/Title/Barcode/Image Src 1..10."""
    import openpyxl
    from openpyxl.utils import get_column_letter
    # Filter: alleen status='ready'
    rows = [r for r in report if r["status"] == "ready"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hextom Foto Update"
    headers = ["Variant SKU","Product Handle","Product Title","Variant Barcode"] + \
              [f"Image Src {i}" for i in range(1, 11)]
    ws.append(headers)
    for r in rows:
        row = [r["sku"], r["handle"] or "", r["title"] or "", r["ean"] or ""]
        photos = r["photos"] + [""] * (10 - len(r["photos"]))
        row += photos
        ws.append(row)
    # Barcode als tekst
    ean_col = 4  # 1-based
    for row in ws.iter_rows(min_row=2, min_col=ean_col, max_col=ean_col):
        for cell in row:
            cell.number_format = "@"
            if cell.value is not None:
                cell.value = str(cell.value)
    wb.save(output)
    print(f"Hextom export: {output} ({len(rows)} ready producten)")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--share", required=True, choices=["1", "2"])
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--execute", action="store_true")
    ap.add_argument("--limit", type=int, help="Alleen eerste N SKUs")
    ap.add_argument("--skus", help="Comma-separated lijst van SKUs")
    ap.add_argument("--output", help="Output Hextom xlsx pad")
    args = ap.parse_args()

    if not (args.dry_run or args.execute):
        print("Kies --dry-run of --execute"); sys.exit(1)

    share_root = ROOT / ".tmp" / "bynder_fotos" / f"share{args.share}"
    if not share_root.exists():
        print(f"FOUT: {share_root} bestaat niet"); sys.exit(1)

    # Zoek de sub-map (Bynder pakt vaak één extra directory uit)
    subdirs = [p for p in share_root.iterdir() if p.is_dir()]
    if len(subdirs) == 1:
        scan_dir = subdirs[0]
    else:
        scan_dir = share_root
    print(f"Scan: {scan_dir}")

    plan = plan_per_sku(scan_dir)
    print(f"Plan: {len(plan)} unieke SKUs in share {args.share}")

    # Filter lijst
    filter_skus = None
    if args.skus:
        filter_skus = [s.strip() for s in args.skus.split(",")]
    elif args.limit:
        filter_skus = sorted(plan.keys())[:args.limit]

    # Supabase
    sb = get_supabase()
    skus_to_check = filter_skus if filter_skus else list(plan.keys())
    db_products = fetch_products(sb, skus_to_check)
    print(f"In Supabase gevonden: {len(db_products)}/{len(skus_to_check)}")

    # Plan tonen
    print("\n" + "="*90)
    print(print_plan(plan, db_products, filter_skus))
    print("="*90)

    if args.dry_run:
        # Save plan-JSON voor referentie
        plan_out = ROOT / ".tmp" / "bynder_fotos" / f"share{args.share}_plan.json"
        with plan_out.open("w") as f:
            json.dump(plan, f, indent=2)
        print(f"\n[DRY RUN] Plan opgeslagen in {plan_out}")
        print("Herstart met --execute om daadwerkelijk te resizen/uploaden/DB update.")
        return

    # Execute
    report = execute_plan(sb, plan, db_products, filter_skus, dry=False)

    # Hextom export
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out = Path(args.output) if args.output else ROOT / "exports" / f"Hextom_Foto_Update_Bynder_{ts}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    write_hextom_export(report, out)

    # Summary
    ready = [r for r in report if r["status"] == "ready"]
    raw = [r for r in report if r["status"] == "raw"]
    print(f"\n{'='*60}\nKLAAR\n{'='*60}")
    print(f"  Totaal verwerkt:  {len(report)}")
    print(f"  Status=ready:     {len(ready)}  -> in Hextom xlsx")
    print(f"  Status=raw:       {len(raw)}    -> alleen DB+Storage")


if __name__ == "__main__":
    main()
