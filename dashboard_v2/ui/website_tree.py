"""Website structure helpers (read-only).

Ported from streamlit_app.py. Reads ``Master Files/Website indeling (1).xlsx``
to produce:
  - ``load_website_tree()`` → {hoofdcat: {subcat: [sub_subcats]}} dict
  - ``load_active_subsubcategories()`` → set of sub-subcat names that are
    actually live on the Shopify store (green cells in the xlsx).
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
import streamlit as st


_WEBSITE_XLSX = Path("Master Files/Website indeling (1).xlsx")


@st.cache_data(show_spinner=False)
def load_active_subsubcategories() -> set[str]:
    """Lowercase set of sub-subcategorieën that are live on the webshop.

    Greenness is derived from cell fill colour in the xlsx (GROEN = live).
    Adds the manual 'drinkglazen' override (confirmed live, not coloured).
    """
    from openpyxl import load_workbook

    if not _WEBSITE_XLSX.exists():
        return set()

    GROEN = {"FF92D050", "FF93C47D"}
    wb = load_workbook(_WEBSITE_XLSX, data_only=True)
    ws = wb["Blad1"]

    actief: set[str] = set()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or cell.row < 3:
                continue
            fg = cell.fill.fgColor if cell.fill else None
            if not fg or fg.type != "rgb" or not fg.rgb or fg.rgb == "00000000":
                continue
            if fg.rgb in GROEN:
                actief.add(str(cell.value).strip().lower())
    actief.add("drinkglazen")
    return actief


@st.cache_data(show_spinner=False)
def load_website_tree() -> dict:
    """{hoofdcat: {subcat: [sub_subcats, ...], ...}, ...}.

    Layout van de xlsx: rij 0 = hoofdcategorieën, rij 1 = subcategorieën,
    rij 2+ = sub-subcategorieën. Cellen kunnen leeg zijn; een hoofdcat 'spant'
    horizontaal over zijn kolommen.
    """
    if not _WEBSITE_XLSX.exists():
        return {}

    df = pd.read_excel(_WEBSITE_XLSX, sheet_name="Blad1", header=None, dtype=str)

    def cell(r, c):
        if r >= len(df) or c >= len(df.columns):
            return None
        v = df.iat[r, c]
        if pd.isna(v):
            return None
        s = str(v).strip()
        return s if s else None

    hoofdcats = {}
    for c in range(len(df.columns)):
        v = cell(0, c)
        if v:
            hoofdcats[c] = v
    hoofdcat_cols = sorted(hoofdcats.keys())

    def hoofdcat_for_col(c):
        parent = None
        for hc in hoofdcat_cols:
            if hc <= c:
                parent = hc
            else:
                break
        return hoofdcats.get(parent)

    subcats = {}
    for c in range(len(df.columns)):
        v = cell(1, c)
        if v:
            subcats[c] = (hoofdcat_for_col(c), v)

    tree: dict = {}
    for c, (hc, sc) in subcats.items():
        if hc not in tree:
            tree[hc] = {}
        if sc not in tree[hc]:
            tree[hc][sc] = []
        for r in range(2, len(df)):
            v = cell(r, c)
            if v:
                tree[hc][sc].append(v)
    return tree
