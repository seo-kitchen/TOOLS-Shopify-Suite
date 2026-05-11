"""Tab — Archief herverwerken.

Primaire bron: shopify_meta_audit (heeft ALLE Shopify-producten met hun status).
Verrijkt met products_curated (pipeline-data) en products_raw (foto's, EAN, afmetingen).
"""
from __future__ import annotations

import io
import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st
from dotenv import load_dotenv

load_dotenv()


@st.cache_resource
def _get_sb():
    from supabase import create_client
    url = os.getenv("SUPABASE_NEW_URL", "")
    key = os.getenv("SUPABASE_NEW_SERVICE_KEY", "") or os.getenv("SUPABASE_NEW_KEY", "")
    if not url or not key:
        raise RuntimeError("SUPABASE_NEW_URL of SUPABASE_NEW_SERVICE_KEY ontbreekt.")
    return create_client(url, key)


LEVERANCIERS = ["Pottery Pots", "Serax", "Salt & Pepper", "Printworks",
                "BONBISTRO", "ONA", "Urban Nature Culture"]
SHOPIFY_STATUSSEN = ["archived", "active", "draft"]

HEXTOM_COLUMNS = [
    "Variant SKU", "", "", "Product Handle", "Product Title",
    "Product Vendor", "Product Type", "Variant Barcode", "Variant Price",
    "Variant Cost", "Product Description", "", "", "", "Product Tags",
    "Variant Metafield custom.collectie", "Product Metafield custom.designer",
    "Product Metafield custom.materiaal", "Product Metafield custom.kleur",
    "Product Metafield custom.hoogte_filter", "Product Metafield custom.lengte_filter",
    "Product Metafield custom.breedte_filter",
    "photo_packshot_1", "photo_packshot_2", "photo_packshot_3",
    "photo_packshot_4", "photo_packshot_5",
    "photo_lifestyle_1", "photo_lifestyle_2", "photo_lifestyle_3",
    "photo_lifestyle_4", "photo_lifestyle_5",
    "Product Metafield custom.ean",
    "Product Metafield custom.artikelnummer",
    "Product Metafield custom.meta_description",
]
TEXT_FORMAT_COLUMNS = {"Variant Barcode", "Product Metafield custom.ean"}
STATUS_FILL = {
    "active":   PatternFill("solid", fgColor="CCFFCC"),
    "archived": PatternFill("solid", fgColor="FFE4B5"),
    "draft":    PatternFill("solid", fgColor="E0E0E0"),
}


def _clean_decimal(value) -> str:
    if value is None:
        return ""
    s = str(value).replace(",", ".")
    try:
        f = float(s)
        return f"{f:.10f}".rstrip("0").rstrip(".")
    except ValueError:
        return s


def _build_hextom_excel(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col_idx, col_name in enumerate(HEXTOM_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name if col_name else "")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, p in enumerate(rows, start=2):
        row_data = {
            "Variant SKU":                               p.get("sku", ""),
            "":                                          "",
            "Product Handle":                            p.get("handle", ""),
            "Product Title":                             p.get("product_title_nl") or p.get("product_title", ""),
            "Product Vendor":                            p.get("vendor", ""),
            "Product Type":                              p.get("hoofdcategorie") or p.get("product_type", ""),
            "Variant Barcode":                           str(p.get("ean_shopify", "") or ""),
            "Variant Price":                             _clean_decimal(p.get("verkoopprijs") or p.get("price")),
            "Variant Cost":                              _clean_decimal(p.get("inkoopprijs")),
            "Product Description":                       p.get("meta_description") or p.get("current_meta_description", "") or "",
            "Product Tags":                              p.get("tags", "") or "",
            "Variant Metafield custom.collectie":        p.get("collectie", "") or "",
            "Product Metafield custom.designer":         p.get("designer", "") or "",
            "Product Metafield custom.materiaal":        p.get("materiaal_nl", "") or "",
            "Product Metafield custom.kleur":            p.get("kleur_nl", "") or "",
            "Product Metafield custom.hoogte_filter":    _clean_decimal(p.get("hoogte_cm")),
            "Product Metafield custom.lengte_filter":    _clean_decimal(p.get("lengte_cm")),
            "Product Metafield custom.breedte_filter":   _clean_decimal(p.get("breedte_cm")),
            "photo_packshot_1":                          p.get("photo_packshot_1", "") or "",
            "photo_packshot_2":                          p.get("photo_packshot_2", "") or "",
            "photo_packshot_3":                          p.get("photo_packshot_3", "") or "",
            "photo_packshot_4":                          p.get("photo_packshot_4", "") or "",
            "photo_packshot_5":                          p.get("photo_packshot_5", "") or "",
            "photo_lifestyle_1":                         p.get("photo_lifestyle_1", "") or "",
            "photo_lifestyle_2":                         p.get("photo_lifestyle_2", "") or "",
            "photo_lifestyle_3":                         p.get("photo_lifestyle_3", "") or "",
            "photo_lifestyle_4":                         p.get("photo_lifestyle_4", "") or "",
            "photo_lifestyle_5":                         p.get("photo_lifestyle_5", "") or "",
            "Product Metafield custom.ean":              str(p.get("ean_piece", "") or ""),
            "Product Metafield custom.artikelnummer":    p.get("sku", "") or "",
            "Product Metafield custom.meta_description": p.get("meta_description") or p.get("current_meta_description", "") or "",
        }
        row_fill = STATUS_FILL.get(p.get("product_status", ""))
        for col_idx, col_name in enumerate(HEXTOM_COLUMNS, start=1):
            value = row_data.get(col_name, "") if col_name else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_name in TEXT_FORMAT_COLUMNS and value:
                cell.value = str(value)
                cell.number_format = "@"
            if row_fill:
                cell.fill = row_fill

    col_widths = {1: 18, 4: 40, 5: 50, 8: 16, 11: 60, 15: 50}
    for col_idx in range(1, len(HEXTOM_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_idx, 20)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _shopify_producten(vendor: str, shopify_status: str, zoek: str, limit: int) -> list[dict]:
    """Haal producten op via Shopify REST API — werkt voor alle statussen."""
    import requests
    store = os.getenv("SHOPIFY_STORE", "")
    token = os.getenv("SHOPIFY_ACCESS_TOKEN", "")
    if not store or not token:
        raise RuntimeError("SHOPIFY_STORE of SHOPIFY_ACCESS_TOKEN ontbreekt in de environment variables.")

    base    = f"https://{store}/admin/api/2026-04"
    headers = {"X-Shopify-Access-Token": token}
    fields  = "id,title,handle,vendor,product_type,status,tags,variants,images"

    params: dict = {"limit": 250, "fields": fields}
    if shopify_status != "Alle":
        params["status"] = shopify_status
    if vendor != "Alle":
        params["vendor"] = vendor

    rows: list[dict] = []
    url = f"{base}/products.json"

    while url and len(rows) < limit:
        r = requests.get(url, headers=headers, params=params, timeout=15)
        if not r.ok:
            raise RuntimeError(f"Shopify API fout {r.status_code}: {r.text[:200]}")
        products = r.json().get("products", [])
        for p in products:
            variant  = p.get("variants", [{}])[0] if p.get("variants") else {}
            image    = p.get("images", [{}])[0] if p.get("images") else {}
            rows.append({
                "shopify_id":    p.get("id", ""),
                "handle":        p.get("handle", ""),
                "product_title": p.get("title", ""),
                "vendor":        p.get("vendor", ""),
                "product_type":  p.get("product_type", ""),
                "product_status": p.get("status", ""),
                "tags":          p.get("tags", ""),
                "sku":           variant.get("sku", ""),
                "ean_shopify":   variant.get("barcode", ""),
                "price":         variant.get("price", ""),
                "has_image":     bool(image),
                "photo_packshot_1": image.get("src", "") if image else "",
            })
        # Volgende pagina via Link-header (cursor paginatie)
        link = r.headers.get("Link", "")
        next_url = None
        for part in link.split(","):
            if 'rel="next"' in part:
                next_url = part.split(";")[0].strip().strip("<>")
                break
        url = next_url
        params = {}  # params zitten al in de next URL

    # Zoekfilter (Shopify API ondersteunt geen vrije tekst)
    if zoek:
        zoek_l = zoek.lower()
        rows = [r for r in rows if zoek_l in r.get("handle", "").lower()
                or zoek_l in r.get("product_title", "").lower()
                or zoek_l in r.get("sku", "").lower()]

    return rows[:limit]


@st.cache_data(ttl=120, show_spinner=False)
def _load(vendor: str, shopify_status: str, zoek: str, limit: int) -> list[dict]:
    """
    Primaire bron: Shopify API (alle statussen: active, archived, draft).
    Verrijkt met Supabase (curated-data, pipeline-status, foto's).
    """
    # Stap 1: Shopify API
    shopify_rows = _shopify_producten(vendor, shopify_status, zoek, limit)
    if not shopify_rows:
        return []

    handles = [r["handle"] for r in shopify_rows if r.get("handle")]
    skus    = [r["sku"] for r in shopify_rows if r.get("sku")]

    sb = _get_sb()

    # Stap 2: verrijk met shopify_meta_audit (SEO-titels, description)
    audit_by_handle: dict[str, dict] = {}
    if handles:
        try:
            res = sb.table("shopify_meta_audit").select(
                "handle,current_meta_title,current_meta_description,title_status,desc_status"
            ).in_("handle", handles[:500]).execute().data or []
            audit_by_handle = {r["handle"]: r for r in res}
        except Exception:
            pass

    # Stap 3: verrijk met products_curated (NL-titel, categorie, pipeline-status)
    curated_by_handle: dict[str, dict] = {}
    if handles:
        try:
            res = sb.table("products_curated").select(
                "handle,sku,supplier,fase,product_title_nl,hoofdcategorie,"
                "collectie,materiaal_nl,kleur_nl,meta_description,"
                "verkoopprijs,inkoopprijs,pipeline_status"
            ).in_("handle", handles[:500]).execute().data or []
            curated_by_handle = {r["handle"]: r for r in res}
        except Exception:
            pass

    # Stap 4: verrijk met products_raw (foto's, EAN, afmetingen, fase, naam) via SKU
    raw_by_sku: dict[str, dict] = {}
    if skus:
        try:
            res = sb.table("products_raw").select(
                "sku,supplier,fase,product_name_raw,ean_piece,designer,"
                "hoogte_cm,lengte_cm,breedte_cm,"
                "leverancier_category,leverancier_item_cat,"
                "photo_packshot_1,photo_packshot_2,photo_packshot_3,"
                "photo_packshot_4,photo_packshot_5,"
                "photo_lifestyle_1,photo_lifestyle_2,photo_lifestyle_3,"
                "photo_lifestyle_4,photo_lifestyle_5"
            ).in_("sku", skus[:500]).execute().data or []
            raw_by_sku = {r["sku"]: r for r in res}
        except Exception:
            pass

    # Samenvoegen (Shopify als basis, Supabase voegt toe)
    merged = []
    for s in shopify_rows:
        handle = s.get("handle", "")
        sku    = s.get("sku", "")
        audit  = audit_by_handle.get(handle, {})
        cur    = curated_by_handle.get(handle, {})
        raw    = raw_by_sku.get(sku, {})
        merged.append({**s, **raw, **audit, **cur})

    return merged


def _run_inline_transform(rows: list[dict]) -> None:
    """Genereer NL-titels + meta descriptions en sla op in products_curated."""
    import anthropic

    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if not api_key:
        st.error("ANTHROPIC_API_KEY ontbreekt in de environment variables.")
        return

    client = anthropic.Anthropic(api_key=api_key)
    sb = _get_sb()
    n = len(rows)

    st.divider()
    st.markdown("### Transformeren bezig...")
    bar = st.progress(0, text="Voorbereiden...")
    log = st.empty()

    # Stap 1: batch naam-vertaling via Haiku (één call voor alle namen)
    namen_raw = [r.get("product_title", "") for r in rows]
    uniek = list(dict.fromkeys(nm.strip() for nm in namen_raw if nm.strip()))
    vertaling_map: dict[str, str] = {}

    bar.progress(0.0, text=f"Stap 1/2 — {len(uniek)} namen vertalen (Haiku)...")
    try:
        prompt = (
            f"Vertaal deze {len(uniek)} Engelse productnamen naar het Nederlands "
            "voor een design webshop.\n\n"
            "REGELS:\n"
            "- Behoud eigennamen (collectie/designer) onveranderd\n"
            "- Title Case (eerste letter van elk woord groot, behalve van/de/het/en)\n"
            "- Vertaal productwoorden: Plate→Bord, Bowl→Kom, Vase→Vaas, Pot→Pot, "
            "Cup→Kopje, Mug→Mok, Tray→Dienblad, Jug→Kan, Mirror→Spiegel\n"
            "- Maatcodes blijven uppercase: XS, S, M, L, XL\n"
            "- Één regel per naam, zelfde volgorde, geen nummering\n\n"
            "INPUT:\n"
            + "\n".join(uniek)
            + "\n\nOUTPUT:"
        )
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=max(2000, len(uniek) * 16),
            messages=[{"role": "user", "content": prompt}],
        )
        lines = [l.strip() for l in resp.content[0].text.strip().split("\n") if l.strip()]
        if len(lines) == len(uniek):
            vertaling_map = dict(zip(uniek, lines))
        log.caption(f"✅ {len(vertaling_map)} namen vertaald")
    except Exception as e:
        log.caption(f"⚠️ Naam-vertaling overgeslagen: {e}")

    # Stap 2: meta description per product + upsert in products_curated
    success = errors = 0
    resultaten = []

    for i, p in enumerate(rows):
        frac = (i + 1) / n
        bar.progress(frac, text=f"Stap 2/2 — {i+1}/{n}: {p.get('sku') or p.get('handle', '')}")

        raw_title = (p.get("product_title") or "").strip()
        title_nl  = vertaling_map.get(raw_title, raw_title)
        vendor    = p.get("vendor", "")
        handle    = p.get("handle", "")
        sku       = p.get("sku", "")

        try:
            meta_prompt = (
                f"Schrijf een Nederlandse SEO meta description (120–155 tekens).\n"
                f"Product: {title_nl}\nMerk: {vendor}\n\n"
                "Regels: gebruik 'je'-vorm, eindig met een CTA (bijv. 'Bestel nu', "
                "'Bekijk het aanbod'), vermeld gratis verzending vanaf €75 als dat past.\n"
                "Geef alleen de meta description terug, geen uitleg."
            )
            meta_resp = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=200,
                messages=[{"role": "user", "content": meta_prompt}],
            )
            meta = meta_resp.content[0].text.strip()[:155]

            # Prijs veilig converteren
            prijs = None
            try:
                prijs = float(p["price"]) if p.get("price") else None
            except (ValueError, TypeError):
                pass

            curated_data = {
                "handle":           handle,
                "sku":              sku,
                "supplier":         vendor,
                "product_title_nl": title_nl,
                "meta_description": meta,
                "pipeline_status":  "ready",
            }
            if prijs is not None:
                curated_data["verkoopprijs"] = prijs

            # Upsert: update als handle al bestaat, anders insert
            existing = sb.table("products_curated").select("id") \
                         .eq("handle", handle).execute().data
            if existing:
                sb.table("products_curated").update(curated_data) \
                  .eq("handle", handle).execute()
            else:
                sb.table("products_curated").insert(curated_data).execute()

            resultaten.append({**p, "product_title_nl": title_nl, "meta_description": meta})
            success += 1

        except Exception as e:
            errors += 1
            resultaten.append({**p, "fout": str(e)})

    bar.progress(1.0, text="Klaar.")

    st.success(f"✅ {success} producten getransformeerd en opgeslagen in de database.")
    if errors:
        st.warning(f"⚠️ {errors} producten mislukt.")

    # Cache legen zodat de tabel de nieuwe data toont
    _load.clear()

    # Download Hextom Excel van getransformeerde producten
    if resultaten:
        xlsx = _build_hextom_excel(resultaten)
        sup_label = (resultaten[0].get("vendor") or "producten").replace(" ", "_")
        st.download_button(
            label=f"📥 Download {success} getransformeerde producten als Hextom Excel",
            data=xlsx,
            file_name=f"hextom_{sup_label}_getransformeerd_{success}st.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="hv_transform_dl",
        )


def _load_skus_from_file(uploaded) -> list[str]:
    """Lees SKUs uit CSV/Excel/TXT — accepteert kolom 'sku', 'SKU', of eerste kolom."""
    naam = uploaded.name.lower()
    raw = uploaded.read()
    if naam.endswith((".xlsx", ".xls")):
        df = pd.read_excel(io.BytesIO(raw), dtype=str, engine="openpyxl")
    elif naam.endswith(".csv"):
        for enc in ("utf-8-sig", "utf-8", "cp1252"):
            try:
                df = pd.read_csv(io.BytesIO(raw), dtype=str, encoding=enc, keep_default_na=False)
                break
            except UnicodeDecodeError:
                continue
        else:
            return []
    else:
        # txt: één SKU per regel
        try:
            text = raw.decode("utf-8")
        except UnicodeDecodeError:
            text = raw.decode("cp1252", errors="ignore")
        return [ln.strip() for ln in text.splitlines() if ln.strip()]

    # Zoek SKU-kolom
    for kandidaat in ("sku", "SKU", "Sku", "brand_id", "Variant SKU", "artikelnummer"):
        if kandidaat in df.columns:
            return [str(s).strip() for s in df[kandidaat].dropna().tolist() if str(s).strip()]
    # Fallback: eerste kolom
    return [str(s).strip() for s in df.iloc[:, 0].dropna().tolist() if str(s).strip()]


def _load_by_skus(skus: list[str]) -> list[dict]:
    """Laad producten op basis van een SKU-lijst — direct uit Supabase, geen Shopify-search nodig."""
    if not skus:
        return []
    sb = _get_sb()

    # products_raw — basis + foto's + afmetingen
    raw_map: dict[str, dict] = {}
    for i in range(0, len(skus), 200):
        chunk = skus[i:i + 200]
        res = sb.table("products_raw").select(
            "sku,supplier,fase,product_name_raw,ean_piece,ean_shopify,designer,"
            "kleur_en,materiaal_raw,hoogte_cm,lengte_cm,breedte_cm,"
            "leverancier_category,leverancier_item_cat,"
            "photo_packshot_1,photo_packshot_2,photo_packshot_3,photo_packshot_4,photo_packshot_5,"
            "photo_lifestyle_1,photo_lifestyle_2,photo_lifestyle_3,photo_lifestyle_4,photo_lifestyle_5"
        ).in_("sku", chunk).execute().data or []
        for r in res:
            raw_map[r["sku"]] = r

    # products_curated — eventueel bestaande NL-data
    cur_map: dict[str, dict] = {}
    for i in range(0, len(skus), 200):
        chunk = skus[i:i + 200]
        res = sb.table("products_curated").select(
            "sku,handle,product_title_nl,hoofdcategorie,subcategorie,sub_subcategorie,"
            "collectie,materiaal_nl,kleur_nl,meta_description,pipeline_status,"
            "verkoopprijs,inkoopprijs"
        ).in_("sku", chunk).execute().data or []
        for r in res:
            cur_map[r["sku"]] = r

    # shopify_meta_audit — huidige live meta title/description op Shopify (via handle)
    handles = [r["handle"] for r in cur_map.values() if r.get("handle")]
    audit_by_handle: dict[str, dict] = {}
    if handles:
        for i in range(0, len(handles), 200):
            chunk = handles[i:i + 200]
            try:
                res = sb.table("shopify_meta_audit").select(
                    "handle,current_meta_title,current_meta_description"
                ).in_("handle", chunk).execute().data or []
                for r in res:
                    audit_by_handle[r["handle"]] = r
            except Exception:
                pass

    merged = []
    for sku in skus:
        raw = raw_map.get(sku, {})
        cur = cur_map.get(sku, {})
        if not raw and not cur:
            continue
        handle = cur.get("handle", "")
        audit = audit_by_handle.get(handle, {}) if handle else {}
        merged.append({
            **raw,
            **cur,
            **audit,
            "sku": sku,
            "product_title": raw.get("product_name_raw", "") or cur.get("product_title_nl", ""),
            "vendor": (raw.get("supplier", "") or "").title(),
            "price": cur.get("verkoopprijs"),
            "has_image": bool(raw.get("photo_packshot_1")),
        })
    return merged


def render() -> None:
    st.subheader("Archief herverwerken")
    st.caption(
        "Selecteer producten direct uit Shopify-data — geen Hextom-export nodig. "
        "Filter op merk en status, selecteer, en exporteer of herstart de pipeline."
    )

    # ── Bron-toggle: nog niet verwerkte producten uit Supabase ───────────────
    with st.expander("📦 Of: nog niet verwerkte producten uit Supabase (niet via Shopify)", expanded=False):
        st.caption(
            "Laadt producten die wel in `products_raw` zitten maar nog niet in de pipeline "
            "(geen `ready` record). Ideaal voor verse leverancier-imports die nog niet "
            "in Shopify staan."
        )
        sup_keuze = st.selectbox(
            "Leverancier",
            ["pottery_pots", "serax", "salt_pepper", "printworks"],
            key="hv_raw_supplier",
        )
        # Beschikbare fases voor deze supplier ophalen
        try:
            fase_res = _get_sb().table("products_raw").select("fase") \
                .eq("supplier", sup_keuze).execute().data or []
            fases_avail = sorted(set(r["fase"] for r in fase_res if r.get("fase")))
        except Exception:
            fases_avail = []
        fc1, fc2 = st.columns([2, 2])
        with fc1:
            fase_keuze = st.selectbox(
                "Fase (optioneel)",
                ["Alle"] + fases_avail,
                key="hv_raw_fase",
            )
        with fc2:
            raw_limit = st.number_input("Max", 50, 2000, 500, 50, key="hv_raw_lim")

        if st.button("🔍 Laad nog niet verwerkte producten", type="primary", key="hv_raw_load"):
            with st.spinner("Zoeken in products_raw..."):
                sb = _get_sb()
                # Haal alle raw-SKUs op voor supplier + (optioneel) fase
                q = sb.table("products_raw").select("sku").eq("supplier", sup_keuze)
                if fase_keuze != "Alle":
                    q = q.eq("fase", fase_keuze)
                raw_rows = q.limit(int(raw_limit) * 3).execute().data or []
                raw_skus = [r["sku"] for r in raw_rows if r.get("sku")]
                # Haal SKUs op die al 'ready' zijn in curated
                ready_skus: set[str] = set()
                if raw_skus:
                    for i in range(0, len(raw_skus), 200):
                        chunk = raw_skus[i:i + 200]
                        r2 = sb.table("products_curated").select("sku") \
                            .eq("pipeline_status", "ready").in_("sku", chunk).execute().data or []
                        ready_skus.update(x["sku"] for x in r2)
                te_doen = [s for s in raw_skus if s not in ready_skus][:int(raw_limit)]

            if not te_doen:
                st.warning(f"Geen niet-verwerkte SKUs gevonden voor {sup_keuze} / {fase_keuze}.")
            else:
                with st.spinner(f"Ophalen van {len(te_doen)} producten..."):
                    rows = _load_by_skus(te_doen)
                st.session_state["hv_geladen"] = True
                st.session_state["hv_rows_override"] = rows
                st.success(f"✅ {len(rows)} producten geladen (van {len(raw_skus)} totaal, "
                           f"{len(ready_skus)} al verwerkt overgeslagen).")
                st.rerun()

    # ── SKU-upload (alternatief voor filter) ─────────────────────────────────
    with st.expander("📋 Of: upload SKU-lijst (CSV / Excel / TXT)", expanded=False):
        st.caption(
            "Upload een bestand met SKUs (kolom 'sku', 'brand_id', of eerste kolom). "
            "Producten worden direct uit Supabase geladen — Shopify-filters worden genegeerd."
        )
        sku_file = st.file_uploader("Bestand", type=["csv", "xlsx", "xls", "txt"], key="hv_sku_up")
        if sku_file:
            try:
                skus = _load_skus_from_file(sku_file)
                st.caption(f"{len(skus)} SKUs gelezen: {', '.join(skus[:5])}{' …' if len(skus) > 5 else ''}")
                if st.button(f"🔍 Laad deze {len(skus)} producten", type="primary", key="hv_sku_load"):
                    with st.spinner("Ophalen uit Supabase..."):
                        rows = _load_by_skus(skus)
                    if not rows:
                        st.error(f"Geen producten gevonden voor deze {len(skus)} SKUs in products_raw.")
                    else:
                        gemist = len(skus) - len(rows)
                        if gemist:
                            st.warning(f"{gemist} SKUs niet gevonden in Supabase (overgeslagen).")
                        st.session_state["hv_geladen"] = True
                        st.session_state["hv_rows_override"] = rows
                        st.success(f"✅ {len(rows)} producten geladen — scroll naar beneden voor de tabel.")
                        st.rerun()
            except Exception as e:
                st.error(f"Bestand niet leesbaar: {e}")

    f1, f2, f3, f4 = st.columns([2, 2, 2, 3])
    with f1:
        vendor = st.selectbox("Merk", ["Alle"] + LEVERANCIERS, key="hv_vendor")
    with f2:
        shopify_status = st.selectbox(
            "Status in Shopify", ["Alle"] + SHOPIFY_STATUSSEN, index=2, key="hv_sh"
        )  # default: archived
    with f3:
        pipeline_filter = st.selectbox(
            "Pipeline-status",
            ["Niet 'ready' (default)", "Alle", "Nog niet verwerkt", "In behandeling", "Klaar (ready)"],
            index=0,
            key="hv_pipe",
            help="'Niet ready' = alles behalve wat al door de pipeline is. Zo zie je geen dubbel werk.",
        )
    with f4:
        zoek = st.text_input("Zoek (handle / productnaam)", placeholder="bijv. pottery-pots-vaas", key="hv_zoek")

    col_lim, col_btn, col_ref = st.columns([2, 1, 1])
    with col_lim:
        limit = st.number_input("Max te laden", 50, 2000, 500, 50, key="hv_limit")
    with col_btn:
        st.caption("&nbsp;")
        laden = st.button("🔍 Laad producten", type="primary", key="hv_laden")
    with col_ref:
        st.caption("&nbsp;")
        if st.button("🔄 Ververs", key="hv_refresh"):
            _load.clear()
            st.rerun()

    if laden:
        st.session_state["hv_geladen"] = True
        st.session_state.pop("hv_rows_override", None)  # filter overrult SKU-upload
    if not st.session_state.get("hv_geladen"):
        st.info("Stel filters in en klik **Laad producten** om te beginnen.")
        return

    # Als SKU-upload route is gebruikt, gebruik die data
    if st.session_state.get("hv_rows_override"):
        rows = st.session_state["hv_rows_override"]
    else:
        with st.spinner("Ophalen uit Supabase..."):
            try:
                rows = _load(vendor, shopify_status, zoek.strip(), int(limit))
            except Exception as e:
                st.error(f"❌ Fout: {e}")
                return

    # Pipeline-status filter toepassen
    voor_filter = len(rows)
    if pipeline_filter == "Niet 'ready' (default)":
        rows = [r for r in rows if (r.get("pipeline_status") or "") != "ready"]
    elif pipeline_filter == "Nog niet verwerkt":
        rows = [r for r in rows if not (r.get("pipeline_status") or "").strip()
                or (r.get("pipeline_status") or "") == "raw"]
    elif pipeline_filter == "In behandeling":
        rows = [r for r in rows if (r.get("pipeline_status") or "") == "in_process"]
    elif pipeline_filter == "Klaar (ready)":
        rows = [r for r in rows if (r.get("pipeline_status") or "") == "ready"]

    verborgen = voor_filter - len(rows)
    if verborgen > 0:
        st.caption(f"⚙️ {verborgen} producten verborgen door pipeline-filter — verander filter naar 'Alle' om ze te zien.")

    if not rows:
        st.warning(f"Geen producten gevonden — vendor='{vendor}', status='{shopify_status}', pipeline='{pipeline_filter}'.")
        return

    df = pd.DataFrame(rows)
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Gevonden", len(rows))
    m2.metric("Met NL-titel", int(df["product_title_nl"].notna().sum()) if "product_title_nl" in df else 0)
    m3.metric("Met meta-desc", int(df["current_meta_description"].notna().sum()) if "current_meta_description" in df else 0)
    m4.metric("Met foto", int(df["has_image"].sum()) if "has_image" in df else 0)

    # Bouw afmetingen-string per rij en heeft_meta/heeft_titel-indicators
    def _afm(r):
        h = r.get("hoogte_cm")
        l = r.get("lengte_cm")
        b = r.get("breedte_cm")
        delen = []
        if h not in (None, "", 0): delen.append(f"H{h}")
        if l not in (None, "", 0): delen.append(f"L{l}")
        if b not in (None, "", 0): delen.append(f"B{b}")
        return " · ".join(str(d) for d in delen) if delen else ""

    df["afmetingen"] = df.apply(lambda r: _afm(r), axis=1)
    # heeft_X: ✅ als gevuld, anders leeg
    for src, tgt in [
        ("current_meta_title", "heeft_title"),
        ("current_meta_description", "heeft_desc"),
        ("meta_description", "heeft_nl_desc"),
    ]:
        if src in df.columns:
            df[tgt] = df[src].apply(lambda v: "✅" if str(v or "").strip() else "—")
        else:
            df[tgt] = "—"

    # Tags inkorten voor display zodat de cel niet te breed wordt
    if "tags" in df.columns:
        df["tags_kort"] = df["tags"].apply(
            lambda v: (str(v)[:60] + "…") if v and len(str(v)) > 60 else str(v or "")
        )
    else:
        df["tags_kort"] = ""

    TOON = ["handle", "product_title", "vendor", "fase", "product_status",
            "pipeline_status", "hoofdcategorie", "tags_kort",
            "heeft_title", "heeft_desc", "afmetingen", "price"]
    for col in TOON:
        if col not in df.columns:
            df[col] = None

    # ── Selectie-knoppen ──────────────────────────────────────────────────────
    sel_col1, sel_col2, sel_col3 = st.columns([1, 1, 4])
    with sel_col1:
        if st.button(f"Selecteer alles ({len(rows)})", key="hv_sel_all"):
            st.session_state["hv_sel_flag"] = True
            st.session_state["hv_editor_v"] = st.session_state.get("hv_editor_v", 0) + 1
            st.rerun()
    with sel_col2:
        if st.button("Wis selectie", key="hv_desel"):
            st.session_state["hv_sel_flag"] = False
            st.session_state["hv_editor_v"] = st.session_state.get("hv_editor_v", 0) + 1
            st.rerun()

    # Editor-key verandert bij selecteer/wis zodat de widget opnieuw rendert
    editor_key = f"hv_editor_{st.session_state.get('hv_editor_v', 0)}"
    init_select = st.session_state.get("hv_sel_flag", False)

    edited = st.data_editor(
        df[["handle"] + TOON[1:]].assign(_select=init_select),
        column_config={
            "_select":         st.column_config.CheckboxColumn("✔", default=False, width="small"),
            "handle":          st.column_config.TextColumn("Handle", disabled=True, width="medium"),
            "product_title":   st.column_config.TextColumn("Titel (Shopify)", disabled=True, width="large"),
            "vendor":          st.column_config.TextColumn("Merk", disabled=True, width="small"),
            "fase":            st.column_config.TextColumn("Fase", disabled=True, width="small",
                                  help="Bron-fase uit products_raw (bv. fase4, august2026)"),
            "product_status":  st.column_config.TextColumn("Shopify", disabled=True, width="small"),
            "pipeline_status": st.column_config.TextColumn("Pipeline", disabled=True, width="small"),
            "hoofdcategorie":  st.column_config.TextColumn("Categorie", disabled=True, width="medium"),
            "tags_kort":       st.column_config.TextColumn("Tags", disabled=True, width="large",
                                  help="Huidige Shopify tags (eerste 60 tekens)"),
            "heeft_title":     st.column_config.TextColumn("Meta title", disabled=True, width="small",
                                  help="✅ = Shopify heeft al een meta title"),
            "heeft_desc":      st.column_config.TextColumn("Meta desc", disabled=True, width="small",
                                  help="✅ = Shopify heeft al een meta description"),
            "afmetingen":      st.column_config.TextColumn("Afmetingen (cm)", disabled=True, width="medium"),
            "price":           st.column_config.NumberColumn("Prijs", disabled=True, width="small", format="€ %.2f"),
        },
        column_order=["_select", "handle", "product_title", "vendor", "fase",
                      "product_status", "pipeline_status", "hoofdcategorie", "tags_kort",
                      "heeft_title", "heeft_desc", "afmetingen", "price"],
        hide_index=True,
        disabled=["handle"] + TOON[1:],
        width="stretch",
        key=editor_key,
    )

    selected_handles = edited.loc[edited["_select"], "handle"].tolist()
    selected_rows = [r for r in rows if r.get("handle") in selected_handles]
    st.caption(f"**{len(selected_handles)} van {len(rows)} geselecteerd**")

    if not selected_handles:
        st.info("Selecteer producten via de checkboxes hierboven.")
        return

    st.divider()
    st.markdown(f"### ⚡ Acties voor {len(selected_handles)} geselecteerde producten")

    a1, a2, a3 = st.columns(3)

    with a1:
        st.markdown("**A — Exporteer naar Hextom**")
        st.caption("Download als Hextom Excel met beschikbare data.")
        compleet = sum(1 for r in selected_rows if r.get("handle") and r.get("product_title"))
        if st.button(f"📥 Download {len(selected_handles)} als Hextom Excel", key="hv_export"):
            with st.spinner("Excel bouwen..."):
                xlsx = _build_hextom_excel(selected_rows)
            vendor_label = vendor.replace(" ", "_").replace("/", "-")
            sh_label = shopify_status if shopify_status != "Alle" else "mix"
            st.download_button(
                label=f"💾 Download ({len(selected_handles)} producten)",
                data=xlsx,
                file_name=f"hextom_{vendor_label}_{sh_label}_{len(selected_handles)}st.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="hv_dl",
            )

    with a2:
        st.markdown("**B — Reset pipeline-status**")
        st.caption("Alleen voor producten die al in products_curated staan.")
        nieuwe_ps = st.selectbox("Nieuwe status", ["matched", "raw", "ready"], key="hv_new_ps")
        curated_ids = [r["id"] for r in selected_rows if r.get("id") and r.get("pipeline_status")]
        if curated_ids:
            if st.button(f"🔄 Reset {len(curated_ids)} → {nieuwe_ps}", type="primary", key="hv_reset"):
                try:
                    sb = _get_sb()
                    sb.table("products_curated").update({"pipeline_status": nieuwe_ps}) \
                      .in_("id", curated_ids).execute()
                    _load.clear()
                    st.success(f"✅ {len(curated_ids)} producten op `{nieuwe_ps}` gezet.")
                except Exception as e:
                    st.error(f"Fout: {e}")
        else:
            st.caption("_Geen van de geselecteerde producten zit in products_curated._")

    with a3:
        st.markdown("**C — Transformeer & bewerk**")
        st.caption(
            "Ga naar het bewerkscherm: AI genereert NL-titels en meta descriptions, "
            "daarna kun je alles aanpassen (categorie, afmetingen, teksten) "
            "en exporteer je naar Hextom."
        )
        st.caption(f"Geschatte kosten: ~€{len(selected_handles) * 0.002:.2f} ({len(selected_handles)} producten)")
        if st.button(
            f"Transformeer & bewerk {len(selected_handles)} producten",
            type="primary", key="hv_transform_inline"
        ):
            st.session_state["hv_pipeline_rows"] = selected_rows
            st.session_state.pop("hvp_ai_klaar", None)
            st.switch_page("pages/09_Herverwerk_Review.py")

    st.divider()
    with st.expander(f"🔍 Volledigheidscheck ({len(selected_handles)} producten)"):
        check = []
        for r in selected_rows:
            check.append({
                "Handle":        r.get("handle", "—"),
                "SKU":           "✅" if r.get("sku") else "❌",
                "NL-titel":      "✅" if r.get("product_title_nl") else "❌",
                "Meta desc":     "✅" if (r.get("meta_description") or r.get("current_meta_description")) else "❌",
                "Categorie":     "✅" if r.get("hoofdcategorie") else "❌",
                "Foto":          "✅" if r.get("photo_packshot_1") or r.get("has_image") else "❌",
                "In pipeline":   "✅" if r.get("pipeline_status") else "❌",
            })
        st.dataframe(pd.DataFrame(check), hide_index=True, use_container_width=True)
