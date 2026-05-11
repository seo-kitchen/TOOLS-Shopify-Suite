"""Herverwerk-pipeline — stap-voor-stap wizard.

Stap 1 — Namen:        Haiku vertaalt namen → jij keurt goed / past aan
Stap 2 — Categorieën:  mapping-tabel + materiaal/kleur → jij keurt goed / koppelt
Stap 3 — Meta:         Sonnet schrijft descriptions → jij keurt goed / past aan
Stap 4 — Opslaan:      schrijf naar products_curated + download Hextom Excel

Foto's, EAN en barcodes worden nooit aangeraakt.
"""
from __future__ import annotations

import io
import json
import os
import re
import sys
import uuid
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st
from dotenv import load_dotenv

load_dotenv()

_ROOT = Path(__file__).resolve().parent.parent / "dashboard_v2"
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))


@st.cache_resource
def _get_sb():
    from supabase import create_client
    url = os.getenv("SUPABASE_NEW_URL", "")
    key = os.getenv("SUPABASE_NEW_SERVICE_KEY", "") or os.getenv("SUPABASE_NEW_KEY", "")
    if not url or not key:
        raise RuntimeError("SUPABASE_NEW_URL ontbreekt.")
    return create_client(url, key)


# ── Pipeline-status bijhouden (in_process / ready) ───────────────────────────

def _mark_in_process(skus: list[str], data: list[dict] | None = None) -> int:
    """Zet pipeline_status='in_process' in products_curated.

    Behoudt bestaande 'ready' status — alleen 'raw' of nieuwe records worden bijgewerkt.
    Schrijft ook product_title_nl / handle / supplier als die er nog niet zijn,
    zodat je in de dashboard ziet om welke producten het gaat.
    """
    if not skus:
        return 0
    try:
        sb = _get_sb()
        data_by_sku = {r.get("sku"): r for r in (data or []) if r.get("sku")}
        # Bestaande records ophalen in batches
        bestaand: dict[str, dict] = {}
        for i in range(0, len(skus), 200):
            chunk = skus[i:i + 200]
            res = sb.table("products_curated").select("id,sku,pipeline_status") \
                .in_("sku", chunk).execute().data or []
            for r in res:
                bestaand[r["sku"]] = r

        count = 0
        for sku in skus:
            row = data_by_sku.get(sku, {})
            if sku in bestaand:
                if bestaand[sku].get("pipeline_status") == "ready":
                    continue  # niet downgraden
                sb.table("products_curated").update({
                    "pipeline_status": "in_process",
                }).eq("sku", sku).execute()
            else:
                payload = {
                    "sku": sku,
                    "pipeline_status": "in_process",
                    "supplier": row.get("vendor") or row.get("supplier") or "",
                    "product_title_nl": row.get("product_title_nl") or row.get("product_title") or "",
                }
                sb.table("products_curated").insert(payload).execute()
            count += 1
        return count
    except Exception as e:
        # Stil falen: we willen de pipeline niet blokkeren
        return 0


# ── Opslaan / hervatten van pipeline-state ────────────────────────────────────

def _save_draft(naam: str) -> bool:
    """Sla huidige pipeline-state op als draft in seo_learnings."""
    try:
        payload = {
            "hvp_data": st.session_state.get("hvp_data", []),
            "hvp_stap": st.session_state.get("hvp_stap", 1),
            "hvp_s1_gerund": st.session_state.get("hvp_s1_gerund", False),
            "hvp_s2_gerund": st.session_state.get("hvp_s2_gerund", False),
            "hvp_s3_gerund": st.session_state.get("hvp_s3_gerund", False),
            "saved_at": datetime.utcnow().isoformat(),
            "n_producten": len(st.session_state.get("hvp_data", [])),
        }
        _get_sb().table("seo_learnings").insert({
            "stap": "pipeline",
            "rule_type": "pipeline_draft",
            "scope": naam[:80],
            "input_text": naam[:200],
            "action": payload,
            "raw_response": "",
            "status": "draft",
            "applied_at": datetime.utcnow().isoformat(),
            "applied_by": "chef@seokitchen.nl",
        }).execute()
        return True
    except Exception as e:
        st.error(f"Opslaan mislukt: {e}")
        return False


def _list_drafts() -> list[dict]:
    """Haal lijst drafts op (nieuwste eerst, max 20)."""
    try:
        rows = _get_sb().table("seo_learnings").select("*") \
            .eq("rule_type", "pipeline_draft").eq("status", "draft") \
            .order("applied_at", desc=True).limit(20).execute().data or []
        return rows
    except Exception:
        return []


def _restore_draft(draft: dict) -> None:
    """Zet draft terug in session_state."""
    payload = draft.get("action") or {}
    st.session_state["hvp_data"] = payload.get("hvp_data", [])
    st.session_state["hvp_stap"] = payload.get("hvp_stap", 1)
    for k in ("hvp_s1_gerund", "hvp_s2_gerund", "hvp_s3_gerund"):
        if payload.get(k):
            st.session_state[k] = True
        else:
            st.session_state.pop(k, None)
    # hv_pipeline_rows wordt verwacht door render() — vul met laatst opgeslagen data
    st.session_state["hv_pipeline_rows"] = payload.get("hvp_data", [])
    # Markeer als in_process zodat status klopt
    skus = [r.get("sku") for r in (payload.get("hvp_data") or []) if r.get("sku")]
    if skus:
        _mark_in_process(skus, payload.get("hvp_data"))


def _delete_draft(draft_id: str) -> bool:
    try:
        _get_sb().table("seo_learnings").delete().eq("id", draft_id).execute()
        return True
    except Exception:
        return False


def _draft_balk() -> None:
    """Render opslaan/hervat-balk bovenin."""
    n = len(st.session_state.get("hvp_data", []))
    with st.container(border=True):
        c1, c2, c3 = st.columns([3, 2, 2])
        with c1:
            if st.session_state.pop("hvp_drnaam_clear", False):
                st.session_state["hvp_drnaam"] = ""
            naam = st.text_input(
                "Naam voor deze sessie",
                key="hvp_drnaam",
                placeholder=f"bv. Pottery Pots batch {datetime.now().strftime('%d-%m %H:%M')}",
                label_visibility="collapsed",
            )
        with c2:
            if st.button(f"💾 Opslaan voortgang ({n})",
                          disabled=(n == 0),
                          key="hvp_drsave"):
                naam_eff = naam.strip() or f"Sessie {datetime.now().strftime('%d-%m %H:%M')}"
                if _save_draft(naam_eff):
                    st.success(f"✅ Opgeslagen: {naam_eff}")
                    st.session_state["hvp_drnaam_clear"] = True
        with c3:
            drafts = _list_drafts()
            if drafts:
                opties = ["—"] + [
                    f"{(d.get('input_text') or 'naamloos')[:40]} "
                    f"({(d.get('action') or {}).get('n_producten', 0)} prod, "
                    f"stap {(d.get('action') or {}).get('hvp_stap', '?')})"
                    for d in drafts
                ]
                keuze = st.selectbox("Hervat draft", opties, key="hvp_drchoose",
                                       label_visibility="collapsed")
                if keuze != "—":
                    idx = opties.index(keuze) - 1
                    bc1, bc2 = st.columns(2)
                    with bc1:
                        if st.button("📂 Laad", key="hvp_drload"):
                            _restore_draft(drafts[idx])
                            st.success(f"✅ Geladen: {keuze}")
                            st.rerun()
                    with bc2:
                        if st.button("🗑 Verwijder", key="hvp_drdel"):
                            if _delete_draft(drafts[idx]["id"]):
                                st.rerun()


# ── Voortgangsbalk ────────────────────────────────────────────────────────────

def _voortgang(stap: int) -> None:
    stappen = ["1. Namen", "2. Categorie & kleur", "3. Meta description", "4. Opslaan"]
    cols = st.columns(len(stappen))
    for i, (col, label) in enumerate(zip(cols, stappen), 1):
        if i < stap:
            col.markdown(f"<div style='text-align:center;color:#4F7A4A;font-size:13px'>✅ {label}</div>",
                         unsafe_allow_html=True)
        elif i == stap:
            col.markdown(f"<div style='text-align:center;font-weight:600;font-size:13px'>▶ {label}</div>",
                         unsafe_allow_html=True)
        else:
            col.markdown(f"<div style='text-align:center;color:#aaa;font-size:13px'>{label}</div>",
                         unsafe_allow_html=True)
    st.divider()


# ── Chat-correctie per stap ───────────────────────────────────────────────────

_CHAT_PROMPTS = {
    1: """Stap: titel-vertaling.
Beschikbare rule_types:
- title_strip — verwijder woorden/namen uit alle titels (bv. collectie-namen die per ongeluk in de Engelse titel staan)
  action: {"strip": ["Ferd Ridge", "Horace Ridge"]}
- title_replace — vervang X door Y in alle titels
  action: {"replace": [{"from": "...", "to": "..."}]}
- title_instruction — vrije regel voor toekomstige Haiku-vertalingen
  action: {"instruction": "Behoud nooit de collectie-naam in de titel"}""",
    2: """Stap: categorie + materiaal + kleur.
Beschikbare rule_types:
- name_rule — als productnaam zoekwoord bevat → volledige categorie
  action: {"zoekwoord": "...", "hoofdcategorie": "...", "subcategorie": "...", "sub_subcategorie": "...", "is_extra": false}
  BELANGRIJK: VUL ALLE DRIE de niveaus in (hoofd + sub + sub_sub) zodat de regel ook
  werkt voor producten ZONDER leverancier_category. Kies uit de bestaande categorie-boom.
  (is_extra=true: voeg toe als tweede categorie i.p.v. overschrijven)
- name_rule_bulk — meerdere regels tegelijk
  action: {"regels": [{"zoekwoord": "...", "hoofdcategorie": "...", "subcategorie": "...", "sub_subcategorie": "...", "is_extra": false}, ...]}
- translation — en→nl voor materiaal of kleur
  action: {"veld": "materiaal" of "kleur", "en": "...", "nl": "..."}
- category_override — voor specifieke SKU's één categorie zetten (eventueel met 2e subcat)
  action: {"skus": ["..."], "hoofdcategorie": "...", "subcategorie": "...", "sub_subcategorie": "...", "sub_subcategorie_2": "..."}
- bulk_classify — gebruik Sonnet om PER PRODUCT iets te bepalen op basis van data uit products_raw
  Use deze als de gebruiker zegt 'op basis van beschrijving/leverancier-info bepaal X voor alle producten'.
  Beschikbare bron-velden uit products_raw: leverancier_category, leverancier_item_cat,
    materiaal_raw, product_name_raw, designer, kleur_en, hoogte_cm, lengte_cm, breedte_cm
  action: {
    "source_fields": ["leverancier_category", "leverancier_item_cat", "materiaal_raw", "product_name_raw"],
    "target_field": "sub_subcategorie_2",  // of sub_subcategorie / kleur_nl / materiaal_nl
    "criterion": "Bepaal of dit een binnen-pot, buiten-pot of beide is...",
    "allowed_values": ["Bloempotten binnen", "Bloempotten buiten", "beide", ""],
    "skip_if_set": false  // true = sla over als target_field al gevuld is
  }""",
    3: """Stap: meta description.
Beschikbare rule_types:
- meta_replace — vervang X door Y in alle meta descriptions
  action: {"replace": [{"from": "...", "to": "..."}]}
- meta_instruction — vrije regel voor toekomstige Sonnet meta-generatie
  action: {"instruction": "Begin nooit met 'Ontdek'"}""",
}

_CHAT_STAP_NAAM = {1: "titel", 2: "categorie", 3: "meta"}


def _interpret_chat(stap: int, input_text: str, voorbeelden: list[str]) -> dict | None:
    """Vraag Sonnet om NL-feedback te parsen naar een gestructureerde regel."""
    import anthropic
    client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY", ""))

    # Bij stap 2: geef bestaande categorie-boom mee zodat Sonnet kan kiezen
    extra_context = ""
    if stap == 2:
        try:
            cats = _laad_cats()
            tree: dict[str, dict[str, set[str]]] = {}
            for c in cats:
                h = c.get("hoofdcategorie") or ""
                s = c.get("subcategorie") or ""
                ss = c.get("sub_subcategorie") or ""
                if not h:
                    continue
                tree.setdefault(h, {}).setdefault(s, set()).add(ss)
            lines = []
            for h in sorted(tree.keys()):
                lines.append(f"• {h}")
                for s in sorted(tree[h].keys()):
                    sss = sorted(x for x in tree[h][s] if x)
                    if s:
                        lines.append(f"   └ {s}: {', '.join(sss) if sss else '(geen sub-sub)'}")
            if lines:
                extra_context = (
                    "\n\nBESCHIKBARE CATEGORIEËN — kies hieruit voor name_rule "
                    "(hoofd → sub → sub_sub):\n" + "\n".join(lines[:200])
                )
        except Exception:
            pass

    system = f"""Je parset gebruikerfeedback naar JSON-regels voor de SEOkitchen pipeline.

🚨 KRITIEKE REGEL — LETTERLIJK ZIJN:
Als de gebruiker iets expliciet zegt ('X = Y', 'verwijder X', 'verander X in Y',
'alle Z moeten naar W'), MOET je de EXACTE strings die ze opgeven gebruiken in de
action. Vertaal ze NIET. Interpreteer ze NIET. Pas ze niet aan naar 'mooiere'
varianten. Hallucineren = grote fout.

Voorbeelden:
- Gebruiker zegt: 'Plastic Pot inserts = Plastic inzetpot'
  → from MOET zijn 'Plastic Pot inserts' (niet 'kunststof' of 'plastic potten')
  → to MOET zijn 'Plastic inzetpot' (niet 'plasticinzetten')
- Gebruiker zegt: 'alle bijzettafels horen in Bijzettafels'
  → zoekwoord MOET zijn 'bijzettafels' (niet 'tafels' of 'tafel')
  → sub_subcategorie MOET zijn 'Bijzettafels'
- Gebruiker zegt: 'verwijder Ferd Ridge uit titels'
  → strip MOET zijn ['Ferd Ridge'] (precies zo)

Je MAG je verstand gebruiken om:
- Te bepalen welk rule_type past
- De juiste hoofdcategorie/subcategorie te zoeken in de categorieën-boom
  als de gebruiker alleen de sub-subcategorie noemt
- Edge cases af te handelen

Je MAG NIET:
- Strings die de gebruiker letterlijk geeft veranderen, vertalen, of "mooier" maken
- Een breder zoekwoord kiezen dan wat de gebruiker zei
  ('bijzettafels' → 'bijzettafels', NOOIT 'tafel' of 'tafels')

{_CHAT_PROMPTS[stap]}{extra_context}

Output JSON:
{{
  "rule_type": "...",
  "action": {{...}},
  "scope": "alle",
  "explanation": "korte uitleg in 1 zin van wat je gaat doen"
}}

Geef ALLEEN valide JSON, geen markdown of uitleg eromheen."""

    sample = "\n".join(f"- {v}" for v in voorbeelden[:10])
    user = f"Huidige voorbeelden uit de batch:\n{sample}\n\nFeedback van gebruiker:\n{input_text}"

    try:
        resp = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=800,
            system=system,
            messages=[{"role": "user", "content": user}],
        )
        text = resp.content[0].text.strip()
        if text.startswith("```"):
            text = text.split("```")[1]
            if text.startswith("json"):
                text = text[4:]
        return json.loads(text.strip())
    except Exception as e:
        st.error(f"Parse-fout: {e}")
        return None


_RAW_VELDEN_TOEGESTAAN = {
    "leverancier_category", "leverancier_item_cat", "materiaal_raw",
    "product_name_raw", "designer", "kleur_en",
    "hoogte_cm", "lengte_cm", "breedte_cm",
}


def _run_bulk_classify(action: dict, data: list[dict]) -> int:
    """Vraag Sonnet PER PRODUCT iets te bepalen op basis van velden uit products_raw.

    Batch van 25 producten per call, terug-mapping op SKU. Past target_field in
    hvp_data aan, herbouwt tags als sub_subcategorie* geraakt wordt.
    """
    import anthropic
    from execution.transform_v2 import build_tags

    src_fields = [f for f in (action.get("source_fields") or []) if f in _RAW_VELDEN_TOEGESTAAN]
    if not src_fields:
        src_fields = ["leverancier_category", "leverancier_item_cat", "materiaal_raw", "product_name_raw"]
    target = action.get("target_field") or "sub_subcategorie_2"
    criterion = (action.get("criterion") or "").strip()
    allowed = action.get("allowed_values") or []
    skip_if_set = bool(action.get("skip_if_set", False))

    if not criterion:
        st.warning("bulk_classify zonder criterion overgeslagen.")
        return 0

    # Welke SKU's moeten we classificeren?
    sku_lijst = [
        r.get("sku", "") for r in data
        if r.get("sku") and (not skip_if_set or not (r.get(target) or "").strip())
    ]
    if not sku_lijst:
        st.info("Geen producten om te classificeren.")
        return 0

    # Haal source fields op uit products_raw (in batches van 200 voor PostgREST IN-limit)
    sb = _get_sb()
    select = "sku," + ",".join(src_fields)
    raw_map: dict[str, dict] = {}
    for i in range(0, len(sku_lijst), 200):
        chunk = sku_lijst[i:i + 200]
        try:
            res = sb.table("products_raw").select(select).in_("sku", chunk).execute().data or []
            for r in res:
                raw_map[r["sku"]] = r
        except Exception as e:
            st.error(f"Fout bij ophalen products_raw: {e}")
            return 0

    # Sonnet-batches van 25
    client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY", ""))
    BATCH = 25
    classifications: dict[str, str] = {}
    bar = st.progress(0.0, text=f"Sonnet classificeert {len(sku_lijst)} producten...")

    for bi in range(0, len(sku_lijst), BATCH):
        batch_skus = sku_lijst[bi:bi + BATCH]
        items = []
        for sku in batch_skus:
            raw = raw_map.get(sku, {})
            felden = {f: raw.get(f, "") for f in src_fields}
            items.append({"sku": sku, **felden})

        allowed_txt = (
            f"\nToegestane waarden (kies exact één): {json.dumps(allowed, ensure_ascii=False)}"
            if allowed else ""
        )
        system = (
            "Je classificeert producten op basis van leveranciers-velden.\n"
            f"Doel: {criterion}{allowed_txt}\n\n"
            "Output JSON: {\"results\": [{\"sku\": \"...\", \"value\": \"...\"}, ...]}\n"
            "Eén entry per SKU, exact zoals in input. Geen markdown, alleen JSON."
        )
        try:
            resp = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=2000,
                system=system,
                messages=[{"role": "user", "content": json.dumps(items, ensure_ascii=False)}],
            )
            text = resp.content[0].text.strip()
            if text.startswith("```"):
                text = text.split("```")[1]
                if text.startswith("json"):
                    text = text[4:]
            parsed = json.loads(text.strip())
            for entry in (parsed.get("results") or []):
                sku = entry.get("sku", "")
                val = entry.get("value", "")
                if sku:
                    classifications[sku] = val
        except Exception as e:
            st.warning(f"Batch {bi // BATCH + 1} mislukt: {e}")

        bar.progress(min(1.0, (bi + BATCH) / len(sku_lijst)),
                      text=f"Sonnet classificeert {min(bi + BATCH, len(sku_lijst))}/{len(sku_lijst)}...")

    bar.progress(1.0, text="Klaar.")

    # Pas classificaties toe + herbouw tags waar nodig
    raakt = 0
    tag_velden = {"sub_subcategorie", "sub_subcategorie_2"}
    for r in data:
        sku = r.get("sku", "")
        if sku not in classifications:
            continue
        val = classifications[sku]
        if val and (not skip_if_set or not (r.get(target) or "").strip()):
            r[target] = val
            raakt += 1
            if target in tag_velden:
                extra = [r.get("sub_subcategorie_2")] if r.get("sub_subcategorie_2") else None
                r["tags"] = build_tags(
                    r.get("hoofdcategorie", ""),
                    r.get("subcategorie", ""),
                    r.get("sub_subcategorie", ""),
                    r.get("fase", ""),
                    extra_tags=extra,
                )

    return raakt


def _apply_rule_locally(stap: int, rule_type: str, action: dict, data: list[dict]) -> int:
    """Pas regel direct toe op de in-view data. Return: aantal records geraakt."""
    raakt = 0
    act = action or {}

    if stap == 1 and rule_type == "title_strip":
        woorden = [w for w in (act.get("strip") or []) if w]
        for r in data:
            oud = r.get("product_title_nl", "") or ""
            nieuw = oud
            for w in woorden:
                nieuw = re.sub(rf"\s*[-–—]?\s*{re.escape(w)}\s*[-–—]?\s*", " ", nieuw, flags=re.IGNORECASE)
            nieuw = re.sub(r"\s{2,}", " ", nieuw).strip(" -–—")
            if nieuw != oud:
                r["product_title_nl"] = nieuw
                raakt += 1

    elif stap == 1 and rule_type == "title_replace":
        paren = [(p.get("from", ""), p.get("to", "")) for p in (act.get("replace") or [])]
        for r in data:
            oud = r.get("product_title_nl", "") or ""
            nieuw = oud
            for fr, to in paren:
                if fr:
                    nieuw = re.sub(re.escape(fr), to, nieuw, flags=re.IGNORECASE)
            if nieuw != oud:
                r["product_title_nl"] = nieuw
                raakt += 1

    elif stap == 1 and rule_type == "title_instruction":
        # alleen opslaan voor toekomstige Haiku-runs, geen lokale wijziging
        pass

    elif stap == 2 and rule_type == "translation":
        veld = (act.get("veld") or "").lower()
        en = (act.get("en") or "").strip().lower()
        nl = (act.get("nl") or "").strip()
        if veld == "materiaal":
            for r in data:
                if (r.get("materiaal_nl", "") or "").strip().lower() == en:
                    r["materiaal_nl"] = nl
                    raakt += 1
        elif veld == "kleur":
            for r in data:
                if (r.get("kleur_nl", "") or "").strip().lower() == en:
                    r["kleur_nl"] = nl
                    raakt += 1

    elif stap == 2 and rule_type in ("name_rule", "name_rule_bulk"):
        from execution.transform_v2 import build_tags
        regels = [act] if rule_type == "name_rule" else (act.get("regels") or [])
        for rl in regels:
            zoek = (rl.get("zoekwoord") or "").strip().lower()
            sub = rl.get("sub_subcategorie") or ""
            hc = rl.get("hoofdcategorie") or ""
            sc = rl.get("subcategorie") or ""
            is_extra = bool(rl.get("is_extra"))
            if not zoek or not (sub or hc or sc):
                continue
            for r in data:
                naam = (r.get("product_title_nl", "") or r.get("product_title", "") or "").lower()
                # WORD BOUNDARY zodat 'tafels' niet ook 'bijzettafels' raakt
                if re.search(rf"\b{re.escape(zoek)}\b", naam):
                    if is_extra:
                        if sub:
                            if r.get("sub_subcategorie") and r.get("sub_subcategorie") != sub:
                                r["sub_subcategorie_2"] = sub
                            else:
                                r["sub_subcategorie"] = sub
                    else:
                        if hc: r["hoofdcategorie"] = hc
                        if sc:
                            r["subcategorie"] = sc
                            r["collectie"] = sc
                        if sub: r["sub_subcategorie"] = sub
                        if r.get("hoofdcategorie") and r.get("subcategorie") and r.get("sub_subcategorie"):
                            r["_cat_gemapt"] = True
                    # Herbouw tags
                    extra_t = [r.get("sub_subcategorie_2")] if r.get("sub_subcategorie_2") else None
                    r["tags"] = build_tags(
                        r.get("hoofdcategorie", ""),
                        r.get("subcategorie", ""),
                        r.get("sub_subcategorie", ""),
                        r.get("fase", ""),
                        extra_tags=extra_t,
                    )
                    raakt += 1

    elif stap == 2 and rule_type == "category_override":
        skus = set((act.get("skus") or []))
        for r in data:
            if r.get("sku") in skus:
                for k in ("hoofdcategorie", "subcategorie", "sub_subcategorie", "sub_subcategorie_2"):
                    if act.get(k):
                        r[k] = act[k]
                r["collectie"] = act.get("subcategorie", r.get("collectie", ""))
                r["_cat_gemapt"] = True
                raakt += 1

    elif stap == 2 and rule_type == "bulk_classify":
        raakt = _run_bulk_classify(act, data)

    elif stap == 3 and rule_type == "meta_replace":
        paren = [(p.get("from", ""), p.get("to", "")) for p in (act.get("replace") or [])]
        for r in data:
            oud = r.get("meta_description", "") or ""
            nieuw = oud
            for fr, to in paren:
                if fr:
                    nieuw = re.sub(re.escape(fr), to, nieuw, flags=re.IGNORECASE)
            if nieuw != oud:
                r["meta_description"] = nieuw[:160]
                raakt += 1

    elif stap == 3 and rule_type == "meta_instruction":
        pass  # alleen opslaan

    return raakt


def _save_rule(stap: int, rule_type: str, action: dict, scope: str, input_text: str,
               explanation: str, raakt: int) -> bool:
    """Sla regel direct op als 'applied' in seo_learnings."""
    try:
        _get_sb().table("seo_learnings").insert({
            "stap": _CHAT_STAP_NAAM.get(stap, str(stap)),
            "rule_type": rule_type,
            "scope": scope or "alle",
            "input_text": input_text,
            "action": action or {},
            "raw_response": json.dumps({"rule_type": rule_type, "action": action,
                                        "explanation": explanation}),
            "status": "applied",
            "applied_at": datetime.utcnow().isoformat(),
            "applied_by": "chef@seokitchen.nl",
            "applied_rows": raakt,
            "example_before": None,
            "example_after": None,
        }).execute()
        return True
    except Exception as e:
        st.warning(f"Regel toegepast maar niet opgeslagen: {e}")
        return False


def _list_active_rules(stap: int) -> list[dict]:
    """Haal actieve regels op voor een stap, nieuwste eerst."""
    stap_naam = _CHAT_STAP_NAAM.get(stap)
    if not stap_naam:
        return []
    try:
        rows = _get_sb().table("seo_learnings").select("*") \
            .eq("status", "applied").eq("stap", stap_naam) \
            .order("applied_at", desc=True).execute().data or []
        return [r for r in rows if r.get("rule_type") != "pipeline_draft"]
    except Exception:
        return []


def _deactivate_rule(rule_id: str) -> bool:
    try:
        _get_sb().table("seo_learnings").update({"status": "superseded"}) \
          .eq("id", rule_id).execute()
        return True
    except Exception:
        return False


def _rule_samenvatting(L: dict) -> str:
    """Maak een korte 1-regel samenvatting van een regel."""
    rt = L.get("rule_type", "?")
    act = L.get("action") or {}
    if rt == "title_strip":
        woorden = act.get("strip") or []
        return f"strip: {', '.join(woorden[:5])}{' …' if len(woorden) > 5 else ''}"
    if rt == "title_replace" or rt == "meta_replace":
        paren = act.get("replace") or []
        if paren:
            p = paren[0]
            return f"vervang '{p.get('from','')}' → '{p.get('to','')}'"
        return "vervang (leeg)"
    if rt in ("title_instruction", "meta_instruction"):
        return (act.get("instruction") or "")[:80]
    if rt == "name_rule":
        return f"als naam '{act.get('zoekwoord','')}' → {act.get('sub_subcategorie','')}"
    if rt == "name_rule_bulk":
        regels = act.get("regels") or []
        return f"{len(regels)} naam-regels"
    if rt == "translation":
        return f"{act.get('veld','')}: {act.get('en','')} → {act.get('nl','')}"
    if rt == "category_override":
        return f"{len(act.get('skus') or [])} SKU's → {act.get('subcategorie','')}/{act.get('sub_subcategorie','')}"
    if rt == "bulk_classify":
        return f"classificeer {act.get('target_field','?')}: {(act.get('criterion') or '')[:60]}"
    return (L.get("input_text") or "")[:80]


def _normalize_parsed(parsed) -> list[dict]:
    """Sonnet geeft soms dict, soms list, soms {'rules': [...]} terug — normaliseer."""
    if isinstance(parsed, list):
        return [p for p in parsed if isinstance(p, dict)]
    if isinstance(parsed, dict):
        if isinstance(parsed.get("rules"), list):
            return [p for p in parsed["rules"] if isinstance(p, dict)]
        if isinstance(parsed.get("operations"), list):
            return [p for p in parsed["operations"] if isinstance(p, dict)]
        return [parsed]
    return []


def _preview_match_count(stap: int, rt: str, act: dict, data: list[dict]) -> tuple[int, list[str]]:
    """Tel hoeveel records geraakt worden + return enkele voorbeeld-namen."""
    geraakt = []
    if stap == 1 and rt == "title_strip":
        woorden = [w.lower() for w in (act.get("strip") or []) if w]
        for r in data:
            n = (r.get("product_title_nl") or "").lower()
            if any(w in n for w in woorden):
                geraakt.append(r.get("product_title_nl", ""))
    elif stap == 1 and rt == "title_replace":
        paren = [(p.get("from", "").lower(), p.get("to", "")) for p in (act.get("replace") or [])]
        for r in data:
            n = (r.get("product_title_nl") or "").lower()
            if any(fr in n for fr, _ in paren):
                geraakt.append(r.get("product_title_nl", ""))
    elif stap == 2 and rt in ("name_rule", "name_rule_bulk"):
        regels = [act] if rt == "name_rule" else (act.get("regels") or [])
        for r in data:
            naam = (r.get("product_title_nl") or r.get("product_title") or "").lower()
            for rl in regels:
                zoek = (rl.get("zoekwoord") or "").strip().lower()
                if zoek and re.search(rf"\b{re.escape(zoek)}\b", naam):
                    geraakt.append(f"{r.get('sku','')} — {r.get('product_title_nl','')}")
                    break
    elif stap == 2 and rt == "category_override":
        skus = set(act.get("skus") or [])
        for r in data:
            if r.get("sku") in skus:
                geraakt.append(f"{r.get('sku','')} — {r.get('product_title_nl','')}")
    elif stap == 2 and rt == "bulk_classify":
        skip = bool(act.get("skip_if_set", False))
        target = act.get("target_field", "")
        for r in data:
            if not skip or not (r.get(target) or "").strip():
                geraakt.append(f"{r.get('sku','')} — {r.get('product_title_nl','')}")
    elif stap == 2 and rt == "translation":
        veld = (act.get("veld") or "").lower()
        en = (act.get("en") or "").strip().lower()
        key = "materiaal_nl" if veld == "materiaal" else "kleur_nl"
        for r in data:
            if (r.get(key) or "").strip().lower() == en:
                geraakt.append(f"{r.get('sku','')} — {r.get(key,'')}")
    elif stap == 3 and rt == "meta_replace":
        paren = [(p.get("from", "").lower(), p.get("to", "")) for p in (act.get("replace") or [])]
        for r in data:
            n = (r.get("meta_description") or "").lower()
            if any(fr in n for fr, _ in paren):
                geraakt.append(f"{r.get('sku','')} — {(r.get('meta_description') or '')[:60]}")
    return len(geraakt), geraakt[:10]


def _render_preview(stap: int, data: list[dict], pending_key: str,
                    pending_txt_key: str, flag_clear: str) -> None:
    """Toon Sonnet's voorstel met edit-mogelijkheid + bevestig/annuleer."""
    parsed = st.session_state.get(pending_key) or {}
    txt = st.session_state.get(pending_txt_key, "")
    regels = _normalize_parsed(parsed)

    if not regels:
        st.warning("Sonnet gaf niets bruikbaars terug.")
        if st.button("← Terug", key=f"hvp_pv_back_{stap}"):
            st.session_state.pop(pending_key, None)
            st.session_state.pop(pending_txt_key, None)
            st.session_state[flag_clear] = True
            st.rerun()
        return

    st.markdown(f"**🔎 Sonnet stelt {len(regels)} regel(s) voor — controleer vóór toepassen:**")
    st.caption(f"Op basis van: _{txt}_")

    # Cache bewerkte regels per index
    bewerkt: list[dict] = []
    cats = _laad_cats() if stap == 2 else []

    for i, rl in enumerate(regels):
        rt = rl.get("rule_type", "?")
        act = dict(rl.get("action") or {})
        expl = rl.get("explanation", "")

        with st.container(border=True):
            st.markdown(f"**Regel {i+1}** · `{rt}`")
            if expl:
                st.caption(expl)

            # Preview match-count + voorbeelden
            n_hit, voorbeelden = _preview_match_count(stap, rt, act, data)
            if rt in ("title_instruction", "meta_instruction"):
                st.info("Geen lokale wijziging — alleen opgeslagen voor toekomstige runs.")
            else:
                st.markdown(f"**→ {n_hit} producten worden geraakt**")
                if voorbeelden:
                    with st.expander(f"Voorbeelden ({min(len(voorbeelden), 10)})", expanded=False):
                        for v in voorbeelden:
                            st.text(v)

            # ── Edit-UI voor stap 2 name_rule (categorie kiezen / nieuw maken) ──
            if stap == 2 and rt in ("name_rule", "name_rule_bulk"):
                NEW = "+ Nieuwe…"
                hoofd_keuzes = sorted(set(c["hoofdcategorie"] for c in cats if c.get("hoofdcategorie")))
                if rt == "name_rule":
                    subregels = [act]
                else:
                    subregels = act.get("regels") or []
                nieuw_subregels = []
                for j, sr in enumerate(subregels):
                    st.markdown(f"_Sub-regel {j+1}_")
                    zoek = st.text_input(
                        "Zoekwoord", value=sr.get("zoekwoord", ""),
                        key=f"hvp_pv_zoek_{stap}_{i}_{j}",
                    )
                    cA, cB, cC = st.columns(3)
                    with cA:
                        huidig_hc = sr.get("hoofdcategorie", "")
                        opties_hc = hoofd_keuzes + [NEW]
                        idx_hc = opties_hc.index(huidig_hc) if huidig_hc in opties_hc else len(opties_hc) - 1
                        hc_sel = st.selectbox(
                            "Hoofdcat",
                            opties_hc + ([huidig_hc] if huidig_hc and huidig_hc not in opties_hc else []),
                            index=(opties_hc.index(huidig_hc) if huidig_hc in opties_hc
                                    else (len(opties_hc) if huidig_hc else idx_hc)),
                            key=f"hvp_pv_hc_{stap}_{i}_{j}",
                        )
                        if hc_sel == NEW:
                            hc = st.text_input("Nieuwe hoofd", value=huidig_hc if huidig_hc not in hoofd_keuzes else "",
                                                 key=f"hvp_pv_hcnew_{stap}_{i}_{j}").strip()
                        else:
                            hc = hc_sel
                    with cB:
                        sub_keuzes = sorted(set(c["subcategorie"] for c in cats
                                                  if c.get("hoofdcategorie") == hc and c.get("subcategorie")))
                        huidig_sc = sr.get("subcategorie", "")
                        if hc and hc not in hoofd_keuzes:
                            sc = st.text_input("Nieuwe sub", value=huidig_sc,
                                                key=f"hvp_pv_scnew_{stap}_{i}_{j}").strip()
                        else:
                            opties_sc = sub_keuzes + [NEW]
                            if huidig_sc and huidig_sc not in opties_sc:
                                opties_sc.append(huidig_sc)
                            sc_sel = st.selectbox(
                                "Subcat",
                                opties_sc,
                                index=opties_sc.index(huidig_sc) if huidig_sc in opties_sc else 0,
                                key=f"hvp_pv_sc_{stap}_{i}_{j}",
                            )
                            if sc_sel == NEW:
                                sc = st.text_input("Naam nieuwe sub", value="",
                                                    key=f"hvp_pv_scinp_{stap}_{i}_{j}").strip()
                            else:
                                sc = sc_sel
                    with cC:
                        ssub_keuzes = sorted(set(c["sub_subcategorie"] for c in cats
                                                   if c.get("hoofdcategorie") == hc
                                                   and c.get("subcategorie") == sc
                                                   and c.get("sub_subcategorie")))
                        huidig_ss = sr.get("sub_subcategorie", "")
                        nieuwe_hc_of_sc = (hc and hc not in hoofd_keuzes) or (sc and sc not in sub_keuzes)
                        if nieuwe_hc_of_sc:
                            ss = st.text_input("Nieuwe sub-sub", value=huidig_ss,
                                                 key=f"hvp_pv_ssnew_{stap}_{i}_{j}").strip()
                        else:
                            opties_ss = ssub_keuzes + [NEW]
                            if huidig_ss and huidig_ss not in opties_ss:
                                opties_ss.append(huidig_ss)
                            ss_sel = st.selectbox(
                                "Sub-subcat",
                                opties_ss,
                                index=opties_ss.index(huidig_ss) if huidig_ss in opties_ss else 0,
                                key=f"hvp_pv_ss_{stap}_{i}_{j}",
                            )
                            if ss_sel == NEW:
                                ss = st.text_input("Naam nieuwe sub-sub", value="",
                                                    key=f"hvp_pv_ssinp_{stap}_{i}_{j}").strip()
                            else:
                                ss = ss_sel

                    nieuw_subregels.append({
                        "zoekwoord": zoek,
                        "hoofdcategorie": hc,
                        "subcategorie": sc,
                        "sub_subcategorie": ss,
                        "is_extra": sr.get("is_extra", False),
                    })

                if rt == "name_rule":
                    new_act = nieuw_subregels[0] if nieuw_subregels else act
                else:
                    new_act = {"regels": nieuw_subregels}
                bewerkt.append({"rule_type": rt, "action": new_act,
                                "explanation": expl, "scope": rl.get("scope", "alle")})
            else:
                # Andere rule_types: toon JSON, niet bewerkbaar
                with st.expander("Action (JSON)", expanded=False):
                    st.json(act)
                bewerkt.append(rl)

    # Bevestig / Annuleer
    cA, cB = st.columns([1, 1])
    with cA:
        if st.button("✅ Bevestig + pas toe", type="primary", key=f"hvp_pv_ok_{stap}"):
            total = 0
            for rl in bewerkt:
                rt = rl.get("rule_type")
                act = rl.get("action") or {}
                if rt == "unclear" or not rt:
                    continue
                raakt = _apply_rule_locally(stap, rt, act, data)
                total += raakt
                _save_rule(stap, rt, act, rl.get("scope", "alle"), txt, rl.get("explanation", ""), raakt)
            st.session_state["hvp_data"] = data
            st.session_state.pop(pending_key, None)
            st.session_state.pop(pending_txt_key, None)
            st.session_state[flag_clear] = True
            st.success(f"✅ {len(bewerkt)} regel(s) toegepast · {total} records aangepast")
            st.rerun()
    with cB:
        if st.button("❌ Annuleer", key=f"hvp_pv_cancel_{stap}"):
            st.session_state.pop(pending_key, None)
            st.session_state.pop(pending_txt_key, None)
            st.session_state[flag_clear] = True
            st.rerun()


def _chat_box(stap: int, kolom_voorbeeld: str) -> None:
    """Render chat-input onderaan een stap.

    kolom_voorbeeld: welke key uit hvp_data tonen als sample (bv. 'product_title_nl').
    """
    data: list[dict] = st.session_state["hvp_data"]
    voorbeelden = [r.get(kolom_voorbeeld, "") for r in data if r.get(kolom_voorbeeld)][:10]

    key_in = f"hvp_chat_in_{stap}"
    key_btn = f"hvp_chat_btn_{stap}"
    key_clr = f"hvp_chat_clr_{stap}"
    flag_clear = f"hvp_chat_clear_{stap}"

    # Clear-flag uitvoeren VOOR de widget wordt aangemaakt (Streamlit-regel)
    if st.session_state.pop(flag_clear, False):
        st.session_state[key_in] = ""

    actief = _list_active_rules(stap)

    with st.expander(
        f"💬 Correctie voor deze stap ({len(actief)} actieve regels)",
        expanded=False,
    ):
        # ── Actieve regels met deactiveer-knop ──
        if actief:
            st.markdown("**Actieve regels (geldt voor toekomstige batches):**")
            for L in actief[:30]:
                sam = _rule_samenvatting(L)
                rt = L.get("rule_type", "?")
                ts = (L.get("applied_at") or "")[:16].replace("T", " ")
                cA, cB = st.columns([8, 1])
                with cA:
                    st.markdown(
                        f"<div style='font-family:Inter;font-size:13px;'>"
                        f"<code style='background:rgba(174,205,246,0.18);padding:1px 6px;border-radius:3px;font-size:11px'>"
                        f"{rt}</code> {sam} "
                        f"<span style='color:#888;font-size:11px'>· {ts}</span></div>",
                        unsafe_allow_html=True,
                    )
                with cB:
                    if st.button("🗑", key=f"hvp_rd_{stap}_{L['id']}",
                                  help="Deactiveer deze regel"):
                        if _deactivate_rule(L["id"]):
                            st.rerun()
            if len(actief) > 30:
                st.caption(f"… + {len(actief) - 30} oudere regels (niet getoond)")
            st.divider()

        st.caption(
            "Typ in normaal Nederlands wat er mis gaat. De fix wordt nu toegepast én "
            "opgeslagen zodat het in toekomstige runs ook automatisch gebeurt."
        )

        txt = st.text_area(
            "Feedback",
            height=80,
            key=key_in,
            placeholder={
                1: "bv. 'Verwijder Ferd Ridge en Horace Ridge uit alle titels — dat zijn collectie-namen'",
                2: "bv. 'Producten met storage_pot in de naam zijn altijd Voorraadpotten'",
                3: "bv. 'Begin nooit met Ontdek, gebruik liever de productnaam'",
            }.get(stap, ""),
        )
        c1, c2 = st.columns([1, 5])
        with c1:
            doe = st.button("Pas toe + onthou", type="primary",
                            disabled=not txt.strip(), key=key_btn)
        with c2:
            if st.button("Wis", key=key_clr):
                st.session_state[flag_clear] = True
                st.rerun()

        pending_key = f"hvp_chat_pending_{stap}"
        pending_txt_key = f"hvp_chat_pending_txt_{stap}"

        # ── Preview-modus: toon eerst wat Sonnet heeft besloten ──
        if st.session_state.get(pending_key):
            _render_preview(stap, data, pending_key, pending_txt_key, flag_clear)
            return

        if doe and txt.strip():
            with st.spinner("Sonnet parset feedback..."):
                parsed = _interpret_chat(stap, txt.strip(), voorbeelden)
            if not parsed:
                return
            st.session_state[pending_key] = parsed
            st.session_state[pending_txt_key] = txt.strip()
            st.rerun()


# ── Stap 1: Namen ─────────────────────────────────────────────────────────────

def _stap_namen() -> None:
    data: list[dict] = st.session_state["hvp_data"]
    n = len(data)

    st.markdown(f"### Namen vertalen ({n} producten)")
    st.caption("Haiku vertaalt alle namen in één batch. Pas aan waar nodig, dan goedkeuren.")

    if not st.session_state.get("hvp_s1_gerund"):
        kosten = max(n * 0.0001, 0.01)
        st.caption(f"Geschatte kosten: ~€{kosten:.2f} (Haiku batch)")
        if st.button("Vertaal namen via Haiku", type="primary", key="hvp_s1_run"):
            try:
                from execution.transform_v2 import (
                    vertaal_productnamen_batch, get_claude, load_active_learnings,
                )
                sb = _get_sb()
                claude = get_claude()
                title_lr = [L for L in load_active_learnings(sb) if L.get("stap") == "titel"]
                namen = [r.get("product_title", "") or r.get("product_name_raw", "") for r in data]
                with st.spinner(f"Haiku vertaalt ({len(title_lr)} actieve titel-regels)..."):
                    vertaling = vertaal_productnamen_batch(namen, claude, title_learnings=title_lr)
                for r in data:
                    raw = r.get("product_title", "") or r.get("product_name_raw", "") or ""
                    r["product_title_nl"] = vertaling.get(raw.strip(), raw)
                st.session_state["hvp_data"] = data
                st.session_state["hvp_s1_gerund"] = True
                st.rerun()
            except Exception as e:
                st.error(f"Fout: {e}")
        return

    # ── Snelle vervanging: letterlijk X → Y, zoekt in EN én NL titel ──
    with st.expander("🎯 Snelle vervanging: X → Y in titels (geen AI)", expanded=False):
        st.caption(
            "Letterlijke tekst-vervanging in titels. Zoekt eerst in de NL-titel; "
            "als daar geen match is, zoekt 'ie ook in de originele Engelse titel "
            "(handig om foute vertalingen van Haiku te corrigeren). "
            "Hoofdletter-ongevoelig en tolerant voor copy-paste verschillen (spaties, NBSP, leestekens)."
        )
        cV1, cV2 = st.columns(2)
        with cV1:
            qr_from = st.text_input("Van", key="hvp_qrv_from",
                                     placeholder="bv. Plastic Pot inserts")
        with cV2:
            qr_to = st.text_input("Naar", key="hvp_qrv_to",
                                   placeholder="bv. Plastic inzetpot")

        # Normalisatie-functie: maakt match tolerant voor whitespace/NBSP/etc.
        def _norm(s: str) -> str:
            s = (s or "").replace("\xa0", " ").replace(" ", " ").replace("​", "")
            s = re.sub(r"\s+", " ", s).strip()
            return s.lower()

        def _build_fuzzy_pattern(needle: str) -> re.Pattern:
            """Bouwt een regex die tolerant is voor whitespace-verschillen.
            'Plastic Pot inserts' matcht ook 'Plastic  Pot Inserts' of met NBSP."""
            woorden = re.split(r"\s+", needle.strip())
            return re.compile(r"\s+".join(re.escape(w) for w in woorden), re.IGNORECASE)

        from_raw = (qr_from or "").strip()
        to_v = (qr_to or "").strip()
        n_nl = 0
        n_en = 0
        voorbeelden = []
        n_hit = 0
        if from_raw:
            patroon = _build_fuzzy_pattern(from_raw)
            for r in data:
                nl = (r.get("product_title_nl", "") or "").replace("\xa0", " ")
                en = (r.get("product_title", "") or r.get("product_name_raw", "") or "").replace("\xa0", " ")
                if patroon.search(nl):
                    n_nl += 1
                    if len(voorbeelden) < 5:
                        voorbeelden.append(("NL", nl, patroon.sub(to_v, nl)))
                elif patroon.search(en):
                    n_en += 1
                    if len(voorbeelden) < 5:
                        voorbeelden.append(("EN→NL", nl, patroon.sub(to_v, en)))
            n_hit = n_nl + n_en
            if n_hit == 0:
                st.warning(
                    f"📊 0 titels gevonden met '{from_raw}'. "
                    "Tip: kopieer-plak een stuk van een bestaande titel uit de tabel hieronder, of typ alleen kernwoorden."
                )
                # Toon 3 voorbeelden uit huidige data zodat user ziet hoe ze er ECHT uit zien
                with st.expander("Zo zien je huidige NL-titels eruit (eerste 5):", expanded=True):
                    for r in data[:5]:
                        st.text(f"NL: {r.get('product_title_nl','')}")
                        st.text(f"EN: {r.get('product_title','') or r.get('product_name_raw','')}")
                        st.divider()
            else:
                st.info(f"📊 {n_hit} titels worden aangepast"
                        + (f" ({n_nl} match in NL, {n_en} match in EN→NL)" if n_en else ""))
                if voorbeelden:
                    with st.expander("Voorbeeld vóór → na", expanded=True):
                        for soort, v_oud, v_nw in voorbeelden:
                            st.caption(soort)
                            st.text(f"OUD:  {v_oud}")
                            st.text(f"NIEUW: {v_nw}")
                            st.divider()

        onthouden = st.checkbox("Onthouden voor toekomstige batches", value=True, key="hvp_qrv_mem")

        if st.button(
            f"✅ Pas toe op {n_hit} titels",
            type="primary",
            disabled=not (from_raw and n_hit > 0),
            key="hvp_qrv_apply",
        ):
            patroon = _build_fuzzy_pattern(from_raw)
            for r in data:
                nl = (r.get("product_title_nl", "") or "").replace("\xa0", " ")
                en = (r.get("product_title", "") or r.get("product_name_raw", "") or "").replace("\xa0", " ")
                if patroon.search(nl):
                    r["product_title_nl"] = patroon.sub(to_v, nl)
                elif patroon.search(en):
                    r["product_title_nl"] = patroon.sub(to_v, en)
            st.session_state["hvp_data"] = data
            if onthouden:
                try:
                    _get_sb().table("seo_learnings").insert({
                        "stap": "titel",
                        "rule_type": "title_replace",
                        "scope": "alle",
                        "input_text": f"snelle vervanging: {from_raw} → {to_v}",
                        "action": {"replace": [{"from": from_raw, "to": to_v}], "match_in_en": True, "fuzzy_ws": True},
                        "status": "applied",
                        "applied_at": datetime.utcnow().isoformat(),
                        "applied_by": "chef@seokitchen.nl",
                        "applied_rows": n_hit,
                    }).execute()
                except Exception as e:
                    st.warning(f"Toegepast maar niet opgeslagen: {e}")
            msg = f"✅ {n_hit} titels aangepast"
            msg += " · onthouden" if onthouden else " · niet onthouden"
            st.success(msg)
            st.rerun()

    # Tabel om namen te bewerken
    df = pd.DataFrame([{
        "sku":           r.get("sku", ""),
        "vendor":        r.get("vendor", "") or r.get("supplier", ""),
        "naam_origineel": r.get("product_title", "") or r.get("product_name_raw", ""),
        "naam_nl":       r.get("product_title_nl", ""),
    } for r in data])

    edited = st.data_editor(
        df,
        column_config={
            "sku":            st.column_config.TextColumn("SKU",        disabled=True, width="small"),
            "vendor":         st.column_config.TextColumn("Merk",       disabled=True, width="small"),
            "naam_origineel": st.column_config.TextColumn("Origineel",  disabled=True, width="large"),
            "naam_nl":        st.column_config.TextColumn("Naam NL ✏️", disabled=False, width="large"),
        },
        hide_index=True,
        disabled=["sku", "vendor", "naam_origineel"],
        width="stretch",
        key="hvp_edit_s1",
    )

    c1, c2, c3 = st.columns([2, 2, 2])
    with c1:
        if st.button("↺ Opnieuw vertalen", key="hvp_s1_reset"):
            st.session_state.pop("hvp_s1_gerund", None)
            st.rerun()
    with c3:
        if st.button("Goedkeuren → Stap 2", type="primary", key="hvp_s1_ok"):
            # Sla bewerkte namen op
            for _, row in edited.iterrows():
                for r in st.session_state["hvp_data"]:
                    if r.get("sku") == row["sku"]:
                        r["product_title_nl"] = row["naam_nl"]
                        break
            st.session_state["hvp_stap"] = 2
            _save_draft(f"Auto na stap 1 — {datetime.now().strftime('%d-%m %H:%M')}")
            st.rerun()

    _chat_box(stap=1, kolom_voorbeeld="product_title_nl")


# ── Stap 2: Categorie, materiaal, kleur ──────────────────────────────────────

@st.cache_data(ttl=300, show_spinner=False)
def _laad_cats() -> list[dict]:
    try:
        rows = _get_sb().table("seo_category_mapping").select(
            "hoofdcategorie,subcategorie,sub_subcategorie"
        ).execute().data or []
        seen, result = set(), []
        for r in rows:
            k = (r.get("hoofdcategorie",""), r.get("subcategorie",""), r.get("sub_subcategorie",""))
            if k not in seen and k[0]:
                seen.add(k)
                result.append({"hoofdcategorie": k[0], "subcategorie": k[1], "sub_subcategorie": k[2]})
        return sorted(result, key=lambda x: (x["hoofdcategorie"], x["subcategorie"], x["sub_subcategorie"]))
    except Exception:
        return []


def _stap_categorie_kleur() -> None:
    data: list[dict] = st.session_state["hvp_data"]
    n = len(data)

    st.markdown(f"### Categorie, materiaal & kleur ({n} producten)")
    st.caption(
        "Categorisatie via mapping-tabel. Materiaal en kleur via vertaaltabellen "
        "(Sonnet alleen als fallback). Producten zonder mapping krijgen een categorie-kiezer."
    )

    if not st.session_state.get("hvp_s2_gerund"):
        kosten = n * 0.0005
        st.caption(f"Geschatte kosten: ~€{kosten:.2f} (alleen Sonnet voor onbekende materialen/kleuren)")
        if st.button("Run categorisatie + vertaling", type="primary", key="hvp_s2_run"):
            try:
                from execution.transform_v2 import (
                    lookup_category, translate_material, translate_color,
                    build_tags, apply_name_rules, apply_translation_learnings,
                    load_active_learnings, get_claude, get_supabase
                )
                sb = _get_sb()
                claude = get_claude()

                all_learnings = load_active_learnings(sb)
                cat_learnings = [L for L in all_learnings if L.get("stap") == "categorie"]
                extra_mat, extra_kl = apply_translation_learnings(
                    [L for L in all_learnings if L.get("stap") == "vertaling"]
                )

                bar = st.progress(0.0, text="Bezig...")
                for idx, r in enumerate(data):
                    bar.progress((idx + 1) / n, text=f"{idx+1}/{n}: {r.get('sku','')}")
                    sku = r.get("sku", "")

                    # Haal raw data op
                    raw_data = {}
                    if sku:
                        try:
                            res = sb.table("products_raw").select(
                                "sku,leverancier_category,leverancier_item_cat,"
                                "materiaal_raw,kleur_en,designer,giftbox,giftbox_qty,fase"
                            ).eq("sku", sku).execute().data or []
                            if res:
                                raw_data = res[0]
                        except Exception:
                            pass

                    # Categorie lookup
                    cat_row = lookup_category(
                        sb,
                        raw_data.get("leverancier_category", ""),
                        raw_data.get("leverancier_item_cat", ""),
                    )
                    if cat_row:
                        r["hoofdcategorie"] = cat_row["hoofdcategorie"]
                        r["subcategorie"] = cat_row["subcategorie"]
                        r["sub_subcategorie"] = cat_row["sub_subcategorie"]
                        r["collectie"] = cat_row["subcategorie"]
                        r["_cat_gemapt"] = True
                    else:
                        r["_cat_gemapt"] = False
                        r["_leverancier_category"] = raw_data.get("leverancier_category", "")
                        r["_leverancier_item_cat"] = raw_data.get("leverancier_item_cat", "")

                    # Name-rule learnings — kunnen nu de volledige triple zetten
                    updates = {
                        "hoofdcategorie":   r.get("hoofdcategorie", ""),
                        "subcategorie":     r.get("subcategorie", ""),
                        "sub_subcategorie": r.get("sub_subcategorie", ""),
                        "collectie":        r.get("collectie", ""),
                    }
                    # Geef ook product_title_nl mee zodat name_rule kan matchen op NL-naam
                    raw_data["product_title_nl"] = r.get("product_title_nl", "")
                    applied = apply_name_rules(raw_data, updates, cat_learnings)
                    if applied > 0:
                        for k in ("hoofdcategorie", "subcategorie", "sub_subcategorie", "collectie"):
                            if updates.get(k):
                                r[k] = updates[k]
                        if updates.get("hoofdcategorie") and updates.get("subcategorie") and updates.get("sub_subcategorie"):
                            r["_cat_gemapt"] = True

                    # Tags
                    r["tags"] = build_tags(
                        r.get("hoofdcategorie", ""),
                        r.get("subcategorie", ""),
                        r.get("sub_subcategorie", ""),
                        raw_data.get("fase", ""),
                        extra_tags=updates.get("_extra_tags"),
                    )

                    # Materiaal + kleur
                    mat_raw = raw_data.get("materiaal_raw", "") or ""
                    r["materiaal_nl"] = translate_material(mat_raw, claude, extra_mat) if mat_raw else ""
                    kl_en = raw_data.get("kleur_en", "") or ""
                    naam_raw = raw_data.get("product_name_raw", "") or r.get("product_title", "")
                    kl_filter, _ = translate_color(kl_en, naam_raw, claude, extra_kl) if kl_en else ("", "")
                    r["kleur_nl"] = kl_filter

                    # Designer, fase bewaren
                    if raw_data.get("designer"):
                        r["designer"] = raw_data["designer"]
                    if raw_data.get("fase"):
                        r["fase"] = raw_data["fase"]
                    if raw_data.get("giftbox"):
                        r["giftbox"] = raw_data["giftbox"]
                    if raw_data.get("giftbox_qty"):
                        r["giftbox_qty"] = raw_data["giftbox_qty"]

                bar.progress(1.0, text="Klaar.")
                st.session_state["hvp_data"] = data
                st.session_state["hvp_s2_gerund"] = True
                st.rerun()
            except Exception as e:
                st.error(f"Fout: {e}")
                import traceback
                with st.expander("Traceback"):
                    st.code(traceback.format_exc())
        return

    # ── Categorie-koppeling voor producten zonder mapping ──
    ongemapt = [r for r in data if not r.get("_cat_gemapt")]
    if ongemapt:
        cats = _laad_cats()
        hoofdcats = sorted(set(c["hoofdcategorie"] for c in cats))
        combo_to_rows: dict[tuple, list] = defaultdict(list)
        for r in ongemapt:
            k = (r.get("_leverancier_category",""), r.get("_leverancier_item_cat",""))
            combo_to_rows[k].append(r)

        with st.expander(
            f"⚠️ {len(ongemapt)} producten zonder categorie-mapping — koppel hier",
            expanded=True
        ):
            st.caption("Koppel de juiste categorie. Die wordt ook opgeslagen in de mapping-tabel voor toekomstige batches.")
            for idx, ((lc, lic), combo_rows) in enumerate(combo_to_rows.items()):
                st.markdown(f"**{lc}** / {lic or '(leeg)'}  —  {len(combo_rows)} producten")
                # Toon de daadwerkelijke productnamen zodat je ziet waar het over gaat
                namen_lijst = [
                    f"{r.get('sku','')} — {r.get('product_title_nl','') or r.get('product_title','')}"
                    for r in combo_rows
                ]
                with st.container(border=True):
                    for n in namen_lijst[:8]:
                        st.text(n)
                    if len(namen_lijst) > 8:
                        st.caption(f"+ {len(namen_lijst) - 8} meer…")
                NEW = "+ Nieuwe categorie…"

                c1, c2, c3 = st.columns(3)
                with c1:
                    hc_sel = st.selectbox(
                        "Hoofdcategorie",
                        hoofdcats + [NEW],
                        key=f"hck2_{idx}",
                    )
                    if hc_sel == NEW:
                        hc = st.text_input(
                            "Naam nieuwe hoofdcategorie",
                            key=f"hck2new_{idx}",
                            placeholder="bv. Tuin & Buiten",
                        ).strip()
                    else:
                        hc = hc_sel
                with c2:
                    if hc and hc not in hoofdcats:
                        # Nieuwe hoofdcat: dus sub ook nieuw
                        sc = st.text_input(
                            "Subcategorie (nieuw)",
                            key=f"sck2new_{idx}",
                            placeholder="bv. Bloempotten buiten",
                        ).strip()
                    else:
                        subcats = sorted(set(c["subcategorie"] for c in cats if c["hoofdcategorie"] == hc and c["subcategorie"]))
                        sc_sel = st.selectbox(
                            "Subcategorie",
                            (subcats or []) + [NEW],
                            key=f"sck2_{idx}",
                        )
                        if sc_sel == NEW:
                            sc = st.text_input(
                                "Naam nieuwe subcategorie",
                                key=f"sck2new_{idx}",
                                placeholder="bv. Bloempotten buiten",
                            ).strip()
                        else:
                            sc = sc_sel
                with c3:
                    if (hc and hc not in hoofdcats) or (sc and (hc not in hoofdcats or sc not in [c["subcategorie"] for c in cats if c["hoofdcategorie"] == hc])):
                        ssc = st.text_input(
                            "Sub-subcategorie (nieuw)",
                            key=f"ssck2new_{idx}",
                            placeholder="bv. Bloempotten buiten",
                        ).strip()
                    else:
                        subsubs = sorted(set(c["sub_subcategorie"] for c in cats if c["hoofdcategorie"] == hc and c["subcategorie"] == sc and c["sub_subcategorie"]))
                        ssc_sel = st.selectbox(
                            "Sub-subcategorie",
                            (subsubs or []) + [NEW],
                            key=f"ssck2_{idx}",
                        )
                        if ssc_sel == NEW:
                            ssc = st.text_input(
                                "Naam nieuwe sub-subcategorie",
                                key=f"ssck2new_{idx}",
                                placeholder="bv. Bloempotten buiten",
                            ).strip()
                        else:
                            ssc = ssc_sel

                # Optionele tweede sub-subcategorie (bv. Bloempotten binnen + buiten)
                alle_subsubs = sorted(set(c["sub_subcategorie"] for c in cats if c["sub_subcategorie"] and c["sub_subcategorie"] != ssc))
                ssc2_sel = st.selectbox(
                    "+ Tweede sub-subcategorie (optioneel)",
                    ["—"] + alle_subsubs + [NEW],
                    key=f"ssck2b_{idx}",
                    help="Bv. een product hoort bij zowel Bloempotten binnen als buiten",
                )
                if ssc2_sel == NEW:
                    ssc2 = st.text_input(
                        "Naam tweede sub-subcategorie",
                        key=f"ssck2bnew_{idx}",
                        placeholder="bv. Bloempotten op pootjes",
                    ).strip()
                else:
                    ssc2 = ssc2_sel

                # Validatie + koppel-knop
                ongeldig = []
                if not hc: ongeldig.append("hoofdcategorie")
                if not sc or sc == "—": ongeldig.append("subcategorie")
                if not ssc or ssc == "—": ongeldig.append("sub-subcategorie")

                koppel_btn = st.button(
                    f"Koppel ({len(combo_rows)} producten)",
                    key=f"koppel2_{idx}",
                    disabled=bool(ongeldig),
                    help=("Vul nog in: " + ", ".join(ongeldig)) if ongeldig else None,
                )

                if koppel_btn:
                    for r in combo_rows:
                        r["hoofdcategorie"] = hc
                        r["subcategorie"] = sc
                        r["sub_subcategorie"] = ssc
                        if ssc2 and ssc2 != "—":
                            r["sub_subcategorie_2"] = ssc2
                        r["collectie"] = sc
                        r["_cat_gemapt"] = True
                    try:
                        sb = _get_sb()
                        sb.table("seo_category_mapping").insert({
                            "id": str(uuid.uuid4()),
                            "leverancier_category": lc,
                            "leverancier_item_cat": lic or "?",
                            "hoofdcategorie": hc,
                            "subcategorie": sc,
                            "sub_subcategorie": ssc,
                        }).execute()
                        _laad_cats.clear()
                    except Exception:
                        pass
                    st.session_state["hvp_data"] = data
                    st.success(f"✅ Gekoppeld + opgeslagen in mapping-tabel.")
                    st.rerun()
                st.markdown("---")

    # ── Snelle regel — deterministische zoekwoord → categorie toewijzing ──
    with st.expander("🎯 Snelle regel: zoekwoord → categorie (geen AI, 100% voorspelbaar)", expanded=False):
        st.caption(
            "Typ exact het woord dat in productnamen moet voorkomen, kies de categorie, "
            "zie wat er gaat gebeuren, klik toepassen. Geen LLM, geen verrassingen."
        )

        cZ, cM = st.columns([3, 1])
        with cZ:
            zk = st.text_input("Zoekwoord (komt in productnaam)",
                                key="hvp_qr_zoek",
                                placeholder="bv. bijzettafels")
        with cM:
            heel_woord = st.checkbox("Alleen heel woord", value=True, key="hvp_qr_wb",
                                       help="AAN: 'tafels' matcht NIET 'bijzettafels'. UIT: substring-match.")

        # Live preview matches
        zk_lower = (zk or "").strip().lower()
        matches: list[dict] = []
        if zk_lower:
            patroon = re.compile(rf"\b{re.escape(zk_lower)}\b") if heel_woord else None
            for r in data:
                naam = (r.get("product_title_nl") or r.get("product_title") or "").lower()
                if heel_woord:
                    if patroon.search(naam):
                        matches.append(r)
                else:
                    if zk_lower in naam:
                        matches.append(r)
            st.info(f"📊 {len(matches)} producten matchen op '{zk}'"
                    + (" (heel woord)" if heel_woord else " (substring)"))
            if matches:
                with st.expander(f"Toon {min(10, len(matches))} voorbeelden", expanded=False):
                    for m in matches[:10]:
                        st.text(f"{m.get('sku','')} — {m.get('product_title_nl','')}")

        # Categorie-keuze
        cats_qr = _laad_cats()
        hoofd_keuzes = sorted(set(c["hoofdcategorie"] for c in cats_qr if c.get("hoofdcategorie")))
        NEW = "+ Nieuwe…"

        cA, cB, cC = st.columns(3)
        with cA:
            hc_sel = st.selectbox("Hoofdcategorie", hoofd_keuzes + [NEW], key="hvp_qr_hc")
            if hc_sel == NEW:
                hc_qr = st.text_input("Naam nieuwe hoofd", key="hvp_qr_hcnew",
                                        placeholder="bv. Meubels").strip()
            else:
                hc_qr = hc_sel
        with cB:
            if hc_qr and hc_qr not in hoofd_keuzes:
                sc_qr = st.text_input("Subcategorie (nieuw)", key="hvp_qr_scnew",
                                        placeholder="bv. Tafels").strip()
            else:
                sub_keuzes = sorted(set(c["subcategorie"] for c in cats_qr
                                         if c.get("hoofdcategorie") == hc_qr and c.get("subcategorie")))
                sc_sel = st.selectbox("Subcategorie", sub_keuzes + [NEW], key="hvp_qr_sc")
                if sc_sel == NEW:
                    sc_qr = st.text_input("Naam nieuwe sub", key="hvp_qr_scinp",
                                            placeholder="bv. Bijzettafels").strip()
                else:
                    sc_qr = sc_sel
        with cC:
            nieuw_hc_sc = (hc_qr and hc_qr not in hoofd_keuzes) or \
                           (sc_qr and sc_qr not in [c["subcategorie"] for c in cats_qr
                                                       if c.get("hoofdcategorie") == hc_qr])
            if nieuw_hc_sc:
                ss_qr = st.text_input("Sub-subcategorie (nieuw)", key="hvp_qr_ssnew",
                                        placeholder="bv. Bijzettafels").strip()
            else:
                ssub_keuzes = sorted(set(c["sub_subcategorie"] for c in cats_qr
                                           if c.get("hoofdcategorie") == hc_qr
                                           and c.get("subcategorie") == sc_qr
                                           and c.get("sub_subcategorie")))
                ss_sel = st.selectbox("Sub-subcategorie", ssub_keuzes + [NEW], key="hvp_qr_ss")
                if ss_sel == NEW:
                    ss_qr = st.text_input("Naam nieuwe sub-sub", key="hvp_qr_ssinp",
                                            placeholder="bv. Bijzettafels").strip()
                else:
                    ss_qr = ss_sel

        ongeldig = []
        if not zk_lower: ongeldig.append("zoekwoord")
        if not hc_qr: ongeldig.append("hoofdcategorie")
        if not sc_qr: ongeldig.append("subcategorie")
        if not ss_qr: ongeldig.append("sub-subcategorie")
        if not matches and zk_lower: ongeldig.append("geen matches")

        if st.button(
            f"✅ Pas toe op {len(matches)} producten + onthou",
            type="primary",
            disabled=bool(ongeldig),
            help=("Vul nog in: " + ", ".join(ongeldig)) if ongeldig else None,
            key="hvp_qr_apply",
        ):
            from execution.transform_v2 import build_tags
            for r in matches:
                r["hoofdcategorie"] = hc_qr
                r["subcategorie"] = sc_qr
                r["sub_subcategorie"] = ss_qr
                r["collectie"] = sc_qr
                r["_cat_gemapt"] = True
                extra_t = [r.get("sub_subcategorie_2")] if r.get("sub_subcategorie_2") else None
                r["tags"] = build_tags(hc_qr, sc_qr, ss_qr, r.get("fase",""), extra_tags=extra_t)
            st.session_state["hvp_data"] = data
            # Sla als applied name_rule op
            try:
                _get_sb().table("seo_learnings").insert({
                    "stap": "categorie",
                    "rule_type": "name_rule",
                    "scope": "alle",
                    "input_text": f"snelle regel: {zk} → {hc_qr}/{sc_qr}/{ss_qr}",
                    "action": {
                        "zoekwoord": zk_lower,
                        "hoofdcategorie": hc_qr,
                        "subcategorie": sc_qr,
                        "sub_subcategorie": ss_qr,
                        "is_extra": False,
                        "whole_word": heel_woord,
                    },
                    "status": "applied",
                    "applied_at": datetime.utcnow().isoformat(),
                    "applied_by": "chef@seokitchen.nl",
                    "applied_rows": len(matches),
                }).execute()
            except Exception as e:
                st.warning(f"Regel toegepast maar niet opgeslagen: {e}")
            st.success(f"✅ {len(matches)} producten naar {hc_qr} / {sc_qr} / {ss_qr} · regel onthouden")
            st.rerun()

    # ── Bewerkbare tabel ──
    st.markdown("**Controleer en pas aan:**")
    df = pd.DataFrame([{
        "sku":             r.get("sku", ""),
        "naam_nl":         r.get("product_title_nl", ""),
        "hoofdcategorie":  r.get("hoofdcategorie", ""),
        "subcategorie":    r.get("subcategorie", ""),
        "sub_subcategorie": r.get("sub_subcategorie", ""),
        "sub_subcategorie_2": r.get("sub_subcategorie_2", ""),
        "materiaal_nl":    r.get("materiaal_nl", ""),
        "kleur_nl":        r.get("kleur_nl", ""),
    } for r in data])

    edited = st.data_editor(
        df,
        column_config={
            "sku":              st.column_config.TextColumn("SKU",          disabled=True,  width="small"),
            "naam_nl":          st.column_config.TextColumn("Naam NL",      disabled=True,  width="medium"),
            "hoofdcategorie":   st.column_config.TextColumn("Hoofdcat ✏️",  disabled=False, width="medium"),
            "subcategorie":     st.column_config.TextColumn("Subcat ✏️",    disabled=False, width="medium"),
            "sub_subcategorie": st.column_config.TextColumn("Sub-subcat ✏️", disabled=False, width="medium"),
            "sub_subcategorie_2": st.column_config.TextColumn("+ Tweede sub-subcat ✏️", disabled=False, width="medium",
                                   help="Optioneel: tweede categorie (bv. Bloempotten binnen + buiten)"),
            "materiaal_nl":     st.column_config.TextColumn("Materiaal ✏️", disabled=False, width="small"),
            "kleur_nl":         st.column_config.TextColumn("Kleur ✏️",     disabled=False, width="small"),
        },
        hide_index=True,
        disabled=["sku", "naam_nl"],
        width="stretch",
        key="hvp_edit_s2",
    )

    c1, c2, c3 = st.columns([2, 2, 2])
    with c1:
        if st.button("← Terug naar namen", key="hvp_s2_back"):
            st.session_state["hvp_stap"] = 1
            st.session_state.pop("hvp_s2_gerund", None)
            st.rerun()
    with c2:
        if st.button("↺ Opnieuw runnen", key="hvp_s2_reset"):
            st.session_state.pop("hvp_s2_gerund", None)
            st.rerun()
    with c3:
        if st.button("Goedkeuren → Stap 3", type="primary", key="hvp_s2_ok"):
            from execution.transform_v2 import build_tags
            for _, row in edited.iterrows():
                for r in st.session_state["hvp_data"]:
                    if r.get("sku") == row["sku"]:
                        r["hoofdcategorie"] = row["hoofdcategorie"]
                        r["subcategorie"] = row["subcategorie"]
                        r["sub_subcategorie"] = row["sub_subcategorie"]
                        r["sub_subcategorie_2"] = (row.get("sub_subcategorie_2") or "").strip()
                        r["materiaal_nl"] = row["materiaal_nl"]
                        r["kleur_nl"] = row["kleur_nl"]
                        # Herbouw tags zodat tweede subcat als tag wordt meegenomen
                        extra = [r["sub_subcategorie_2"]] if r["sub_subcategorie_2"] else None
                        r["tags"] = build_tags(
                            r.get("hoofdcategorie", ""),
                            r.get("subcategorie", ""),
                            r.get("sub_subcategorie", ""),
                            r.get("fase", ""),
                            extra_tags=extra,
                        )
                        break
            st.session_state["hvp_stap"] = 3
            _save_draft(f"Auto na stap 2 — {datetime.now().strftime('%d-%m %H:%M')}")
            st.rerun()

    _chat_box(stap=2, kolom_voorbeeld="product_title_nl")


# ── Stap 3: Meta descriptions ─────────────────────────────────────────────────

def _stap_meta() -> None:
    data: list[dict] = st.session_state["hvp_data"]
    n = len(data)

    st.markdown(f"### Meta descriptions schrijven ({n} producten)")
    st.caption(
        "Sonnet schrijft een meta description per product op basis van naam, categorie, "
        "materiaal en kleur uit de vorige stappen. 120-155 tekens, je-vorm, eindigt met CTA."
    )

    if not st.session_state.get("hvp_s3_gerund"):
        kosten = n * 0.002
        st.caption(f"Geschatte kosten: ~€{kosten:.2f} (Sonnet, {n} calls)")
        if st.button(f"Schrijf {n} meta descriptions", type="primary", key="hvp_s3_run"):
            try:
                import anthropic
                from concurrent.futures import ThreadPoolExecutor, as_completed
                from execution.transform_v2 import load_active_learnings
                sb = _get_sb()
                meta_lr = [L for L in load_active_learnings(sb) if L.get("stap") == "meta"]

                # Verzamel meta_instruction's en meta_replace's
                meta_instructies: list[str] = []
                meta_replaces: list[tuple[str, str]] = []
                for L in meta_lr:
                    act = L.get("action") or {}
                    if L.get("rule_type") == "meta_instruction":
                        inst = (act.get("instruction") or "").strip()
                        if inst:
                            meta_instructies.append(inst)
                    elif L.get("rule_type") == "meta_replace":
                        for p in (act.get("replace") or []):
                            fr, to = (p.get("from") or "").strip(), (p.get("to") or "").strip()
                            if fr:
                                meta_replaces.append((fr, to))

                extra_regels = ""
                if meta_instructies:
                    extra_regels = "\nEXTRA REGELS (uit eerdere feedback):\n- " + "\n- ".join(meta_instructies)

                client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY", ""))

                def _gen_meta(idx_r):
                    idx, r = idx_r
                    title  = r.get("product_title_nl", "") or r.get("product_title", "")
                    vendor = r.get("vendor", "") or r.get("supplier", "")
                    subcat = r.get("subcategorie", "")
                    mat    = r.get("materiaal_nl", "")
                    kleur  = r.get("kleur_nl", "")
                    h = r.get("hoogte_cm","") or ""
                    l = r.get("lengte_cm","") or ""
                    b = r.get("breedte_cm","") or ""
                    afm = f"H {h} x L {l} x B {b} cm" if all([h, l, b]) else ""
                    extra = "\n".join(filter(None, [
                        f"Materiaal: {mat}" if mat else "",
                        f"Kleur: {kleur}" if kleur else "",
                        f"Subcategorie: {subcat}" if subcat else "",
                        f"Afmetingen: {afm}" if afm else "",
                    ]))
                    try:
                        resp = client.messages.create(
                            model="claude-sonnet-4-6",
                            max_tokens=200,
                            messages=[{"role": "user", "content":
                                f"Schrijf een Nederlandse SEO meta description (120–155 tekens).\n"
                                f"Product: {title}\nMerk: {vendor}\n{extra}\n\n"
                                "Regels: 'je'-vorm, eindig met CTA, vermeld gratis verzending €75 als dat past."
                                f"{extra_regels}\n\n"
                                "Geef alleen de meta description terug."}],
                        )
                        meta = resp.content[0].text.strip()[:155]
                        for fr, to in meta_replaces:
                            meta = re.sub(re.escape(fr), to, meta, flags=re.IGNORECASE)
                        return idx, meta[:155], None
                    except Exception as e:
                        return idx, "", str(e)

                # Parallel — 10 workers. Anthropic API verdraagt dit ruim;
                # bij echt grote batches kan je dit verhogen tot 20.
                WORKERS = 10
                bar = st.progress(0.0, text=f"Bezig ({len(meta_lr)} actieve meta-regels, {WORKERS} parallel)...")
                klaar = 0
                fouten = 0
                with ThreadPoolExecutor(max_workers=WORKERS) as ex:
                    futures = {ex.submit(_gen_meta, (i, r)): i for i, r in enumerate(data)}
                    for fut in as_completed(futures):
                        idx, meta, err = fut.result()
                        if err:
                            fouten += 1
                            data[idx]["meta_description"] = ""
                        else:
                            data[idx]["meta_description"] = meta
                        klaar += 1
                        bar.progress(klaar / n, text=f"{klaar}/{n} klaar ({fouten} fouten)")

                bar.progress(1.0, text=f"Klaar — {n - fouten} ok, {fouten} fouten.")
                st.session_state["hvp_data"] = data
                st.session_state["hvp_s3_gerund"] = True
                st.rerun()
            except Exception as e:
                st.error(f"Fout: {e}")
        return

    # ── Snelle vervanging in meta descriptions ──
    with st.expander("🎯 Snelle vervanging: X → Y in alle meta descriptions (geen AI)", expanded=False):
        st.caption("Letterlijke tekst-vervanging in alle meta descriptions. Hoofdletter-ongevoelig.")
        cV1, cV2 = st.columns(2)
        with cV1:
            mv_from = st.text_input("Van", key="hvp_mv_from",
                                     placeholder="bv. Ontdek de")
        with cV2:
            mv_to = st.text_input("Naar", key="hvp_mv_to",
                                   placeholder="bv. Bekijk de")
        from_l = (mv_from or "").strip()
        to_v = (mv_to or "").strip()
        n_hit = 0
        voorbeelden = []
        if from_l:
            patroon = re.compile(re.escape(from_l), re.IGNORECASE)
            for r in data:
                m = r.get("meta_description", "") or ""
                if patroon.search(m):
                    n_hit += 1
                    if len(voorbeelden) < 3:
                        voorbeelden.append((m, patroon.sub(to_v, m)))
            st.info(f"📊 {n_hit} meta descriptions worden aangepast")
            for v_oud, v_nw in voorbeelden:
                st.text(f"OUD:  {v_oud[:120]}")
                st.text(f"NIEUW: {v_nw[:120]}")
                st.divider()
        mv_mem = st.checkbox("Onthouden voor toekomstige batches", value=True, key="hvp_mv_mem")
        if st.button(f"✅ Pas toe op {n_hit} meta's", type="primary",
                      disabled=not (from_l and n_hit > 0), key="hvp_mv_apply"):
            patroon = re.compile(re.escape(from_l), re.IGNORECASE)
            for r in data:
                m = r.get("meta_description", "") or ""
                if patroon.search(m):
                    r["meta_description"] = patroon.sub(to_v, m)[:160]
            st.session_state["hvp_data"] = data
            if mv_mem:
                try:
                    _get_sb().table("seo_learnings").insert({
                        "stap": "meta",
                        "rule_type": "meta_replace",
                        "scope": "alle",
                        "input_text": f"snelle vervanging: {from_l} → {to_v}",
                        "action": {"replace": [{"from": from_l, "to": to_v}]},
                        "status": "applied",
                        "applied_at": datetime.utcnow().isoformat(),
                        "applied_by": "chef@seokitchen.nl",
                        "applied_rows": n_hit,
                    }).execute()
                except Exception as e:
                    st.warning(f"Toegepast maar niet opgeslagen: {e}")
            st.success(f"✅ {n_hit} meta's aangepast" + (" · onthouden" if mv_mem else " · niet onthouden"))
            st.rerun()

    # Bewerkbare tabel met tekenteller
    df = pd.DataFrame([{
        "sku":              r.get("sku", ""),
        "naam_nl":          r.get("product_title_nl", ""),
        "meta_description": r.get("meta_description", ""),
        "tekens":           len(r.get("meta_description", "") or ""),
    } for r in data])

    edited = st.data_editor(
        df,
        column_config={
            "sku":              st.column_config.TextColumn("SKU",              disabled=True,  width="small"),
            "naam_nl":          st.column_config.TextColumn("Naam NL",          disabled=True,  width="medium"),
            "meta_description": st.column_config.TextColumn("Meta description ✏️", disabled=False, width="large"),
            "tekens":           st.column_config.NumberColumn("Tekens",          disabled=True,  width="small"),
        },
        hide_index=True,
        disabled=["sku", "naam_nl", "tekens"],
        width="stretch",
        key="hvp_edit_s3",
    )

    # Live tekentellers
    ok = int(((df["tekens"] >= 120) & (df["tekens"] <= 155)).sum()) if len(df) else 0
    m1, m2, m3 = st.columns(3)
    m1.metric("120-155 tekens ✅", ok)
    m2.metric("Te kort (<120)", int(((df["tekens"] < 120) & (df["tekens"] > 0)).sum()) if len(df) else 0)
    m3.metric("Leeg", int((df["tekens"] == 0).sum()) if len(df) else 0)

    c1, c2, c3 = st.columns([2, 2, 2])
    with c1:
        if st.button("← Terug naar categorie", key="hvp_s3_back"):
            st.session_state["hvp_stap"] = 2
            st.session_state.pop("hvp_s3_gerund", None)
            st.rerun()
    with c2:
        if st.button("↺ Opnieuw genereren", key="hvp_s3_reset"):
            st.session_state.pop("hvp_s3_gerund", None)
            st.rerun()
    with c3:
        if st.button("Goedkeuren → Opslaan", type="primary", key="hvp_s3_ok"):
            for _, row in edited.iterrows():
                for r in st.session_state["hvp_data"]:
                    if r.get("sku") == row["sku"]:
                        r["meta_description"] = row["meta_description"]
                        break
            st.session_state["hvp_stap"] = 4
            _save_draft(f"Auto na stap 3 — {datetime.now().strftime('%d-%m %H:%M')}")
            st.rerun()

    _chat_box(stap=3, kolom_voorbeeld="meta_description")


# ── Stap 4: Opslaan + Export ──────────────────────────────────────────────────

HEXTOM_COLUMNS = [
    "Variant SKU", "", "", "Product Handle", "Product Title",
    "Product Vendor", "Product Type", "Variant Barcode", "Variant Price",
    "Variant Cost", "Product Description", "", "", "", "Product Tags",
    "Variant Metafield custom.collectie", "Product Metafield custom.designer",
    "Product Metafield custom.materiaal", "Product Metafield custom.kleur",
    "Product Metafield custom.hoogte_filter", "Product Metafield custom.lengte_filter",
    "Product Metafield custom.breedte_filter",
    "photo_packshot_1", "photo_packshot_2", "photo_packshot_3",
    "photo_packshot_4", "photo_packshot_5",
    "photo_lifestyle_1", "photo_lifestyle_2", "photo_lifestyle_3",
    "photo_lifestyle_4", "photo_lifestyle_5",
    "Product Metafield custom.ean",
    "Product Metafield custom.artikelnummer",
    "Product Metafield custom.meta_description",
]
TEXT_COLS = {"Variant Barcode", "Product Metafield custom.ean"}


def _clean(v) -> str:
    if v is None: return ""
    s = str(v).replace(",", ".")
    try:
        f = float(s)
        return f"{f:.10f}".rstrip("0").rstrip(".")
    except ValueError:
        return s


def _build_excel(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    hf = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    for ci, col in enumerate(HEXTOM_COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=col if col else "")
        c.fill = hf; c.font = hfont
        c.alignment = Alignment(horizontal="center")
    for ri, p in enumerate(rows, 2):
        row_data = {
            "Variant SKU":    p.get("sku",""),
            "Product Handle": p.get("handle","") or p.get("sku",""),
            "Product Title":  p.get("product_title_nl","") or p.get("product_title",""),
            "Product Vendor": p.get("vendor","") or p.get("supplier",""),
            "Product Type":   p.get("hoofdcategorie","") or p.get("product_type",""),
            "Variant Barcode": str(p.get("ean_shopify","") or ""),
            "Variant Price":  _clean(p.get("verkoopprijs") or p.get("price")),
            "Variant Cost":   _clean(p.get("inkoopprijs")),
            "Product Description": p.get("meta_description","") or "",
            "Product Tags":   p.get("tags","") or "",
            "Variant Metafield custom.collectie": p.get("collectie","") or "",
            "Product Metafield custom.designer":  p.get("designer","") or "",
            "Product Metafield custom.materiaal": p.get("materiaal_nl","") or "",
            "Product Metafield custom.kleur":     p.get("kleur_nl","") or "",
            "Product Metafield custom.hoogte_filter": _clean(p.get("hoogte_cm")),
            "Product Metafield custom.lengte_filter": _clean(p.get("lengte_cm")),
            "Product Metafield custom.breedte_filter": _clean(p.get("breedte_cm")),
            "photo_packshot_1": p.get("photo_packshot_1","") or "",
            "photo_packshot_2": p.get("photo_packshot_2","") or "",
            "photo_packshot_3": p.get("photo_packshot_3","") or "",
            "photo_packshot_4": p.get("photo_packshot_4","") or "",
            "photo_packshot_5": p.get("photo_packshot_5","") or "",
            "photo_lifestyle_1": p.get("photo_lifestyle_1","") or "",
            "photo_lifestyle_2": p.get("photo_lifestyle_2","") or "",
            "photo_lifestyle_3": p.get("photo_lifestyle_3","") or "",
            "photo_lifestyle_4": p.get("photo_lifestyle_4","") or "",
            "photo_lifestyle_5": p.get("photo_lifestyle_5","") or "",
            "Product Metafield custom.ean": str(p.get("ean_piece","") or ""),
            "Product Metafield custom.artikelnummer": p.get("sku","") or "",
            "Product Metafield custom.meta_description": p.get("meta_description","") or "",
        }
        for ci, col in enumerate(HEXTOM_COLUMNS, 1):
            val = row_data.get(col,"") if col else ""
            cell = ws.cell(row=ri, column=ci, value=val)
            if col in TEXT_COLS and val:
                cell.value = str(val); cell.number_format = "@"
    for ci in range(1, len(HEXTOM_COLUMNS)+1):
        ws.column_dimensions[get_column_letter(ci)].width = {1:18,4:40,5:50,8:16,11:60,15:50}.get(ci,20)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _stap_export() -> None:
    data: list[dict] = st.session_state["hvp_data"]
    n = len(data)

    st.markdown(f"### Opslaan & exporteren ({n} producten)")

    # Samenvatting
    df_sum = pd.DataFrame([{
        "SKU": r.get("sku",""),
        "Naam NL": r.get("product_title_nl",""),
        "Categorie": f"{r.get('hoofdcategorie','')} / {r.get('sub_subcategorie','')}",
        "Materiaal": r.get("materiaal_nl",""),
        "Kleur": r.get("kleur_nl",""),
        "Meta (tekens)": len(r.get("meta_description","") or ""),
    } for r in data])
    st.dataframe(df_sum, hide_index=True, use_container_width=True)

    c1, c2, c3 = st.columns([2, 2, 2])

    with c1:
        if st.button("← Terug naar meta", key="hvp_s4_back"):
            st.session_state["hvp_stap"] = 3
            st.rerun()

    with c2:
        if st.button(f"Opslaan in database ({n})", type="primary", key="hvp_s4_save"):
            sb = _get_sb()
            saved = errors = 0
            prog = st.progress(0.0)
            for idx, r in enumerate(data):
                prog.progress((idx+1)/n)
                sku = r.get("sku","")
                if not sku:
                    continue
                try:
                    from execution.transform_v2 import generate_handle, build_title
                    titel = r.get("product_title_nl","")
                    handle = r.get("handle","") or generate_handle(titel) if titel else ""
                    upd = {
                        "sku": sku,
                        "supplier": r.get("vendor","") or r.get("supplier",""),
                        "product_title_nl": titel,
                        "handle": handle,
                        "hoofdcategorie": r.get("hoofdcategorie",""),
                        "subcategorie": r.get("subcategorie",""),
                        "sub_subcategorie": r.get("sub_subcategorie",""),
                        "collectie": r.get("collectie","") or r.get("subcategorie",""),
                        "tags": r.get("tags",""),
                        "materiaal_nl": r.get("materiaal_nl",""),
                        "kleur_nl": r.get("kleur_nl",""),
                        "meta_description": r.get("meta_description",""),
                        "pipeline_status": "ready",
                        "review_reden": None,
                    }
                    for veld in ("verkoopprijs","inkoopprijs"):
                        val = r.get(veld)
                        if val is not None:
                            try: upd[veld] = float(val)
                            except (ValueError, TypeError): pass
                    existing = sb.table("products_curated").select("id").eq("sku",sku).execute().data
                    if existing:
                        sb.table("products_curated").update(upd).eq("sku",sku).execute()
                    else:
                        sb.table("products_curated").insert(upd).execute()
                    saved += 1
                except Exception as e:
                    errors += 1
            if errors:
                st.warning(f"✅ {saved} opgeslagen · ⚠️ {errors} fouten")
            else:
                st.success(f"✅ {saved} producten opgeslagen in products_curated.")

    with c3:
        if st.button("Download Hextom Excel", key="hvp_s4_excel"):
            with st.spinner("Excel bouwen..."):
                xlsx = _build_excel(data)
            vendor = (data[0].get("vendor","") or data[0].get("supplier","")).replace(" ","_") if data else "export"
            st.download_button(
                f"💾 Download ({n} producten)",
                data=xlsx,
                file_name=f"hextom_{vendor}_herverwerkt_{n}st.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="hvp_s4_dl",
            )

    st.divider()
    if st.button("Opnieuw beginnen (nieuwe selectie)", key="hvp_s4_nieuw"):
        for k in ["hvp_data","hvp_stap","hvp_s1_gerund","hvp_s2_gerund","hvp_s3_gerund","hvp_ai_klaar"]:
            st.session_state.pop(k, None)
        st.switch_page("pages/08_Herverwerk.py")


# ── Render ────────────────────────────────────────────────────────────────────

def render() -> None:
    st.subheader("Herverwerk — stap-voor-stap pipeline")

    rows: list[dict] = st.session_state.get("hv_pipeline_rows", [])

    # Als er geen rows zijn: bied hervat-balk aan, anders terug
    if not rows and "hvp_data" not in st.session_state:
        st.warning("Geen producten geladen. Ga terug naar **Archief herverwerken** of hervat een opgeslagen sessie hieronder.")
        _draft_balk()
        if st.button("← Terug naar Archief herverwerken"):
            st.switch_page("pages/08_Herverwerk.py")
        return

    # Initialiseer data + stap — markeer ook als in_process bij eerste laadbeurt
    eerste_keer = "hvp_data" not in st.session_state
    if eerste_keer:
        st.session_state["hvp_data"] = list(rows)
    if "hvp_stap" not in st.session_state:
        st.session_state["hvp_stap"] = 1

    n = len(st.session_state["hvp_data"])

    if eerste_keer and rows:
        skus = [r.get("sku") for r in rows if r.get("sku")]
        if skus:
            aantal = _mark_in_process(skus, rows)
            if aantal:
                st.caption(f"{aantal} producten gemarkeerd als 'in_process' in products_curated.")

    st.caption(f"{n} producten geladen — foto's, EAN en barcodes worden niet aangeraakt.")

    # Opslaan/hervat-balk altijd zichtbaar
    _draft_balk()

    _voortgang(st.session_state["hvp_stap"])

    stap = st.session_state["hvp_stap"]
    if stap == 1:
        _stap_namen()
    elif stap == 2:
        _stap_categorie_kleur()
    elif stap == 3:
        _stap_meta()
    elif stap == 4:
        _stap_export()
