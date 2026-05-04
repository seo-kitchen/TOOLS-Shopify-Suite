"""
SEOkitchen — visueel Streamlit dashboard voor batch-verwerking.

Flow:
  1. Upload publicatie-batch Excel
  2. Match tegen seo_products (Supabase) op SKU of EAN
  3. Categorie-toewijzing
  4. Vertalen materiaal/kleur
  5. Producttitel + meta description (Claude)
  6. Eindreview
  7. Export naar Shopify

Start met:
    streamlit run streamlit_app.py
"""

import os
import re
from pathlib import Path

import pandas as pd
import streamlit as st
from dotenv import load_dotenv

load_dotenv()

# ── Page config ───────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="SEOkitchen Pipeline",
    page_icon="🍳",
    layout="wide",
)

STAPPEN = [
    "📦 Masterdata beheren",
    "1. Upload publicatie-batch",
    "2. Match tegen Supabase",
    "3. Categorie-toewijzing",
    "4. Vertalen materiaal/kleur",
    "5. Producttitel + meta description (Claude)",
    "6. Eindreview",
    "7. Export naar Shopify",
    "📊 Overzicht & vragen",
    "🌐 Website structuur",
    "🏷️ Categoriestatus",
    "🔍 Product zoeken",
]

# ── Session state defaults ────────────────────────────────────────────────────

DEFAULTS = {
    "uploaded_filename": None,
    "uploaded_df": None,
    "sku_col": None,
    "ean_col": None,
    "naam_col": None,
    "merk": "Serax",          # merknaam voor producttitel (Serax / Pottery Pots / Printworks)
    "batch_tag": "structuur_fase4",  # hoofdtag voor deze batch
    "matches_df": None,  # resultaat van stap 2
    "match_stats": None,
    "selected_for_enrichment": None,  # set van DB-id's die door mogen
    "batch_products": None,  # volledige product-records voor enrichment-stappen
    "stap3_done": False,
    "stap4_done": False,
    "stap5_done": False,
    "stap6_committed": False,
    "exported_path": None,
}
for k, v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ── Helpers ───────────────────────────────────────────────────────────────────

@st.cache_resource
def get_supabase():
    from supabase import create_client
    return create_client(os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_KEY"))


@st.cache_resource
def get_supabase_new():
    from supabase import create_client
    return create_client(os.getenv("SUPABASE_NEW_URL"), os.getenv("SUPABASE_NEW_SERVICE_KEY"))


# Aliases hergebruiken uit setup_masterdata SCHEMA — losse import om dependency
# op het exacte path te vermijden, en omdat we hier alleen SKU/EAN/naam nodig hebben.
SKU_ALIASES = {
    "sku", "variant sku", "brand_id", "artikel", "artikelnummer", "product id",
    "item number", "code", "articlecode", "article code", "artikelcode",
}
EAN_ALIASES = {
    "ean", "ean piece", "ean code piece", "ean stuk", "barcode piece", "ean los",
    "ean code", "barcode", "ean packaging/giftbox", "ean packaging", "ean giftbox",
    "ean box", "ean-ucc _code", "ean-ucc code", "ean code per stuk",
}
NAAM_ALIASES = {
    "product name", "name", "title", "omschrijving", "naam", "description en",
    "short product name piece (english)", "short product name piece (dutch)",
    "description", "productnaam",
}


def _norm(s) -> str:
    import re
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def detect_column(columns: list, aliases: set) -> str | None:
    """Vind de eerste kolom waarvan de genormaliseerde naam in de aliases zit."""
    for col in columns:
        if _norm(col) in aliases:
            return col
    return None


def safe_index(seq, value):
    try:
        return list(seq).index(value)
    except ValueError:
        return 0


def kleur_status(status: str) -> str:
    if status.startswith("🟢"):
        return "background-color: #d4edda"
    if status.startswith("🟡"):
        return "background-color: #fff3cd"
    if status.startswith("🔴"):
        return "background-color: #f8d7da"
    return ""


def load_full_batch(ids: list) -> list[dict]:
    """Haalt volledige product-records op voor de geselecteerde DB-ids (batches van 200)."""
    if not ids:
        return []
    sb = get_supabase()
    out = []
    ids_list = list(ids)
    BATCH = 200
    for i in range(0, len(ids_list), BATCH):
        chunk = ids_list[i:i + BATCH]
        res = sb.table("seo_products").select("*").in_("id", chunk).execute()
        out.extend(res.data or [])
    return out


@st.cache_data(show_spinner=False)
def load_active_subsubcategories() -> set[str]:
    """
    Lees Master Files/Website indeling (1).xlsx, parse cell-fill colors,
    en geef terug welke sub-subcategorieën GROEN zijn (= al online op de webshop).

    Plus user-overrides: 'Drinkglazen' is handmatig bevestigd als actief.
    Returns: set van lowercase sub-subcat namen.
    """
    from openpyxl import load_workbook

    pad = Path("Master Files/Website indeling (1).xlsx")
    if not pad.exists():
        return set()

    GROEN = {"FF92D050", "FF93C47D"}
    wb = load_workbook(pad, data_only=True)
    ws = wb["Blad1"]

    actief = set()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or cell.row < 3:  # alleen sub-subcat rijen (rij 3+)
                continue
            fg = cell.fill.fgColor if cell.fill else None
            if not fg or fg.type != "rgb" or not fg.rgb or fg.rgb == "00000000":
                continue
            if fg.rgb in GROEN:
                actief.add(str(cell.value).strip().lower())

    # User-overrides (zie eerdere conversatie)
    actief.add("drinkglazen")
    return actief


@st.cache_data(show_spinner=False)
def load_website_tree() -> dict:
    """
    Lees de website indeling xlsx en bouw een hierarchische dict:
        { hoofdcat: { subcat: [sub_subcat, ...], ... }, ... }

    Gebruikt voor cascading dropdowns in stap 3.
    """
    pad = Path("Master Files/Website indeling (1).xlsx")
    if not pad.exists():
        return {}

    df = pd.read_excel(pad, sheet_name="Blad1", header=None, dtype=str)

    def cell(r, c):
        if r >= len(df) or c >= len(df.columns):
            return None
        v = df.iat[r, c]
        if pd.isna(v):
            return None
        s = str(v).strip()
        return s if s else None

    # Hoofdcats op rij 0
    hoofdcats = {}
    for c in range(len(df.columns)):
        v = cell(0, c)
        if v:
            hoofdcats[c] = v
    hoofdcat_cols = sorted(hoofdcats.keys())

    def hoofdcat_for_col(c):
        parent = None
        for hc in hoofdcat_cols:
            if hc <= c:
                parent = hc
            else:
                break
        return hoofdcats.get(parent)

    # Subcats op rij 1
    subcats = {}
    for c in range(len(df.columns)):
        v = cell(1, c)
        if v:
            subcats[c] = (hoofdcat_for_col(c), v)

    # Bouw boom
    tree = {}
    for c, (hc, sc) in subcats.items():
        if hc not in tree:
            tree[hc] = {}
        if sc not in tree[hc]:
            tree[hc][sc] = []
        for r in range(2, len(df)):
            v = cell(r, c)
            if v:
                tree[hc][sc].append(v)

    return tree


def ensure_batch_loaded() -> list[dict]:
    """Laadt batch_products als die nog leeg is. Roep aan bij entry van stap 3+."""
    if not st.session_state.batch_products:
        ids = st.session_state.selected_for_enrichment or set()
        if not ids:
            return []
        with st.spinner(f"Volledige records ophalen voor {len(ids)} producten..."):
            products = load_full_batch(ids)
            # Zet merknaam en batch-tag op elk product
            merk = st.session_state.get("merk") or "Serax"
            batch_tag = st.session_state.get("batch_tag") or "structuur_fase4"
            for p in products:
                p["_merk"] = merk
                p["_batch_tag"] = batch_tag
            st.session_state.batch_products = products
    return st.session_state.batch_products


# ── Learnings log ─────────────────────────────────────────────────────────────

LEARNINGS_PATH = Path("config/learnings.json")


def _load_learnings() -> list[dict]:
    if LEARNINGS_PATH.exists():
        import json
        with open(LEARNINGS_PATH, encoding="utf-8") as f:
            return json.load(f)
    return []


def _save_learning(entry: dict):
    import json
    from datetime import datetime
    learnings = _load_learnings()
    entry["timestamp"] = datetime.now().isoformat()
    learnings.append(entry)
    LEARNINGS_PATH.parent.mkdir(exist_ok=True)
    with open(LEARNINGS_PATH, "w", encoding="utf-8") as f:
        json.dump(learnings, f, ensure_ascii=False, indent=2)


def learning_chatbox(stap_naam: str, stap_context: str, key: str):
    """
    Herbruikbare chatbox die per stap onderaan verschijnt. De user typt een correctie
    in gewone taal. Claude Haiku interpreteert de correctie, geeft terug wat er moet
    veranderen, en het systeem voert het uit.

    Ondersteunde acties:
      - stap3: upsert naar seo_category_mapping in Supabase
      - stap4: nieuwe vertaling toevoegen aan MATERIAAL_NL of KLEUR_FILTER in transform.py
      - stap5: titel-patroon of naamregel opslaan als learning

    Returns: None (alles gaat via st.session_state en Supabase side-effects)
    """
    st.divider()
    st.subheader("💡 Correctie doorvoeren")
    st.caption(
        "Typ hieronder een correctie in gewone taal. Het systeem leert ervan zodat "
        "dezelfde fout volgende keer niet meer gemaakt wordt."
    )

    # Voorbeelden per stap
    voorbeelden = {
        "stap3": (
            "Voorbeelden:\n"
            "- *\"deep plate moet bij Diepe borden, niet Dinerborden\"*\n"
            "- *\"Interior Accessories / storage & organisation hoort bij Keuken & Eetkamer > Keukenorganisatie > Voorraadpotten\"*\n"
            "- *\"alle Pottery pots met 'Round' in de item cat zijn Bloempotten binnen\"*"
        ),
        "stap4": (
            "Voorbeelden:\n"
            "- *\"ash wood moet vertaald worden als Essenhout\"*\n"
            "- *\"kleur 'rust' is Roestbruin, niet Rood\"*\n"
            "- *\"samengesteld materiaal 'glass + metal' = Glas & Metaal\"*"
        ),
        "stap5": (
            "Voorbeelden:\n"
            "- *\"voor Pottery pots producten zonder designer, gebruik 'Pottery Pots' als merk in de titel ipv 'Serax'\"*\n"
            "- *\"SOPH moet altijd aangevuld worden tot SOPHIA\"*\n"
            "- *\"productnaam 'Fertilized Soil' vertalen als 'Bemeste Aarde'\"*"
        ),
    }
    st.markdown(voorbeelden.get(key, ""))

    user_input = st.text_area(
        "Jouw correctie",
        placeholder="Typ hier wat er anders moet...",
        key=f"learning_input_{key}",
        height=80,
    )

    if st.button("🧠 Toepassen", type="primary", key=f"learning_apply_{key}"):
        if not user_input or not user_input.strip():
            st.warning("Typ eerst een correctie.")
            return

        try:
            claude = get_claude_client()
        except Exception as e:
            st.error(f"Kon Claude niet initialiseren. Check ANTHROPIC_API_KEY in .env. Fout: {e}")
            return

        # Bouw prompt op basis van de stap
        system_prompt = _build_learning_prompt(key, stap_context)

        with st.spinner("Claude interpreteert je correctie..."):
            try:
                resp = claude.messages.create(
                    model="claude-haiku-4-5-20251001",
                    max_tokens=1000,
                    system=system_prompt,
                    messages=[
                        {"role": "user", "content": user_input}
                    ],
                )
                result_text = resp.content[0].text.strip()
            except Exception as e:
                st.error(f"Claude-fout: {e}")
                return

        # Parse het JSON-antwoord
        import json
        try:
            # Zoek JSON block in de response — pak het eerste { ... } blok
            # Gebruik een simpelere regex die niet crasht op geneste braces
            brace_start = result_text.find("{")
            if brace_start == -1:
                raise ValueError("Geen JSON gevonden")
            # Vind de matchende sluitende brace
            depth = 0
            brace_end = -1
            for i in range(brace_start, len(result_text)):
                if result_text[i] == "{":
                    depth += 1
                elif result_text[i] == "}":
                    depth -= 1
                    if depth == 0:
                        brace_end = i + 1
                        break
            if brace_end == -1:
                raise ValueError("Onvolledige JSON")
            action = json.loads(result_text[brace_start:brace_end])
        except (json.JSONDecodeError, ValueError) as e:
            st.error(f"Kon Claude's antwoord niet parsen. Raw output:\n\n{result_text}")
            # Log het toch zodat de user het kan zien
            _save_learning({
                "stap": stap_naam,
                "input": user_input,
                "actie": {"type": "parse_error", "beschrijving": str(e)},
                "raw_response": result_text,
            })
            return

        # Sla de actie op in session_state voor preview (nog NIET uitvoeren)
        st.session_state[f"_learning_preview_{key}"] = {
            "action": action,
            "raw_response": result_text,
            "user_input": user_input,
            "stap_naam": stap_naam,
        }
        st.rerun()

    # ── Preview + akkoord ─────────────────────────────────────────────────
    preview_key = f"_learning_preview_{key}"
    if st.session_state.get(preview_key):
        preview = st.session_state[preview_key]
        action = preview["action"]
        action_type = action.get("type", "")

        st.divider()
        st.markdown("### 📋 Preview — dit gaat er veranderen:")

        # Toon wat de actie gaat doen
        if action_type == "name_rule":
            zoekwoord = action.get("zoekwoord", "")
            subsub = action.get("sub_subcategorie", "")
            is_extra = action.get("is_extra", False)

            # Zoek voorbeeldproducten die matchen
            batch = st.session_state.get("batch_products") or []
            matches = [p for p in batch if zoekwoord in (p.get("product_name_raw") or "").lower()]

            if is_extra:
                st.info(f'**Regel:** Producten met **"{zoekwoord}"** in de naam krijgen **"{subsub}"** als extra sub-subcategorie (2e tag)')
            else:
                st.info(f'**Regel:** Producten met **"{zoekwoord}"** in de naam krijgen **"{subsub}"** als primaire sub-subcategorie')

            st.markdown(f"**{len(matches)} producten** in de huidige batch worden aangepast:")
            if matches:
                preview_rows = []
                for p in matches[:10]:
                    huidige = p.get("sub_subcategorie") or ""
                    extra = ", ".join(p.get("_extra_tags") or [])
                    preview_rows.append({
                        "SKU": p.get("sku", ""),
                        "Productnaam": (p.get("product_name_raw") or "")[:40],
                        "Huidige sub-subcat": huidige,
                        "Huidige extra": extra,
                        "Na wijziging": f"{huidige} + {subsub}" if is_extra else subsub,
                    })
                st.dataframe(pd.DataFrame(preview_rows), width="stretch", hide_index=True)
                if len(matches) > 10:
                    st.caption(f"... en {len(matches) - 10} meer")
            else:
                st.warning("Geen producten in de huidige batch matchen dit zoekwoord.")

        elif action_type == "category_mapping":
            st.info(
                f'**Mapping:** `({action.get("leverancier_category", "?")}, {action.get("leverancier_item_cat", "?")})` '
                f'→ **{action.get("hoofdcategorie", "")} > {action.get("subcategorie", "")} > {action.get("sub_subcategorie", "")}**'
            )
            st.caption("Wordt opgeslagen in seo_category_mapping — geldt voor alle toekomstige batches.")

        elif action_type == "translation":
            en_term = action.get("en", "")
            nl_term = action.get("nl", "")
            veld = action.get("veld", "")
            st.info(f'**Vertaling ({veld}):** `{en_term}` → **{nl_term}**')

            # Toon voorbeeldproducten die matchen
            batch = st.session_state.get("batch_products") or []
            if veld == "materiaal":
                matches = [p for p in batch if (p.get("materiaal_nl") or "").strip().lower() == en_term.lower()]
                st.markdown(f"**{len(matches)} producten** in de batch met materiaal `{en_term}`")
            elif veld == "kleur":
                matches = [p for p in batch if (p.get("kleur_en") or "").strip().lower() == en_term.lower()]
                st.markdown(f"**{len(matches)} producten** in de batch met kleur `{en_term}`")
            else:
                matches = []

            if matches:
                preview_rows = [{"SKU": p.get("sku",""), "Productnaam": (p.get("product_name_raw") or "")[:40],
                                 f"Huidig ({veld})": en_term, "Na wijziging": nl_term} for p in matches[:8]]
                st.dataframe(pd.DataFrame(preview_rows), width="stretch", hide_index=True)
                if len(matches) > 8:
                    st.caption(f"... en {len(matches) - 8} meer")

            st.caption("Wordt permanent opgeslagen in transform.py — geldt voor alle toekomstige batches.")

        elif action_type == "name_translation":
            st.info(f'**Productnaam-vertaling:** `{action.get("en", "")}` → **{action.get("nl", "")}**')
            batch = st.session_state.get("batch_products") or []
            en = (action.get("en") or "").lower()
            matches = [p for p in batch if en in (p.get("product_name_raw") or "").lower()]
            st.markdown(f"**{len(matches)} producten** in de batch bevatten `{action.get('en', '')}`")
            if matches:
                preview_rows = [{"SKU": p.get("sku",""), "Productnaam": (p.get("product_name_raw") or "")[:40]} for p in matches[:8]]
                st.dataframe(pd.DataFrame(preview_rows), width="stretch", hide_index=True)
            st.caption("Wordt opgeslagen als learning — toekomstige Claude-vertalingen houden hier rekening mee.")

        elif action_type == "title_rule":
            st.info(f'**Titel-regel:** {action.get("beschrijving", "")}')
            st.caption("Wordt opgeslagen als learning — toekomstige titel-generatie houdt hier rekening mee.")

        else:
            st.info(f'**Actie:** {action.get("beschrijving", str(action)[:100])}')

        # Akkoord-knoppen
        ak1, ak2 = st.columns(2)
        if ak1.button("✅ Akkoord — voer uit", type="primary", key=f"learning_ok_{key}"):
            try:
                success = _execute_learning(key, action)
            except Exception as e:
                st.error(f"Fout bij uitvoeren: {e}")
                success = False

            if success:
                _save_learning({
                    "stap": preview["stap_naam"],
                    "input": preview["user_input"],
                    "actie": action,
                    "raw_response": preview["raw_response"],
                })
                st.session_state.pop(preview_key, None)
                st.success(f"✅ **Geleerd en toegepast!** {action.get('beschrijving', '')}")
                st.rerun()
            else:
                st.warning(f"Kon niet worden uitgevoerd: {action.get('beschrijving', '')}")

        if ak2.button("❌ Annuleren", key=f"learning_cancel_{key}"):
            st.session_state.pop(preview_key, None)
            st.rerun()

    # Toon recente learnings voor deze stap
    learnings = _load_learnings()
    stap_learnings = [l for l in learnings if l.get("stap") == stap_naam]
    if stap_learnings:
        with st.expander(f"📚 Eerdere correcties ({len(stap_learnings)})", expanded=False):
            for l in reversed(stap_learnings[-10:]):
                ts = l.get("timestamp", "?")[:16]
                st.markdown(f"**{ts}** — {l.get('input', '')}")
                actie = l.get("actie", {})
                if actie.get("beschrijving"):
                    st.caption(f"→ {actie['beschrijving']}")
                st.divider()


def _build_learning_prompt(stap_key: str, context: str) -> str:
    """Bouwt de Claude system-prompt per stap-type."""
    base = (
        "Je bent een JSON-generator voor het categorie- en vertaalsysteem van een Belgische "
        "interieur-webshop (Serax, Pottery Pots, S&P/Bonbistro, Printworks). "
        "Je geeft ALLEEN een JSON-object terug, GEEN uitleg of vragen. "
        "De gebruiker praat in gewone taal over producten en categorieën — vertaal dat naar een actie.\n\n"
        f"CONTEXT: {context}\n\n"
    )

    if stap_key == "stap3":
        # Bouw de website-boom als referentie voor Claude
        tree = load_website_tree()
        boom_tekst = ""
        if tree:
            lines = []
            for hc in sorted(tree.keys()):
                for sc in sorted(tree[hc].keys()):
                    subcats = ", ".join(sorted(tree[hc][sc]))
                    lines.append(f"  {hc} > {sc} > [{subcats}]")
            boom_tekst = "\n\nBESCHIKBARE CATEGORIEËN (hoofdcat > subcat > [sub-subcats]):\n" + "\n".join(lines)

        return base + (
            "De gebruiker geeft feedback over categorisering. Dit kan zijn:\n\n"
            "1. Een CATEGORIE-MAPPING: koppeling van leverancier-codes naar website-categorieën\n"
            "2. Een PRODUCTNAAM-REGEL: producten met een bepaald woord in de naam moeten een bepaalde categorie krijgen\n"
            "3. Een EXTRA TAG: producten moeten een extra sub-subcategorie als tag krijgen\n\n"
            "BELANGRIJK:\n"
            "- De gebruiker praat vaak in PRODUCTNAMEN (bijv. 'deep plate') niet in leverancier-codes\n"
            "- Als de gebruiker een productnaam noemt, vertaal dat naar een naam-patroon regel\n"
            "- Vraag NOOIT om leverancier-codes — los het op met wat de gebruiker zegt\n"
            f"{boom_tekst}\n\n"
            "MOGELIJKE JSON OUTPUTS:\n\n"
            "A) Categorie-mapping (als de gebruiker leverancier-codes noemt):\n"
            '{"type": "category_mapping", "leverancier_category": "...", "leverancier_item_cat": "?",\n'
            ' "hoofdcategorie": "...", "subcategorie": "...", "sub_subcategorie": "...", "beschrijving": "..."}\n\n'
            "B) Productnaam-regel (als de gebruiker een productnaam of -woord noemt):\n"
            '{"type": "name_rule", "zoekwoord": "het woord dat in de productnaam moet zitten (lowercase)",\n'
            ' "sub_subcategorie": "de sub-subcategorie die het moet krijgen",\n'
            ' "is_extra": true als het een EXTRA/2e sub-subcategorie is (naast de bestaande), false als het de primaire vervangt,\n'
            ' "beschrijving": "..."}\n\n'
            "C) Bulk (meerdere regels tegelijk):\n"
            '{"type": "name_rule_bulk", "regels": [{"zoekwoord": "...", "sub_subcategorie": "...", "is_extra": true/false}, ...], "beschrijving": "..."}\n\n'
            "VOORBEELDEN:\n"
            '- "Deep plate moet bij Diepe borden" → {"type": "name_rule", "zoekwoord": "deep plate", "sub_subcategorie": "Diepe borden", "is_extra": false, "beschrijving": "..."}\n'
            '- "Deep plate heeft ook als extra subcategorie Diepe borden" → {"type": "name_rule", "zoekwoord": "deep plate", "sub_subcategorie": "Diepe borden", "is_extra": true, "beschrijving": "..."}\n'
            '- "Alle Wally producten zijn Hangpotten" → {"type": "name_rule", "zoekwoord": "wally", "sub_subcategorie": "Hangpotten", "is_extra": false, "beschrijving": "..."}\n'
        )
    elif stap_key == "stap4":
        return base + (
            "De gebruiker corrigeert een VERTALING (materiaal of kleur, Engels → Nederlands).\n\n"
            "Geef JSON terug:\n"
            "{\n"
            '  "type": "translation",\n'
            '  "veld": "materiaal" of "kleur",\n'
            '  "en": "engelse term (lowercase)",\n'
            '  "nl": "nederlandse vertaling",\n'
            '  "beschrijving": "korte uitleg"\n'
            "}\n\n"
            "Als je de input niet begrijpt, geef: {\"type\": \"unclear\", \"beschrijving\": \"...\"}"
        )
    elif stap_key == "stap5":
        return base + (
            "De gebruiker corrigeert een PRODUCTNAAM-VERTALING of TITEL-REGEL.\n\n"
            "Geef JSON terug:\n"
            "{\n"
            '  "type": "name_translation" of "title_rule",\n'
            '  "en": "engelse term",\n'
            '  "nl": "nederlandse vertaling",\n'
            '  "beschrijving": "korte uitleg"\n'
            "}\n\n"
            "Als je de input niet begrijpt, geef: {\"type\": \"unclear\", \"beschrijving\": \"...\"}"
        )
    return base + 'Geef JSON terug: {"type": "general", "beschrijving": "..."}'


def _execute_learning(stap_key: str, action: dict) -> bool:
    """Voert de actie uit die Claude heeft geïnterpreteerd. Returns True bij succes."""
    action_type = action.get("type", "")

    if action_type == "unclear":
        return False

    if action_type == "category_mapping":
        sb = get_supabase()
        # Null/None → "?" voor leverancier_item_cat (die is soms leeg bij Printworks etc.)
        if not action.get("leverancier_item_cat"):
            action["leverancier_item_cat"] = "?"
        required = ["leverancier_category", "leverancier_item_cat", "hoofdcategorie", "subcategorie", "sub_subcategorie"]
        missing = [k for k in required if not action.get(k)]
        if missing:
            action["beschrijving"] = f"Velden ontbreken: {missing}. Herformuleer je correctie specifieker."
            return False
        row = {k: action[k] for k in required}
        sb.table("seo_category_mapping").upsert(
            row, on_conflict="leverancier_category,leverancier_item_cat"
        ).execute()
        st.session_state.pop("_cat_cache", None)
        return True

    elif action_type == "category_mapping_bulk":
        sb = get_supabase()
        mappings = action.get("mappings", [])
        if not mappings:
            return False
        for m in mappings:
            required = ["leverancier_category", "leverancier_item_cat", "hoofdcategorie", "subcategorie", "sub_subcategorie"]
            if all(m.get(k) for k in required):
                sb.table("seo_category_mapping").upsert(
                    {k: m[k] for k in required},
                    on_conflict="leverancier_category,leverancier_item_cat"
                ).execute()
        st.session_state.pop("_cat_cache", None)
        return True

    elif action_type == "translation":
        veld = action.get("veld", "")
        en_term = (action.get("en") or "").strip().lower()
        nl_term = (action.get("nl") or "").strip()
        if not en_term or not nl_term or veld not in ("materiaal", "kleur"):
            action["beschrijving"] = f"Ongeldig: veld={veld!r}, en={en_term!r}, nl={nl_term!r}"
            return False
        return _add_translation_to_transform(veld, en_term, nl_term)

    elif action_type == "name_rule":
        # Productnaam-regel: pas direct toe op huidige batch EN sla op als learning
        zoekwoord = (action.get("zoekwoord") or "").strip().lower()
        subsub = (action.get("sub_subcategorie") or "").strip()
        hoofd = (action.get("hoofdcategorie") or "").strip()
        sub = (action.get("subcategorie") or "").strip()
        is_extra = action.get("is_extra", False)
        if not zoekwoord or not subsub:
            action["beschrijving"] = f"Zoekwoord of sub-subcategorie ontbreekt"
            return False

        # Pas toe op huidige batch
        batch = st.session_state.get("batch_products") or []
        toegepast = 0
        for p in batch:
            naam = (p.get("product_name_raw") or "").lower()
            if zoekwoord in naam:
                if is_extra:
                    extra = p.get("_extra_tags") or []
                    if subsub not in extra:
                        extra.append(subsub)
                    p["_extra_tags"] = extra
                else:
                    p["sub_subcategorie"] = subsub
                    if hoofd:
                        p["hoofdcategorie"] = hoofd
                    if sub:
                        p["subcategorie"] = sub
                        p["collectie"] = sub
                toegepast += 1
        if batch:
            st.session_state.batch_products = batch

        # Sla ook op als mapping in Supabase zodat het permanent werkt
        if hoofd and sub and not is_extra:
            try:
                sb = get_supabase()
                sb.table("seo_category_mapping").upsert({
                    "leverancier_category": f"_namecontains_{zoekwoord}",
                    "leverancier_item_cat": "?",
                    "hoofdcategorie": hoofd,
                    "subcategorie": sub,
                    "sub_subcategorie": subsub,
                }, on_conflict="leverancier_category,leverancier_item_cat").execute()
            except Exception:
                pass

        action["beschrijving"] = (action.get("beschrijving") or "") + f" (toegepast op {toegepast} producten)"
        return True

    elif action_type == "name_rule_bulk":
        batch = st.session_state.get("batch_products") or []
        totaal = 0
        for regel in action.get("regels", []):
            zoekwoord = (regel.get("zoekwoord") or "").strip().lower()
            subsub = (regel.get("sub_subcategorie") or "").strip()
            is_extra = regel.get("is_extra", False)
            if not zoekwoord or not subsub:
                continue
            for p in batch:
                naam = (p.get("product_name_raw") or "").lower()
                if zoekwoord in naam:
                    if is_extra:
                        extra = p.get("_extra_tags") or []
                        if subsub not in extra:
                            extra.append(subsub)
                        p["_extra_tags"] = extra
                    else:
                        p["sub_subcategorie"] = subsub
                    totaal += 1
        if batch:
            st.session_state.batch_products = batch
        action["beschrijving"] = (action.get("beschrijving") or "") + f" (toegepast op {totaal} producten)"
        return True

    elif action_type in ("name_translation", "title_rule"):
        return True  # logging gebeurt in de caller

    # Onbekend type — toch loggen als succes zodat de learning bewaard wordt
    return True


def _add_translation_to_transform(veld: str, en_term: str, nl_term: str) -> bool:
    """Voegt een vertaling toe aan MATERIAAL_NL of KLEUR_FILTER in transform.py."""
    transform_path = Path("execution/transform.py")
    if not transform_path.exists():
        return False

    content = transform_path.read_text(encoding="utf-8")

    if veld == "materiaal":
        # Zoek het einde van MATERIAAL_NL dict en voeg entry toe
        marker = "}\n\n# Kleur: filter-weergave"
        if marker in content and f'"{en_term}"' not in content:
            insert = f'    "{en_term}":' + " " * max(1, 18 - len(en_term)) + f'"{nl_term}",\n'
            content = content.replace(marker, insert + marker)
            transform_path.write_text(content, encoding="utf-8")
            return True
    elif veld == "kleur":
        marker = "}\n\n# Kleuren die bewaard"
        if marker in content and f'"{en_term}"' not in content:
            insert = f'    "{en_term}":' + " " * max(1, 18 - len(en_term)) + f'"{nl_term}",\n'
            content = content.replace(marker, insert + marker)
            transform_path.write_text(content, encoding="utf-8")
            return True

    return False


def lookup_categorie(sb, lev_cat: str, lev_item: str) -> dict | None:
    """Zoek exacte match in seo_category_mapping."""
    if not lev_cat or not lev_item:
        return None
    res = sb.table("seo_category_mapping").select("*").eq(
        "leverancier_category", lev_cat
    ).eq("leverancier_item_cat", lev_item).execute()
    return res.data[0] if res.data else None


def get_claude_client():
    """Lazy Claude client — alleen instantiëren als gebruiker op een knop drukt."""
    import anthropic
    return anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))


# ── Sidebar ───────────────────────────────────────────────────────────────────

st.sidebar.title("🍳 SEOkitchen")
st.sidebar.caption("Batch-verwerking pipeline")

stap = st.sidebar.radio("Pipeline-stap", STAPPEN, index=0)

st.sidebar.divider()
st.sidebar.markdown("**Sessie-status**")
if st.session_state.uploaded_df is not None:
    st.sidebar.success(f"📄 Excel: {st.session_state.uploaded_filename} ({len(st.session_state.uploaded_df)} rijen)")
else:
    st.sidebar.info("📄 Geen Excel geladen")
if st.session_state.matches_df is not None:
    s = st.session_state.match_stats or {}
    st.sidebar.success(
        f"🔗 Match: {s.get('groen', 0)}🟢 / {s.get('rood', 0)}🔴"
    )
else:
    st.sidebar.info("🔗 Nog niet gematcht")

if st.sidebar.button("🔄 Reset sessie", type="secondary"):
    for k in DEFAULTS:
        st.session_state[k] = DEFAULTS[k]
    st.rerun()


# ═════════════════════════════════════════════════════════════════════════════
# MASTERDATA BEHEREN
# ═════════════════════════════════════════════════════════════════════════════

if stap == STAPPEN[0]:
    st.title("📦 Masterdata beheren")
    st.markdown(
        "Beheer de leverancier-masterdata in Supabase. Upload nieuwe bestanden, bekijk wat er al in staat, "
        "en pas data aan waar nodig."
    )

    sb = get_supabase()

    # ── Overzicht huidige data ────────────────────────────────────────────────
    st.subheader("Huidige database")

    @st.cache_data(ttl=30, show_spinner=False)
    def _db_stats():
        total = sb.table("seo_products").select("id", count="exact", head=True).execute()
        runs = sb.table("seo_import_runs").select("*").order("id", desc=True).limit(10).execute()
        return total.count or 0, runs.data or []

    total_count, import_runs = _db_stats()

    st.metric("Totaal producten in Supabase", total_count)

    if import_runs:
        st.markdown("**Laatste imports:**")
        runs_df = pd.DataFrame([{
            "Bestand": r.get("bestandsnaam", ""),
            "Fase": r.get("fase", ""),
            "Aantal": r.get("aantal_producten", 0),
            "Warnings": r.get("aantal_warnings", 0),
            "Datum": (r.get("created_at") or "")[:16],
        } for r in import_runs])
        st.dataframe(runs_df, width="stretch", hide_index=True)

    st.divider()

    # ── Nieuw bestand uploaden ────────────────────────────────────────────────
    st.subheader("📤 Nieuw masterbestand uploaden")
    st.caption(
        "Upload een leverancier-Excel (.xlsx). Het systeem detecteert automatisch de kolommen, "
        "toont een preview, en laadt de data in Supabase."
    )

    mc1, mc2, mc3 = st.columns(3)
    with mc1:
        master_leverancier = st.selectbox(
            "Leverancier",
            ["serax", "potterypots", "printworks", "anders"],
            key="master_leverancier",
            help="Gebruikt de opgeslagen kolom-mapping als die er is.",
        )
    with mc2:
        master_fase = st.text_input(
            "Fase / batch-naam",
            value="3april",
            key="master_fase",
            help="Wordt opgeslagen in het 'fase'-veld in Supabase.",
        )
    with mc3:
        master_auto = st.checkbox(
            "Auto-mapping gebruiken",
            value=True,
            key="master_auto",
            help="Gebruik opgeslagen kolom-mapping zonder handmatige bevestiging.",
        )

    master_file = st.file_uploader(
        "Excel bestand (.xlsx)",
        type=["xlsx"],
        key="master_upload",
    )

    if master_file is not None:
        import tempfile
        from pathlib import Path as _Path

        # Sla tijdelijk op zodat setup_masterdata het kan lezen
        tmp_dir = _Path(".tmp")
        tmp_dir.mkdir(exist_ok=True)
        tmp_path = tmp_dir / master_file.name
        with open(tmp_path, "wb") as f:
            f.write(master_file.getbuffer())

        # Preview
        try:
            preview_df = pd.read_excel(tmp_path, dtype=str, nrows=50)
            st.success(f"✅ Bestand gelezen: {len(preview_df.columns)} kolommen")
            st.dataframe(preview_df.head(20), width="stretch", height=300)
        except Exception as e:
            st.error(f"Kon bestand niet lezen: {e}")
            st.stop()

        # Kolom-mapping preview
        sys_path_backup = __import__("sys").path[:]
        __import__("sys").path.insert(0, "execution")
        from setup_masterdata import detect_mapping, detecteer_header_rij, laad_opgeslagen_mapping

        header_rij = detecteer_header_rij(str(tmp_path))
        df_full = pd.read_excel(tmp_path, header=header_rij, nrows=0, dtype=str)
        kolommen = list(df_full.columns)

        opgeslagen = laad_opgeslagen_mapping(master_leverancier)
        if opgeslagen and master_auto:
            mapping = opgeslagen
            st.info(f"📋 Opgeslagen mapping voor **{master_leverancier}** geladen ({len(mapping)} velden)")
        else:
            mapping = detect_mapping(kolommen)
            st.info(f"🔍 Auto-detectie: {len(mapping)} velden herkend")

        # Toon mapping
        with st.expander("Kolom-mapping preview", expanded=True):
            mapping_rows = []
            for db_veld, kolom in mapping.items():
                mapping_rows.append({"DB veld": db_veld, "→": "→", "Kolom in bestand": kolom})
            niet_gemapt = [k for k in kolommen if k not in mapping.values()]
            st.dataframe(pd.DataFrame(mapping_rows), width="stretch", hide_index=True)
            if niet_gemapt:
                st.caption(f"{len(niet_gemapt)} kolommen niet gekoppeld: {', '.join(niet_gemapt[:8])}{'...' if len(niet_gemapt)>8 else ''}")

        # Laden
        if st.button("🚀 Laden in Supabase", type="primary", key="master_load"):
            from setup_masterdata import laad_masterdata, sla_mapping_op
            sla_mapping_op(master_leverancier, mapping)
            with st.spinner("Producten laden in Supabase..."):
                try:
                    n = laad_masterdata(str(tmp_path), master_leverancier, master_fase, mapping)
                except Exception as e:
                    st.error(f"Fout bij laden: {e}")
                    n = 0
            if n > 0:
                st.success(f"✅ **{n} producten** geladen in Supabase (fase={master_fase}, status=raw)")
                # Dimensie-fix voor Serax (gecombineerde kolom)
                if master_leverancier == "serax" and "Product dimensions cm" in kolommen:
                    st.info("🔧 Serax-afmetingen worden geparsed (L/W/H uit gecombineerde kolom)...")
                    from fix_serax_dimensions import parse_dimensions
                    df_dims = pd.read_excel(tmp_path, header=header_rij, dtype=str)
                    updated_dims = 0
                    for _, row in df_dims.iterrows():
                        sku = str(row.get("Brand_id") or "").strip()
                        raw_dim = row.get("Product dimensions cm")
                        if not sku or pd.isna(raw_dim):
                            continue
                        dims = parse_dimensions(str(raw_dim))
                        if dims:
                            sb.table("seo_products").update(dims).eq("sku", sku).execute()
                            updated_dims += 1
                    st.success(f"✅ Afmetingen geparsed voor {updated_dims} Serax-producten")
                # Diameter-fix voor Pottery Pots
                if master_leverancier == "potterypots":
                    diam_col_name = next((c for c in kolommen if "diameter" in c.lower() and "single" in c.lower()), None)
                    l_col_name = mapping.get("lengte_cm")
                    if diam_col_name:
                        st.info("🔧 Pottery Pots: ronde potten krijgen diameter als L/B...")
                        df_pp = pd.read_excel(tmp_path, header=header_rij, dtype=str)
                        sku_col_name = mapping.get("sku")
                        updated_diam = 0
                        LEEG = {"-", "", "nan", "n/a", "none"}
                        for _, row in df_pp.iterrows():
                            sku = str(row.get(sku_col_name) or "").strip()
                            if not sku or sku.lower() in LEEG:
                                continue
                            l_val = str(row.get(l_col_name) or "").strip() if l_col_name else ""
                            d_val = str(row.get(diam_col_name) or "").strip()
                            if l_val.lower() in LEEG and d_val.lower() not in LEEG:
                                try:
                                    d_num = float(d_val.replace(",", "."))
                                    sb.table("seo_products").update(
                                        {"lengte_cm": d_num, "breedte_cm": d_num}
                                    ).eq("sku", sku).execute()
                                    updated_diam += 1
                                except (ValueError, Exception):
                                    pass
                        st.success(f"✅ Diameter als L/B ingevuld voor {updated_diam} ronde potten")
                # Clear stats cache
                _db_stats.clear()
                st.rerun()
            else:
                st.warning("Geen producten geladen. Check de mapping en het bestand.")

    st.divider()

    # ── Bestaande data bekijken / aanpassen ───────────────────────────────────
    st.subheader("🔍 Bestaande data bekijken & aanpassen")

    search_sku = st.text_input("Zoek op SKU (of deel ervan)", key="master_search", placeholder="bv. B2026105 of 399234")

    if search_sku and search_sku.strip():
        results = sb.table("seo_products").select(
            "sku, product_name_raw, leverancier_category, leverancier_item_cat, "
            "kleur_en, materiaal_nl, rrp_stuk_eur, inkoopprijs_stuk_eur, "
            "hoogte_cm, lengte_cm, breedte_cm, ean_shopify, fase, status"
        ).ilike("sku", f"%{search_sku.strip()}%").limit(20).execute()

        if results.data:
            st.success(f"{len(results.data)} resultaten gevonden")
            search_df = pd.DataFrame(results.data)
            edited_search = st.data_editor(
                search_df,
                width="stretch",
                hide_index=True,
                disabled=["sku", "fase", "status"],
                key="master_edit",
            )

            if st.button("💾 Wijzigingen opslaan naar Supabase", key="master_save_edits"):
                saved = 0
                for _, row in edited_search.iterrows():
                    sku = row["sku"]
                    updates = {}
                    for col in ["product_name_raw", "leverancier_category", "leverancier_item_cat",
                                "kleur_en", "materiaal_nl", "rrp_stuk_eur", "inkoopprijs_stuk_eur",
                                "hoogte_cm", "lengte_cm", "breedte_cm", "ean_shopify"]:
                        val = row.get(col)
                        if pd.notna(val):
                            updates[col] = val
                        else:
                            updates[col] = None
                    sb.table("seo_products").update(updates).eq("sku", sku).execute()
                    saved += 1
                st.success(f"✅ {saved} producten bijgewerkt in Supabase")
                _db_stats.clear()
        else:
            st.warning(f"Geen producten gevonden voor '{search_sku}'")

    # ── Bulk verwijderen ──────────────────────────────────────────────────────
    with st.expander("🗑️ Data verwijderen", expanded=False):
        st.warning("⚠️ Let op: dit verwijdert producten permanent uit Supabase.")
        del_fase = st.text_input("Verwijder alle producten met fase:", key="master_del_fase")
        if st.button("🗑️ Verwijderen", key="master_del_btn", type="secondary"):
            if del_fase and del_fase.strip():
                count_before = sb.table("seo_products").select("id", count="exact", head=True).eq("fase", del_fase.strip()).execute()
                n_del = count_before.count or 0
                if n_del > 0:
                    sb.table("seo_products").delete().eq("fase", del_fase.strip()).execute()
                    st.success(f"✅ {n_del} producten met fase='{del_fase.strip()}' verwijderd")
                    _db_stats.clear()
                    st.rerun()
                else:
                    st.info(f"Geen producten gevonden met fase='{del_fase.strip()}'")


# ═════════════════════════════════════════════════════════════════════════════
# STAP 1 — Upload publicatie-batch
# ═════════════════════════════════════════════════════════════════════════════

elif stap == STAPPEN[1]:
    st.title("Stap 1 — Upload publicatie-batch")
    st.markdown(
        "Upload de Excel met producten die gepubliceerd moeten worden op de webshop. "
        "Het systeem zal deze rijen in stap 2 matchen tegen `seo_products` in Supabase op **SKU** of **EAN**."
    )

    uploaded = st.file_uploader(
        "Excel bestand (.xlsx)",
        type=["xlsx"],
        help="De Excel mag elke kolomstructuur hebben. Het systeem detecteert automatisch welke kolom de SKU en EAN bevat.",
    )

    if uploaded is not None:
        try:
            df = pd.read_excel(uploaded, dtype=str)
        except Exception as e:
            st.error(f"Kon Excel niet lezen: {e}")
            st.stop()

        st.session_state.uploaded_df = df
        st.session_state.uploaded_filename = uploaded.name
        st.success(f"✅ {len(df)} rijen ingelezen uit **{uploaded.name}** ({len(df.columns)} kolommen)")

        st.divider()
        st.subheader("Batch-instellingen")
        bc1, bc2 = st.columns(2)
        with bc1:
            st.session_state.merk = st.text_input(
                "Merknaam (voor producttitel)",
                value=st.session_state.merk,
                help="Dit wordt het eerste segment in de titel: '{Merk} - Designer - Product'. "
                     "Bijv. Serax, Pottery Pots, Printworks.",
            )
        with bc2:
            st.session_state.batch_tag = st.text_input(
                "Batch-tag (voor Shopify tags)",
                value=st.session_state.batch_tag,
                help="Wordt als laatste tag toegevoegd. Bijv. structuur_fase4, tijdelijke_import, test_batch.",
            )

        st.subheader("Preview (eerste 50 rijen)")
        st.dataframe(df.head(50), width="stretch", height=400)

        st.divider()
        st.subheader("Welke kolom is wat?")
        st.caption("Auto-detectie heeft een gok gedaan. Pas aan als nodig.")

        cols = list(df.columns)
        c1, c2, c3 = st.columns(3)

        sku_default = detect_column(cols, SKU_ALIASES)
        ean_default = detect_column(cols, EAN_ALIASES)
        naam_default = detect_column(cols, NAAM_ALIASES)

        with c1:
            st.session_state.sku_col = st.selectbox(
                "SKU / Artikelnummer-kolom",
                cols,
                index=safe_index(cols, sku_default) if sku_default else 0,
            )
        with c2:
            st.session_state.ean_col = st.selectbox(
                "EAN-kolom",
                cols,
                index=safe_index(cols, ean_default) if ean_default else 0,
            )
        with c3:
            st.session_state.naam_col = st.selectbox(
                "Naam-kolom (optioneel, voor visuele check)",
                ["(geen)"] + cols,
                index=(safe_index(cols, naam_default) + 1) if naam_default else 0,
            )

        st.divider()
        # Toon waardes uit de gekozen kolommen — dedup zodat dezelfde kolom
        # niet twee keer in de slice komt (anders crasht pyarrow)
        sample_cols_raw = [st.session_state.sku_col, st.session_state.ean_col, st.session_state.naam_col]
        sample_cols = []
        for c in sample_cols_raw:
            if c and c != "(geen)" and c not in sample_cols:
                sample_cols.append(c)
        if sample_cols:
            st.subheader("Steekproef gekozen kolommen")
            st.dataframe(df[sample_cols].head(10), width="stretch")

        st.success("Klaar om te matchen — ga in de zijbalk naar **stap 2**.")

    elif st.session_state.uploaded_df is not None:
        st.info(
            f"📄 Vorige upload nog in sessie: **{st.session_state.uploaded_filename}** "
            f"({len(st.session_state.uploaded_df)} rijen). "
            "Kies hierboven een nieuwe upload of ga in de zijbalk naar stap 2."
        )


# ═════════════════════════════════════════════════════════════════════════════
# STAP 2 — Match tegen Supabase
# ═════════════════════════════════════════════════════════════════════════════

elif stap == STAPPEN[2]:
    st.title("Stap 2 — Match tegen Supabase")

    if st.session_state.uploaded_df is None:
        st.warning("⚠️ Nog geen Excel geüpload. Ga eerst naar **stap 1**.")
        st.stop()

    df_upload = st.session_state.uploaded_df
    sku_col = st.session_state.sku_col
    ean_col = st.session_state.ean_col
    naam_col = st.session_state.naam_col if st.session_state.naam_col != "(geen)" else None

    if not sku_col:
        st.warning("⚠️ SKU-kolom niet gekozen. Ga eerst naar **stap 1**.")
        st.stop()

    st.markdown(
        f"**{len(df_upload)} rijen** uit `{st.session_state.uploaded_filename}` "
        f"worden gematcht tegen `seo_products` in Supabase."
    )
    st.markdown(
        f"- SKU-kolom Excel: `{sku_col}`\n"
        f"- Match-strategie: **alleen exacte SKU-match** tegen `seo_products.sku`\n"
        f"- EAN wordt wel in de side-by-side getoond ter visuele controle, maar speelt geen rol in de matching"
    )

    run = st.button("▶️ Start matching", type="primary")

    if run:
        with st.spinner("Producten ophalen uit Supabase..."):
            sb = get_supabase()

            skus = sorted({str(s).strip() for s in df_upload[sku_col].dropna() if str(s).strip()})

            sb_by_sku = {}

            BATCH = 200
            for i in range(0, len(skus), BATCH):
                batch = skus[i:i + BATCH]
                res = sb.table("seo_products").select(
                    "id,sku,ean_shopify,ean_piece,product_name_raw,designer,"
                    "leverancier_category,leverancier_item_cat,materiaal_nl,kleur_en,"
                    "rrp_stuk_eur,inkoopprijs_stuk_eur,fase,status"
                ).in_("sku", batch).execute()
                for r in res.data:
                    sb_by_sku[r["sku"]] = r

        # Bouw match-resultaten per Excel rij
        rows = []
        for _, row in df_upload.iterrows():
            ex_sku = str(row.get(sku_col, "") or "").strip()
            ex_ean = str(row.get(ean_col, "") or "").strip() if ean_col else ""
            ex_naam = str(row.get(naam_col, "") or "").strip() if naam_col else ""

            best = sb_by_sku.get(ex_sku) if ex_sku else None

            if best:
                status = "🟢 SKU match"
            else:
                status = "🔴 geen match"

            rows.append({
                "Status": status,
                "Excel SKU": ex_sku,
                "Excel EAN": ex_ean,
                "Excel naam": ex_naam[:60],
                "→": "→",
                "DB SKU": (best or {}).get("sku", "") or "",
                "DB EAN (shopify)": (best or {}).get("ean_shopify", "") or "",
                "DB EAN (piece)": (best or {}).get("ean_piece", "") or "",
                "DB naam": ((best or {}).get("product_name_raw") or "")[:60],
                "DB Designer": (best or {}).get("designer", "") or "",
                "DB Materiaal": (best or {}).get("materiaal_nl", "") or "",
                "DB Kleur (EN)": (best or {}).get("kleur_en", "") or "",
                "DB Lev.cat.": (best or {}).get("leverancier_category", "") or "",
                "DB Lev.item": (best or {}).get("leverancier_item_cat", "") or "",
                "DB Prijs €": (best or {}).get("rrp_stuk_eur", "") or "",
                "DB id": (best or {}).get("id", "") or "",
            })

        match_df = pd.DataFrame(rows)
        groen = sum(1 for r in rows if r["Status"].startswith("🟢"))
        rood  = sum(1 for r in rows if r["Status"].startswith("🔴"))

        st.session_state.matches_df = match_df
        st.session_state.match_stats = {"groen": groen, "geel": 0, "rood": rood, "totaal": len(rows)}
        # Reset enrichment-state — anders blijft oude batch hangen na hermatch
        st.session_state.batch_products = None
        st.session_state.stap3_done = False
        st.session_state.stap4_done = False
        st.session_state.stap5_done = False
        st.session_state.stap6_committed = False
        st.session_state.exported_path = None

    # Toon resultaat (zowel bij verse run als bij terugkeer naar de pagina)
    if st.session_state.matches_df is not None:
        match_df = st.session_state.matches_df
        s = st.session_state.match_stats

        st.divider()
        c1, c2, c3 = st.columns(3)
        c1.metric("📦 Totaal Excel-rijen", s["totaal"])
        c2.metric("🟢 SKU match", s["groen"])
        c3.metric("🔴 Geen match", s["rood"])

        st.divider()
        st.subheader("Side-by-side vergelijking")

        # Filter
        filter_keuze = st.radio(
            "Filter",
            ["Alles", "🟢 Alleen matches", "🔴 Alleen geen-match"],
            horizontal=True,
        )
        if filter_keuze == "🟢 Alleen matches":
            view_df = match_df[match_df["Status"].str.startswith("🟢")]
        elif filter_keuze == "🔴 Alleen geen-match":
            view_df = match_df[match_df["Status"].str.startswith("🔴")]
        else:
            view_df = match_df

        # Toon met kleurcodering via pandas Styler
        styled = view_df.style.map(kleur_status, subset=["Status"])
        st.dataframe(styled, width="stretch", height=600)

        st.caption(f"Toont {len(view_df)} van {len(match_df)} rijen")

        st.divider()
        st.subheader("Doorlaten naar enrichment")
        st.markdown(
            "Alleen 🟢 SKU-matches gaan door naar de volgende stappen. "
            "🔴 geen-match wordt overgeslagen want er is geen DB-record om te enrichen."
        )

        ids_door = match_df[match_df["Status"].str.startswith("🟢")]["DB id"].tolist()
        ids_door = [i for i in ids_door if i]
        st.session_state.selected_for_enrichment = set(ids_door)
        st.success(f"✅ {len(ids_door)} producten staan klaar voor stap 3 (categorie-toewijzing).")

        # Download knoppen
        dc1, dc2 = st.columns(2)

        # Volledig rapport
        csv_all = match_df.to_csv(index=False).encode("utf-8-sig")
        dc1.download_button(
            "📥 Download volledig match-rapport (CSV)",
            csv_all,
            file_name=f"match_rapport_{st.session_state.uploaded_filename or 'batch'}.csv",
            mime="text/csv",
        )

        # Alleen niet-gematchte rijen exporteren
        no_match_df = match_df[match_df["Status"].str.startswith("🔴")]
        if len(no_match_df) > 0:
            # Pak ook de originele Excel-rijen voor de niet-gematchte SKUs
            no_match_skus = set(no_match_df["Excel SKU"].tolist())
            df_upload = st.session_state.uploaded_df
            sku_col = st.session_state.sku_col
            original_no_match = df_upload[df_upload[sku_col].astype(str).str.strip().isin(no_match_skus)]

            csv_nomatch = original_no_match.to_csv(index=False).encode("utf-8-sig")
            dc2.download_button(
                f"📥 Export {len(no_match_df)} niet-gematchte producten (CSV)",
                csv_nomatch,
                file_name=f"niet_gematcht_{st.session_state.uploaded_filename or 'batch'}.csv",
                mime="text/csv",
            )
        else:
            dc2.success("Geen niet-gematchte producten")

    else:
        st.info("Klik op **Start matching** om de Supabase-vergelijking te draaien.")


# ═════════════════════════════════════════════════════════════════════════════
# STAP 3-7 — STUBS (worden gebouwd nadat 1+2 zijn getest)
# ═════════════════════════════════════════════════════════════════════════════

elif stap == STAPPEN[3]:
    st.title("Stap 3 — Categorie-toewijzing")

    if not st.session_state.selected_for_enrichment:
        st.warning("⚠️ Geen producten geselecteerd. Doe eerst **stap 1+2**.")
        st.stop()

    batch = ensure_batch_loaded()
    if not batch:
        st.error("Kon geen producten ophalen uit Supabase. Ga terug naar stap 2 en draai matching opnieuw.")
        st.stop()

    st.markdown(
        f"Voor elk van de **{len(batch)} gematchte producten** zoekt het systeem in "
        f"`seo_category_mapping` op `(leverancier_category, leverancier_item_cat)`. "
        f"Niet gevonden? Vul handmatig in (de tabel is editbaar)."
    )

    sb = get_supabase()

    # Bouw cache van alle bestaande mappings (één query)
    if "_cat_cache" not in st.session_state:
        all_maps = sb.table("seo_category_mapping").select("*").execute().data or []
        st.session_state._cat_cache = {
            (m["leverancier_category"], m["leverancier_item_cat"]): m for m in all_maps
        }
    cat_cache = st.session_state._cat_cache

    # Lees actieve sub-subcategorieën uit de website indeling (groene cellen)
    actieve_subsubcats = load_active_subsubcategories()

    # Bouw editor-rijen — auto-fill uit mapping waar mogelijk
    # AUTO-COMMIT: schrijf direct terug naar batch_products zodat tags in stap 5 niet leeg blijven
    # SLIM: outdoor-detectie voor potten/vazen (op basis van leverancier_category + productnaam)
    rows = []
    auto_committed = 0
    for p in batch:
        lev_cat = p.get("leverancier_category") or ""
        lev_item = p.get("leverancier_item_cat") or ""
        naam_lower = (p.get("product_name_raw") or "").lower()
        match = cat_cache.get((lev_cat, lev_item))

        if match:
            status = "🟢 mapping gevonden"
            hoofd = p.get("hoofdcategorie") or match.get("hoofdcategorie") or ""
            sub = p.get("subcategorie") or match.get("subcategorie") or ""
            subsub = p.get("sub_subcategorie") or match.get("sub_subcategorie") or ""

            # Productnaam-gebaseerde sub-subcat overrides
            if subsub in ("Serax bloempot", "Bloempotten binnen", "Bloempotten buiten"):
                if "flower pot" in naam_lower or "sierpot" in naam_lower:
                    subsub = "Sierpotten"
                    p["_extra_tags"] = []
                elif "wally" in naam_lower or "hanging" in naam_lower:
                    subsub = "Hangpotten"
                    p["_extra_tags"] = []

            # Stoelen: chair / seat
            if any(w in naam_lower for w in ["chair", "seat", "stoel"]):
                hoofd, sub, subsub = "Wonen & badkamer", "Verlichting & Meubels", "Stoelen"
                p["_extra_tags"] = []

            # Espresso (zonder saucer) → Espressokopjes
            if "espresso" in naam_lower and "saucer" not in naam_lower:
                hoofd, sub, subsub = "servies", "Kommen, Mokken & Bekers", "Espressokopjes"
                p["_extra_tags"] = []

            # Candle holders → Kandelaars (niet Geurkaarsen)
            if "candle holder" in naam_lower or "tea light holder" in naam_lower or "candleholder" in naam_lower:
                hoofd, sub, subsub = "Keuken & Eetkamer", "Tafel & Sfeer", "Kandelaars"
                p["_extra_tags"] = []

            # Deep plate → extra tag "Diepe borden"
            if "deep plate" in naam_lower or "deep serving plate" in naam_lower:
                extra = p.get("_extra_tags") or []
                if "Diepe borden" not in extra:
                    extra.append("Diepe borden")
                p["_extra_tags"] = extra

            # Glazen: naam-gebaseerde sub-subcat op basis van producttype
            if match and lev_cat == "Glassware":
                if "wine" in naam_lower or " vin " in naam_lower:
                    hoofd, sub, subsub = "Glazen", "Wijn & Champagne", "Wijnglazen"
                elif "champagne" in naam_lower or "coupe" in naam_lower or "flute" in naam_lower:
                    hoofd, sub, subsub = "Glazen", "Wijn & Champagne", "Champagneglazen"
                elif "water glass" in naam_lower:
                    hoofd, sub, subsub = "Glazen", "Water & Thee", "Waterglazen"
                elif "tea glass" in naam_lower or "tea cup" in naam_lower or "tea strainer" in naam_lower:
                    hoofd, sub, subsub = "Glazen", "Water & Thee", "Theeglazen"
                elif "longdrink" in naam_lower or "long drink" in naam_lower:
                    hoofd, sub, subsub = "Glazen", "Bar & Cocktail", "Longdrink glazen"
                elif "shot" in naam_lower:
                    hoofd, sub, subsub = "Glazen", "Bar & Cocktail", "Shotglazen"
                elif "whisky" in naam_lower or "whiskey" in naam_lower:
                    hoofd, sub, subsub = "Glazen", "Bar & Cocktail", "Whiskyglazen"
                elif "cocktail" in naam_lower:
                    hoofd, sub, subsub = "Glazen", "Bar & Cocktail", "Cocktailglazen"
                elif "espresso" in naam_lower:
                    hoofd, sub, subsub = "servies", "Kommen, Mokken & Bekers", "Espressokopjes"
                elif "tumbler" in naam_lower:
                    hoofd, sub, subsub = "Glazen", "Water & Thee", "Drinkglazen"
                # else: default Drinkglazen (uit mapping)

            # Bloempotten: bijna alle potten zijn zowel indoor als outdoor geschikt.
            # → sub_subcategorie = "Bloempotten buiten" (primair)
            # → _extra_tags krijgt "Bloempotten binnen" erbij (dual tag)
            # Uitzondering: expliciet indoor-only producten
            if subsub in ("Bloempotten binnen", "Bloempotten buiten"):
                is_indoor_only = (
                    "indoor" in naam_lower and "outdoor" not in naam_lower
                )
                if is_indoor_only:
                    subsub = "Bloempotten binnen"
                    p["_extra_tags"] = []  # alleen indoor
                else:
                    # Default: outdoor + indoor dual tag
                    subsub = "Bloempotten buiten"
                    p["_extra_tags"] = ["Bloempotten binnen"]

            # Direct in batch_products schrijven — altijd overschrijven met
            # de mapping-tabel waarde (ook als er al een categorie stond)
            if hoofd:
                p["hoofdcategorie"] = hoofd
                p["subcategorie"] = sub
                p["sub_subcategorie"] = subsub
                p["collectie"] = sub  # SOP: collectie = subcategorie
                auto_committed += 1
        else:
            status = "🔴 geen mapping"
            hoofd = p.get("hoofdcategorie") or ""
            sub = p.get("subcategorie") or ""
            subsub = p.get("sub_subcategorie") or ""

        # Serviessets & cadeausets: ALTIJD detecteren, ook zonder mapping match
        if "startset" in naam_lower:
            hoofd, sub, subsub = "Servies", "Serviessets", "Serviessets"
        elif "breakfast set" in naam_lower:
            hoofd, sub, subsub = "Servies", "Serviessets", "Ontbijtsets"
        elif "dinner set" in naam_lower:
            hoofd, sub, subsub = "Servies", "Serviessets", "Dinersets"
        elif "cappuccino set" in naam_lower or "tea set" in naam_lower:
            hoofd, sub, subsub = "Servies", "Serviessets", "Koffie- & Theesets"
        elif "giftset" in naam_lower:
            hoofd, sub, subsub = "Servies", "Serviessets", "Cadeausets"

        # Schrijf set-categorieën terug (ook als er geen mapping was)
        if hoofd and not p.get("hoofdcategorie"):
            p["hoofdcategorie"] = hoofd
            p["subcategorie"] = sub
            p["sub_subcategorie"] = subsub
            p["collectie"] = sub

        # Online-status van de sub-subcategorie op de webshop
        if not subsub:
            online = ""
        elif subsub.strip().lower() in actieve_subsubcats:
            online = "🟢 online"
        else:
            online = "🔴 nog niet"

        # Extra sub-subcategorie (dual tag) visueel tonen
        extra_tags = p.get("_extra_tags") or []
        sub2 = extra_tags[0] if extra_tags else ""

        rows.append({
            "Status": status,
            "SKU": p.get("sku") or "",
            "🔍": f"https://www.google.com/search?q={(p.get('sku') or '').replace(' ', '+')}",
            "Productnaam (raw)": (p.get("product_name_raw") or "")[:60],
            "Designer": p.get("designer") or "",
            "Kleur (raw)": p.get("kleur_en") or "",
            "Lev. category": lev_cat,
            "Lev. item cat": lev_item,
            "→": "→",
            "Hoofdcategorie": hoofd,
            "Subcategorie": sub,
            "Sub-subcategorie": subsub,
            "Sub-subcategorie 2": sub2,
            "Online?": online,
            "_id": p.get("id"),
        })

    if auto_committed > 0:
        st.session_state.batch_products = batch

    df_cat = pd.DataFrame(rows)

    # Stats
    n_groen = sum(1 for r in rows if r["Status"].startswith("🟢"))
    n_rood = sum(1 for r in rows if r["Status"].startswith("🔴"))

    c1, c2, c3 = st.columns(3)
    c1.metric("📦 Producten", len(rows))
    c2.metric("🟢 Auto-mapped", n_groen)
    c3.metric("🔴 Handmatig invullen", n_rood)

    if auto_committed > 0:
        st.info(f"ℹ️ {auto_committed} categorieën zijn automatisch gevuld op basis van `seo_category_mapping`. "
                f"Pas aan als nodig en klik **Vastleggen** om eigen wijzigingen te bewaren.")

    tree = load_website_tree()

    # ── Tab 1: gemapte producten / Tab 2: niet-gemapte producten ──────────────
    tab_mapped, tab_unmapped = st.tabs([
        f"🟢 Gemapt ({n_groen})",
        f"🔴 Niet gemapt ({n_rood})",
    ])

    with tab_mapped:
        mapped_df = df_cat[df_cat["Status"].str.startswith("🟢")]
        if len(mapped_df) > 0:
            # Bouw lijst van alle sub-subcats voor de dropdown
            alle_subsubcats = sorted({
                ssc for subs in tree.values() for sscs in subs.values() for ssc in sscs
            }) if tree else []

            mapped_edited = st.data_editor(
                mapped_df,
                width="stretch",
                height=400,
                hide_index=True,
                disabled=["Status", "SKU", "🔍", "Productnaam (raw)", "Designer", "Kleur (raw)",
                          "Lev. category", "Lev. item cat", "→", "Online?", "_id"],
                column_config={
                    "_id": None,
                    "🔍": st.column_config.LinkColumn("🔍", display_text="zoek", width="small"),
                    "Hoofdcategorie": st.column_config.TextColumn(width="medium"),
                    "Subcategorie": st.column_config.TextColumn(width="medium"),
                    "Sub-subcategorie": st.column_config.TextColumn(width="medium"),
                    "Sub-subcategorie 2": st.column_config.TextColumn(width="medium"),
                },
                key="stap3_mapped_editor",
            )

            if st.button("💾 Categorieën opslaan", type="primary", key="stap3_save_cats"):
                saved = 0
                for _, row in mapped_edited.iterrows():
                    pid = row["_id"]
                    for p in batch:
                        if p["id"] == pid:
                            h = (row.get("Hoofdcategorie") or "").strip()
                            s = (row.get("Subcategorie") or "").strip()
                            ss = (row.get("Sub-subcategorie") or "").strip()
                            sub2 = (row.get("Sub-subcategorie 2") or "").strip()
                            if h: p["hoofdcategorie"] = h
                            if s:
                                p["subcategorie"] = s
                                p["collectie"] = s
                            if ss: p["sub_subcategorie"] = ss
                            p["_extra_tags"] = [sub2] if sub2 else []
                            saved += 1
                            break
                st.session_state.batch_products = batch
                st.success(f"✅ Categorieën opgeslagen voor {saved} producten.")
                st.rerun()
        else:
            st.info("Alle producten zijn nog niet gemapt.")

    with tab_unmapped:
        unmapped_df = df_cat[df_cat["Status"].str.startswith("🔴")]
        if len(unmapped_df) > 0:
            st.dataframe(unmapped_df, width="stretch", height=400, hide_index=True,
                         column_config={"_id": None, "🔍": st.column_config.LinkColumn("🔍", display_text="zoek", width="small")})

            # ── Claude batch-categorisatie ────────────────────────────────────
            st.divider()
            st.subheader("🤖 Claude categoriseren")
            st.caption(
                "Laat Claude in 1 batch-call alle niet-gemapte producten categoriseren op basis van "
                "de productnaam en de website-indeling. Je kunt daarna reviewen en aanpassen."
            )

            # Bouw de unieke (lev_cat, lev_item, voorbeeld_naam) combinaties
            unmapped_combos = {}
            for p in batch:
                if p.get("hoofdcategorie"):
                    continue  # al gemapt
                lev_cat = p.get("leverancier_category") or "?"
                lev_item = p.get("leverancier_item_cat") or "?"
                key = (lev_cat, lev_item)
                if key not in unmapped_combos:
                    unmapped_combos[key] = {
                        "voorbeeld": (p.get("product_name_raw") or "")[:50],
                        "count": 0,
                    }
                unmapped_combos[key]["count"] += 1

            st.info(f"**{len(unmapped_combos)} unieke leverancier-combinaties** voor {n_rood} producten.")

            if st.button(f"🤖 Laat Claude {len(unmapped_combos)} combinaties categoriseren (1 call)",
                         type="primary", key="stap3_claude_cat"):
                # Bouw boom-tekst voor de prompt
                boom_lines = []
                if tree:
                    for hc in sorted(tree.keys()):
                        for sc in sorted(tree[hc].keys()):
                            subs = ", ".join(sorted(tree[hc][sc]))
                            boom_lines.append(f"  {hc} > {sc} > [{subs}]")
                boom_tekst = "\n".join(boom_lines)

                # Bouw de input-lijst
                input_lines = []
                combo_keys = list(unmapped_combos.keys())
                for i, (lev_cat, lev_item) in enumerate(combo_keys):
                    info = unmapped_combos[(lev_cat, lev_item)]
                    input_lines.append(f"{i+1}. ({lev_cat} / {lev_item}) voorbeeld: {info['voorbeeld']}")

                prompt = (
                    f"Je bent een categorie-expert voor een Belgische interieur-webshop.\n\n"
                    f"Hieronder staan {len(combo_keys)} leverancier-codes met een voorbeeldproductnaam.\n"
                    f"Kies voor ELKE code de juiste categorie uit de website-boom.\n\n"
                    f"WEBSITE CATEGORIEËN:\n{boom_tekst}\n\n"
                    f"REGELS:\n"
                    f"- Kies ALLEEN uit bovenstaande categorieën\n"
                    f"- Output per regel: nummer|hoofdcategorie|subcategorie|sub-subcategorie\n"
                    f"- GEEN uitleg, ALLEEN de pipe-separated output\n\n"
                    f"INPUT:\n" + "\n".join(input_lines) + "\n\nOUTPUT:"
                )

                claude = get_claude_client()
                with st.spinner(f"Claude categoriseert {len(combo_keys)} combinaties..."):
                    try:
                        resp = claude.messages.create(
                            model="claude-haiku-4-5-20251001",
                            max_tokens=4000,
                            messages=[{"role": "user", "content": prompt}],
                        )
                        output = resp.content[0].text.strip()
                        lines = [l.strip() for l in output.split("\n") if l.strip()]

                        suggesties = {}
                        for line in lines:
                            parts = line.split("|")
                            if len(parts) >= 4:
                                try:
                                    idx = int(parts[0].strip().rstrip(".")) - 1
                                    if 0 <= idx < len(combo_keys):
                                        suggesties[combo_keys[idx]] = {
                                            "hoofdcategorie": parts[1].strip(),
                                            "subcategorie": parts[2].strip(),
                                            "sub_subcategorie": parts[3].strip(),
                                        }
                                except (ValueError, IndexError):
                                    pass

                        st.session_state["_cat_suggesties"] = suggesties
                        st.success(f"✅ {len(suggesties)} suggesties ontvangen. Review hieronder.")
                    except Exception as e:
                        st.error(f"Claude-fout: {e}")

            # ── Review suggesties ─────────────────────────────────────────────
            if st.session_state.get("_cat_suggesties"):
                suggesties = st.session_state["_cat_suggesties"]
                st.divider()
                st.subheader("📋 Review Claude's suggesties")
                st.caption("Pas aan waar nodig en klik **Goedkeuren en opslaan**.")

                sug_rows = []
                for (lev_cat, lev_item), sug in suggesties.items():
                    info = unmapped_combos.get((lev_cat, lev_item), {})
                    online = ""
                    ss = sug.get("sub_subcategorie", "")
                    if ss and ss.strip().lower() in actieve_subsubcats:
                        online = "🟢 online"
                    elif ss:
                        online = "🔴 nog niet"
                    sug_rows.append({
                        "Lev. category": lev_cat,
                        "Lev. item cat": lev_item,
                        "Voorbeeld": info.get("voorbeeld", ""),
                        "Aantal": info.get("count", 0),
                        "Hoofdcategorie": sug.get("hoofdcategorie", ""),
                        "Subcategorie": sug.get("subcategorie", ""),
                        "Sub-subcategorie": sug.get("sub_subcategorie", ""),
                        "Online?": online,
                    })

                sug_df = pd.DataFrame(sug_rows)
                sug_edited = st.data_editor(
                    sug_df,
                    width="stretch",
                    height=400,
                    hide_index=True,
                    disabled=["Lev. category", "Lev. item cat", "Voorbeeld", "Aantal", "Online?"],
                    key="stap3_sug_editor",
                )

                sc1, sc2 = st.columns(2)
                if sc1.button("✅ Goedkeuren en opslaan", type="primary", key="stap3_approve_sug"):
                    sb = get_supabase()
                    saved = 0
                    for _, row in sug_edited.iterrows():
                        lev_cat = row["Lev. category"]
                        lev_item = row["Lev. item cat"]
                        hoofd = (row["Hoofdcategorie"] or "").strip()
                        sub = (row["Subcategorie"] or "").strip()
                        subsub = (row["Sub-subcategorie"] or "").strip()
                        if hoofd and sub and subsub:
                            # 1. Sla op in mapping-tabel voor toekomstige runs
                            sb.table("seo_category_mapping").upsert({
                                "leverancier_category": lev_cat,
                                "leverancier_item_cat": lev_item,
                                "hoofdcategorie": hoofd,
                                "subcategorie": sub,
                                "sub_subcategorie": subsub,
                            }, on_conflict="leverancier_category,leverancier_item_cat").execute()

                            # 2. Pas toe op huidige batch
                            for p in batch:
                                if (p.get("leverancier_category") or "?") == lev_cat and \
                                   (p.get("leverancier_item_cat") or "?") == lev_item:
                                    p["hoofdcategorie"] = hoofd
                                    p["subcategorie"] = sub
                                    p["sub_subcategorie"] = subsub
                                    p["collectie"] = sub
                            saved += 1

                    st.session_state.batch_products = batch
                    st.session_state.pop("_cat_cache", None)
                    st.session_state.pop("_cat_suggesties", None)
                    st.success(f"✅ {saved} mappings opgeslagen in Supabase en toegepast op de batch.")
                    st.rerun()

                if sc2.button("❌ Verwijder suggesties", key="stap3_clear_sug"):
                    st.session_state.pop("_cat_suggesties", None)
                    st.rerun()

        else:
            st.success("Alle producten zijn al gecategoriseerd!")

    st.divider()
    cc1, cc2, cc3 = st.columns([1, 1, 2])
    if cc1.button("💾 Stap 3 afronden", type="primary"):
        st.session_state.batch_products = batch
        st.session_state.stap3_done = True
        st.success(f"✅ Categorieën vastgelegd voor {len(batch)} producten. Ga naar **stap 4**.")

    if cc2.button("🔄 Reset mapping-cache"):
        st.session_state.pop("_cat_cache", None)
        st.rerun()

    if st.session_state.stap3_done:
        cc3.success("Stap 3 afgerond")

    # Leerbox stap 3
    learning_chatbox(
        stap_naam="Categorie-toewijzing",
        stap_context=(
            "Stap 3: categorie-mapping van leverancier-codes naar website-categorieën. "
            "De website heeft 6 hoofdcategorieën: servies, Glazen, Vazen & Potten, "
            "Keuken & Eetkamer, Slapen, Wonen & badkamer. Elke hoofdcat heeft subcats en sub-subcats. "
            "De mapping-tabel in Supabase (seo_category_mapping) koppelt "
            "(leverancier_category, leverancier_item_cat) → (hoofdcategorie, subcategorie, sub_subcategorie)."
        ),
        key="stap3",
    )


elif stap == STAPPEN[4]:
    st.title("Stap 4 — Vertalen materiaal & kleur")

    if not st.session_state.selected_for_enrichment:
        st.warning("⚠️ Geen producten geselecteerd. Doe eerst **stap 1+2**.")
        st.stop()

    batch = ensure_batch_loaded()
    if not batch:
        st.error("Kon geen producten ophalen. Ga terug naar stap 2.")
        st.stop()

    # Importeer de lookup-tabellen uit transform.py
    from execution.transform import MATERIAAL_NL, KLEUR_FILTER, KLEUR_PRESERVE_IN_TITLE, LAMP_EXCEPTIONS

    st.markdown(
        "Vertalingen worden eerst geprobeerd via de **lookup-tabellen** in `transform.py` "
        "(geen Claude). Onbekende termen blijven leeg — je kunt ze handmatig invullen of "
        "expliciet Claude aanroepen via de knop onderaan."
    )

    def vertaal_materiaal(raw):
        if not raw:
            return ("", "leeg")
        lower = raw.lower().strip()
        # Samengesteld materiaal: split op + of &, vertaal elk deel apart
        if "+" in lower or "&" in lower:
            parts = re.split(r"[+&]", lower)
            translated = []
            all_found = True
            for p in parts:
                p = p.strip()
                if p in MATERIAAL_NL:
                    translated.append(MATERIAAL_NL[p])
                else:
                    all_found = False
                    translated.append(p)
            result = " & ".join(translated)
            return (result, "lookup" if all_found else "deels onbekend")
        if lower in MATERIAAL_NL:
            return (MATERIAAL_NL[lower], "lookup")
        return ("", "onbekend")

    def vertaal_kleur(raw):
        if not raw:
            return ("", "", "leeg")
        if "/" in raw:
            return ("", "", "multi — vul handmatig in")
        lower = raw.lower().strip()
        if lower in KLEUR_FILTER:
            kf = KLEUR_FILTER[lower]
            kt = raw.upper() if lower in KLEUR_PRESERVE_IN_TITLE else kf.upper()
            return (kf, kt, "lookup")
        return ("", "", "onbekend")

    rows = []
    for p in batch:
        raw_mat = (p.get("materiaal_nl") or "").strip()  # bevat nog EN-waarde uit masterdata
        raw_kleur = (p.get("kleur_en") or "").strip()
        product_naam_upper = (p.get("product_name_raw") or "").upper()

        # Lamp-uitzondering: kleur staat in productnaam, niet in Color-veld
        is_lamp = any(lamp in product_naam_upper for lamp in LAMP_EXCEPTIONS)

        # Pak vorige edits uit batch_products als die er al zijn (na eerdere vastlegging)
        bestaand_mat_nl = (p.get("_materiaal_nl_translated") or "").strip()
        bestaand_kleur_nl = (p.get("_kleur_nl_translated") or "").strip()
        bestaand_kleur_titel = (p.get("_kleur_titel") or "").strip()

        if bestaand_mat_nl:
            mat_nl, mat_bron = bestaand_mat_nl, "vorige sessie"
        else:
            mat_nl, mat_bron = vertaal_materiaal(raw_mat)

        if bestaand_kleur_nl:
            kleur_nl, kleur_titel, kleur_bron = bestaand_kleur_nl, bestaand_kleur_titel, "vorige sessie"
        else:
            if is_lamp:
                kleur_nl, kleur_titel, kleur_bron = "", "", "lamp — kleur uit productnaam"
            else:
                kleur_nl, kleur_titel, kleur_bron = vertaal_kleur(raw_kleur)

        rows.append({
            "SKU": p.get("sku") or "",
            "🔍": f"https://www.google.com/search?q={(p.get('sku') or '').replace(' ', '+')}",
            "Materiaal (EN)": raw_mat,
            "→": "→",
            "Materiaal (NL)": mat_nl,
            "Mat. bron": mat_bron,
            "Kleur (EN)": raw_kleur,
            "→ ": "→",
            "Kleur (NL filter)": kleur_nl,
            "Kleur (titel)": kleur_titel,
            "Kleur bron": kleur_bron,
            "_id": p["id"],
        })

    df_vert = pd.DataFrame(rows)

    n_mat_ok = sum(1 for r in rows if r["Materiaal (NL)"])
    n_kleur_ok = sum(1 for r in rows if r["Kleur (NL filter)"])

    c1, c2, c3 = st.columns(3)
    c1.metric("📦 Producten", len(rows))
    c2.metric("🟢 Materiaal vertaald", f"{n_mat_ok}/{len(rows)}")
    c3.metric("🟢 Kleur vertaald", f"{n_kleur_ok}/{len(rows)}")

    st.divider()
    st.subheader("Vertalingen — editbaar")
    st.caption("De NL-kolommen zijn editbaar. Lookup-bron wordt getoond zodat je weet of het uit de tabel komt of dat je het zelf hebt ingevuld.")

    edited = st.data_editor(
        df_vert,
        width="stretch",
        height=600,
        hide_index=True,
        disabled=["SKU", "🔍", "Materiaal (EN)", "→", "Mat. bron", "Kleur (EN)", "→ ", "Kleur bron", "_id"],
        column_config={
            "_id": None,
            "🔍": st.column_config.LinkColumn("🔍", display_text="zoek", width="small"),
        },
        key="stap4_editor",
    )

    st.divider()
    st.subheader("Onbekende termen via Claude")
    onbekende_mat = sorted({r["Materiaal (EN)"] for _, r in edited.iterrows()
                            if not r["Materiaal (NL)"] and r["Materiaal (EN)"]})
    onbekende_kleur = sorted({r["Kleur (EN)"] for _, r in edited.iterrows()
                              if not r["Kleur (NL filter)"] and r["Kleur (EN)"]})

    cc1, cc2 = st.columns(2)
    with cc1:
        st.markdown(f"**Onbekende materialen:** {len(onbekende_mat)}")
        if onbekende_mat:
            with st.expander(f"Toon de {len(onbekende_mat)} termen"):
                for t in onbekende_mat:
                    st.code(t)
            if st.button(f"🤖 Vraag Claude voor {len(onbekende_mat)} materialen (1 batch-call)", key="claude_mat"):
                claude = get_claude_client()
                basismaterialen = ("Steengoed, Porselein, Aardewerk, Terracotta, Fiberstone, Zandsteen, Beton, Marmer, Steen, "
                                   "Glas, Borosilicaatglas, Kristal, "
                                   "Metaal, Staal, RVS, Aluminium, Gietijzer, IJzer, Koper, Messing, Zink, "
                                   "Hout, Eikenhout, Notenhout, Acaciahout, "
                                   "Linnen, Katoen, Fluweel, Leer, Zijde, Polyester, "
                                   "Kunststof, Rotan, Bamboe, Papier, Verf, Kaarsvet, Overig")
                prompt = (
                    f"Je krijgt {len(onbekende_mat)} materiaalnamen van producten voor een webshop.\n"
                    f"Kies voor ELK materiaal de dichtstbijzijnde BASISNAAM uit deze vaste lijst:\n"
                    f"{basismaterialen}\n\n"
                    f"REGELS:\n"
                    f"- ALTIJD een van de basismaterialen kiezen\n"
                    f"- Bij samengestelde materialen (bv. 'glass + steel'): geef 'Glas & Staal'\n"
                    f"- Output: ÉÉN materiaalnaam per regel, DEZELFDE volgorde, GEEN uitleg\n\n"
                    f"INPUT:\n" + "\n".join(onbekende_mat) + "\n\nOUTPUT:"
                )
                with st.spinner(f"Claude bevragen voor {len(onbekende_mat)} materialen (1 call)..."):
                    try:
                        resp = claude.messages.create(
                            model="claude-haiku-4-5-20251001",
                            max_tokens=2000,
                            messages=[{"role": "user", "content": prompt}],
                        )
                        output = resp.content[0].text.strip()
                        lines = [l.strip() for l in output.split("\n") if l.strip()]
                        if len(lines) == len(onbekende_mat):
                            st.session_state["_mat_vertalingen"] = dict(zip(onbekende_mat, lines))
                        elif len(lines) > 0:
                            vertalingen_mat = {}
                            for i, mat in enumerate(onbekende_mat):
                                vertalingen_mat[mat] = lines[i] if i < len(lines) else ""
                            st.session_state["_mat_vertalingen"] = vertalingen_mat
                            st.warning(f"Claude gaf {len(lines)} regels voor {len(onbekende_mat)} materialen. "
                                       f"Pas lege velden handmatig aan.")
                        else:
                            st.error("Claude gaf een lege response. Probeer opnieuw.")
                    except Exception as e:
                        st.error(f"Claude-fout: {e}")

            # Editbare preview als vertalingen beschikbaar zijn
            if st.session_state.get("_mat_vertalingen"):
                mat_vertalingen = st.session_state["_mat_vertalingen"]
                st.markdown("**Pas aan waar nodig en klik Opslaan:**")
                mat_preview = pd.DataFrame([
                    {"Materiaal (EN)": en, "Materiaal (NL)": nl} for en, nl in mat_vertalingen.items()
                ])
                mat_edited = st.data_editor(
                    mat_preview, width="stretch", hide_index=True,
                    disabled=["Materiaal (EN)"],
                    key="mat_preview_editor",
                )
                if st.button("✅ Materialen opslaan", key="mat_approve", type="primary"):
                    finale = {row["Materiaal (EN)"]: row["Materiaal (NL)"] for _, row in mat_edited.iterrows()}
                    # Toepassen op batch
                    for idx in edited.index:
                        en_val = edited.at[idx, "Materiaal (EN)"]
                        if not edited.at[idx, "Materiaal (NL)"] and en_val in finale:
                            edited.at[idx, "Materiaal (NL)"] = finale[en_val]
                            edited.at[idx, "Mat. bron"] = "Claude+edit"
                    for _, e in edited.iterrows():
                        pid = e["_id"]
                        for p in batch:
                            if p["id"] == pid:
                                p["_materiaal_nl_translated"] = e["Materiaal (NL)"]
                                break
                    st.session_state.batch_products = batch
                    # Opslaan in MATERIAAL_NL (transform.py)
                    for en_term, nl_term in finale.items():
                        _add_translation_to_transform("materiaal", en_term.lower(), nl_term)
                    st.session_state.pop("_mat_vertalingen", None)
                    st.success(f"✅ {len(finale)} materialen opgeslagen. Volgende keer automatisch herkend.")
                    st.rerun()

    with cc2:
        st.markdown(f"**Onbekende kleuren:** {len(onbekende_kleur)}")
        if onbekende_kleur:
            with st.expander(f"Toon de {len(onbekende_kleur)} termen"):
                for t in onbekende_kleur:
                    st.code(t)
            if st.button(f"🤖 Vraag Claude voor {len(onbekende_kleur)} kleuren (1 batch-call)", key="claude_kleur"):
                claude = get_claude_client()
                basiskleuren = "Wit, Zwart, Grijs, Beige, Bruin, Blauw, Groen, Rood, Roze, Geel, Oranje, Paars, Goud, Zilver, Koper, Transparant, Ivoor, Multi, Terracotta"
                prompt = (
                    f"Je krijgt {len(onbekende_kleur)} kleurnamen van producten voor een webshop.\n"
                    f"Kies voor ELKE kleur de dichtstbijzijnde BASISKLEURNAAM uit deze vaste lijst:\n"
                    f"{basiskleuren}\n\n"
                    f"REGELS:\n"
                    f"- ALTIJD een van de basiskleuren kiezen, NOOIT een eigen vertaling verzinnen\n"
                    f"- 'Imperial Brown' → Bruin (niet 'Imperiaal Bruin')\n"
                    f"- 'Clouded Grey' → Grijs (niet 'Bewolkt Grijs')\n"
                    f"- 'Pine Green' → Groen\n"
                    f"- 'Weathered Black' → Zwart\n"
                    f"- 'Travertine Beige' → Beige\n"
                    f"- Als het een mix is van 2 kleuren → kies de dominante kleur\n"
                    f"- Output: ÉÉN basiskleurnaam per regel, DEZELFDE volgorde, GEEN uitleg\n\n"
                    f"INPUT:\n" + "\n".join(onbekende_kleur) + "\n\nOUTPUT:"
                )
                with st.spinner(f"Claude bevragen voor {len(onbekende_kleur)} kleuren (1 call)..."):
                    try:
                        resp = claude.messages.create(
                            model="claude-haiku-4-5-20251001",
                            max_tokens=2000,
                            messages=[{"role": "user", "content": prompt}],
                        )
                        output = resp.content[0].text.strip()
                        lines = [l.strip() for l in output.split("\n") if l.strip()]
                        if len(lines) == len(onbekende_kleur):
                            vertalingen = dict(zip(onbekende_kleur, lines))
                        elif len(lines) > 0:
                            # Mismatch — pak wat we kunnen en laat de rest leeg
                            vertalingen = {}
                            for i, kleur in enumerate(onbekende_kleur):
                                if i < len(lines):
                                    vertalingen[kleur] = lines[i]
                                else:
                                    vertalingen[kleur] = ""
                            st.warning(f"Claude gaf {len(lines)} regels voor {len(onbekende_kleur)} kleuren. "
                                       f"Pas de lege velden handmatig aan in de preview hieronder.")
                        else:
                            vertalingen = {}
                            st.error("Claude gaf een lege response. Probeer opnieuw.")
                    except Exception as e:
                        st.error(f"Claude-fout: {e}")
                        vertalingen = {}

                if vertalingen:
                    st.session_state["_kleur_vertalingen"] = vertalingen

            # Toon editbare preview als vertalingen beschikbaar zijn
            if st.session_state.get("_kleur_vertalingen"):
                vertalingen = st.session_state["_kleur_vertalingen"]
                st.markdown("**Pas aan waar nodig en klik Opslaan:**")

                basiskleuren_lijst = ["Wit", "Zwart", "Grijs", "Beige", "Bruin", "Blauw",
                                      "Groen", "Rood", "Roze", "Geel", "Oranje", "Paars",
                                      "Goud", "Zilver", "Koper", "Transparant", "Ivoor",
                                      "Multi", "Terracotta"]

                preview_df = pd.DataFrame([
                    {"Kleur (EN)": en, "Basiskleurnaam": nl} for en, nl in vertalingen.items()
                ])
                kleur_edited = st.data_editor(
                    preview_df,
                    width="stretch",
                    hide_index=True,
                    disabled=["Kleur (EN)"],
                    column_config={
                        "Basiskleurnaam": st.column_config.SelectboxColumn(
                            options=basiskleuren_lijst,
                            required=True,
                        ),
                    },
                    key="kleur_preview_editor",
                )

                if st.button("✅ Opslaan", key="kleur_approve", type="primary"):
                    # Lees de (mogelijk aangepaste) waarden
                    finale = {row["Kleur (EN)"]: row["Basiskleurnaam"] for _, row in kleur_edited.iterrows()}

                    # 1. Toepassen op huidige batch
                    for idx in edited.index:
                        en_val = edited.at[idx, "Kleur (EN)"]
                        if not edited.at[idx, "Kleur (NL filter)"] and en_val in finale:
                            nl = finale[en_val]
                            edited.at[idx, "Kleur (NL filter)"] = nl
                            edited.at[idx, "Kleur (titel)"] = nl.upper()
                            edited.at[idx, "Kleur bron"] = "Claude+edit"
                    for _, e in edited.iterrows():
                        pid = e["_id"]
                        for p in batch:
                            if p["id"] == pid:
                                p["_kleur_nl_translated"] = e["Kleur (NL filter)"]
                                p["_kleur_titel"] = e["Kleur (titel)"]
                                break
                    st.session_state.batch_products = batch

                    # 2. Opslaan in KLEUR_FILTER (transform.py)
                    for en_term, nl_term in finale.items():
                        _add_translation_to_transform("kleur", en_term.lower(), nl_term)

                    st.session_state.pop("_kleur_vertalingen", None)
                    st.success(f"✅ {len(finale)} kleuren opgeslagen. Volgende keer automatisch herkend.")
                    st.rerun()

    # ── Afmetingen ─────────────────────────────────────────────────────────────
    st.divider()
    st.subheader("📏 Afmetingen (H × L × B)")

    from execution.transform import clean_decimal

    dim_rows = []
    for p in batch:
        h = p.get("hoogte_cm")
        l = p.get("lengte_cm")
        b = p.get("breedte_cm")
        missend = []
        if h is None: missend.append("H")
        if l is None: missend.append("L")
        if b is None: missend.append("B")

        if not missend:
            status = "🟢 compleet"
        elif len(missend) == 3:
            status = "🔴 alles mist"
        else:
            status = f"🟡 mist: {', '.join(missend)}"

        dim_rows.append({
            "Status": status,
            "SKU": p.get("sku") or "",
            "Productnaam": (p.get("product_name_raw") or "")[:40],
            "Hoogte (cm)": clean_decimal(h) if h is not None else None,
            "Lengte (cm)": clean_decimal(l) if l is not None else None,
            "Breedte (cm)": clean_decimal(b) if b is not None else None,
            "_id": p["id"],
        })

    df_dim = pd.DataFrame(dim_rows)
    n_compleet = sum(1 for r in dim_rows if r["Status"].startswith("🟢"))
    n_missend = sum(1 for r in dim_rows if not r["Status"].startswith("🟢"))

    dc1, dc2, dc3 = st.columns(3)
    dc1.metric("📦 Producten", len(dim_rows))
    dc2.metric("🟢 Afmetingen compleet", n_compleet)
    dc3.metric("⚠️ Data mist", n_missend)

    # Filter
    dim_filter = st.radio(
        "Filter",
        ["Alles", "⚠️ Alleen missende afmetingen", "🟢 Alleen compleet"],
        horizontal=True,
        key="dim_filter",
    )
    if dim_filter == "⚠️ Alleen missende afmetingen":
        dim_view = df_dim[~df_dim["Status"].str.startswith("🟢")]
    elif dim_filter == "🟢 Alleen compleet":
        dim_view = df_dim[df_dim["Status"].str.startswith("🟢")]
    else:
        dim_view = df_dim

    dim_edited = st.data_editor(
        dim_view,
        width="stretch",
        height=400,
        hide_index=True,
        disabled=["Status", "SKU", "Productnaam", "_id"],
        column_config={
            "_id": None,
            "Hoogte (cm)": st.column_config.NumberColumn(format="%.1f", min_value=0),
            "Lengte (cm)": st.column_config.NumberColumn(format="%.1f", min_value=0),
            "Breedte (cm)": st.column_config.NumberColumn(format="%.1f", min_value=0),
        },
        key="dim_editor",
    )

    if st.button("💾 Afmetingen opslaan", key="dim_save"):
        saved_dims = 0
        for _, row in dim_edited.iterrows():
            pid = row["_id"]
            for p in batch:
                if p["id"] == pid:
                    if row.get("Hoogte (cm)") is not None and not pd.isna(row.get("Hoogte (cm)")):
                        p["hoogte_cm"] = float(row["Hoogte (cm)"])
                    if row.get("Lengte (cm)") is not None and not pd.isna(row.get("Lengte (cm)")):
                        p["lengte_cm"] = float(row["Lengte (cm)"])
                    if row.get("Breedte (cm)") is not None and not pd.isna(row.get("Breedte (cm)")):
                        p["breedte_cm"] = float(row["Breedte (cm)"])
                    saved_dims += 1
                    break
        st.session_state.batch_products = batch
        st.success(f"✅ Afmetingen bijgewerkt voor {saved_dims} producten.")

    st.divider()
    if st.button("💾 Stap 4 afronden", type="primary", key="stap4_done_btn"):
        # Sla vertalingen op
        edits_by_id = {row["_id"]: row for _, row in edited.iterrows()}
        for p in batch:
            e = edits_by_id.get(p["id"])
            if e is not None:
                p["_materiaal_nl_translated"] = (e["Materiaal (NL)"] or "").strip()
                p["_kleur_nl_translated"] = (e["Kleur (NL filter)"] or "").strip()
                p["_kleur_titel"] = (e["Kleur (titel)"] or "").strip()
        st.session_state.batch_products = batch
        st.session_state.stap4_done = True
        st.success(f"✅ Vertalingen + afmetingen vastgelegd. Ga naar **stap 5**.")

    # Leerbox stap 4
    learning_chatbox(
        stap_naam="Vertaling materiaal/kleur",
        stap_context=(
            "Stap 4: vertaling van Engelse materiaal- en kleurnamen naar het Nederlands. "
            "Materiaal-lookup staat in MATERIAAL_NL dict in transform.py (bijv. 'stoneware' → 'Steengoed'). "
            "Kleur-lookup staat in KLEUR_FILTER dict (bijv. 'white' → 'Wit'). "
            "Als de user een vertaling corrigeert, voeg die toe aan de juiste dict in transform.py."
        ),
        key="stap4",
    )


elif stap == STAPPEN[5]:
    st.title("Stap 5 — Producttitel + meta description")

    if not st.session_state.selected_for_enrichment:
        st.warning("⚠️ Geen producten geselecteerd. Doe eerst **stap 1+2**.")
        st.stop()

    batch = ensure_batch_loaded()
    if not batch:
        st.error("Kon geen producten ophalen. Ga terug naar stap 2.")
        st.stop()

    from execution.transform import (
        build_title, build_tags, generate_handle, build_meta_description,
        build_page_title, LAMP_EXCEPTIONS, vertaal_productnamen_batch,
        _fix_set_namen,
    )

    # ── Merknaam aanpassen (zonder restart) ───────────────────────────────────
    huidige_merk = st.session_state.get("merk") or "Serax"
    nieuw_merk = st.text_input(
        "Merknaam (wordt gebruikt in producttitel + meta description)",
        value=huidige_merk,
        key="stap5_merk",
    )
    if nieuw_merk != huidige_merk:
        st.session_state.merk = nieuw_merk
        for p in batch:
            p["_merk"] = nieuw_merk
        st.session_state.batch_products = batch
        st.success(f"✅ Merknaam gewijzigd naar **{nieuw_merk}** — titels worden hieronder automatisch bijgewerkt.")

    st.divider()
    st.markdown(
        "**Producttitel** wordt deterministisch opgebouwd via `build_title()` uit "
        "`transform.py`. Default gebruikt 'ie de raw EN productnaam — klik op de knop "
        "hieronder om alle productnamen via Claude (Haiku, 1 batch-call) te vertalen "
        "naar NL. **Meta description** is een aparte stap onderaan."
    )

    # ── Knop: vertaal productnamen via Claude in 1 batch ──────────────────────
    namen_zonder_vertaling = sorted({
        (p.get("product_name_raw") or "").strip()
        for p in batch
        if (p.get("product_name_raw") or "").strip() and not (p.get("_product_name_nl") or "").strip()
    })

    if namen_zonder_vertaling:
        # Cost estimate: input tokens ≈ 200 instr + 10 per naam, output ≈ 10 per naam
        n = len(namen_zonder_vertaling)
        in_tok = 200 + n * 12
        out_tok = n * 12
        # Haiku 4.5: $1/M input, $5/M output
        cost_usd = (in_tok / 1_000_000) * 1.0 + (out_tok / 1_000_000) * 5.0
        cost_eur = cost_usd * 0.92  # ruwe USD→EUR

        st.info(
            f"📝 **{n} unieke productnamen** nog niet vertaald. "
            f"Eén Claude-call (Haiku 4.5) kost ongeveer **€{cost_eur:.4f}** "
            f"(~{in_tok} input + {out_tok} output tokens)."
        )

        if st.button(f"🤖 Vertaal {n} productnamen via Claude (1 batch-call)",
                     type="primary", key="batch_translate_names"):
            claude = get_claude_client()
            try:
                with st.spinner(f"Claude bevragen voor {n} productnamen..."):
                    mapping = vertaal_productnamen_batch(namen_zonder_vertaling, claude)
            except Exception as e:
                st.error(f"Vertaling mislukt: {e}")
                mapping = {}

            # Schrijf vertalingen terug in batch_products
            count = 0
            for p in batch:
                raw = (p.get("product_name_raw") or "").strip()
                if raw in mapping and mapping[raw]:
                    p["_product_name_nl"] = mapping[raw]
                    count += 1
            st.session_state.batch_products = batch
            st.success(f"✅ {len(mapping)} unieke vertalingen toegepast op {count} producten. Pagina ververst.")
            st.rerun()
    else:
        n_already = sum(1 for p in batch if (p.get("_product_name_nl") or "").strip())
        if n_already > 0:
            st.success(f"✅ Alle {n_already} productnamen zijn al vertaald (uit eerdere batch-call).")

    st.divider()

    # Bouw titel + tags + handle voor elk product (deterministisch, no Claude)
    rows = []
    for p in batch:
        # Voeg de stap-4 vertalingen toe als velden voor build_title
        p_voor_titel = dict(p)
        p_voor_titel["materiaal_nl"] = p.get("_materiaal_nl_translated") or p.get("materiaal_nl") or ""
        p_voor_titel["kleur_nl"] = p.get("_kleur_nl_translated") or ""
        p_voor_titel["_kleur_titel"] = p.get("_kleur_titel") or ""
        p_voor_titel["_product_name_nl"] = p.get("_product_name_nl") or ""

        try:
            titel_auto = build_title(p_voor_titel)
        except Exception as e:
            titel_auto = f"FOUT: {e}"

        try:
            tags_auto = build_tags(
                p.get("hoofdcategorie") or "",
                p.get("subcategorie") or "",
                p.get("sub_subcategorie") or "",
                batch_tag=p.get("_batch_tag") or st.session_state.get("batch_tag") or "",
                extra_tags=p.get("_extra_tags"),
            )
        except Exception:
            tags_auto = ""

        try:
            handle_auto = generate_handle(titel_auto)
        except Exception:
            handle_auto = ""

        # Meta description: automatisch opgebouwd uit beschikbare data
        try:
            meta_auto = build_meta_description(p_voor_titel)
        except Exception:
            meta_auto = ""

        # Page title: max 70 tekens, "naam | Interieur Shop"
        try:
            page_title_auto = build_page_title(p_voor_titel)
        except Exception:
            page_title_auto = ""

        # Pak vorige edits als die er zijn
        titel_huidig = p.get("_product_title_nl") or titel_auto
        tags_huidig = p.get("_tags") or tags_auto
        handle_huidig = p.get("_handle") or handle_auto
        meta_huidig = p.get("_meta_description") or meta_auto
        page_title_huidig = p.get("_page_title") or page_title_auto

        is_lamp = any(lamp in (p.get("product_name_raw") or "").upper() for lamp in LAMP_EXCEPTIONS)

        rows.append({
            "SKU": p.get("sku") or "",
            "🔍": f"https://www.google.com/search?q={(p.get('sku') or '').replace(' ', '+')}",
            "Designer": p.get("designer") or "",
            "Productnaam (raw)": (p.get("product_name_raw") or "")[:50],
            "Productnaam (NL)": _fix_set_namen(p.get("_product_name_nl") or ""),
            "→": "→",
            "Producttitel NL": titel_huidig,
            "Page title (SEO)": page_title_huidig,
            "Tags": tags_huidig,
            "Handle": handle_huidig,
            "Meta description": meta_huidig,
            "Meta lengte": len(meta_huidig),
            "Lamp?": "🚨" if is_lamp else "",
            "_id": p["id"],
        })

    df_titel = pd.DataFrame(rows)

    n_titel = sum(1 for r in rows if r["Producttitel NL"] and not r["Producttitel NL"].startswith("FOUT"))
    n_meta = sum(1 for r in rows if r["Meta description"])
    n_meta_te_lang = sum(1 for r in rows if r["Meta lengte"] > 160)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("📦 Producten", len(rows))
    c2.metric("🟢 Titel OK", n_titel)
    c3.metric("📝 Meta gegenereerd", n_meta)
    c4.metric("⚠️ Meta > 160 tekens", n_meta_te_lang)

    st.divider()
    st.subheader("Titels & meta — editbaar")
    st.caption("Titels worden automatisch opgebouwd. Pas aan als nodig. Meta description vereist een Claude-call (zie knop hieronder).")

    edited = st.data_editor(
        df_titel,
        width="stretch",
        height=600,
        hide_index=True,
        disabled=["SKU", "🔍", "Designer", "Productnaam (raw)", "→", "Meta lengte", "Lamp?", "_id"],
        column_config={
            "_id": None,
            "🔍": st.column_config.LinkColumn("🔍", display_text="zoek", width="small"),
            "Productnaam (raw)": st.column_config.TextColumn(width="medium"),
            "Productnaam (NL)": st.column_config.TextColumn(width="medium"),
            "Producttitel NL": st.column_config.TextColumn(width="large"),
            "Tags": st.column_config.TextColumn(width="large"),
            "Meta description": st.column_config.TextColumn(width="large", max_chars=200),
        },
        key="stap5_editor",
    )

    st.divider()
    if st.button("💾 Vastleggen in batch", type="primary", key="stap5_save"):
        edits_by_id = {row["_id"]: row for _, row in edited.iterrows()}
        for p in batch:
            e = edits_by_id.get(p["id"])
            if e is not None:
                p["_product_name_nl"] = (e["Productnaam (NL)"] or "").strip()
                p["_product_title_nl"] = (e["Producttitel NL"] or "").strip()
                p["_page_title"] = (e.get("Page title (SEO)") or "").strip()
                p["_handle"] = (e["Handle"] or "").strip()
                p["_tags"] = (e["Tags"] or "").strip()
                p["_meta_description"] = (e["Meta description"] or "").strip()
        st.session_state.batch_products = batch
        st.session_state.stap5_done = True
        st.success(f"✅ Vastgelegd. Ga naar **stap 6**.")

    # Leerbox stap 5
    learning_chatbox(
        stap_naam="Producttitel & meta",
        stap_context=(
            "Stap 5: producttitels en meta descriptions. Titels worden opgebouwd als: "
            "Serax - {Designer} - {PRODUCTNAAM NL HOOFDLETTERS} - Set van {N}. "
            "Productnamen worden vertaald van EN naar NL via een Claude batch-call. "
            "Als de user een productnaam-vertaling corrigeert of een titel-regel toevoegt, "
            "sla dit op zodat het de volgende keer automatisch goed gaat."
        ),
        key="stap5",
    )


elif stap == STAPPEN[6]:
    st.title("Stap 6 — Eindreview")

    if not st.session_state.selected_for_enrichment:
        st.warning("⚠️ Geen producten geselecteerd. Doe eerst **stap 1+2**.")
        st.stop()

    batch = ensure_batch_loaded()
    if not batch:
        st.error("Kon geen producten ophalen. Ga terug naar stap 2.")
        st.stop()

    # Sync ontbrekende categorieën uit database (bijv. na handmatige DB-update)
    ids_zonder_cat = [p["id"] for p in batch if not p.get("hoofdcategorie")]
    if ids_zonder_cat:
        sb = get_supabase()
        for i in range(0, len(ids_zonder_cat), 50):
            chunk = ids_zonder_cat[i:i+50]
            result = sb.table("seo_products").select(
                "id, hoofdcategorie, subcategorie, sub_subcategorie, collectie"
            ).in_("id", chunk).execute()
            db_map = {r["id"]: r for r in result.data}
            for p in batch:
                if p["id"] in db_map:
                    db = db_map[p["id"]]
                    if db.get("hoofdcategorie"):
                        p["hoofdcategorie"] = db["hoofdcategorie"]
                        p["subcategorie"] = db.get("subcategorie") or ""
                        p["sub_subcategorie"] = db.get("sub_subcategorie") or ""
                        p["collectie"] = db.get("collectie") or ""
        st.session_state.batch_products = batch

    from execution.transform import clean_decimal, build_tags, build_title

    def build_tags_safe(p):
        try:
            return build_tags(
                p.get("hoofdcategorie") or "",
                p.get("subcategorie") or "",
                p.get("sub_subcategorie") or "",
                batch_tag=p.get("_batch_tag") or st.session_state.get("batch_tag") or "",
                extra_tags=p.get("_extra_tags"),
            )
        except Exception:
            return ""

    def build_title_safe(p):
        """Bouwt titel on-the-fly als _product_title_nl leeg is — geen save in stap 5 nodig."""
        if (p.get("_product_title_nl") or "").strip():
            return p["_product_title_nl"].strip()
        try:
            p_voor_titel = dict(p)
            p_voor_titel["materiaal_nl"] = p.get("_materiaal_nl_translated") or p.get("materiaal_nl") or ""
            p_voor_titel["kleur_nl"] = p.get("_kleur_nl_translated") or ""
            p_voor_titel["_kleur_titel"] = p.get("_kleur_titel") or ""
            p_voor_titel["_product_name_nl"] = p.get("_product_name_nl") or ""
            return build_title(p_voor_titel)
        except Exception:
            return ""

    st.markdown(
        "Eindreview vóór commit naar Supabase. Het systeem checkt elk product op blokkers "
        "en waarschuwingen. Pas eventueel terug in stap 3-5 aan en kom hier weer naartoe."
    )

    # Bouw review-rijen met issues-detectie
    rows = []
    for p in batch:
        sku = p.get("sku") or ""
        titel = build_title_safe(p)  # bouw on-the-fly als _product_title_nl leeg is
        meta = p.get("_meta_description") or ""
        mat = p.get("_materiaal_nl_translated") or ""
        kleur = p.get("_kleur_nl_translated") or ""
        hoofdcat = p.get("hoofdcategorie") or ""
        subcat = p.get("subcategorie") or ""
        subsubcat = p.get("sub_subcategorie") or ""
        ean = p.get("ean_shopify") or ""
        prijs_raw = p.get("rrp_stuk_eur")
        inkoop_raw = p.get("inkoopprijs_stuk_eur")
        hoogte = p.get("hoogte_cm")
        lengte = p.get("lengte_cm")
        breedte = p.get("breedte_cm")

        issues = []
        if not titel or titel.startswith("FOUT") or titel == "Serax":
            issues.append("titel ontbreekt")
        if not meta:
            issues.append("meta ontbreekt")
        if len(meta) > 160:
            issues.append(f"meta {len(meta)}>160")
        if not hoofdcat or not subcat or not subsubcat:
            issues.append("categorie incompleet")
        if not ean:
            issues.append("EAN leeg")
        # Prijs: alleen blokker bij echte 0 of negatief — None betekent "nog in te vullen"
        # en wordt als waarschuwing behandeld, niet als blokker
        if prijs_raw is None or prijs_raw == "":
            issues.append("prijs ontbreekt")
        else:
            try:
                if float(prijs_raw) < 0:
                    issues.append("prijs negatief")
                elif float(prijs_raw) == 0:
                    issues.append("prijs ≤0")
            except (ValueError, TypeError):
                issues.append("prijs ongeldig")

        # Blokker = echte data-corruptie. Ontbrekende prijs/EAN = waarschuwing
        if not issues:
            status = "🟢 OK"
        elif any(blok in " ".join(issues) for blok in ["prijs ≤0", "prijs negatief", "prijs ongeldig", "titel ontbreekt"]):
            status = "🔴 blokker"
        else:
            status = "🟡 waarschuwing"

        rows.append({
            "Status": status,
            "SKU": sku,
            "Producttitel NL": titel[:60],
            "Hoofd > Sub > Subsub": f"{hoofdcat} > {subcat} > {subsubcat}",
            "Sub-subcat 2": (p.get("_extra_tags") or [""])[0] if p.get("_extra_tags") else "",
            "Materiaal": mat,
            "Kleur": kleur,
            "Verkoop €": clean_decimal(prijs_raw),
            "Inkoop €":  clean_decimal(inkoop_raw),
            "H (cm)":    clean_decimal(hoogte),
            "L (cm)":    clean_decimal(lengte),
            "B (cm)":    clean_decimal(breedte),
            "EAN": ean,
            "Meta lengte": len(meta),
            "Issues": ", ".join(issues) if issues else "",
            "_id": p["id"],
        })

    df_review = pd.DataFrame(rows)

    n_ok = sum(1 for r in rows if r["Status"].startswith("🟢"))
    n_warn = sum(1 for r in rows if r["Status"].startswith("🟡"))
    n_blok = sum(1 for r in rows if r["Status"].startswith("🔴"))

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("📦 Totaal", len(rows))
    c2.metric("🟢 Klaar", n_ok)
    c3.metric("🟡 Waarschuwing", n_warn)
    c4.metric("🔴 Blokker", n_blok)

    st.divider()

    # ── Twee tabs: issues-check en Shopify-preview ────────────────────────────
    tab_issues, tab_shopify = st.tabs(["🔍 Issues check", "📋 Shopify preview (33 kolommen)"])

    with tab_issues:
        st.caption("Status van elk product op verplichte velden — fix blokkers in stap 3-5 voor je commit.")
        filter_keuze = st.radio(
            "Filter",
            ["Alles", "🟢 Alleen OK", "🟡 Waarschuwingen", "🔴 Blokkers"],
            horizontal=True,
            key="stap6_filter",
        )
        if filter_keuze == "🟢 Alleen OK":
            view_df = df_review[df_review["Status"].str.startswith("🟢")]
        elif filter_keuze == "🟡 Waarschuwingen":
            view_df = df_review[df_review["Status"].str.startswith("🟡")]
        elif filter_keuze == "🔴 Blokkers":
            view_df = df_review[df_review["Status"].str.startswith("🔴")]
        else:
            view_df = df_review

        styled = view_df.style.map(kleur_status, subset=["Status"])
        st.dataframe(styled, width="stretch", height=600, hide_index=True,
                     column_config={"_id": None})
        st.caption(f"Toont {len(view_df)} van {len(df_review)} rijen")

    with tab_shopify:
        st.caption(
            "Exact dezelfde 33 kolommen zoals ze in de Shopify-export-xlsx komen "
            "(`Shopify_Nieuwe_Producten_Fase3_EAN_gecorrigeerd.xlsx`-formaat). "
            "Hier zie je afmetingen, photos en de Shopify-specifieke velden naast elkaar."
        )

        shopify_rows = []
        for p in batch:
            tags = p.get("_tags") or build_tags_safe(p)
            shopify_rows.append({
                "Variant SKU":           p.get("sku") or "",
                "Product ID":            p.get("shopify_product_id") or "",
                "Variant ID":            p.get("shopify_variant_id") or "",
                "Product handle":        p.get("_handle") or "",
                "Product title":         p.get("_product_title_nl") or "",
                "Product vendor":        p.get("_merk") or st.session_state.get("merk") or "Serax",
                "Product type":          p.get("leverancier_category") or "",
                "EAN Code ":             p.get("ean_shopify") or "",
                "Verkoopprijs Shopify":  clean_decimal(p.get("rrp_stuk_eur")) or "",
                "Inkoopprijs Shopify":   clean_decimal(p.get("inkoopprijs_stuk_eur")) or "",
                "Product description":   p.get("product_description") or "",
                "Nieuwe hoofdcategorie": p.get("hoofdcategorie") or "",
                "Nieuwe subcategorie":   p.get("subcategorie") or "",
                "Nieuwe sub-subcategorie": p.get("sub_subcategorie") or "",
                "Nieuwe sub-subcategorie 2": (p.get("_extra_tags") or [""])[0] if p.get("_extra_tags") else "",
                "Nieuwe tag":            tags,
                "collectie":             p.get("collectie") or "",
                "designer":              p.get("designer") or "",
                "materiaal":             p.get("_materiaal_nl_translated") or "",
                "kleur":                 p.get("_kleur_nl_translated") or "",
                "hoogte_cm":             clean_decimal(p.get("hoogte_cm")) or "",
                "lengte_cm":             clean_decimal(p.get("lengte_cm")) or "",
                "breedte_cm":            clean_decimal(p.get("breedte_cm")) or "",
                "meta_description":      p.get("_meta_description") or "",
                "photo_packshot1":       p.get("photo_packshot_1") or "",
                "photo_packshot2":       p.get("photo_packshot_2") or "",
                "photo_packshot3":       p.get("photo_packshot_3") or "",
                "photo_packshot4":       p.get("photo_packshot_4") or "",
                "photo_packshot5":       p.get("photo_packshot_5") or "",
                "photo_lifestyle1":      p.get("photo_lifestyle_1") or "",
                "photo_lifestyle2":      p.get("photo_lifestyle_2") or "",
                "photo_lifestyle3":      p.get("photo_lifestyle_3") or "",
                "photo_lifestyle4":      p.get("photo_lifestyle_4") or "",
                "photo_lifestyle5":      p.get("photo_lifestyle_5") or "",
            })

        df_shopify = pd.DataFrame(shopify_rows)

        # Markeer lege velden visueel
        n_lege_titels    = sum(1 for r in shopify_rows if not r["Product title"])
        n_lege_afmeting  = sum(1 for r in shopify_rows if not (r["hoogte_cm"] and r["lengte_cm"] and r["breedte_cm"]))
        n_lege_photo     = sum(1 for r in shopify_rows if not r["photo_packshot1"])
        n_lege_meta      = sum(1 for r in shopify_rows if not r["meta_description"])

        cc1, cc2, cc3, cc4 = st.columns(4)
        cc1.metric("⚠️ Lege titels", n_lege_titels)
        cc2.metric("⚠️ Lege afmetingen", n_lege_afmeting)
        cc3.metric("⚠️ Geen foto", n_lege_photo)
        cc4.metric("⚠️ Lege meta", n_lege_meta)

        st.dataframe(df_shopify, width="stretch", height=600, hide_index=True)
        st.caption(f"{len(df_shopify)} producten · {len(df_shopify.columns)} kolommen")

        # Download knop
        csv = df_shopify.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            "📥 Download Shopify preview (CSV)",
            csv,
            file_name="shopify_preview.csv",
            mime="text/csv",
            key="download_shopify_preview",
        )

    st.divider()
    st.subheader("Vastleggen in Supabase")

    if n_blok > 0:
        st.error(f"⛔ Er zijn nog {n_blok} **blokkers**. Los die eerst op (stap 3-5) voor je commit.")

    bevestig_commit = st.checkbox(
        f"Ja, ik wil {n_ok + n_warn} producten vastleggen in Supabase met `status='ready'` "
        f"(blokkers worden opgeslagen met `status='review'`)",
        key="commit_confirm",
    )

    if st.button("💾 Vastleggen in Supabase", type="primary", disabled=not bevestig_commit):
        sb = get_supabase()
        rows_by_id = {r["_id"]: r for r in rows}

        commit_count = 0
        review_count = 0
        errors = []

        progress = st.progress(0.0, text="Vastleggen...")
        for i, p in enumerate(batch):
            review_row = rows_by_id.get(p["id"])
            if not review_row:
                continue

            new_status = "ready" if review_row["Status"].startswith("🟢") else (
                "review" if review_row["Status"].startswith("🔴") else "ready"
            )
            review_reden = review_row["Issues"] if review_row["Issues"] else None

            update = {
                "hoofdcategorie": p.get("hoofdcategorie") or "",
                "subcategorie": p.get("subcategorie") or "",
                "sub_subcategorie": p.get("sub_subcategorie") or "",
                "collectie": p.get("collectie") or "",
                "tags": p.get("_tags") or "",
                "materiaal_nl": p.get("_materiaal_nl_translated") or "",
                "kleur_nl": p.get("_kleur_nl_translated") or "",
                "product_title_nl": p.get("_product_title_nl") or "",
                "handle": p.get("_handle") or "",
                "meta_description": p.get("_meta_description") or "",
                "verkoopprijs": p.get("rrp_stuk_eur"),
                "inkoopprijs": p.get("inkoopprijs_stuk_eur"),
                "status": new_status,
                "review_reden": review_reden,
            }

            try:
                sb.table("seo_products").update(update).eq("id", p["id"]).execute()
                if new_status == "ready":
                    commit_count += 1
                else:
                    review_count += 1
            except Exception as e:
                errors.append(f"{p.get('sku')}: {e}")

            progress.progress((i + 1) / len(batch), text=f"{i + 1}/{len(batch)}")

        progress.empty()
        st.session_state.stap6_committed = True
        st.success(f"✅ {commit_count} producten op `status=ready`, {review_count} op `status=review`. Ga naar **stap 7** voor de export.")
        if errors:
            with st.expander(f"⚠️ {len(errors)} fouten"):
                for e in errors:
                    st.code(e)


elif stap == STAPPEN[7]:
    st.title("Stap 7 — Export naar Shopify")

    if not st.session_state.selected_for_enrichment:
        st.warning("⚠️ Geen producten geselecteerd. Doe eerst **stap 1+2**.")
        st.stop()

    if not st.session_state.stap6_committed:
        st.warning(
            "⚠️ Stap 6 (vastleggen in Supabase) is nog niet gebeurd. "
            "Je kunt wel exporteren vanuit de in-memory batch, maar dan staat de "
            "data niet permanent in Supabase. Aanbevolen: doe eerst stap 6."
        )

    batch = ensure_batch_loaded()
    if not batch:
        st.error("Kon geen producten ophalen. Ga terug naar stap 2.")
        st.stop()

    st.markdown(
        "Genereert het Shopify Hextom-bestand met drie tabs (`Shopify_Nieuw`, "
        "`Shopify_Archief`, `Analyse`) volgens het template-format uit "
        "`execution/export_standaard.py`."
    )

    # Bouw producten-lijst in het format dat schrijf_product_tab verwacht
    # We kopiëren de stap-3/4/5 enrichments naar de top-level kolommen
    producten_voor_export = []
    for p in batch:
        copy = dict(p)
        copy["materiaal_nl"] = p.get("_materiaal_nl_translated") or ""
        copy["kleur_nl"] = p.get("_kleur_nl_translated") or ""
        copy["product_title_nl"] = p.get("_product_title_nl") or ""
        copy["handle"] = p.get("_handle") or ""
        copy["tags"] = p.get("_tags") or ""
        copy["meta_description"] = p.get("_meta_description") or ""
        copy["verkoopprijs"] = p.get("rrp_stuk_eur")
        copy["inkoopprijs"] = p.get("inkoopprijs_stuk_eur")
        producten_voor_export.append(copy)

    # Statistieken
    n_nieuw = sum(1 for p in producten_voor_export if (p.get("status_shopify") or "nieuw") == "nieuw")
    n_archief = sum(1 for p in producten_voor_export if p.get("status_shopify") == "archief")
    n_actief = sum(1 for p in producten_voor_export if p.get("status_shopify") == "actief")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("📦 Totaal", len(producten_voor_export))
    c2.metric("🆕 Nieuw", n_nieuw)
    c3.metric("📁 Archief", n_archief)
    c4.metric("✅ Actief", n_actief)

    st.divider()
    st.subheader("Preview eerste 5 rijen")
    preview_df = pd.DataFrame([
        {
            "SKU": p.get("sku") or "",
            "Producttitel": (p.get("product_title_nl") or "")[:50],
            "Categorie": f"{p.get('hoofdcategorie','')} > {p.get('sub_subcategorie','')}",
            "Materiaal": p.get("materiaal_nl") or "",
            "Kleur": p.get("kleur_nl") or "",
            "Prijs": p.get("verkoopprijs") or "",
            "EAN": p.get("ean_shopify") or "",
            "Meta (50)": (p.get("meta_description") or "")[:50],
        }
        for p in producten_voor_export[:5]
    ])
    st.dataframe(preview_df, width="stretch", hide_index=True)

    st.divider()
    if st.button("📤 Genereer Shopify-export bestand", type="primary"):
        from datetime import datetime
        from pathlib import Path as _Path
        import openpyxl as _xl

        # Hergebruik de schrijf_product_tab functie
        from execution.export_standaard import schrijf_product_tab, schrijf_analyse_tab

        try:
            out_dir = _Path("./exports/")
            out_dir.mkdir(parents=True, exist_ok=True)

            ts = datetime.now().strftime("%Y%m%d_%H%M")
            naam = f"Serax_Batch_{ts}.xlsx"
            pad = out_dir / naam

            # Splitsen op Shopify-status
            nieuw = [p for p in producten_voor_export if (p.get("status_shopify") or "nieuw") == "nieuw"]
            archief_actief = [p for p in producten_voor_export
                              if p.get("status_shopify") in ("archief", "actief")]

            wb = _xl.Workbook()
            wb.remove(wb.active)

            ws_n = wb.create_sheet("Shopify_Nieuw")
            schrijf_product_tab(ws_n, nieuw)

            ws_a = wb.create_sheet("Shopify_Archief")
            schrijf_product_tab(ws_a, archief_actief)

            # Analyse-tab heeft Supabase nodig voor filterwaarden — geef sb door
            ws_an = wb.create_sheet("Analyse")
            try:
                sb = get_supabase()
                schrijf_analyse_tab(ws_an, "batch", producten_voor_export, sb)
            except Exception as e:
                ws_an["A1"] = f"Analyse-tab kon niet worden gegenereerd: {e}"

            wb.save(pad)
            st.session_state.exported_path = str(pad)

            st.success(f"✅ Bestand gegenereerd: `{pad}` ({pad.stat().st_size / 1024:.1f} KB)")
        except Exception as e:
            st.error(f"❌ Export mislukt: {e}")
            import traceback
            with st.expander("Stack trace"):
                st.code(traceback.format_exc())

    if st.session_state.exported_path:
        from pathlib import Path as _Path
        pad = _Path(st.session_state.exported_path)
        if pad.exists():
            with open(pad, "rb") as f:
                st.download_button(
                    "📥 Download Shopify Excel",
                    f.read(),
                    file_name=pad.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            st.divider()
            st.markdown(
                "**Pre-import checklist:**\n"
                "1. Open de Excel — check tab **Analyse** voor nieuwe filterwaarden\n"
                "2. Aanmaken in Shopify VOOR je importeert\n"
                "3. Tab **Shopify_Nieuw**: kolom 8 (EAN Code) moet tekst zijn (`5400123456789` niet `5.40E+12`)\n"
                "4. Importeer **Shopify_Nieuw** eerst, daarna **Shopify_Archief**"
            )


# ═════════════════════════════════════════════════════════════════════════════
# OVERZICHT & VRAGEN
# ═════════════════════════════════════════════════════════════════════════════

elif stap == STAPPEN[8]:
    st.title("📊 Overzicht & vragen")
    st.markdown("Stel vragen over de database, bekijk overzichten, en download exports — zonder het dashboard te verlaten.")

    sb = get_supabase()

    # ── Database stats ────────────────────────────────────────────────────────
    st.subheader("Database overzicht")

    all_prods = []
    offset = 0
    while True:
        res = sb.table("seo_products").select(
            "sku, status, hoofdcategorie, subcategorie, sub_subcategorie, "
            "rrp_stuk_eur, hoogte_cm, photo_packshot_1, leverancier_category"
        ).range(offset, offset + 999).execute()
        if not res.data:
            break
        all_prods.extend(res.data)
        if len(res.data) < 1000:
            break
        offset += 1000

    from collections import Counter
    statussen = Counter(p.get("status") or "?" for p in all_prods)
    mc1, mc2, mc3, mc4 = st.columns(4)
    mc1.metric("Totaal producten", len(all_prods))
    mc2.metric("Ready", statussen.get("ready", 0))
    mc3.metric("Raw", statussen.get("raw", 0))
    mc4.metric("Exported", statussen.get("exported", 0))

    st.divider()

    # ── Categorie-overzicht ───────────────────────────────────────────────────
    st.subheader("Categorieën in gebruik")

    actieve_subsubcats = load_active_subsubcategories()

    cat_combos = Counter()
    for p in all_prods:
        if p.get("status") == "ready":
            h = p.get("hoofdcategorie") or ""
            s = p.get("subcategorie") or ""
            ss = p.get("sub_subcategorie") or ""
            if h and s and ss:
                cat_combos[(h, s, ss)] += 1

    if cat_combos:
        cat_rows = []
        for (h, s, ss), n in sorted(cat_combos.items()):
            online = "🟢 online" if ss.strip().lower() in actieve_subsubcats else "🔴 nog niet"
            cat_rows.append({
                "Hoofdcategorie": h,
                "Subcategorie": s,
                "Sub-subcategorie": ss,
                "Aantal": n,
                "Online?": online,
            })

        df_cats = pd.DataFrame(cat_rows)

        online_n = sum(1 for r in cat_rows if r["Online?"].startswith("🟢"))
        offline_n = len(cat_rows) - online_n

        cc1, cc2, cc3 = st.columns(3)
        cc1.metric("Unieke sub-subcats", len(cat_rows))
        cc2.metric("🟢 Al online", online_n)
        cc3.metric("🔴 Nog toevoegen", offline_n)

        cat_filter = st.radio(
            "Filter", ["Alles", "🔴 Nog toevoegen aan website", "🟢 Al online"],
            horizontal=True, key="ov_cat_filter",
        )
        if cat_filter == "🔴 Nog toevoegen aan website":
            df_cats = df_cats[df_cats["Online?"].str.startswith("🔴")]
        elif cat_filter == "🟢 Al online":
            df_cats = df_cats[df_cats["Online?"].str.startswith("🟢")]

        st.dataframe(df_cats, width="stretch", hide_index=True, height=400)

        csv = df_cats.to_csv(index=False).encode("utf-8-sig")
        st.download_button("📥 Download categorie-overzicht (CSV)", csv,
                           file_name="categorie_overzicht.csv", mime="text/csv")

    st.divider()

    # ── Data-kwaliteit ────────────────────────────────────────────────────────
    st.subheader("Data-kwaliteit")

    geen_prijs = sum(1 for p in all_prods if not p.get("rrp_stuk_eur"))
    geen_afm = sum(1 for p in all_prods if not p.get("hoogte_cm"))
    geen_foto = sum(1 for p in all_prods if not p.get("photo_packshot_1"))
    geen_cat = sum(1 for p in all_prods if not p.get("hoofdcategorie"))

    dq1, dq2, dq3, dq4 = st.columns(4)
    dq1.metric("Zonder prijs", geen_prijs)
    dq2.metric("Zonder afmetingen", geen_afm)
    dq3.metric("Zonder foto", geen_foto)
    dq4.metric("Zonder categorie", geen_cat)

    st.divider()

    # ── Chatbox voor vrije vragen ─────────────────────────────────────────────
    st.subheader("💬 Stel een vraag")
    st.caption(
        "Stel een vraag over de database in gewone taal. Bijv.:\n"
        "- *\"Welke S&P producten hebben geen foto?\"*\n"
        "- *\"Hoeveel Serax producten zijn er per categorie?\"*\n"
        "- *\"Geef me een export van alle producten zonder prijs\"*"
    )

    vraag = st.text_area("Jouw vraag", placeholder="Typ hier je vraag...", key="ov_vraag", height=80)

    if st.button("🔍 Beantwoord", type="primary", key="ov_beantwoord"):
        if not vraag or not vraag.strip():
            st.warning("Typ eerst een vraag.")
        else:
            claude = get_claude_client()
            prompt = (
                f"Je bent een data-assistent voor een Belgische interieur-webshop.\n"
                f"De database bevat {len(all_prods)} producten van leveranciers: Serax, Pottery Pots, Printworks, S&P/Bonbistro.\n"
                f"Statussen: {dict(statussen)}\n"
                f"Zonder prijs: {geen_prijs}, zonder afmetingen: {geen_afm}, zonder foto: {geen_foto}, zonder categorie: {geen_cat}\n\n"
                f"De gebruiker stelt een vraag. Beantwoord kort en bondig in het Nederlands.\n"
                f"Als de vraag een SQL-query vereist, schrijf de query zodat de gebruiker die kan draaien.\n"
                f"Als de vraag een export vereist, beschrijf welke data nodig is.\n\n"
                f"VRAAG: {vraag}"
            )
            with st.spinner("Claude denkt na..."):
                try:
                    resp = claude.messages.create(
                        model="claude-haiku-4-5-20251001",
                        max_tokens=1000,
                        messages=[{"role": "user", "content": prompt}],
                    )
                    antwoord = resp.content[0].text.strip()
                    st.markdown(antwoord)
                except Exception as e:
                    st.error(f"Fout: {e}")


# ═════════════════════════════════════════════════════════════════════════════
# WEBSITE STRUCTUUR
# ═════════════════════════════════════════════════════════════════════════════

elif stap == STAPPEN[9]:
    st.title("🌐 Website structuur")
    st.markdown("De volledige categorieboom van de website, met online-status en aantal producten per sub-subcategorie.")

    tree = load_website_tree()
    actieve_subsubcats = load_active_subsubcategories()

    sb = get_supabase()

    # Tel producten per sub-subcat (alleen ready)
    all_prods = []
    offset = 0
    while True:
        res = sb.table("seo_products").select("sub_subcategorie, status").eq("status", "ready").range(offset, offset + 999).execute()
        if not res.data:
            break
        all_prods.extend(res.data)
        if len(res.data) < 1000:
            break
        offset += 1000

    from collections import Counter
    prod_per_subsubcat = Counter(
        (p.get("sub_subcategorie") or "").strip() for p in all_prods if p.get("sub_subcategorie")
    )

    if tree:
        # Bouw de volledige boom als tabel
        boom_rows = []
        for hc in sorted(tree.keys()):
            for sc in sorted(tree[hc].keys()):
                for ssc in sorted(tree[hc][sc]):
                    is_online = ssc.strip().lower() in actieve_subsubcats
                    n_producten = prod_per_subsubcat.get(ssc, 0)
                    # Check of deze sub-subcat in gebruik is maar niet in de boom staat
                    from execution.transform import slug_for_tag
                    # Laad intros
                    intros_path = Path("config/categorie_intros.json")
                    if intros_path.exists() and "_cat_intros" not in st.session_state:
                        import json as _json
                        with open(intros_path, encoding="utf-8") as _f:
                            st.session_state._cat_intros = _json.load(_f)
                    cat_intros = st.session_state.get("_cat_intros", {})

                    intro = cat_intros.get(f"{hc} > {sc} > {ssc}", "")
                    boom_rows.append({
                        "Hoofdcategorie": hc,
                        "Subcategorie": sc,
                        "Sub-subcategorie": ssc,
                        "Tag": f"cat_{slug_for_tag(ssc)}",
                        "Online?": "🟢 online" if is_online else "🔴 nog niet",
                        "Producten (ready)": n_producten,
                        "Intro": intro[:80] + "..." if len(intro) > 80 else intro,
                    })

        # Voeg sub-subcats toe die WEL in gebruik zijn maar NIET in de website-boom
        from execution.transform import slug_for_tag as _sft
        cat_intros = st.session_state.get("_cat_intros", {})
        boom_subsubcats = {r["Sub-subcategorie"] for r in boom_rows}
        for ssc, n in prod_per_subsubcat.items():
            if ssc and ssc not in boom_subsubcats:
                boom_rows.append({
                    "Hoofdcategorie": "⚠️ NIET IN BOOM",
                    "Subcategorie": "",
                    "Sub-subcategorie": ssc,
                    "Tag": f"cat_{_sft(ssc)}",
                    "Online?": "🟢 online" if ssc.strip().lower() in actieve_subsubcats else "🔴 nog niet",
                    "Producten (ready)": n,
                    "Intro": "",
                })

        df_boom = pd.DataFrame(boom_rows)

        # Stats
        totaal_subcats = len(df_boom)
        online = sum(1 for r in boom_rows if r["Online?"].startswith("🟢"))
        offline = totaal_subcats - online
        in_gebruik = sum(1 for r in boom_rows if r["Producten (ready)"] > 0)
        niet_in_boom = sum(1 for r in boom_rows if r["Hoofdcategorie"] == "⚠️ NIET IN BOOM")

        bc1, bc2, bc3, bc4 = st.columns(4)
        bc1.metric("Totaal sub-subcats", totaal_subcats)
        bc2.metric("🟢 Online", online)
        bc3.metric("🔴 Nog niet online", offline)
        bc4.metric("⚠️ Niet in boom", niet_in_boom)

        st.divider()

        # Filter — default op "Met producten" zodat lege categorieën niet afleiden
        boom_filter = st.radio(
            "Filter",
            ["📦 Met producten", "Alles", "🔴 Nog niet online", "🟢 Al online", "⚠️ Niet in boom"],
            horizontal=True,
            key="boom_filter",
        )
        if boom_filter == "🔴 Nog niet online":
            df_boom = df_boom[df_boom["Online?"].str.startswith("🔴")]
        elif boom_filter == "🟢 Al online":
            df_boom = df_boom[df_boom["Online?"].str.startswith("🟢")]
        elif boom_filter == "📦 Met producten":
            df_boom = df_boom[df_boom["Producten (ready)"] > 0]
        elif boom_filter == "⚠️ Niet in boom":
            df_boom = df_boom[df_boom["Hoofdcategorie"] == "⚠️ NIET IN BOOM"]
        # "Alles" toont alles

        def _kleur_online(val):
            if str(val).startswith("🟢"):
                return "background-color: #d4edda"
            if str(val).startswith("🔴"):
                return "background-color: #f8d7da"
            if str(val).startswith("⚠️"):
                return "background-color: #fff3cd"
            return ""

        styled = df_boom.style.map(_kleur_online, subset=["Online?"])
        st.dataframe(styled, width="stretch", height=600, hide_index=True)

        st.caption(f"Toont {len(df_boom)} sub-subcategorieën")

        csv = df_boom.to_csv(index=False).encode("utf-8-sig")
        st.download_button("📥 Download website structuur (CSV)", csv,
                           file_name="website_structuur.csv", mime="text/csv")

        # ── Online-status aanpassen ───────────────────────────────────────────
        st.divider()
        st.subheader("✏️ Online-status aanpassen")
        st.caption(
            "Zet sub-subcategorieën op online (als je ze in Shopify hebt aangemaakt) "
            "of voeg nieuwe toe. Wijzigingen worden opgeslagen in de Website indeling xlsx."
        )

        # Lijst van alle niet-online sub-subcats
        offline_subcats = sorted([r["Sub-subcategorie"] for r in boom_rows if r["Online?"].startswith("🔴")])

        if offline_subcats:
            op_online_zetten = st.multiselect(
                "Zet op online (selecteer sub-subcategorieën die je in Shopify hebt aangemaakt)",
                offline_subcats,
                key="boom_online_select",
            )

            if op_online_zetten and st.button("✅ Op online zetten", type="primary", key="boom_set_online"):
                # Update de Website indeling xlsx: zet de cellen groen
                from openpyxl import load_workbook
                from openpyxl.styles import PatternFill

                xlsx_path = "Master Files/Website indeling (1).xlsx"
                wb = load_workbook(xlsx_path)
                ws = wb["Blad1"]

                groen_fill = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")
                updated = 0

                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and str(cell.value).strip() in op_online_zetten:
                            cell.fill = groen_fill
                            updated += 1

                wb.save(xlsx_path)

                # Clear cache zodat het direct zichtbaar is
                load_active_subsubcategories.clear()

                st.success(f"✅ {updated} sub-subcategorieën op online gezet in de Website indeling. "
                           f"Klik **Rerun** om de tabel te verversen.")
                st.rerun()
        else:
            st.success("Alle sub-subcategorieën staan al online!")

        # Nieuwe sub-subcat toevoegen aan de boom
        st.divider()
        with st.expander("➕ Nieuwe sub-subcategorie toevoegen aan website-boom"):
            nc1, nc2, nc3 = st.columns(3)
            with nc1:
                new_hc = st.selectbox("Hoofdcategorie", sorted(tree.keys()), key="boom_new_hc")
            with nc2:
                new_sc_opts = sorted(tree.get(new_hc, {}).keys()) if new_hc else []
                new_sc = st.selectbox("Subcategorie", new_sc_opts, key="boom_new_sc") if new_sc_opts else None
            with nc3:
                new_ssc = st.text_input("Nieuwe sub-subcategorie", key="boom_new_ssc")

            meteen_online = st.checkbox("Direct op online zetten", value=True, key="boom_new_online")

            if st.button("💾 Toevoegen", key="boom_add_new"):
                if new_hc and new_sc and new_ssc and new_ssc.strip():
                    from openpyxl import load_workbook
                    from openpyxl.styles import PatternFill

                    xlsx_path = "Master Files/Website indeling (1).xlsx"
                    wb = load_workbook(xlsx_path)
                    ws = wb["Blad1"]

                    # Zoek de kolom van de subcategorie
                    target_col = None
                    for cell in ws[2]:  # rij 2 = subcats
                        if cell.value and str(cell.value).strip() == new_sc:
                            target_col = cell.column
                            break

                    if target_col:
                        # Zoek de eerste lege rij in die kolom (na rij 2)
                        target_row = None
                        for r in range(3, ws.max_row + 2):
                            if not ws.cell(row=r, column=target_col).value:
                                target_row = r
                                break

                        if target_row:
                            ws.cell(row=target_row, column=target_col, value=new_ssc.strip())
                            if meteen_online:
                                groen_fill = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")
                                ws.cell(row=target_row, column=target_col).fill = groen_fill

                            wb.save(xlsx_path)
                            load_active_subsubcategories.clear()
                            load_website_tree.clear()
                            st.success(f"✅ **{new_hc} > {new_sc} > {new_ssc.strip()}** toegevoegd aan de website-boom"
                                       f"{' en op online gezet' if meteen_online else ''}.")
                            st.rerun()
                        else:
                            st.error("Kon geen lege rij vinden in de kolom.")
                    else:
                        st.error(f"Subcategorie '{new_sc}' niet gevonden in de xlsx.")
                else:
                    st.warning("Vul alle drie de velden in.")
    else:
        st.warning("Website indeling niet gevonden. Zorg dat `Master Files/Website indeling (1).xlsx` bestaat.")

# ── PAGINA: Categoriestatus ────────────────────────────────────────────────────

elif stap == STAPPEN[10]:
    st.title("🏷️ Categoriestatus")
    st.caption("Welke categorieën staan online in Shopify, welke zijn nog niet gekoppeld?")

    import requests as _req

    sb_new = get_supabase_new()

    with st.spinner("Categorieën laden..."):
        cats = sb_new.table("category_collection_map").select("*").execute().data or []

    with st.spinner("Shopify collecties ophalen..."):
        _token = os.getenv("SHOPIFY_ACCESS_TOKEN")
        _store = os.getenv("SHOPIFY_STORE")
        _h = {"X-Shopify-Access-Token": _token}
        _resp = _req.get(
            f"https://{_store}/admin/api/2026-04/smart_collections.json?limit=250&fields=id,title,handle,published_at",
            headers=_h, timeout=15
        )
        _shopify_cols = {c["handle"]: c for c in _resp.json().get("smart_collections", [])}

    # Verrijken met Shopify status
    for c in cats:
        handle = c.get("shopify_collection_handle")
        if not handle:
            c["_status"] = "Niet gekoppeld"
        elif handle in _shopify_cols:
            c["_status"] = "🟢 Online" if _shopify_cols[handle].get("published_at") else "🔴 Offline"
        else:
            c["_status"] = "❓ Niet gevonden"

    # Filters
    col1, col2 = st.columns([2, 2])
    with col1:
        status_filter = st.selectbox(
            "Filter op status",
            ["Alle", "🟢 Online", "Niet gekoppeld", "🔴 Offline", "❓ Niet gevonden"]
        )
    with col2:
        hoofdcat_filter = st.selectbox(
            "Filter op hoofdcategorie",
            ["Alle"] + sorted(set(c.get("hoofdcategorie", "") for c in cats if c.get("hoofdcategorie")))
        )

    gefilterd = cats
    if status_filter != "Alle":
        gefilterd = [c for c in gefilterd if c["_status"] == status_filter]
    if hoofdcat_filter != "Alle":
        gefilterd = [c for c in gefilterd if c.get("hoofdcategorie") == hoofdcat_filter]

    # Statistieken
    online_n   = sum(1 for c in cats if c["_status"] == "🟢 Online")
    ontkoppeld = sum(1 for c in cats if c["_status"] == "Niet gekoppeld")
    niet_gevonden = sum(1 for c in cats if c["_status"] == "❓ Niet gevonden")

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Totaal categorieën", len(cats))
    m2.metric("🟢 Online", online_n)
    m3.metric("Niet gekoppeld", ontkoppeld)
    m4.metric("❓ Niet gevonden in Shopify", niet_gevonden)

    st.divider()

    # Tabel
    if gefilterd:
        df_cats = pd.DataFrame([{
            "Hoofdcategorie":   c.get("hoofdcategorie", ""),
            "Subcategorie":     c.get("subcategorie", ""),
            "Sub-subcategorie": c.get("sub_subcategorie", "") or "",
            "Shopify handle":   c.get("shopify_collection_handle", "") or "",
            "Cat tag":          c.get("cat_tag", "") or "",
            "Status":           c["_status"],
        } for c in gefilterd])

        st.dataframe(
            df_cats,
            use_container_width=True,
            height=600,
            column_config={
                "Status": st.column_config.TextColumn("Status", width="medium"),
                "Cat tag": st.column_config.TextColumn("Cat tag", width="medium"),
            }
        )
        st.caption(f"{len(gefilterd)} categorieën weergegeven")
    else:
        st.info("Geen categorieën gevonden voor deze filters.")


# ── PAGINA: Product zoeken ─────────────────────────────────────────────────────

elif stap == STAPPEN[11]:
    st.title("🔍 Product zoeken")
    st.caption("Typ een SKU of EAN om alles over dat product te zien — raw data, curated data en Shopify status.")

    import requests as _req

    sb_new = get_supabase_new()

    zoekterm = st.text_input(
        "SKU of EAN",
        placeholder="Bijv. B0126008-008 of 5400959163491",
        help="Voer een SKU of EAN in om het product op te zoeken"
    ).strip()

    if zoekterm:
        with st.spinner("Zoeken..."):
            # Zoek in products_curated op SKU
            res_curated = sb_new.table("products_curated").select("*").eq("sku", zoekterm).execute().data
            # Zoek in products_raw op SKU of EAN
            res_raw_sku = sb_new.table("products_raw").select("*").eq("sku", zoekterm).execute().data
            res_raw_ean = sb_new.table("products_raw").select("*").eq("ean_piece", zoekterm).execute().data
            res_raw = res_raw_sku or res_raw_ean
            # Zoek in shopify_sync
            res_shopify = sb_new.table("shopify_sync").select("*").eq("sku", zoekterm).execute().data

        if not res_curated and not res_raw and not res_shopify:
            st.error(f"Geen product gevonden voor: **{zoekterm}**")
        else:
            st.success(f"Product gevonden: **{zoekterm}**")

            tab1, tab2, tab3 = st.tabs(["📋 Curated (ons werk)", "📦 Raw (leverancier)", "🛒 Shopify live"])

            with tab1:
                if res_curated:
                    p = res_curated[0]
                    c1, c2 = st.columns(2)
                    with c1:
                        st.subheader("Productinfo")
                        st.write(f"**Titel:** {p.get('product_title_nl') or '—'}")
                        st.write(f"**Handle:** {p.get('handle') or '—'}")
                        st.write(f"**Supplier:** {p.get('supplier') or '—'}")
                        st.write(f"**Fase:** {p.get('fase') or '—'}")
                        st.write(f"**Pipeline status:** {p.get('pipeline_status') or '—'}")
                    with c2:
                        st.subheader("Categorie")
                        st.write(f"**Hoofdcategorie:** {p.get('hoofdcategorie') or '—'}")
                        st.write(f"**Subcategorie:** {p.get('subcategorie') or '—'}")
                        st.write(f"**Sub-subcategorie:** {p.get('sub_subcategorie') or '—'}")
                        st.write(f"**Tags:** {p.get('tags') or '—'}")
                    st.divider()
                    c3, c4 = st.columns(2)
                    with c3:
                        st.subheader("Prijs")
                        st.write(f"**Verkoopprijs:** €{p.get('verkoopprijs') or '—'}")
                        st.write(f"**Inkoopprijs:** €{p.get('inkoopprijs') or '—'}")
                    with c4:
                        st.subheader("SEO")
                        st.write(f"**Meta description:** {p.get('meta_description') or '—'}")
                    if p.get("review_reden"):
                        st.warning(f"**Review reden:** {p['review_reden']}")
                else:
                    st.info("Niet gevonden in products_curated.")

            with tab2:
                if res_raw:
                    p = res_raw[0]
                    c1, c2 = st.columns(2)
                    with c1:
                        st.write(f"**Product naam (raw):** {p.get('product_name_raw') or '—'}")
                        st.write(f"**Supplier:** {p.get('supplier') or '—'}")
                        st.write(f"**Import batch:** {p.get('import_batch') or '—'}")
                        st.write(f"**EAN piece:** {p.get('ean_piece') or '—'}")
                        st.write(f"**EAN shopify:** {p.get('ean_shopify') or '—'}")
                        st.write(f"**Designer:** {p.get('designer') or '—'}")
                        st.write(f"**Kleur (EN):** {p.get('kleur_en') or '—'}")
                        st.write(f"**Materiaal (raw):** {p.get('materiaal_raw') or '—'}")
                    with c2:
                        st.write(f"**Leverancier categorie:** {p.get('leverancier_category') or '—'}")
                        st.write(f"**RRP stuk:** €{p.get('rrp_stuk_eur') or '—'}")
                        st.write(f"**Inkoopprijs stuk:** €{p.get('inkoopprijs_stuk_eur') or '—'}")
                        st.write(f"**Giftbox:** {p.get('giftbox') or '—'} ({p.get('giftbox_qty') or 0}x)")
                        st.write(f"**H×L×B:** {p.get('hoogte_cm') or '—'} × {p.get('lengte_cm') or '—'} × {p.get('breedte_cm') or '—'} cm")
                else:
                    st.info("Niet gevonden in products_raw.")

            with tab3:
                if res_shopify:
                    p = res_shopify[0]
                    status_kleur = {"active": "🟢", "draft": "🟡", "archived": "🔴"}.get(p.get("shopify_status", ""), "❓")
                    st.write(f"**Status:** {status_kleur} {p.get('shopify_status') or '—'}")
                    st.write(f"**Titel in Shopify:** {p.get('title') or '—'}")
                    st.write(f"**Handle:** {p.get('handle') or '—'}")
                    st.write(f"**Shopify product ID:** {p.get('shopify_product_id') or '—'}")
                    st.write(f"**Prijs live:** €{p.get('price') or '—'}")
                    st.write(f"**Gepubliceerd op:** {p.get('published_at') or '—'}")
                    st.write(f"**Tags:** {p.get('tags') or '—'}")
                    st.write(f"**Laatste sync:** {p.get('synced_at') or '—'}")

                    # Vergelijk titel curated vs shopify
                    if res_curated:
                        curated_titel = res_curated[0].get("product_title_nl") or ""
                        shopify_titel = p.get("title") or ""
                        if curated_titel and shopify_titel and curated_titel != shopify_titel:
                            st.warning(f"**Titelafwijking:**\n- Curated: {curated_titel}\n- Shopify: {shopify_titel}")
                else:
                    st.info("Niet gevonden in shopify_sync. Draai eerst de Shopify sync voor deze fase.")
